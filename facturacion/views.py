from django.shortcuts import render, redirect, get_object_or_404
from django import forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.forms import modelform_factory, inlineformset_factory
from django.http import Http404, HttpResponse
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.conf import settings
from django.utils import timezone
from django.utils.text import slugify
from django.views.decorators.http import require_POST
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.utils.http import url_has_allowed_host_and_scheme
from django.db import IntegrityError, transaction
from django.db.models import Count, F, Prefetch, Q, Sum
from django.db.models.functions import TruncMonth
from django.urls import reverse
from weasyprint import HTML
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation, ROUND_FLOOR
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference
import os
import logging
import json
import calendar
from pathlib import Path

from core.models import ConfiguracionAvanzadaEmpresa, ConfiguracionPowerBIEmpresa, Empresa, RegistroAuditoria, Usuario
from contabilidad.services import (
    asegurar_cuenta_contable_cliente,
    asegurar_cuentas_financieras_base_honduras,
    cargar_catalogo_base_honduras,
    registrar_asiento_compra_aplicada,
    registrar_asiento_factura_emitida,
    registrar_asiento_nota_credito,
    registrar_asiento_pago_cliente,
    registrar_asiento_pago_proveedor,
    registrar_reversion_documento,
)
from .models import CAI, BitacoraProductoEliminado, BodegaInventario, CategoriaProductoFarmaceutico, CierreCaja, ComprobanteEgresoCompra, CompraInventario, ConfiguracionFacturacionEmpresa, CorreccionNumeroFactura, EMPRESAS_PRECIO_FINAL_CON_IMPUESTO, EntradaInventarioDocumento, ExistenciaLoteBodega, Factura, HistorialCostoRealProducto, InventarioProducto, LineaCompraInventario, LineaEntradaInventario, LineaFactura, LineaNotaCredito, LoteInventario, MovimientoInventario, MovimientoLoteBodega, NotaCredito, PagoCompra, PerfilFarmaceuticoProducto, Producto, ProductoPromocionPuntoVenta, PromocionPuntoVenta, Proveedor, ReciboPago, RegistroCompraFiscal, TipoImpuesto, Cliente, PagoFactura
from .forms import AjusteInventarioForm, CAIForm, CategoriaProductoFarmaceuticoForm, ClienteForm, ConfiguracionFacturacionEmpresaForm, ConfiguracionPowerBIForm, CorreccionNumeroFacturaForm, DATE_INPUT_FORMATS_LATAM, EliminarProductoForm, EntradaInventarioForm, ImportarLibroComprasForm, PagoCompraForm, ProductoForm, PromocionPuntoVentaForm, ProveedorForm, ReciboPagoForm, RegistroCompraFiscalForm, TipoImpuestoForm, configurar_campo_fecha
from .importadores import importar_libro_compras_desde_excel
from contabilidad.models import AsientoContable, ClasificacionCompraFiscal, CuentaFinanciera

logger = logging.getLogger(__name__)

EMPRESAS_COSTO_REAL_INVENTARIO = {"hospital_mia", "medical_spa"}


def _empresa_muestra_costo_real_inventario(empresa):
    return bool(empresa and empresa.slug in EMPRESAS_COSTO_REAL_INVENTARIO)


def _registrar_cambio_costo_real(producto, costo_anterior, usuario):
    anterior = Decimal(costo_anterior or 0).quantize(Decimal("0.0001"))
    nuevo = Decimal(producto.costo_real_inventario or 0).quantize(Decimal("0.0001"))
    if anterior == nuevo:
        return
    HistorialCostoRealProducto.objects.create(
        empresa=producto.empresa,
        producto=producto,
        costo_anterior=anterior,
        costo_nuevo=nuevo,
        usuario=usuario if getattr(usuario, "is_authenticated", False) else None,
    )

POS_CLIENTE_OBLIGATORIO_SLUGS = {"hospital_mia", "medical_spa"}
CAJA_EXCLUYE_FACTURAS_ANULADAS_SLUGS = {"hospital_mia", "medical_spa"}


def _precios_incluyen_impuesto(empresa):
    debe_incluir = empresa.slug in EMPRESAS_PRECIO_FINAL_CON_IMPUESTO
    configuracion, _ = ConfiguracionFacturacionEmpresa.objects.get_or_create(
        empresa=empresa,
        defaults={"precios_incluyen_impuesto": debe_incluir},
    )
    if configuracion.precios_incluyen_impuesto != debe_incluir:
        configuracion.precios_incluyen_impuesto = debe_incluir
        configuracion.save(update_fields=["precios_incluyen_impuesto", "fecha_actualizacion"])
    return debe_incluir


def _empresa_usa_cierre_caja(empresa):
    return bool(
        empresa.tiene_modulo_activo("punto_venta")
        or ConfiguracionAvanzadaEmpresa.para_empresa(empresa).usa_cierre_caja
    )


def _puede_ver_historial_cierres(request, empresa):
    if empresa.slug != "hospital_mia":
        return True
    usuario = request.user
    return bool(usuario.is_superuser or usuario.es_administrador_empresa)


def _fecha_caja_desde_parametro(valor):
    if isinstance(valor, date):
        return valor
    try:
        return datetime.strptime(str(valor), "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return timezone.localdate()


def _monto_neto_caja_pago(pago, excluir_anuladas=False):
    monto = Decimal(pago.monto or 0)
    if excluir_anuladas and pago.factura.estado == "anulada":
        return Decimal("0.00")
    return monto


def _total_neto_caja(pagos):
    return sum((Decimal(getattr(pago, "monto_neto_caja", pago.monto) or 0) for pago in pagos), Decimal("0.00"))


def _preparar_pagos_caja(pagos, empresa=None):
    pagos_preparados = list(pagos)
    excluir_anuladas = bool(empresa and empresa.slug in CAJA_EXCLUYE_FACTURAS_ANULADAS_SLUGS)
    for pago in pagos_preparados:
        pago.es_anulacion_caja = pago.factura.estado == "anulada"
        pago.anulacion_no_suma_caja = excluir_anuladas and pago.es_anulacion_caja
        pago.monto_neto_caja = _monto_neto_caja_pago(pago, excluir_anuladas=excluir_anuladas)
    return pagos_preparados


def _desglose_metodo_caja(pagos, metodo_pago):
    pagos_metodo = [pago for pago in pagos if pago.metodo == metodo_pago]
    ventas = sum((Decimal(pago.monto or 0) for pago in pagos_metodo if not pago.anulacion_no_suma_caja), Decimal("0.00"))
    anulaciones = sum((Decimal(pago.monto or 0) for pago in pagos_metodo if pago.anulacion_no_suma_caja), Decimal("0.00"))
    return {
        "ventas": ventas,
        "anulaciones": anulaciones,
        "neto": ventas,
        "cantidad": len(pagos_metodo),
        "cantidad_anulaciones": sum(1 for pago in pagos_metodo if pago.es_anulacion_caja),
    }


def _cuentas_financieras_activas_para_pago(empresa):
    asegurar_cuentas_financieras_base_honduras(empresa)
    cuentas = CuentaFinanciera.objects.filter(
        empresa=empresa,
        activa=True,
    ).select_related("cuenta_contable").order_by("nombre")
    if not cuentas.exists():
        cargar_catalogo_base_honduras(empresa)
        cuentas = CuentaFinanciera.objects.filter(
            empresa=empresa,
            activa=True,
        ).select_related("cuenta_contable").order_by("nombre")
    return cuentas


def _form_data_pago_por_defecto(cuentas_financieras):
    cuenta = cuentas_financieras.first()
    return {
        "cuenta_financiera": str(cuenta.id) if cuenta else "",
    }


def _cuenta_financiera_por_defecto(cuentas_financieras, metodo=None):
    if metodo == "efectivo":
        return cuentas_financieras.filter(tipo="caja").first() or cuentas_financieras.first()
    if metodo == "tarjeta":
        return cuentas_financieras.filter(tipo="banco").first() or cuentas_financieras.first()
    return cuentas_financieras.filter(tipo="banco").first() or cuentas_financieras.first()


def _empresa_usa_perfil_farmaceutico(empresa):
    config = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)
    return bool(
        empresa.tiene_modulo_activo("clinica_medica")
        and config.usa_inventario_farmaceutico
    )


# =====================================================
# DASHBOARD
# =====================================================


def _parsear_fecha_latam(valor):
    if isinstance(valor, date):
        return valor
    valor = (valor or "").strip()
    if not valor:
        return None
    for formato in DATE_INPUT_FORMATS_LATAM:
        try:
            return datetime.strptime(valor, formato).date()
        except ValueError:
            continue
    return None

@login_required
def facturacion_dashboard(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)

    return render(request, "facturacion/dashboard_modulo_cxp_premium.html", {
        "empresa": empresa,
    })


@login_required
def configuracion_facturacion(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    precios_incluyen_impuesto = _precios_incluyen_impuesto(empresa)
    configuracion, _ = ConfiguracionFacturacionEmpresa.objects.get_or_create(
        empresa=empresa,
        defaults={"precios_incluyen_impuesto": precios_incluyen_impuesto},
    )
    permite_plantilla_notas_extensas = _empresa_permite_plantilla_notas_extensas(empresa)
    permite_plantilla_independiente = _empresa_permite_plantilla_independiente(empresa)
    form = ConfiguracionFacturacionEmpresaForm(
        request.POST or None,
        instance=configuracion,
        permite_plantilla_notas_extensas=permite_plantilla_notas_extensas,
        permite_plantilla_independiente=permite_plantilla_independiente,
    )

    if request.method == "POST" and form.is_valid():
        configuracion = form.save(commit=False)
        configuracion.empresa = empresa
        configuracion.precios_incluyen_impuesto = precios_incluyen_impuesto
        configuracion.save()
        messages.success(request, "Configuracion de facturacion actualizada correctamente.")
        return redirect("configuracion_facturacion", empresa_slug=empresa.slug)

    return render(request, "facturacion/configuracion_facturacion.html", {
        "empresa": empresa,
        "form": form,
        "configuracion": configuracion,
        "permite_plantilla_notas_extensas": permite_plantilla_notas_extensas,
        "permite_plantilla_independiente": permite_plantilla_independiente,
        "precios_incluyen_impuesto": precios_incluyen_impuesto,
    })


@login_required
def facturas_dashboard(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    config_avanzada = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)
    q = request.GET.get("q", "").strip()
    facturas = Factura.objects.filter(empresa=empresa).select_related('cliente').order_by('-fecha_creacion')

    if q:
        facturas = facturas.filter(
            Q(cliente__nombre__icontains=q) |
            Q(numero_factura__icontains=q)
        )

    resumen = {
        "total_documentos": facturas.count(),
        "borradores": facturas.filter(estado='borrador').count(),
        "por_cobrar": facturas.filter(estado_pago__in=['pendiente', 'parcial']).count(),
        "monto_total": sum((factura.total for factura in facturas), Decimal('0.00')),
    }

    return render(request, "facturacion/facturas_premium.html", {
        "empresa": empresa,
        "facturas": facturas,
        "resumen": resumen,
        "q": q,
        "clientes_sugeridos": Cliente.objects.filter(empresa=empresa).order_by('nombre').values_list('nombre', flat=True).distinct(),
        "permite_gestion_fiscal_historica": config_avanzada.permite_gestion_fiscal_historica,
    })


@login_required
def punto_venta(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    precios_incluyen_impuesto = _precios_incluyen_impuesto(empresa)
    solicitud_pos_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if not empresa.tiene_modulo_activo("punto_venta"):
        if solicitud_pos_ajax:
            return JsonResponse(
                {"ok": False, "error": "El modulo Punto de Venta no esta activo para esta empresa."},
                status=403,
            )
        messages.error(request, "El modulo Punto de Venta no esta activo para esta empresa.")
        return redirect("facturacion_dashboard", empresa_slug=empresa.slug)

    productos_qs = (
        Producto.objects.filter(empresa=empresa, activo=True, eliminado=False)
        .select_related("impuesto_predeterminado", "inventario")
        .order_by("nombre")
    )
    impuestos_qs = TipoImpuesto.objects.filter(activo=True).order_by("nombre")
    impuesto_default = impuestos_qs.first()
    cuentas_financieras = _cuentas_financieras_activas_para_pago(empresa)
    cliente_obligatorio = _pos_cliente_obligatorio(empresa)
    bodegas_pos = []
    if empresa.slug in POS_CLIENTE_OBLIGATORIO_SLUGS:
        bodegas_pos = list(
            BodegaInventario.objects.filter(
                empresa=empresa,
                activa=True,
            ).order_by("tipo", "nombre")
        )
        if not bodegas_pos:
            bodegas_pos = list(_asegurar_bodegas_farmaceuticas(empresa).values())

    if request.method == "POST":
        try:
            payload = json.loads(request.POST.get("payload") or "{}")
        except json.JSONDecodeError:
            payload = {}

        items = payload.get("items") or []
        metodo = (payload.get("metodo") or "efectivo").strip()
        cuenta_id = payload.get("cuenta_financiera")
        referencia = (payload.get("referencia") or "").strip()
        monto_recibido_raw = payload.get("monto_recibido")
        cliente_id = payload.get("cliente_id")
        fecha_venta = _parsear_fecha_latam(payload.get("fecha")) or timezone.localdate()

        try:
            if metodo not in dict(PagoFactura.METODOS):
                raise ValidationError("Selecciona un metodo de pago valido.")
            if not impuesto_default and any(not p.impuesto_predeterminado_id for p in productos_qs):
                raise ValidationError("Configura al menos un impuesto activo antes de usar el punto de venta.")
            if not items:
                raise ValidationError("Agrega al menos un producto al carrito.")
            if cliente_obligatorio and not cliente_id:
                raise ValidationError("Selecciona o crea el cliente antes de cobrar esta venta.")

            cuenta_financiera = cuentas_financieras.filter(id=cuenta_id).first() or _cuenta_financiera_por_defecto(cuentas_financieras, metodo)
            if not cuenta_financiera:
                raise ValidationError("Configura una caja o cuenta bancaria activa para cobrar en punto de venta.")

            producto_ids = []
            for item in items:
                try:
                    producto_ids.append(int(item.get("producto_id")))
                except (TypeError, ValueError):
                    raise ValidationError("El carrito contiene un producto invalido.")

            productos = {producto.id: producto for producto in productos_qs.filter(id__in=producto_ids)}
            lineas_preparadas = []
            for item in items:
                producto_id = int(item.get("producto_id"))
                producto = productos.get(producto_id)
                if not producto:
                    raise ValidationError("Uno de los productos del carrito ya no esta disponible.")
                cantidad = Decimal(str(item.get("cantidad") or "0")).quantize(Decimal("0.01"))
                precio_unitario = Decimal(str(item.get("precio_unitario") or producto.precio)).quantize(Decimal("0.01"))
                if cantidad <= 0:
                    raise ValidationError(f"La cantidad de {producto.nombre} debe ser mayor que cero.")
                if precio_unitario < 0:
                    raise ValidationError(f"El precio de {producto.nombre} no puede ser negativo.")
                impuesto = producto.impuesto_predeterminado or impuesto_default
                if not impuesto:
                    raise ValidationError(f"El producto {producto.nombre} no tiene impuesto asignado.")
                lineas_preparadas.append({
                    "producto": producto,
                    "cantidad": cantidad,
                    "precio_unitario": precio_unitario,
                    "impuesto": impuesto,
                })
            regalos_promocion, promocion_pos = _calcular_regalos_promocion_pos(empresa, lineas_preparadas, fecha_venta)

            with transaction.atomic():
                cliente = None
                if cliente_id:
                    cliente = Cliente.objects.filter(empresa=empresa, id=cliente_id, activo=True).first()
                if cliente_obligatorio and not cliente:
                    raise ValidationError("El cliente seleccionado ya no esta disponible. Vuelve a buscarlo o crealo nuevamente.")
                if not cliente:
                    cliente = Cliente.objects.filter(
                        empresa=empresa,
                        nombre__iexact="Consumidor Final",
                    ).first()
                    if not cliente:
                        cliente = Cliente.objects.create(
                            empresa=empresa,
                            nombre="Consumidor Final",
                            rtn="",
                            ciudad=empresa.ciudad or "",
                        )
                asegurar_cuenta_contable_cliente(cliente)

                factura = Factura.objects.create(
                    empresa=empresa,
                    cliente=cliente,
                    vendedor=request.user,
                    estado="emitida",
                    fecha_emision=fecha_venta,
                    fecha_vencimiento=fecha_venta,
                    moneda="HNL",
                    tipo_cambio=Decimal("1.0000"),
                )

                lineas = [
                    LineaFactura(
                        factura=factura,
                        producto=linea["producto"],
                        cantidad=linea["cantidad"],
                        precio_unitario=linea["precio_unitario"],
                        impuesto=linea["impuesto"],
                        precio_incluye_impuesto=precios_incluyen_impuesto,
                    )
                    for linea in lineas_preparadas
                ]
                if promocion_pos and regalos_promocion:
                    for producto_id, cantidad_gratis in regalos_promocion.items():
                        if cantidad_gratis <= 0:
                            continue
                        producto = productos.get(producto_id)
                        if not producto:
                            continue
                        impuesto = producto.impuesto_predeterminado or impuesto_default
                        lineas.append(LineaFactura(
                            factura=factura,
                            producto=producto,
                            cantidad=cantidad_gratis,
                            precio_unitario=producto.precio,
                            impuesto=impuesto,
                            precio_incluye_impuesto=precios_incluyen_impuesto,
                            descuento_porcentaje=Decimal("100.00"),
                            comentario=f"{promocion_pos.nombre}: unidad gratis aplicada automaticamente.",
                        ))
                for linea in lineas:
                    linea.save()

                _validar_stock_disponible_para_lineas(lineas)
                _actualizar_totales_factura(factura)
                _registrar_salida_factura(factura)
                registrar_asiento_factura_emitida(factura)

                monto_recibido = None
                cambio = Decimal("0.00")
                if metodo == "efectivo" and monto_recibido_raw not in (None, ""):
                    monto_recibido = Decimal(str(monto_recibido_raw)).quantize(Decimal("0.01"))
                    if monto_recibido < factura.total_documento_ajustado:
                        raise ValidationError("El efectivo recibido es menor que el total de la venta.")
                    cambio = monto_recibido - factura.total_documento_ajustado

                referencia_pago = referencia or f"POS {factura.numero_factura or factura.id}"
                if monto_recibido is not None:
                    referencia_pago = (
                        f"{referencia_pago} | Recibido L. {monto_recibido:.2f} | Cambio L. {cambio:.2f}"
                    )[:100]

                pago = PagoFactura.objects.create(
                    factura=factura,
                    fecha=fecha_venta,
                    monto=factura.total_documento_ajustado,
                    metodo=metodo,
                    referencia=referencia_pago,
                    cuenta_financiera=cuenta_financiera,
                    cajero=request.user,
                )
                registrar_asiento_pago_cliente(pago)

            recibo_numero = pago.recibo.numero_recibo if hasattr(pago, "recibo") else "sin numero"
            if solicitud_pos_ajax:
                return JsonResponse({
                    "ok": True,
                    "mensaje": "Venta cobrada y facturada correctamente.",
                    "factura_id": factura.id,
                    "numero_factura": factura.numero_factura,
                    "numero_recibo": recibo_numero,
                    "total": f"{factura.total:.2f}",
                    "cambio": f"{cambio:.2f}",
                    "factura_url": reverse("ver_factura", args=[empresa.slug, factura.id]),
                    "ticket_url": reverse("imprimir_factura_pos", args=[empresa.slug, factura.id]),
                })
            messages.success(
                request,
                f"Venta POS registrada. Factura {factura.numero_factura}; recibo {recibo_numero}; total L. {factura.total:.2f}.",
            )
            return redirect("ver_factura", empresa_slug=empresa.slug, factura_id=factura.id)
        except (InvalidOperation, ValueError) as exc:
            if solicitud_pos_ajax:
                return JsonResponse(
                    {"ok": False, "error": f"No se pudo registrar la venta: {exc}"},
                    status=400,
                )
            messages.error(request, f"No se pudo registrar la venta: {exc}")
        except ValidationError as exc:
            if solicitud_pos_ajax:
                return JsonResponse({"ok": False, "error": "; ".join(exc.messages)}, status=400)
            messages.error(request, "; ".join(exc.messages))
        except Exception as exc:
            logger.exception("Error registrando venta POS para empresa %s", empresa.id)
            if solicitud_pos_ajax:
                return JsonResponse(
                    {"ok": False, "error": "No se pudo completar la venta. Revisa la configuracion fiscal y vuelve a intentar."},
                    status=500,
                )
            messages.error(request, f"No se pudo registrar la venta POS: {exc}")

    productos_payload = [_pos_producto_payload(producto, impuesto_default) for producto in productos_qs]
    clientes_payload = [
        _pos_cliente_payload(cliente)
        for cliente in Cliente.objects.filter(empresa=empresa, activo=True)
        .exclude(nombre__iexact="Consumidor Final")
        .order_by("nombre")[:250]
    ]
    promocion_pos_payload = _promocion_pos_payload(empresa, timezone.localdate())

    return render(request, "facturacion/punto_venta.html", {
        "empresa": empresa,
        "productos_payload": productos_payload,
        "clientes_payload": clientes_payload,
        "bodegas_payload": [_pos_bodega_payload(bodega) for bodega in bodegas_pos],
        "promocion_pos_payload": promocion_pos_payload,
        "productos_count": len(productos_payload),
        "cuentas_financieras": cuentas_financieras,
        "fecha_hoy": timezone.localdate().strftime("%Y-%m-%d"),
        "metodos_pago": PagoFactura.METODOS,
        "cliente_obligatorio": cliente_obligatorio,
        "precios_incluyen_impuesto": precios_incluyen_impuesto,
    })


@login_required
def configurar_promocion_pos(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    if not empresa.tiene_modulo_activo("punto_venta"):
        messages.error(request, "El modulo de punto de venta no esta activo para esta empresa.")
        return redirect("facturacion_dashboard", empresa_slug=empresa.slug)

    promocion = PromocionPuntoVenta.objects.filter(empresa=empresa).order_by("-activa", "-fecha_actualizacion").first()
    if not promocion:
        promocion = PromocionPuntoVenta.objects.create(
            empresa=empresa,
            nombre="Promocion 3 + 1 gratis",
            cantidad_pagada=3,
            cantidad_gratis=1,
        )
    productos = Producto.objects.filter(
        empresa=empresa,
        activo=True,
        eliminado=False,
        tipo_item="producto",
    ).order_by("nombre")
    productos_seleccionados = set(
        promocion.productos_configurados.filter(activo=True).values_list("producto_id", flat=True)
    )

    if request.method == "POST":
        form = PromocionPuntoVentaForm(request.POST, instance=promocion)
        productos_post = {
            int(producto_id)
            for producto_id in request.POST.getlist("productos")
            if str(producto_id).isdigit()
        }
        productos_validos = set(productos.filter(id__in=productos_post).values_list("id", flat=True))
        if form.is_valid():
            promocion = form.save()
            ProductoPromocionPuntoVenta.objects.filter(promocion=promocion).exclude(
                producto_id__in=productos_validos
            ).update(activo=False)
            for producto_id in productos_validos:
                ProductoPromocionPuntoVenta.objects.update_or_create(
                    promocion=promocion,
                    producto_id=producto_id,
                    defaults={"activo": True},
                )
            messages.success(request, "Promocion de punto de venta actualizada correctamente.")
            return redirect("configurar_promocion_pos", empresa_slug=empresa.slug)
    else:
        form = PromocionPuntoVentaForm(instance=promocion)

    return render(request, "facturacion/promociones_pos.html", {
        "empresa": empresa,
        "form": form,
        "promocion": promocion,
        "productos": productos,
        "productos_seleccionados": productos_seleccionados,
    })


@login_required
def pos_buscar_clientes(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    if empresa.slug not in POS_CLIENTE_OBLIGATORIO_SLUGS:
        return JsonResponse({"ok": False, "error": "La busqueda rapida POS no esta activa para esta empresa."}, status=403)

    termino = request.GET.get("q", "")
    clientes = [_pos_cliente_payload(cliente) for cliente in _buscar_clientes_pos(empresa, termino)]
    return JsonResponse({"ok": True, "clientes": clientes})


@login_required
@require_POST
def pos_crear_cliente_rapido(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    if empresa.slug not in POS_CLIENTE_OBLIGATORIO_SLUGS:
        return JsonResponse({"ok": False, "error": "La creacion rapida POS no esta activa para esta empresa."}, status=403)

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        payload = {}

    nombre = (payload.get("nombre") or "").strip()
    rtn = (payload.get("rtn") or "").strip()
    telefono = (payload.get("telefono") or "").strip()
    correo = (payload.get("correo") or "").strip().lower()
    telefono_whatsapp = (payload.get("telefono_whatsapp") or telefono).strip()
    ciudad = (payload.get("ciudad") or empresa.ciudad or "").strip()
    direccion = (payload.get("direccion") or "").strip()
    fecha_nacimiento = _parsear_fecha_latam(payload.get("fecha_nacimiento"))

    errores = []
    if not nombre:
        errores.append("Ingresa el nombre del cliente.")
    if not rtn:
        errores.append("Ingresa el numero de identidad o RTN.")
    if not telefono:
        errores.append("Ingresa el numero de telefono.")
    if errores:
        return JsonResponse({"ok": False, "error": " ".join(errores)}, status=400)

    try:
        with transaction.atomic():
            # Serializa las altas rápidas de una misma empresa para que dos cajas
            # no intenten reservar simultáneamente el mismo código contable.
            Empresa.objects.select_for_update().get(pk=empresa.pk)
            cliente = Cliente.objects.create(
                empresa=empresa,
                nombre=nombre,
                rtn=rtn,
                telefono=telefono,
                telefono_whatsapp=telefono_whatsapp,
                correo=correo,
                ciudad=ciudad,
                direccion=direccion,
                fecha_nacimiento=fecha_nacimiento,
                acepta_promociones=bool(payload.get("acepta_promociones", True)),
                canal_preferido=payload.get("canal_preferido") or "whatsapp",
                activo=True,
            )
            asegurar_cuenta_contable_cliente(cliente)
    except ValidationError as exc:
        return JsonResponse({"ok": False, "error": "; ".join(exc.messages)}, status=400)
    except IntegrityError:
        logger.exception("Conflicto de integridad creando cliente rápido POS para %s", empresa.slug)
        return JsonResponse(
            {
                "ok": False,
                "error": (
                    "No se pudo guardar porque otro cliente o cuenta contable utiliza "
                    "los mismos datos. Verifica el DNI/RTN e intenta nuevamente."
                ),
            },
            status=409,
        )
    except Exception:
        logger.exception("Error creando cliente rápido POS para %s", empresa.slug)
        return JsonResponse(
            {
                "ok": False,
                "error": (
                    "El servidor no pudo completar el registro del cliente. "
                    "Intenta nuevamente; si continúa, revisa el registro de errores."
                ),
            },
            status=500,
        )

    return JsonResponse({"ok": True, "cliente": _pos_cliente_payload(cliente)})


@login_required
@require_POST
def pos_crear_producto_rapido(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    if empresa.slug not in POS_CLIENTE_OBLIGATORIO_SLUGS:
        return JsonResponse({"ok": False, "error": "La creacion rapida POS no esta activa para esta empresa."}, status=403)

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        payload = {}

    nombre = (payload.get("nombre") or "").strip()
    codigo = (payload.get("codigo") or "").strip()
    try:
        precio = Decimal(str(payload.get("precio") or "0")).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return JsonResponse({"ok": False, "error": "El precio del producto no es valido."}, status=400)
    if not nombre:
        return JsonResponse({"ok": False, "error": "Ingresa el nombre del producto."}, status=400)
    if precio < 0:
        return JsonResponse({"ok": False, "error": "El precio no puede ser negativo."}, status=400)

    impuesto_default = TipoImpuesto.objects.filter(activo=True).order_by("nombre").first()
    impuesto = impuesto_default
    impuesto_id = payload.get("impuesto_id")
    if impuesto_id:
        impuesto = TipoImpuesto.objects.filter(id=impuesto_id, activo=True).first() or impuesto_default
    if not impuesto:
        return JsonResponse({"ok": False, "error": "Configura al menos un impuesto activo antes de crear productos desde POS."}, status=400)

    distribucion = payload.get("bodegas") or {}
    lote_inicial = (payload.get("lote_inicial") or "").strip()
    fecha_vencimiento = _parsear_fecha_latam(payload.get("vencimiento_lote"))

    try:
        with transaction.atomic():
            producto = Producto.objects.create(
                empresa=empresa,
                nombre=nombre,
                codigo=codigo or None,
                tipo_item=payload.get("tipo_item") or "producto",
                unidad_medida=payload.get("unidad_medida") or "unidad",
                precio=precio,
                impuesto_predeterminado=impuesto,
                controla_inventario=bool(payload.get("controla_inventario", True)),
                descripcion=(payload.get("descripcion") or "").strip(),
                activo=True,
            )
            if producto.controla_inventario:
                _obtener_inventario_producto(producto)
                for bodega_id, cantidad_raw in distribucion.items():
                    try:
                        cantidad = Decimal(str(cantidad_raw or "0")).quantize(Decimal("0.01"))
                    except (InvalidOperation, ValueError):
                        raise ValidationError("Una de las cantidades por bodega no es valida.")
                    if cantidad <= 0:
                        continue
                    bodega = BodegaInventario.objects.filter(empresa=empresa, activa=True, id=bodega_id).first()
                    if not bodega:
                        raise ValidationError("Una de las bodegas seleccionadas ya no esta disponible.")
                    _ajustar_existencia_producto_bodega(
                        producto,
                        bodega,
                        cantidad,
                        numero_lote=lote_inicial,
                        fecha_vencimiento=fecha_vencimiento,
                        observacion="Alta rapida de producto desde punto de venta.",
                    )
                _sincronizar_existencia_general_producto(producto)
    except ValidationError as exc:
        return JsonResponse({"ok": False, "error": "; ".join(exc.messages)}, status=400)

    return JsonResponse({"ok": True, "producto": _pos_producto_payload(producto, impuesto_default)})


@login_required
def prefijo_factura_manual(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    config_avanzada = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)
    fecha_referencia = _parsear_fecha_latam(request.GET.get("fecha")) or timezone.localdate()

    if not config_avanzada.permite_gestion_fiscal_historica:
        return JsonResponse({"ok": False, "error": "Empresa sin numeracion manual historica."}, status=403)

    factura_tmp = Factura(empresa=empresa, fecha_emision=fecha_referencia)
    cai = factura_tmp._obtener_queryset_cai_factura(
        fecha_referencia=factura_tmp._obtener_fecha_referencia_cai()
    ).first()

    if not cai:
        return JsonResponse({
            "ok": False,
            "prefijo": "",
            "detalle": "No existe un CAI disponible para la fecha seleccionada.",
        })

    prefijo = (
        f"{cai.establecimiento}-"
        f"{cai.punto_emision}-"
        f"{cai.tipo_documento}-"
        f"{str(cai.rango_inicial).zfill(8)[:5]}"
    )
    return JsonResponse({
        "ok": True,
        "prefijo": prefijo,
        "detalle": (
            f"CAI {cai.numero_cai} | Rango {str(cai.rango_inicial).zfill(8)} a "
            f"{str(cai.rango_final).zfill(8)} | Vigencia {cai.fecha_activacion.strftime('%d/%m/%Y')} "
            f"al {cai.fecha_limite.strftime('%d/%m/%Y')}"
        ),
    })


def _redirect_seguro(request, fallback, *args, **kwargs):
    next_url = request.POST.get("next") or request.GET.get("next")
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure()
    ):
        return redirect(next_url)
    return redirect(fallback, *args, **kwargs)


def _recalcular_correlativo_cai_factura(cai_id):
    cai = CAI.objects.select_for_update().get(pk=cai_id)
    maximo_utilizado = cai.rango_inicial - 1

    numeros = (
        Factura.objects.filter(cai_id=cai_id, estado="emitida")
        .exclude(numero_factura__isnull=True)
        .exclude(numero_factura__exact="")
        .values_list("numero_factura", flat=True)
    )

    for numero in numeros:
        coincidencia = Factura.NUMERO_FACTURA_REGEX.match((numero or "").strip())
        if not coincidencia:
            continue
        correlativo = int(coincidencia.group("correlativo"))
        if correlativo > maximo_utilizado:
            maximo_utilizado = correlativo

    if cai.correlativo_actual != maximo_utilizado:
        cai.correlativo_actual = maximo_utilizado
        cai.save(update_fields=["correlativo_actual"])


def _puede_configurar_power_bi(usuario):
    return bool(
        usuario.is_superuser
        or usuario.es_administrador_empresa
    )


def _construir_bi_interno_facturacion(facturas):
    hoy = timezone.localdate()
    inicio_ventana = date(hoy.year, hoy.month, 1) - timedelta(days=150)

    facturas_emitidas = facturas.exclude(estado="anulada").select_related("cliente")
    lineas = LineaFactura.objects.filter(factura__in=facturas_emitidas).select_related("producto")

    ventas_mensuales_qs = (
        facturas_emitidas.filter(fecha_emision__gte=inicio_ventana)
        .annotate(periodo=TruncMonth("fecha_emision"))
        .values("periodo")
        .annotate(
            total=Sum("total"),
            documentos=Count("id"),
        )
        .order_by("periodo")
    )

    ventas_mensuales = []
    max_total_mes = Decimal("0.00")
    for item in ventas_mensuales_qs:
        total_mes = item["total"] or Decimal("0.00")
        periodo = item["periodo"]
        saldo_mes = sum(
            (
                factura.saldo_pendiente
                for factura in facturas_emitidas.filter(
                    fecha_emision__year=periodo.year,
                    fecha_emision__month=periodo.month,
                )
            ),
            Decimal("0.00"),
        )
        max_total_mes = max(max_total_mes, total_mes)
        ventas_mensuales.append(
            {
                "periodo": periodo,
                "total": total_mes,
                "saldo": saldo_mes,
                "documentos": item["documentos"],
            }
        )
    for item in ventas_mensuales:
        item["ancho"] = float((item["total"] / max_total_mes) * 100) if max_total_mes else 0

    top_clientes_qs = (
        facturas_emitidas.values("cliente__nombre")
        .annotate(total=Sum("total"), documentos=Count("id"))
        .order_by("-total")[:6]
    )
    top_clientes = []
    max_cliente = Decimal("0.00")
    for item in top_clientes_qs:
        total_cliente = item["total"] or Decimal("0.00")
        max_cliente = max(max_cliente, total_cliente)
        top_clientes.append(
            {
                "nombre": item["cliente__nombre"] or "Cliente sin nombre",
                "total": total_cliente,
                "documentos": item["documentos"],
            }
        )
    for item in top_clientes:
        item["ancho"] = float((item["total"] / max_cliente) * 100) if max_cliente else 0

    top_productos_qs = (
        lineas.values("producto__nombre")
        .annotate(
            subtotal=Sum("subtotal"),
            impuesto_total=Sum("impuesto_monto"),
            cantidad=Sum("cantidad"),
        )
        .order_by("-subtotal")[:6]
    )
    top_productos = []
    max_producto = Decimal("0.00")
    for item in top_productos_qs:
        total_producto = (item["subtotal"] or Decimal("0.00")) + (item["impuesto_total"] or Decimal("0.00"))
        max_producto = max(max_producto, total_producto)
        top_productos.append(
            {
                "nombre": item["producto__nombre"] or "Producto sin nombre",
                "total": total_producto,
                "cantidad": item["cantidad"] or Decimal("0.00"),
            }
        )
    for item in top_productos:
        item["ancho"] = float((item["total"] / max_producto) * 100) if max_producto else 0

    estados = [
        ("Pagado", facturas_emitidas.filter(estado_pago="pagado")),
        ("Parcial", facturas_emitidas.filter(estado_pago="parcial")),
        ("Pendiente", facturas_emitidas.filter(estado_pago="pendiente")),
    ]
    estado_cobro = []
    max_estado = Decimal("0.00")
    for etiqueta, queryset in estados:
        total_estado = queryset.aggregate(total=Sum("total"))["total"] or Decimal("0.00")
        max_estado = max(max_estado, total_estado)
        estado_cobro.append(
            {
                "etiqueta": etiqueta,
                "cantidad": queryset.count(),
                "total": total_estado,
            }
        )
    for item in estado_cobro:
        item["ancho"] = float((item["total"] / max_estado) * 100) if max_estado else 0

    return {
        "ventas_mensuales": ventas_mensuales,
        "top_clientes": top_clientes,
        "top_productos": top_productos,
        "estado_cobro": estado_cobro,
    }


def _calcular_variacion_porcentual(actual, anterior):
    actual = actual or Decimal("0.00")
    anterior = anterior or Decimal("0.00")
    if anterior == 0:
        if actual == 0:
            return Decimal("0.00")
        return Decimal("100.00")
    return ((actual - anterior) / anterior) * Decimal("100.00")


def _nombre_archivo_pdf(prefijo, numero_documento, nombre_cliente):
    numero = numero_documento or "sin_numero"
    cliente = slugify(nombre_cliente or "cliente")
    return f"{prefijo}_{numero} ({cliente}).pdf"


def _empresa_permite_plantilla_notas_extensas(empresa):
    identificadores = [
        slugify(empresa.nombre or ""),
        (empresa.slug or "").strip().lower(),
    ]
    return any(
        "digital-planning" in identificador or "diggital-planning" in identificador
        for identificador in identificadores
    )


def _empresa_permite_plantilla_independiente(empresa):
    config = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)
    return bool(config.permite_plantilla_factura_independiente)


def _resolver_plantilla_factura(configuracion, empresa, plantilla_forzada=None):
    plantilla_activa = plantilla_forzada or configuracion.plantilla_factura_pdf
    if plantilla_activa == "termica_80mm" and empresa.slug in {"hospital_mia", "medical_spa", "demo_1"}:
        return "facturacion/factura_pdf_termica_80mm.html"
    if plantilla_activa == "independiente" and _empresa_permite_plantilla_independiente(empresa):
        return "facturacion/factura_pdf_independiente.html"
    if plantilla_activa == "alternativa":
        return "facturacion/factura_pdf_alternativa.html"
    if plantilla_activa == "notas_extensas" and _empresa_permite_plantilla_notas_extensas(empresa):
        return "facturacion/factura_pdf_notas_extensas.html"
    return "facturacion/factura_pdf.html"


def _obtener_logo_url(empresa):
    if not empresa.logo:
        return None
    try:
        return Path(settings.MEDIA_ROOT, empresa.logo.name).resolve().as_uri()
    except Exception:
        logo_path = os.path.join(settings.MEDIA_ROOT, empresa.logo.name)
        logo_path = logo_path.replace("\\", "/")
        return "file:///" + logo_path


def _resumen_detallado(subtotal_neto, resumen_fiscal):
    subtotal_bruto = subtotal_neto + Decimal(str(resumen_fiscal.get("descuento_total", 0) or 0))
    return {
        "subtotal_bruto": subtotal_bruto,
        "descuento_total": Decimal(str(resumen_fiscal.get("descuento_total", 0) or 0)),
        "subtotal_antes_impuesto": subtotal_neto,
        "base_15": Decimal(str(resumen_fiscal.get("base_15", 0) or 0)),
        "base_18": Decimal(str(resumen_fiscal.get("base_18", 0) or 0)),
        "base_exento": Decimal(str(resumen_fiscal.get("base_exento", 0) or 0)),
        "base_exonerado": Decimal(str(resumen_fiscal.get("base_exonerado", 0) or 0)),
        "isv_15": Decimal(str(resumen_fiscal.get("isv_15", 0) or 0)),
        "isv_18": Decimal(str(resumen_fiscal.get("isv_18", 0) or 0)),
    }


def _filtrar_facturas_reporte(empresa, params):
    facturas = Factura.objects.filter(empresa=empresa).select_related("cliente").prefetch_related("lineas", "lineas__impuesto")

    cliente_id = (params.get("cliente") or "").strip()
    estado_pago = (params.get("estado_pago") or "").strip()
    fecha_desde = (params.get("fecha_desde") or "").strip()
    fecha_hasta = (params.get("fecha_hasta") or "").strip()
    impuesto = (params.get("impuesto") or "").strip()

    if cliente_id:
        facturas = facturas.filter(cliente_id=cliente_id)

    if estado_pago:
        facturas = facturas.filter(estado_pago=estado_pago)

    if fecha_desde:
        facturas = facturas.filter(fecha_emision__gte=fecha_desde)

    if fecha_hasta:
        facturas = facturas.filter(fecha_emision__lte=fecha_hasta)

    if impuesto:
        facturas = facturas.filter(lineas__impuesto__porcentaje=impuesto).distinct()

    return facturas


def _construir_reporte_cxc(empresa, params):
    facturas = (
        Factura.objects.filter(empresa=empresa, estado="emitida")
        .select_related("cliente")
        .order_by("fecha_vencimiento", "fecha_emision", "id")
    )

    hoy = date.today()
    cliente_id = (params.get("cliente") or "").strip()
    q = (params.get("q") or "").strip()
    q_normalizado = q.lower()

    facturas_con_saldo = []
    for factura in facturas:
        saldo = factura.saldo_pendiente
        if saldo <= 0:
            continue
        if cliente_id and str(factura.cliente_id) != cliente_id:
            continue
        if q_normalizado and q_normalizado not in factura.cliente.nombre.lower():
            continue
        facturas_con_saldo.append(factura)

    data = {}
    for factura in facturas_con_saldo:
        saldo = factura.saldo_pendiente
        cliente = factura.cliente
        dias = (hoy - (factura.fecha_vencimiento or factura.fecha_emision)).days

        if cliente.id not in data:
            data[cliente.id] = {
                "cliente": cliente,
                "0_30": Decimal("0.00"),
                "31_60": Decimal("0.00"),
                "61_90": Decimal("0.00"),
                "90_mas": Decimal("0.00"),
                "total": Decimal("0.00"),
            }

        if dias <= 30:
            data[cliente.id]["0_30"] += saldo
        elif dias <= 60:
            data[cliente.id]["31_60"] += saldo
        elif dias <= 90:
            data[cliente.id]["61_90"] += saldo
        else:
            data[cliente.id]["90_mas"] += saldo

        data[cliente.id]["total"] += saldo

    cliente_seleccionado = None
    facturas_pendientes_cliente = None
    cliente_resumen = None

    if cliente_id:
        try:
            cliente_id_int = int(cliente_id)
            cliente_seleccionado = Cliente.objects.filter(id=cliente_id_int, empresa=empresa).first()
            if cliente_seleccionado:
                facturas_pendientes_cliente = [f for f in facturas_con_saldo if f.cliente_id == cliente_id_int]
                if facturas_pendientes_cliente:
                    cliente_resumen = {
                        "facturas": len(facturas_pendientes_cliente),
                        "saldo_total": sum((f.saldo_pendiente for f in facturas_pendientes_cliente), Decimal("0.00")),
                        "vencidas": sum(1 for f in facturas_pendientes_cliente if (f.fecha_vencimiento or f.fecha_emision) < hoy),
                        "proxima_fecha": min((f.fecha_vencimiento or f.fecha_emision for f in facturas_pendientes_cliente)),
                    }
        except (TypeError, ValueError):
            cliente_seleccionado = None
            facturas_pendientes_cliente = None
            cliente_resumen = None

    data_lista = list(data.values())
    total_cartera = sum((item["total"] for item in data_lista), Decimal("0.00"))
    resumen = {
        "clientes_con_saldo": len(data_lista),
        "total_cartera": total_cartera,
        "facturas_pendientes": len(facturas_con_saldo),
    }

    return {
        "data": data_lista,
        "resumen": resumen,
        "cliente_seleccionado": cliente_seleccionado,
        "cliente_resumen": cliente_resumen,
        "facturas_pendientes_cliente": facturas_pendientes_cliente,
        "facturas_con_saldo": facturas_con_saldo,
        "q": q,
        "cliente_id": cliente_id,
        "clientes_sugeridos": list(
            Cliente.objects.filter(empresa=empresa)
            .order_by("nombre")
            .values_list("nombre", flat=True)
            .distinct()
        ),
        "clientes_preview": list(
            Cliente.objects.filter(empresa=empresa)
            .order_by("nombre")
            .values("id", "nombre", "rtn")
        ),
    }


def _actualizar_totales_factura(factura):
    factura.calcular_totales()
    factura.save(update_fields=[
        'subtotal',
        'impuesto',
        'total',
        'total_lempiras',
    ])


def _emitir_factura_desde_borrador(factura):
    if factura.estado == 'anulada':
        raise ValidationError("No se puede validar una factura anulada.")
    if factura.estado == 'emitida':
        raise ValidationError("Esta factura ya fue validada previamente.")

    lineas = list(factura.lineas.select_related('producto').all())
    if not lineas:
        raise ValidationError("La factura debe tener al menos una linea para poder validarse.")

    _validar_stock_disponible_para_lineas(lineas)

    factura.estado = 'emitida'
    factura.save(update_fields=['estado'])
    _actualizar_totales_factura(factura)

    _registrar_salida_factura(factura)
    registrar_asiento_factura_emitida(factura)


def _contexto_documento_factura(empresa, factura):
    configuracion, _ = ConfiguracionFacturacionEmpresa.objects.get_or_create(empresa=empresa)
    resumen = factura.resumen_fiscal()
    resumen_detallado = _resumen_detallado(factura.subtotal, resumen)
    lineas = list(factura.lineas.select_related("producto", "impuesto").all())
    bloques_comentario = sum(
        max(1, len((linea.comentario or "").strip()) // 42 + 1)
        for linea in lineas
        if (linea.comentario or "").strip()
    )
    alto_ticket_mm = min(2000, max(132, 122 + (len(lineas) * 8) + (bloques_comentario * 4)))
    return {
        "empresa": empresa,
        "factura": factura,
        "resumen": resumen,
        "resumen_detallado": resumen_detallado,
        "lineas_factura": lineas,
        "alto_ticket_mm": alto_ticket_mm,
        "logo_url": _obtener_logo_url(empresa),
        "configuracion_facturacion": configuracion,
    }


def _pos_cliente_obligatorio(empresa):
    return empresa.slug in POS_CLIENTE_OBLIGATORIO_SLUGS


def _pos_cliente_payload(cliente):
    return {
        "id": cliente.id,
        "nombre": cliente.nombre,
        "rtn": cliente.rtn or "",
        "telefono": cliente.telefono or "",
        "telefono_whatsapp": cliente.telefono_whatsapp or "",
        "correo": cliente.correo or "",
        "ciudad": cliente.ciudad or "",
    }


def _solo_digitos(valor):
    return "".join(caracter for caracter in str(valor or "") if caracter.isdigit())


def _buscar_clientes_pos(empresa, termino, limite=12):
    termino = (termino or "").strip()
    if len(termino) < 2:
        return []

    clientes_base = (
        Cliente.objects.filter(empresa=empresa, activo=True)
        .exclude(nombre__iexact="Consumidor Final")
        .order_by("nombre", "id")
    )
    clientes = []
    ids = set()
    termino_digitos = _solo_digitos(termino)

    if termino_digitos:
        for cliente in clientes_base.iterator(chunk_size=500):
            valores = [
                cliente.rtn,
                cliente.telefono,
                cliente.telefono_whatsapp,
            ]
            if any(_solo_digitos(valor) == termino_digitos for valor in valores):
                clientes.append(cliente)
                ids.add(cliente.id)
                if len(clientes) >= limite:
                    break

    filtro = (
        Q(nombre__icontains=termino)
        | Q(rtn__icontains=termino)
        | Q(telefono__icontains=termino)
        | Q(telefono_whatsapp__icontains=termino)
        | Q(correo__icontains=termino)
    )
    for cliente in clientes_base.filter(filtro)[:limite]:
        if cliente.id not in ids:
            clientes.append(cliente)
            ids.add(cliente.id)
            if len(clientes) >= limite:
                break

    if termino_digitos and len(clientes) < limite:
        for cliente in clientes_base.iterator(chunk_size=500):
            if cliente.id in ids:
                continue
            valores = [
                cliente.rtn,
                cliente.telefono,
                cliente.telefono_whatsapp,
            ]
            if any(termino_digitos in _solo_digitos(valor) for valor in valores):
                clientes.append(cliente)
                ids.add(cliente.id)
                if len(clientes) >= limite:
                    break

    if len(clientes) < limite:
        try:
            from clinica.models import Paciente
        except Exception:
            Paciente = None
        if Paciente:
            pacientes_base = Paciente.objects.filter(empresa=empresa, activo=True).order_by("nombre", "id")
            pacientes = []
            paciente_ids = set()
            if termino_digitos:
                for paciente in pacientes_base.iterator(chunk_size=500):
                    valores = [paciente.identidad, paciente.telefono, paciente.whatsapp, paciente.celular_2]
                    if any(_solo_digitos(valor) == termino_digitos for valor in valores):
                        pacientes.append(paciente)
                        paciente_ids.add(paciente.id)
                        if len(pacientes) >= limite:
                            break
            filtro_paciente = (
                Q(nombre__icontains=termino)
                | Q(primer_nombre__icontains=termino)
                | Q(segundo_nombre__icontains=termino)
                | Q(primer_apellido__icontains=termino)
                | Q(segundo_apellido__icontains=termino)
                | Q(identidad__icontains=termino)
                | Q(telefono__icontains=termino)
                | Q(whatsapp__icontains=termino)
                | Q(celular_2__icontains=termino)
                | Q(correo__icontains=termino)
            )
            for paciente in pacientes_base.filter(filtro_paciente)[:limite]:
                if paciente.id not in paciente_ids:
                    pacientes.append(paciente)
                    paciente_ids.add(paciente.id)
                    if len(pacientes) >= limite:
                        break
            if termino_digitos and len(pacientes) < limite:
                for paciente in pacientes_base.iterator(chunk_size=500):
                    if paciente.id in paciente_ids:
                        continue
                    valores = [paciente.identidad, paciente.telefono, paciente.whatsapp, paciente.celular_2]
                    if any(termino_digitos in _solo_digitos(valor) for valor in valores):
                        pacientes.append(paciente)
                        paciente_ids.add(paciente.id)
                        if len(pacientes) >= limite:
                            break

            for paciente in pacientes:
                if len(clientes) >= limite:
                    break
                cliente = _cliente_desde_paciente_pos(paciente)
                if cliente.id not in ids:
                    clientes.append(cliente)
                    ids.add(cliente.id)

    return clientes[:limite]


def _cliente_desde_paciente_pos(paciente):
    cliente = paciente.cliente if paciente.cliente_id and paciente.cliente.empresa_id == paciente.empresa_id else None
    identidad = (paciente.identidad or "").strip()
    if not cliente and identidad:
        cliente = Cliente.objects.filter(empresa=paciente.empresa, rtn__iexact=identidad).first()
    if not cliente and paciente.nombre:
        cliente = Cliente.objects.filter(empresa=paciente.empresa, nombre__iexact=paciente.nombre.strip()).first()

    datos = {
        "nombre": paciente.nombre or "Paciente sin nombre",
        "rtn": identidad,
        "telefono": paciente.telefono or paciente.whatsapp or paciente.celular_2 or "",
        "telefono_whatsapp": paciente.whatsapp or paciente.telefono or "",
        "correo": paciente.correo or "",
        "fecha_nacimiento": paciente.fecha_nacimiento,
        "acepta_promociones": paciente.acepta_promociones,
        "direccion": paciente.direccion or "",
        "ciudad": paciente.municipio or paciente.departamento or "",
        "canal_preferido": "correo" if paciente.recibir_email and paciente.correo else "whatsapp",
        "activo": paciente.activo,
    }

    if cliente:
        cambios = []
        for campo, valor in datos.items():
            if campo == "rtn" and valor:
                existe = Cliente.objects.filter(
                    empresa=paciente.empresa,
                    rtn__iexact=valor,
                ).exclude(pk=cliente.pk).exists()
                if existe:
                    continue
            if campo == "nombre" and valor:
                existe = Cliente.objects.filter(
                    empresa=paciente.empresa,
                    nombre__iexact=valor,
                ).exclude(pk=cliente.pk).exists()
                if existe:
                    continue
            if getattr(cliente, campo) != valor:
                setattr(cliente, campo, valor)
                cambios.append(campo)
        if cambios:
            cliente.save(update_fields=cambios)
    else:
        cliente = Cliente.objects.create(empresa=paciente.empresa, **datos)

    if paciente.cliente_id != cliente.id:
        paciente.cliente = cliente
        paciente.save(update_fields=["cliente"])
    asegurar_cuenta_contable_cliente(cliente)
    return cliente


def _pos_producto_payload(producto, impuesto_default=None):
    impuesto = producto.impuesto_predeterminado or impuesto_default
    return {
        "id": producto.id,
        "nombre": producto.nombre,
        "codigo": producto.codigo or "",
        "foto_url": producto.foto.url if producto.foto else "",
        "precio": float(producto.precio or Decimal("0.00")),
        "impuesto": float(impuesto.porcentaje if impuesto else Decimal("0.00")),
        "stock": float(producto.stock_actual),
        "unidad": producto.get_unidad_medida_display(),
        "controla_inventario": producto.controla_inventario,
    }


def _pos_bodega_payload(bodega):
    return {
        "id": bodega.id,
        "nombre": bodega.nombre,
        "tipo": bodega.tipo,
    }


def _promocion_pos_vigente(empresa, fecha):
    promocion = (
        PromocionPuntoVenta.objects.filter(empresa=empresa)
        .prefetch_related("productos_configurados")
        .order_by("-activa", "-fecha_actualizacion")
        .first()
    )
    if not promocion or not promocion.vigente_en(fecha):
        return None
    if not promocion.productos_configurados.filter(
        activo=True,
        producto__activo=True,
        producto__eliminado=False,
        producto__tipo_item="producto",
    ).exists():
        return None
    return promocion


def _promocion_pos_payload(empresa, fecha):
    promocion = _promocion_pos_vigente(empresa, fecha)
    if not promocion:
        return {
            "activa": False,
            "nombre": "",
            "cantidad_pagada": 3,
            "cantidad_gratis": 1,
            "productos": [],
        }
    productos_ids = list(
        promocion.productos_configurados.filter(
            activo=True,
            producto__activo=True,
            producto__eliminado=False,
            producto__tipo_item="producto",
        ).values_list("producto_id", flat=True)
    )
    return {
        "activa": True,
        "nombre": promocion.nombre,
        "cantidad_pagada": promocion.cantidad_pagada,
        "cantidad_gratis": promocion.cantidad_gratis,
        "productos": productos_ids,
    }


def _calcular_regalos_promocion_pos(empresa, lineas_preparadas, fecha_venta):
    promocion = _promocion_pos_vigente(empresa, fecha_venta)
    if not promocion:
        return {}, None

    productos_permitidos = set(
        promocion.productos_configurados.filter(
            activo=True,
            producto__activo=True,
            producto__eliminado=False,
            producto__tipo_item="producto",
        ).values_list("producto_id", flat=True)
    )
    elegibles = [
        linea
        for linea in lineas_preparadas
        if linea["producto"].id in productos_permitidos and linea["cantidad"] > 0
    ]
    if not elegibles:
        return {}, promocion

    cantidad_pagada = Decimal(str(promocion.cantidad_pagada))
    cantidad_gratis = Decimal(str(promocion.cantidad_gratis))
    total_pagado = sum((linea["cantidad"] for linea in elegibles), Decimal("0.00"))
    grupos = (total_pagado / cantidad_pagada).to_integral_value(rounding=ROUND_FLOOR)
    total_gratis = grupos * cantidad_gratis
    if total_gratis <= 0:
        return {}, promocion

    regalos = {}
    pendiente = total_gratis
    for linea in sorted(elegibles, key=lambda item: (item["precio_unitario"], item["producto"].nombre)):
        if pendiente <= 0:
            break
        cantidad_base = linea["cantidad"].to_integral_value(rounding=ROUND_FLOOR)
        if cantidad_base <= 0:
            continue
        cantidad_producto = min(cantidad_base, pendiente)
        regalos[linea["producto"].id] = regalos.get(linea["producto"].id, Decimal("0.00")) + cantidad_producto
        pendiente -= cantidad_producto

    return regalos, promocion


def _render_factura_pdf_response(empresa, factura, plantilla, inline=False, prefijo_archivo="Factura"):
    contexto = _contexto_documento_factura(empresa, factura)

    html_string = render_to_string(
        plantilla,
        contexto,
    )

    pdf_file = HTML(
        string=html_string,
        base_url=str(settings.BASE_DIR)
    ).write_pdf()

    nombre_archivo = _nombre_archivo_pdf(
        prefijo_archivo,
        factura.numero_factura or f"factura_{factura.id}",
        factura.cliente.nombre,
    )

    response = HttpResponse(pdf_file, content_type="application/pdf")
    disposition = "inline" if inline else "attachment"
    response["Content-Disposition"] = f'{disposition}; filename="{nombre_archivo}"'
    return response


def _obtener_inventario_producto(producto):
    return InventarioProducto.objects.get_or_create(
        empresa=producto.empresa,
        producto=producto,
        defaults={'stock_minimo': Decimal('0.00')}
    )[0]


def _asegurar_bodegas_farmaceuticas(empresa):
    if empresa.slug == "luque_aestetic":
        bodegas_base = [
            ("Bodega Principal", "principal"),
            ("Bodega Clinica", "provisional"),
        ]
    elif empresa.slug in {"hospital_mia", "medical_spa"}:
        bodegas_base = [
            ("Bodega General", "principal"),
            ("Bodega Hospital", "provisional"),
            ("Vitrina", "vitrina"),
        ]
    else:
        bodegas_base = [
            ("Bodega principal", "principal"),
            ("Bodega provisional", "provisional"),
            ("Vitrina", "vitrina"),
        ]
    bodegas = {}
    for nombre, tipo in bodegas_base:
        bodega = (
            BodegaInventario.objects.filter(empresa=empresa, nombre__iexact=nombre).first()
            or BodegaInventario.objects.filter(empresa=empresa, tipo=tipo).order_by("id").first()
        )
        if not bodega:
            bodega = BodegaInventario.objects.create(
                empresa=empresa,
                nombre=nombre,
                tipo=tipo,
                activa=True,
            )
        cambios = []
        if bodega.nombre != nombre and not BodegaInventario.objects.filter(
            empresa=empresa,
            nombre__iexact=nombre,
        ).exclude(id=bodega.id).exists():
            bodega.nombre = nombre
            cambios.append("nombre")
        if bodega.tipo != tipo:
            bodega.tipo = tipo
            cambios.append("tipo")
        if not bodega.activa:
            bodega.activa = True
            cambios.append("activa")
        if cambios:
            bodega.save(update_fields=cambios)
        bodegas[tipo] = bodega
    if empresa.slug == "luque_aestetic":
        for bodega_extra in BodegaInventario.objects.filter(
            empresa=empresa,
            activa=True,
        ).exclude(id__in=[bodega.id for bodega in bodegas.values()]):
            if not ExistenciaLoteBodega.objects.filter(
                empresa=empresa,
                bodega=bodega_extra,
                cantidad__gt=0,
            ).exists():
                bodega_extra.activa = False
                bodega_extra.save(update_fields=["activa"])
    return bodegas


def _obtener_bodega_venta(empresa):
    bodegas = _asegurar_bodegas_farmaceuticas(empresa)
    return bodegas.get("vitrina")


def _obtener_lote_generico(producto):
    lote, _ = LoteInventario.objects.get_or_create(
        empresa=producto.empresa,
        producto=producto,
        numero_lote=f"SIN-LOTE-{producto.id}",
        defaults={"activo": True},
    )
    return lote


def _registrar_movimiento_lote_bodega(*, empresa, bodega, lote, tipo, cantidad, referencia="", observacion="", factura=None):
    existencia, _ = ExistenciaLoteBodega.objects.get_or_create(
        empresa=empresa,
        bodega=bodega,
        lote=lote,
        defaults={"cantidad": Decimal("0.00")},
    )
    anterior = existencia.cantidad
    if tipo in ["entrada", "traslado_entrada", "reversion", "ajuste"]:
        resultante = anterior + cantidad
    else:
        resultante = anterior - cantidad
    if resultante < 0:
        raise ValidationError(f"Stock insuficiente en {bodega.nombre} para el lote {lote.numero_lote}.")

    existencia.cantidad = resultante
    existencia.save(update_fields=["cantidad", "fecha_actualizacion"])
    return MovimientoLoteBodega.objects.create(
        empresa=empresa,
        bodega=bodega,
        lote=lote,
        tipo=tipo,
        cantidad=cantidad,
        existencia_anterior=anterior,
        existencia_resultante=resultante,
        referencia=referencia,
        observacion=observacion,
        factura=factura,
    )


def _entrada_farmaceutica_generica(empresa, producto, cantidad, referencia="", observacion=""):
    if not _empresa_usa_perfil_farmaceutico(empresa) or not producto.controla_inventario:
        return
    bodega = _asegurar_bodegas_farmaceuticas(empresa)["principal"]
    lote = _obtener_lote_generico(producto)
    _registrar_movimiento_lote_bodega(
        empresa=empresa,
        bodega=bodega,
        lote=lote,
        tipo="entrada",
        cantidad=cantidad,
        referencia=referencia,
        observacion=observacion or "Entrada automatica a bodega principal.",
    )


def _registrar_entrada_inicial_producto(producto, *, bodega, cantidad, numero_lote="", fecha_vencimiento=None):
    if not producto.controla_inventario or not bodega or cantidad <= 0:
        return
    inventario = _obtener_inventario_producto(producto)
    inventario.existencias += cantidad
    inventario.save(update_fields=["existencias", "fecha_actualizacion"])

    lote, _ = LoteInventario.objects.get_or_create(
        empresa=producto.empresa,
        producto=producto,
        numero_lote=numero_lote or f"INICIAL-{producto.id}",
        defaults={
            "fecha_vencimiento": fecha_vencimiento,
            "activo": True,
        },
    )
    if fecha_vencimiento and lote.fecha_vencimiento != fecha_vencimiento:
        lote.fecha_vencimiento = fecha_vencimiento
        lote.save(update_fields=["fecha_vencimiento"])

    _registrar_movimiento_lote_bodega(
        empresa=producto.empresa,
        bodega=bodega,
        lote=lote,
        tipo="entrada",
        cantidad=cantidad,
        referencia=f"Producto {producto.id}",
        observacion="Entrada inicial registrada al crear producto.",
    )


def _sincronizar_existencia_general_producto(producto):
    inventario = _obtener_inventario_producto(producto)
    total = ExistenciaLoteBodega.objects.filter(
        empresa=producto.empresa,
        lote__producto=producto,
    ).aggregate(total=Sum("cantidad"))["total"] or Decimal("0.00")
    inventario.existencias = total
    inventario.save(update_fields=["existencias", "fecha_actualizacion"])
    return total


def _registrar_ajuste_existencia_bodega(*, empresa, bodega, lote, cantidad_resultante, referencia, observacion):
    existencia, _ = ExistenciaLoteBodega.objects.get_or_create(
        empresa=empresa,
        bodega=bodega,
        lote=lote,
        defaults={"cantidad": Decimal("0.00")},
    )
    anterior = existencia.cantidad
    if cantidad_resultante < 0:
        raise ValidationError(f"La existencia en {bodega.nombre} no puede quedar negativa.")
    delta = cantidad_resultante - anterior
    if delta == 0:
        return None
    existencia.cantidad = cantidad_resultante
    existencia.save(update_fields=["cantidad", "fecha_actualizacion"])
    return MovimientoLoteBodega.objects.create(
        empresa=empresa,
        bodega=bodega,
        lote=lote,
        tipo="ajuste",
        cantidad=delta,
        existencia_anterior=anterior,
        existencia_resultante=cantidad_resultante,
        referencia=referencia,
        observacion=observacion,
    )


def _ajustar_existencia_producto_bodega(producto, bodega, cantidad_objetivo, *, numero_lote="", fecha_vencimiento=None, observacion=""):
    cantidad_objetivo = cantidad_objetivo or Decimal("0.00")
    if cantidad_objetivo < 0:
        raise ValidationError("La cantidad por bodega no puede ser negativa.")

    actual = ExistenciaLoteBodega.objects.filter(
        empresa=producto.empresa,
        bodega=bodega,
        lote__producto=producto,
    ).aggregate(total=Sum("cantidad"))["total"] or Decimal("0.00")
    delta = cantidad_objetivo - actual
    if delta == 0:
        return

    referencia = f"Producto {producto.id}"
    observacion = observacion or "Ajuste de distribucion por bodega desde producto."
    if delta > 0:
        lote, _ = LoteInventario.objects.get_or_create(
            empresa=producto.empresa,
            producto=producto,
            numero_lote=numero_lote or f"AJUSTE-{producto.id}",
            defaults={
                "fecha_vencimiento": fecha_vencimiento,
                "activo": True,
            },
        )
        if fecha_vencimiento and lote.fecha_vencimiento != fecha_vencimiento:
            lote.fecha_vencimiento = fecha_vencimiento
            lote.save(update_fields=["fecha_vencimiento"])
        existencia, _ = ExistenciaLoteBodega.objects.get_or_create(
            empresa=producto.empresa,
            bodega=bodega,
            lote=lote,
            defaults={"cantidad": Decimal("0.00")},
        )
        _registrar_ajuste_existencia_bodega(
            empresa=producto.empresa,
            bodega=bodega,
            lote=lote,
            cantidad_resultante=existencia.cantidad + delta,
            referencia=referencia,
            observacion=observacion,
        )
        return

    restante = abs(delta)
    existencias = (
        ExistenciaLoteBodega.objects.filter(
            empresa=producto.empresa,
            bodega=bodega,
            lote__producto=producto,
            cantidad__gt=0,
        )
        .select_related("lote")
        .order_by("lote__fecha_vencimiento", "lote__fecha_creacion", "id")
    )
    for existencia in existencias:
        if restante <= 0:
            break
        salida = min(restante, existencia.cantidad)
        _registrar_ajuste_existencia_bodega(
            empresa=producto.empresa,
            bodega=bodega,
            lote=existencia.lote,
            cantidad_resultante=existencia.cantidad - salida,
            referencia=referencia,
            observacion=observacion,
        )
        restante -= salida
    if restante > 0:
        raise ValidationError(f"No hay suficiente existencia en {bodega.nombre} para ajustar {producto.nombre}.")


def _aplicar_distribucion_bodegas_producto(producto, form):
    if not getattr(form, "mostrar_bodega_inicial", False):
        return
    if not producto.controla_inventario:
        return
    numero_lote = (form.cleaned_data.get("lote_inicial") or "").strip()
    fecha_vencimiento = form.cleaned_data.get("vencimiento_lote_inicial")
    for item in form.distribucion_bodegas():
        _ajustar_existencia_producto_bodega(
            producto,
            item["bodega"],
            item["cantidad"] or Decimal("0.00"),
            numero_lote=numero_lote,
            fecha_vencimiento=fecha_vencimiento,
        )
    _sincronizar_existencia_general_producto(producto)


def _parse_vencimiento_lote_rapido(valor):
    valor = (valor or "").strip().lower().replace(".", "").replace("/", "-").replace(" ", "-")
    if not valor:
        return None
    meses = {
        "ene": 1, "enero": 1,
        "feb": 2, "febrero": 2,
        "mar": 3, "marzo": 3,
        "abr": 4, "abril": 4,
        "may": 5, "mayo": 5,
        "jun": 6, "junio": 6,
        "jul": 7, "julio": 7,
        "ago": 8, "agos": 8, "agosto": 8,
        "sep": 9, "sept": 9, "septiembre": 9,
        "oct": 10, "octubre": 10,
        "nov": 11, "noviembre": 11,
        "dic": 12, "diciembre": 12,
    }
    partes = [parte for parte in valor.split("-") if parte]
    if len(partes) != 2:
        raise ValidationError("Usa el formato de vencimiento rapido como ago-27 o dic-26.")
    mes = meses.get(partes[0])
    if not mes:
        raise ValidationError("El mes del vencimiento rapido no es valido.")
    try:
        anio = int(partes[1])
    except ValueError as exc:
        raise ValidationError("El anio del vencimiento rapido no es valido.") from exc
    if anio < 100:
        anio += 2000
    ultimo_dia = calendar.monthrange(anio, mes)[1]
    return date(anio, mes, ultimo_dia)


def _validar_stock_vitrina_para_lineas(empresa, cantidades_por_producto):
    config = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)
    if not config.ventas_solo_desde_vitrina:
        return []
    bodega_venta = _obtener_bodega_venta(empresa)
    faltantes = []
    for item in cantidades_por_producto.values():
        producto = item["producto"]
        cantidad = item["cantidad"]
        disponible = ExistenciaLoteBodega.objects.filter(
            empresa=empresa,
            bodega=bodega_venta,
            lote__producto=producto,
            lote__activo=True,
        ).aggregate(total=Sum("cantidad"))["total"] or Decimal("0.00")
        if cantidad > disponible:
            faltantes.append(
                f"{producto.nombre}: en Vitrina {disponible:.2f}, solicitado {cantidad:.2f}"
            )
    return faltantes


def _registrar_salida_vitrina_factura(factura):
    config = ConfiguracionAvanzadaEmpresa.para_empresa(factura.empresa)
    if not config.ventas_solo_desde_vitrina:
        return
    if MovimientoLoteBodega.objects.filter(factura=factura, tipo="salida_factura").exists():
        return

    bodega_venta = _obtener_bodega_venta(factura.empresa)
    for linea in factura.lineas.select_related("producto").all():
        if not linea.producto_id or not linea.producto:
            continue
        if not linea.producto.controla_inventario:
            continue
        pendiente = linea.cantidad
        existencias = (
            ExistenciaLoteBodega.objects.filter(
                empresa=factura.empresa,
                bodega=bodega_venta,
                lote__producto=linea.producto,
                lote__activo=True,
                cantidad__gt=0,
            )
            .select_related("lote")
            .order_by("lote__fecha_vencimiento", "lote__fecha_creacion", "id")
        )
        for existencia in existencias:
            if pendiente <= 0:
                break
            salida = min(pendiente, existencia.cantidad)
            _registrar_movimiento_lote_bodega(
                empresa=factura.empresa,
                bodega=bodega_venta,
                lote=existencia.lote,
                tipo="salida_factura",
                cantidad=salida,
                referencia=factura.numero_factura or f"Factura {factura.id}",
                observacion="Salida automatica desde Vitrina por factura emitida.",
                factura=factura,
            )
            pendiente -= salida
        if pendiente > 0:
            raise ValidationError(f"No hay suficiente inventario en Vitrina para {linea.producto.nombre}.")


def _registrar_movimiento_inventario(
    *,
    empresa,
    producto,
    tipo,
    cantidad,
    referencia="",
    observacion="",
    factura=None,
    nota_credito=None,
    entrada_documento=None,
    compra_documento=None,
    bodega=None,
):
    if not producto.controla_inventario:
        return None

    inventario = _obtener_inventario_producto(producto)
    existencia_anterior = inventario.existencias

    if tipo in ['entrada', 'entrada_compra', 'ajuste_entrada', 'devolucion_nota_credito', 'reversion_factura']:
        existencia_resultante = existencia_anterior + cantidad
    else:
        existencia_resultante = existencia_anterior - cantidad

    inventario.existencias = existencia_resultante
    inventario.save(update_fields=['existencias', 'fecha_actualizacion'])

    return MovimientoInventario.objects.create(
        empresa=empresa,
        producto=producto,
        bodega=bodega,
        tipo=tipo,
        cantidad=cantidad,
        existencia_anterior=existencia_anterior,
        existencia_resultante=existencia_resultante,
        referencia=referencia,
        observacion=observacion,
        factura=factura,
        nota_credito=nota_credito,
        entrada_documento=entrada_documento,
        compra_documento=compra_documento,
    )


def _registrar_salida_factura(factura):
    if MovimientoInventario.objects.filter(factura=factura, tipo='salida_factura').exists():
        return

    _registrar_salida_vitrina_factura(factura)

    for linea in factura.lineas.select_related('producto').all():
        if not linea.producto_id or not linea.producto:
            continue
        if not linea.producto.controla_inventario:
            continue
        _registrar_movimiento_inventario(
            empresa=factura.empresa,
            producto=linea.producto,
            tipo='salida_factura',
            cantidad=linea.cantidad,
            referencia=factura.numero_factura or f"Factura {factura.id}",
            observacion='Salida generada automaticamente por emision de factura.',
            factura=factura,
        )


def _revertir_salida_factura(factura):
    if not MovimientoInventario.objects.filter(factura=factura, tipo='salida_factura').exists():
        return
    if MovimientoInventario.objects.filter(factura=factura, tipo='reversion_factura').exists():
        return

    movimientos_lote = MovimientoLoteBodega.objects.filter(factura=factura, tipo="salida_factura").select_related("bodega", "lote")
    for movimiento in movimientos_lote:
        _registrar_movimiento_lote_bodega(
            empresa=factura.empresa,
            bodega=movimiento.bodega,
            lote=movimiento.lote,
            tipo="reversion",
            cantidad=movimiento.cantidad,
            referencia=factura.numero_factura or f"Factura {factura.id}",
            observacion="Reversion automatica de lote por anulacion de factura.",
            factura=factura,
        )

    for linea in factura.lineas.select_related('producto').all():
        if not linea.producto_id or not linea.producto:
            continue
        if not linea.producto.controla_inventario:
            continue
        _registrar_movimiento_inventario(
            empresa=factura.empresa,
            producto=linea.producto,
            tipo='reversion_factura',
            cantidad=linea.cantidad,
            referencia=factura.numero_factura or f"Factura {factura.id}",
            observacion='Reversion automatica por anulacion de factura.',
            factura=factura,
        )


def _registrar_entrada_nota_credito(nota):
    if MovimientoInventario.objects.filter(nota_credito=nota, tipo='devolucion_nota_credito').exists():
        return

    for linea in nota.lineas.select_related('producto').all():
        if not linea.producto.controla_inventario:
            continue
        _registrar_movimiento_inventario(
            empresa=nota.empresa,
            producto=linea.producto,
            tipo='devolucion_nota_credito',
            cantidad=linea.cantidad,
            referencia=nota.numero_nota or f"NC {nota.id}",
            observacion='Entrada generada automaticamente por nota de credito emitida.',
            nota_credito=nota,
        )


def _revertir_entrada_nota_credito(nota):
    if not MovimientoInventario.objects.filter(nota_credito=nota, tipo='devolucion_nota_credito').exists():
        return
    if MovimientoInventario.objects.filter(nota_credito=nota, tipo='reversion_nota_credito').exists():
        return

    for linea in nota.lineas.select_related('producto').all():
        if not linea.producto.controla_inventario:
            continue
        _registrar_movimiento_inventario(
            empresa=nota.empresa,
            producto=linea.producto,
            tipo='reversion_nota_credito',
            cantidad=linea.cantidad,
            referencia=nota.numero_nota or f"NC {nota.id}",
            observacion='Reversion automatica por anulacion de nota de credito.',
            nota_credito=nota,
        )


def _reconstruir_nota_credito_emitida(nota):
    MovimientoInventario.objects.filter(
        nota_credito=nota,
        tipo__in=["devolucion_nota_credito", "reversion_nota_credito"],
    ).delete()
    AsientoContable.objects.filter(
        empresa=nota.empresa,
        documento_tipo="nota_credito",
        documento_id=nota.id,
        evento="emision",
    ).delete()
    _registrar_entrada_nota_credito(nota)
    registrar_asiento_nota_credito(nota)


def _aplicar_entrada_documento(entrada):
    if MovimientoInventario.objects.filter(entrada_documento=entrada).exists():
        return

    for linea in entrada.lineas.select_related('producto').all():
        if not linea.producto.controla_inventario:
            continue
        _registrar_movimiento_inventario(
            empresa=entrada.empresa,
            producto=linea.producto,
            tipo='entrada',
            cantidad=linea.cantidad,
            referencia=entrada.referencia,
            observacion=linea.comentario or entrada.observacion or 'Entrada formal de inventario.',
            entrada_documento=entrada,
        )
        _entrada_farmaceutica_generica(
            empresa=entrada.empresa,
            producto=linea.producto,
            cantidad=linea.cantidad,
            referencia=entrada.referencia,
            observacion=linea.comentario or entrada.observacion,
        )


def _aplicar_compra_documento(compra):
    if MovimientoInventario.objects.filter(compra_documento=compra).exists():
        return

    for linea in compra.lineas.select_related('producto').all():
        if not linea.producto.controla_inventario:
            continue
        inventario = _obtener_inventario_producto(linea.producto)
        cantidad_anterior = Decimal(inventario.existencias or 0)
        valor_anterior = cantidad_anterior * Decimal(linea.producto.costo_promedio or 0)
        cantidad_nueva = cantidad_anterior + Decimal(linea.cantidad or 0)
        if cantidad_nueva > 0:
            costo_nuevo = (
                (valor_anterior + (Decimal(linea.cantidad) * Decimal(linea.costo_unitario)))
                / cantidad_nueva
            ).quantize(Decimal("0.0001"))
            Producto.objects.filter(pk=linea.producto_id).update(costo_promedio=costo_nuevo)
            linea.producto.costo_promedio = costo_nuevo
        _registrar_movimiento_inventario(
            empresa=compra.empresa,
            producto=linea.producto,
            tipo='entrada_compra',
            cantidad=linea.cantidad,
            referencia=compra.numero_compra or compra.referencia_documento or f"Compra {compra.id}",
            observacion=linea.comentario or compra.observacion or f'Ingreso por compra a {compra.proveedor_nombre}.',
            compra_documento=compra,
        )
        _entrada_farmaceutica_generica(
            empresa=compra.empresa,
            producto=linea.producto,
            cantidad=linea.cantidad,
            referencia=compra.numero_compra or compra.referencia_documento or f"Compra {compra.id}",
            observacion=linea.comentario or compra.observacion,
        )


def _revertir_compra_documento(compra):
    if not MovimientoInventario.objects.filter(compra_documento=compra, tipo='entrada_compra').exists():
        return
    if MovimientoInventario.objects.filter(compra_documento=compra, tipo='reversion_compra').exists():
        return

    for linea in compra.lineas.select_related('producto').all():
        if not linea.producto.controla_inventario:
            continue
        inventario = _obtener_inventario_producto(linea.producto)
        cantidad_anterior = Decimal(inventario.existencias or 0)
        cantidad_nueva = cantidad_anterior - Decimal(linea.cantidad or 0)
        valor_nuevo = (
            (cantidad_anterior * Decimal(linea.producto.costo_promedio or 0))
            - (Decimal(linea.cantidad or 0) * Decimal(linea.costo_unitario or 0))
        )
        costo_nuevo = (
            max(valor_nuevo, Decimal("0.00")) / cantidad_nueva
            if cantidad_nueva > 0 else Decimal("0.00")
        ).quantize(Decimal("0.0001"))
        Producto.objects.filter(pk=linea.producto_id).update(costo_promedio=costo_nuevo)
        linea.producto.costo_promedio = costo_nuevo
        _registrar_movimiento_inventario(
            empresa=compra.empresa,
            producto=linea.producto,
            tipo='reversion_compra',
            cantidad=linea.cantidad,
            referencia=compra.numero_compra or compra.referencia_documento or f"Compra {compra.id}",
            observacion=f'Reversion automatica por anulacion de compra a {compra.proveedor_nombre}.',
            compra_documento=compra,
        )


def _validar_stock_disponible_para_lineas(lineas):
    cantidades_por_producto = {}

    for linea in lineas:
        producto = getattr(linea, 'producto', None)
        cantidad = getattr(linea, 'cantidad', None)

        if not producto or not producto.controla_inventario:
            continue

        if cantidad is None or cantidad <= 0:
            continue

        cantidades_por_producto.setdefault(producto.id, {
            'producto': producto,
            'cantidad': Decimal('0.00'),
        })
        cantidades_por_producto[producto.id]['cantidad'] += cantidad

    faltantes = []
    faltantes_vitrina = _validar_stock_vitrina_para_lineas(
        empresa=lineas[0].producto.empresa if lineas else None,
        cantidades_por_producto=cantidades_por_producto,
    ) if cantidades_por_producto else []
    if faltantes_vitrina:
        raise ValueError(
            "Stock insuficiente en Vitrina para emitir la factura. "
            "Revisa estas lineas: " + "; ".join(faltantes_vitrina) + "."
        )

    for item in cantidades_por_producto.values():
        producto = item['producto']
        cantidad_solicitada = item['cantidad']
        inventario = _obtener_inventario_producto(producto)
        disponible = inventario.existencias

        if cantidad_solicitada > disponible:
            faltantes.append(
                f"{producto.nombre}: disponible {disponible:.2f}, solicitado {cantidad_solicitada:.2f}"
            )

    if faltantes:
        raise ValueError(
            "Stock insuficiente para emitir la factura. "
            "Revisa estas lineas: " + "; ".join(faltantes) + "."
        )


def _factura_bloqueada_para_edicion(factura):
    return factura.estado == 'emitida' and not factura.puede_editar_emitida


@login_required
def clientes_facturacion(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    q = request.GET.get("q", "").strip()
    clientes = Cliente.objects.filter(empresa=empresa).order_by('nombre')
    if q:
        clientes = clientes.filter(
            Q(nombre__icontains=q) |
            Q(rtn__icontains=q) |
            Q(ciudad__icontains=q)
        )
    resumen = {
        "total": clientes.count(),
        "activos": clientes.filter(activo=True).count(),
        "inactivos": clientes.filter(activo=False).count(),
        "con_rtn": clientes.exclude(rtn__isnull=True).exclude(rtn__exact="").count(),
    }

    return render(request, "facturacion/clientes_premium.html", {
        "empresa": empresa,
        "clientes": clientes,
        "resumen": resumen,
        "q": q,
        "clientes_sugeridos": Cliente.objects.filter(empresa=empresa).order_by('nombre').values_list('nombre', flat=True).distinct(),
    })


@login_required
@xframe_options_sameorigin
def crear_cliente(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    quick_mode = request.GET.get("modal") == "1" or request.POST.get("quick_mode") == "1"

    if request.method == "POST":
        form = ClienteForm(request.POST, empresa=empresa)
        if form.is_valid():
            try:
                cliente = form.save(commit=False)
                cliente.empresa = empresa
                cliente.save()
                asegurar_cuenta_contable_cliente(cliente)
                if quick_mode:
                    return render(request, "facturacion/crear_cliente_rapido_modal.html", {
                        "empresa": empresa,
                        "success_payload": {
                            "type": "erp-cliente-creado",
                            "cliente": {
                                "id": cliente.id,
                                "nombre": cliente.nombre,
                                "rtn": cliente.rtn or "",
                                "direccion": cliente.direccion or "",
                            },
                        },
                    })
                messages.success(request, "Cliente creado correctamente.")
                return _redirect_seguro(request, "clientes_facturacion", empresa_slug=empresa.slug)
            except ValidationError as exc:
                if hasattr(exc, "message_dict"):
                    for campo, errores in exc.message_dict.items():
                        destino = campo if campo in form.fields else None
                        for error in errores:
                            form.add_error(destino, error)
                else:
                    form.add_error(None, str(exc))
    else:
        form = ClienteForm(empresa=empresa)
        nombre_prefill = (request.GET.get("nombre") or "").strip()
        if nombre_prefill:
            form.fields["nombre"].initial = nombre_prefill

    template_name = "facturacion/crear_cliente_rapido_modal.html" if quick_mode else "facturacion/crear_cliente.html"
    return render(request, template_name, {
        "empresa": empresa,
        "form": form,
        "quick_mode": quick_mode,
        "next": request.GET.get("next", ""),
        "titulo": "Nuevo Cliente",
        "texto_boton": "Guardar Cliente",
    })


@login_required
def editar_cliente(request, empresa_slug, cliente_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    cliente = get_object_or_404(Cliente, id=cliente_id, empresa=empresa)

    if request.method == "POST":
        form = ClienteForm(request.POST, instance=cliente, empresa=empresa)
        if form.is_valid():
            try:
                cliente = form.save()
                asegurar_cuenta_contable_cliente(cliente)
                messages.success(request, "Cliente actualizado correctamente.")
                return redirect("clientes_facturacion", empresa_slug=empresa.slug)
            except ValidationError as exc:
                if hasattr(exc, "message_dict"):
                    for campo, errores in exc.message_dict.items():
                        destino = campo if campo in form.fields else None
                        for error in errores:
                            form.add_error(destino, error)
                else:
                    form.add_error(None, str(exc))
    else:
        form = ClienteForm(instance=cliente, empresa=empresa)

    return render(request, "facturacion/crear_cliente.html", {
        "empresa": empresa,
        "form": form,
        "titulo": "Editar Cliente",
        "texto_boton": "Guardar Cambios",
    })


@login_required
@require_POST
def eliminar_cliente(request, empresa_slug, cliente_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    cliente = get_object_or_404(Cliente, id=cliente_id, empresa=empresa)

    if Factura.objects.filter(empresa=empresa, cliente=cliente).exists():
        messages.error(request, "No se puede eliminar este cliente porque tiene facturas registradas.")
    else:
        cliente.delete()
        messages.success(request, "Cliente eliminado correctamente.")

    return redirect("clientes_facturacion", empresa_slug=empresa.slug)


@login_required
def productos_facturacion(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    q = request.GET.get("q", "").strip()
    categoria_id = request.GET.get("categoria", "").strip()
    productos = Producto.objects.filter(empresa=empresa, eliminado=False).select_related(
        'impuesto_predeterminado',
        'perfil_farmaceutico__categoria',
    ).order_by('nombre')
    if q:
        productos = productos.filter(
            Q(nombre__icontains=q) |
            Q(codigo__icontains=q) |
            Q(descripcion__icontains=q) |
            Q(perfil_farmaceutico__principio_activo__icontains=q) |
            Q(perfil_farmaceutico__laboratorio__icontains=q) |
            Q(perfil_farmaceutico__registro_sanitario__icontains=q)
        )
    if categoria_id:
        productos = productos.filter(perfil_farmaceutico__categoria_id=categoria_id)
    resumen = {
        "total": productos.count(),
        "activos": productos.filter(activo=True).count(),
        "inactivos": productos.filter(activo=False).count(),
        "valor_catalogo": sum((producto.precio for producto in productos), Decimal('0.00')),
        "servicios": productos.filter(tipo_item='servicio').count(),
        "con_foto": productos.exclude(foto__isnull=True).exclude(foto__exact="").count(),
        "con_inventario": productos.filter(controla_inventario=True).count(),
        "controlados": productos.filter(perfil_farmaceutico__producto_controlado=True).count(),
        "refrigerados": productos.filter(perfil_farmaceutico__requiere_refrigeracion=True).count(),
    }

    return render(request, "facturacion/productos_premium.html", {
        "empresa": empresa,
        "productos": productos,
        "resumen": resumen,
        "q": q,
        "categoria_id": categoria_id,
        "categorias_farmaceuticas": CategoriaProductoFarmaceutico.objects.filter(empresa=empresa, activa=True).order_by('nombre'),
        "productos_sugeridos": Producto.objects.filter(empresa=empresa, eliminado=False).order_by('nombre').values_list('nombre', flat=True).distinct(),
        "ultimas_bajas": BitacoraProductoEliminado.objects.filter(empresa=empresa).select_related('usuario', 'producto')[:8],
    })


@login_required
def codigos_barras_productos(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)

    if request.method == "POST":
        producto = get_object_or_404(
            Producto,
            id=request.POST.get("producto_id"),
            empresa=empresa,
            eliminado=False,
        )
        codigo = (request.POST.get("codigo") or "").strip()
        if not codigo:
            return JsonResponse({"ok": False, "error": "Escanea o escribe un codigo de barras."}, status=400)
        if len(codigo) > 50:
            return JsonResponse({"ok": False, "error": "El codigo no puede superar 50 caracteres."}, status=400)

        producto.codigo = codigo
        try:
            producto.save(update_fields=["codigo"])
        except ValidationError as exc:
            errores = exc.message_dict.get("codigo", exc.messages) if hasattr(exc, "message_dict") else exc.messages
            return JsonResponse({"ok": False, "error": " ".join(errores)}, status=400)

        return JsonResponse({
            "ok": True,
            "producto_id": producto.id,
            "producto": producto.nombre,
            "codigo": producto.codigo,
        })

    q = (request.GET.get("q") or "").strip()
    solo_sin_codigo = request.GET.get("estado", "sin-codigo") != "todos"
    productos = Producto.objects.filter(
        empresa=empresa,
        eliminado=False,
    ).order_by("nombre")
    if q:
        productos = productos.filter(
            Q(nombre__icontains=q) |
            Q(codigo__icontains=q) |
            Q(descripcion__icontains=q)
        )
    if solo_sin_codigo:
        productos = productos.filter(Q(codigo__isnull=True) | Q(codigo__exact=""))

    total_catalogo = Producto.objects.filter(empresa=empresa, eliminado=False).count()
    sin_codigo = Producto.objects.filter(
        empresa=empresa,
        eliminado=False,
    ).filter(Q(codigo__isnull=True) | Q(codigo__exact="")).count()

    return render(request, "facturacion/codigos_barras_productos.html", {
        "empresa": empresa,
        "productos": productos,
        "q": q,
        "solo_sin_codigo": solo_sin_codigo,
        "total_catalogo": total_catalogo,
        "sin_codigo": sin_codigo,
        "con_codigo": total_catalogo - sin_codigo,
    })


@login_required
def proveedores_facturacion(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    q = request.GET.get("q", "").strip()
    estado = request.GET.get("estado", "").strip()
    proveedores = Proveedor.objects.filter(empresa=empresa).order_by('nombre')

    if q:
        proveedores = proveedores.filter(
            Q(nombre__icontains=q) |
            Q(rtn__icontains=q) |
            Q(contacto__icontains=q) |
            Q(ciudad__icontains=q)
        )
    if estado == "activos":
        proveedores = proveedores.filter(activo=True)
    elif estado == "inactivos":
        proveedores = proveedores.filter(activo=False)

    resumen = {
        "total": proveedores.count(),
        "activos": proveedores.filter(activo=True).count(),
        "inactivos": proveedores.filter(activo=False).count(),
        "con_rtn": proveedores.exclude(rtn__isnull=True).exclude(rtn__exact="").count(),
    }

    return render(request, "facturacion/proveedores_filtros_premium.html", {
        "empresa": empresa,
        "proveedores": proveedores,
        "resumen": resumen,
        "q": q,
        "estado": estado,
    })


@login_required
def ver_proveedor(request, empresa_slug, proveedor_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    proveedor = get_object_or_404(Proveedor, id=proveedor_id, empresa=empresa)
    compras = (
        CompraInventario.objects.filter(empresa=empresa, proveedor=proveedor)
        .prefetch_related('lineas', 'pagos_compra')
        .order_by('-fecha_documento', '-id')
    )
    resumen = {
        "compras": compras.count(),
        "aplicadas": compras.filter(estado='aplicada').count(),
        "borradores": compras.filter(estado='borrador').count(),
        "monto": sum((compra.total_documento for compra in compras), Decimal('0.00')),
        "saldo_pendiente": sum((compra.saldo_pendiente for compra in compras), Decimal('0.00')),
        "por_pagar": sum(1 for compra in compras if compra.saldo_pendiente > 0),
        "vencidas": sum(1 for compra in compras if compra.esta_vencida),
    }

    return render(request, "facturacion/ver_proveedor_cxp_premium.html", {
        "empresa": empresa,
        "proveedor": proveedor,
        "compras": compras[:20],
        "resumen": resumen,
    })


@login_required
def crear_proveedor(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)

    if request.method == "POST":
        post_data = request.POST.copy()
        post_data.setdefault('condicion_pago', 'contado')
        post_data.setdefault('dias_credito', '0')
        form = ProveedorForm(post_data)
        if form.is_valid():
            try:
                proveedor = form.save(commit=False)
                proveedor.empresa = empresa
                proveedor.save()
                messages.success(request, "Proveedor creado correctamente.")
                return _redirect_seguro(request, "proveedores_facturacion", empresa_slug=empresa.slug)
            except ValidationError as exc:
                if hasattr(exc, "message_dict"):
                    for campo, errores in exc.message_dict.items():
                        destino = campo if campo in form.fields else None
                        for error in errores:
                            form.add_error(destino, error)
                else:
                    form.add_error(None, str(exc))
    else:
        form = ProveedorForm()

    return render(request, "facturacion/crear_proveedor.html", {
        "empresa": empresa,
        "form": form,
        "next": request.GET.get("next", ""),
        "titulo": "Nuevo Proveedor",
        "texto_boton": "Guardar Proveedor",
    })


@login_required
def editar_proveedor(request, empresa_slug, proveedor_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    proveedor = get_object_or_404(Proveedor, id=proveedor_id, empresa=empresa)

    if request.method == "POST":
        post_data = request.POST.copy()
        post_data.setdefault('condicion_pago', proveedor.condicion_pago or 'contado')
        post_data.setdefault('dias_credito', str(proveedor.dias_credito or 0))
        form = ProveedorForm(post_data, instance=proveedor)
        if form.is_valid():
            form.save()
            messages.success(request, "Proveedor actualizado correctamente.")
            return redirect("proveedores_facturacion", empresa_slug=empresa.slug)
    else:
        form = ProveedorForm(instance=proveedor)

    return render(request, "facturacion/crear_proveedor.html", {
        "empresa": empresa,
        "form": form,
        "titulo": "Editar Proveedor",
        "texto_boton": "Guardar Cambios",
    })


@login_required
def inventario_facturacion(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    mostrar_costo_real = _empresa_muestra_costo_real_inventario(empresa)
    if request.method == "POST" and request.POST.get("accion") == "actualizar_costo_real":
        if not mostrar_costo_real:
            messages.error(request, "El costo real de inventario no esta habilitado para esta empresa.")
            return redirect("inventario_facturacion", empresa_slug=empresa.slug)
        producto = get_object_or_404(
            Producto,
            id=request.POST.get("producto_id"),
            empresa=empresa,
            controla_inventario=True,
            eliminado=False,
        )
        valor_raw = (request.POST.get("costo_real_inventario") or "0").strip()
        try:
            costo_real = Decimal(valor_raw or "0").quantize(Decimal("0.0001"))
        except (InvalidOperation, ValueError):
            messages.error(request, "Ingrese un costo real valido.")
            return redirect("inventario_facturacion", empresa_slug=empresa.slug)
        if costo_real < 0:
            messages.error(request, "El costo real no puede ser negativo.")
            return redirect("inventario_facturacion", empresa_slug=empresa.slug)
        costo_anterior = producto.costo_real_inventario
        with transaction.atomic():
            producto.costo_real_inventario = costo_real
            producto.save(update_fields=["costo_real_inventario"])
            _registrar_cambio_costo_real(producto, costo_anterior, request.user)
        messages.success(request, f"Costo real actualizado para {producto.nombre}.")
        return redirect(f"{reverse('inventario_facturacion', args=[empresa.slug])}?producto={producto.id}")

    productos = Producto.objects.filter(
        empresa=empresa,
        controla_inventario=True,
        eliminado=False,
    ).select_related('impuesto_predeterminado').order_by('nombre')

    for producto in productos:
        _obtener_inventario_producto(producto)

    productos = Producto.objects.filter(
        empresa=empresa,
        controla_inventario=True,
        eliminado=False,
    ).select_related('impuesto_predeterminado', 'inventario').prefetch_related(
        Prefetch(
            'historial_costos_reales',
            queryset=HistorialCostoRealProducto.objects.select_related('usuario'),
            to_attr='historial_costos_recientes',
        )
    ).order_by('nombre')

    producto_id = request.GET.get('producto')
    producto_seleccionado = None
    movimientos = None

    if producto_id:
        try:
            producto_seleccionado = productos.get(id=int(producto_id))
            movimientos = MovimientoInventario.objects.filter(
                empresa=empresa,
                producto=producto_seleccionado
            ).select_related('factura', 'nota_credito')[:20]
        except (Producto.DoesNotExist, ValueError, TypeError):
            producto_seleccionado = None
            movimientos = None

    resumen = {
        'productos_controlados': productos.count(),
        'stock_total': sum((p.stock_actual for p in productos), Decimal('0.00')),
        'con_alerta': sum(
            1 for p in productos
            if hasattr(p, 'inventario') and p.inventario.existencias <= p.inventario.stock_minimo
        ),
        'agotados': sum(
            1 for p in productos
            if hasattr(p, 'inventario') and p.inventario.existencias <= 0
        ),
        'movimientos': MovimientoInventario.objects.filter(empresa=empresa).count(),
    }
    productos_alerta = [
        p for p in productos
        if hasattr(p, 'inventario') and p.inventario.existencias <= p.inventario.stock_minimo
    ][:8]

    return render(request, "facturacion/inventario_premium.html", {
        "empresa": empresa,
        "productos": productos,
        "resumen": resumen,
        "producto_seleccionado": producto_seleccionado,
        "movimientos": movimientos,
        "productos_alerta": productos_alerta,
        "mostrar_costo_real_inventario": mostrar_costo_real,
    })


@login_required
def inventario_farmaceutico(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    if not _empresa_usa_perfil_farmaceutico(empresa):
        messages.error(request, "El inventario farmaceutico no esta activo para esta empresa.")
        return redirect("inventario_facturacion", empresa_slug=empresa.slug)

    bodegas = _asegurar_bodegas_farmaceuticas(empresa)
    hoy = timezone.localdate()
    fecha_alerta = hoy + timezone.timedelta(days=60)
    existencias = (
        ExistenciaLoteBodega.objects.filter(empresa=empresa, cantidad__gt=0)
        .select_related("bodega", "lote", "lote__producto")
        .order_by("lote__fecha_vencimiento", "lote__producto__nombre", "bodega__tipo")
    )
    lotes_alerta = existencias.filter(
        lote__fecha_vencimiento__isnull=False,
        lote__fecha_vencimiento__lte=fecha_alerta,
    )
    resumen_bodegas = []
    for bodega in BodegaInventario.objects.filter(empresa=empresa, activa=True):
        total = existencias.filter(bodega=bodega).aggregate(total=Sum("cantidad"))["total"] or Decimal("0.00")
        resumen_bodegas.append({"bodega": bodega, "total": total})

    return render(request, "facturacion/inventario_farmaceutico.html", {
        "empresa": empresa,
        "bodegas_base": bodegas,
        "existencias": existencias[:250],
        "lotes_alerta": lotes_alerta[:30],
        "resumen_bodegas": resumen_bodegas,
        "resumen": {
            "lotes": LoteInventario.objects.filter(empresa=empresa, activo=True).count(),
            "bodegas": BodegaInventario.objects.filter(empresa=empresa, activa=True).count(),
            "unidades": existencias.aggregate(total=Sum("cantidad"))["total"] or Decimal("0.00"),
            "por_vencer": lotes_alerta.count(),
        },
    })


@login_required
def bodegas_dashboard(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    config_avanzada = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)
    if not config_avanzada.usa_bodegas_internas:
        messages.error(request, "El modulo de bodegas internas no esta activo para esta empresa.")
        return redirect("inventario_facturacion", empresa_slug=empresa.slug)

    _asegurar_bodegas_farmaceuticas(empresa)
    bodegas = BodegaInventario.objects.filter(empresa=empresa, activa=True).order_by("tipo", "nombre")
    resumen_bodegas = []
    productos_unicos = set()
    for bodega in bodegas:
        existencias = list(
            ExistenciaLoteBodega.objects.filter(
                empresa=empresa,
                bodega=bodega,
                cantidad__gt=0,
            ).select_related("lote__producto")
        )
        producto_ids = {existencia.lote.producto_id for existencia in existencias}
        productos_unicos.update(producto_ids)
        unidades = sum((existencia.cantidad for existencia in existencias), Decimal("0.00"))
        costo_real = sum(
            (
                existencia.cantidad * Decimal(existencia.lote.producto.costo_real_inventario or 0)
                for existencia in existencias
            ),
            Decimal("0.00"),
        )
        valor_venta = sum(
            (
                existencia.cantidad * Decimal(existencia.lote.producto.precio or 0)
                for existencia in existencias
            ),
            Decimal("0.00"),
        )
        resumen_bodegas.append({
            "bodega": bodega,
            "unidades": unidades,
            "productos": len(producto_ids),
            "lotes": len({existencia.lote_id for existencia in existencias}),
            "costo_real": costo_real,
            "valor_venta": valor_venta,
        })

    servicios = Producto.objects.filter(
        empresa=empresa,
        tipo_item="servicio",
        activo=True,
        eliminado=False,
    ).count()
    return render(request, "facturacion/bodegas_dashboard.html", {
        "empresa": empresa,
        "resumen_bodegas": resumen_bodegas,
        "mostrar_resumen_comercial": empresa.slug in {"medical_spa", "luque_aestetic"},
        "resumen": {
            "bodegas": bodegas.count(),
            "unidades": sum((item["unidades"] for item in resumen_bodegas), Decimal("0.00")),
            "productos": len(productos_unicos),
            "servicios": servicios,
            "lotes": sum((item["lotes"] for item in resumen_bodegas), 0),
            "costo_real": sum((item["costo_real"] for item in resumen_bodegas), Decimal("0.00")),
            "valor_venta": sum((item["valor_venta"] for item in resumen_bodegas), Decimal("0.00")),
        },
    })


@login_required
def ver_bodega_inventario(request, empresa_slug, bodega_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    config_avanzada = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)
    if not config_avanzada.usa_bodegas_internas:
        messages.error(request, "El modulo de bodegas internas no esta activo para esta empresa.")
        return redirect("inventario_facturacion", empresa_slug=empresa.slug)

    _asegurar_bodegas_farmaceuticas(empresa)
    bodega = get_object_or_404(BodegaInventario, empresa=empresa, id=bodega_id, activa=True)
    existencias = (
        ExistenciaLoteBodega.objects.filter(
            empresa=empresa,
            bodega=bodega,
            cantidad__gt=0,
        )
        .select_related("lote", "lote__producto")
        .order_by("lote__producto__nombre", "lote__fecha_vencimiento", "lote__numero_lote")
    )
    productos = (
        existencias.values(
            "lote__producto",
            "lote__producto__nombre",
            "lote__producto__codigo",
            "lote__producto__unidad_medida",
        )
        .annotate(total=Sum("cantidad"), lotes=Count("lote", distinct=True))
        .order_by("lote__producto__nombre")
    )

    return render(request, "facturacion/ver_bodega_inventario.html", {
        "empresa": empresa,
        "bodega": bodega,
        "productos": productos,
        "existencias": existencias,
        "resumen": {
            "unidades": existencias.aggregate(total=Sum("cantidad"))["total"] or Decimal("0.00"),
            "productos": productos.count(),
            "lotes": existencias.values("lote").distinct().count(),
        },
    })


@login_required
def crear_lote_inventario(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    if not _empresa_usa_perfil_farmaceutico(empresa):
        messages.error(request, "El inventario farmaceutico no esta activo para esta empresa.")
        return redirect("inventario_facturacion", empresa_slug=empresa.slug)

    bodegas = _asegurar_bodegas_farmaceuticas(empresa)
    productos = Producto.objects.filter(empresa=empresa, activo=True, controla_inventario=True).order_by("nombre")
    proveedores = Proveedor.objects.filter(empresa=empresa, activo=True).order_by("nombre")
    producto_preseleccionado = productos.filter(id=request.GET.get("producto")).first()
    next_url = request.POST.get("next") or request.GET.get("next") or ""
    if request.method == "POST":
        producto = productos.filter(id=request.POST.get("producto")).first()
        bodega = BodegaInventario.objects.filter(empresa=empresa, id=request.POST.get("bodega")).first()
        numero_lote = request.POST.get("numero_lote", "").strip()
        cantidad_raw = request.POST.get("cantidad", "").strip()
        fecha_vencimiento = request.POST.get("fecha_vencimiento") or None
        vencimiento_rapido = request.POST.get("fecha_vencimiento_rapida", "").strip()
        proveedor = proveedores.filter(id=request.POST.get("proveedor")).first()
        try:
            cantidad = Decimal(cantidad_raw)
        except InvalidOperation:
            cantidad = Decimal("0.00")
        try:
            fecha_vencimiento = fecha_vencimiento or _parse_vencimiento_lote_rapido(vencimiento_rapido)
        except ValidationError as exc:
            messages.error(request, str(exc))
            fecha_vencimiento = "__invalida__"

        if fecha_vencimiento == "__invalida__":
            pass
        elif not producto or not bodega or not numero_lote or cantidad <= 0:
            messages.error(request, "Completa producto, bodega, lote y cantidad mayor que cero.")
        else:
            lote, _ = LoteInventario.objects.get_or_create(
                empresa=empresa,
                producto=producto,
                numero_lote=numero_lote,
                defaults={
                    "fecha_vencimiento": fecha_vencimiento,
                    "proveedor": proveedor,
                    "activo": True,
                },
            )
            lote.fecha_vencimiento = fecha_vencimiento
            lote.proveedor = proveedor
            lote.activo = True
            lote.save(update_fields=["fecha_vencimiento", "proveedor", "activo"])
            _registrar_movimiento_lote_bodega(
                empresa=empresa,
                bodega=bodega,
                lote=lote,
                tipo="entrada",
                cantidad=cantidad,
                referencia=f"Lote {numero_lote}",
                observacion=request.POST.get("observacion", "").strip() or "Entrada manual de lote farmaceutico.",
            )
            _registrar_movimiento_inventario(
                empresa=empresa,
                producto=producto,
                tipo="entrada",
                cantidad=cantidad,
                referencia=f"Lote {numero_lote}",
                observacion=request.POST.get("observacion", "").strip() or "Entrada manual de lote farmaceutico.",
            )
            messages.success(request, "Lote registrado correctamente.")
            if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
                return redirect(next_url)
            return redirect("inventario_farmaceutico", empresa_slug=empresa.slug)

    return render(request, "facturacion/crear_lote_inventario.html", {
        "empresa": empresa,
        "productos": productos,
        "proveedores": proveedores,
        "bodegas": BodegaInventario.objects.filter(empresa=empresa, activa=True).order_by("tipo", "nombre"),
        "bodega_principal": bodegas["principal"],
        "producto_preseleccionado": producto_preseleccionado,
        "next": next_url,
    })


@login_required
def traslado_inventario_farmaceutico(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    config_avanzada = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)
    if not config_avanzada.usa_bodegas_internas:
        messages.error(request, "Los traslados internos no estan activos para esta empresa.")
        return redirect("inventario_facturacion", empresa_slug=empresa.slug)

    _asegurar_bodegas_farmaceuticas(empresa)
    bodegas = BodegaInventario.objects.filter(empresa=empresa, activa=True).order_by("tipo", "nombre")
    existencias = (
        ExistenciaLoteBodega.objects.filter(empresa=empresa, cantidad__gt=0)
        .select_related("bodega", "lote", "lote__producto")
        .order_by("lote__producto__nombre", "lote__fecha_vencimiento")
    )

    if request.method == "POST":
        existencia = existencias.filter(id=request.POST.get("existencia")).first()
        bodega_destino = bodegas.filter(id=request.POST.get("bodega_destino")).first()
        cantidad_raw = request.POST.get("cantidad", "").strip()
        try:
            cantidad = Decimal(cantidad_raw)
        except InvalidOperation:
            cantidad = Decimal("0.00")

        if not existencia or not bodega_destino or cantidad <= 0:
            messages.error(request, "Selecciona lote, destino y cantidad mayor que cero.")
        elif existencia.bodega_id == bodega_destino.id:
            messages.error(request, "La bodega destino debe ser diferente a la bodega origen.")
        elif cantidad > existencia.cantidad:
            messages.error(request, "No puedes trasladar mas unidades que las disponibles.")
        else:
            referencia = f"Traslado {existencia.bodega.nombre} -> {bodega_destino.nombre}"
            observacion = request.POST.get("observacion", "").strip()
            with transaction.atomic():
                _registrar_movimiento_lote_bodega(
                    empresa=empresa,
                    bodega=existencia.bodega,
                    lote=existencia.lote,
                    tipo="traslado_salida",
                    cantidad=cantidad,
                    referencia=referencia,
                    observacion=observacion,
                )
                _registrar_movimiento_lote_bodega(
                    empresa=empresa,
                    bodega=bodega_destino,
                    lote=existencia.lote,
                    tipo="traslado_entrada",
                    cantidad=cantidad,
                    referencia=referencia,
                    observacion=observacion,
                )
            messages.success(request, "Traslado registrado correctamente.")
            return redirect("inventario_farmaceutico", empresa_slug=empresa.slug)

    return render(request, "facturacion/traslado_inventario_farmaceutico.html", {
        "empresa": empresa,
        "existencias": existencias,
        "bodegas": bodegas,
    })


@login_required
def traslado_rapido_farmaceutico(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    config_avanzada = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)
    if not config_avanzada.usa_bodegas_internas:
        messages.error(request, "Los traslados internos no estan activos para esta empresa.")
        return redirect("inventario_facturacion", empresa_slug=empresa.slug)

    bodegas = _asegurar_bodegas_farmaceuticas(empresa)
    if empresa.slug == "luque_aestetic":
        rutas = {
            "principal_clinica": {
                "label": "Principal -> Clinica",
                "origen": bodegas["principal"],
                "destino": bodegas["provisional"],
            },
            "clinica_principal": {
                "label": "Clinica -> Principal",
                "origen": bodegas["provisional"],
                "destino": bodegas["principal"],
            },
        }
        ruta_default = "principal_clinica"
    else:
        rutas = {
            "principal_provisional": {
                "label": "Principal -> Provisional",
                "origen": bodegas["principal"],
                "destino": bodegas["provisional"],
            },
            "principal_vitrina": {
                "label": "Principal -> Vitrina",
                "origen": bodegas["principal"],
                "destino": bodegas["vitrina"],
            },
            "provisional_vitrina": {
                "label": "Provisional -> Vitrina",
                "origen": bodegas["provisional"],
                "destino": bodegas["vitrina"],
            },
        }
        ruta_default = "principal_vitrina"

    ruta_key = request.POST.get("ruta") or request.GET.get("ruta") or ruta_default
    ruta = rutas.get(ruta_key, rutas[ruta_default])
    existencias = (
        ExistenciaLoteBodega.objects.filter(
            empresa=empresa,
            bodega=ruta["origen"],
            cantidad__gt=0,
        )
        .select_related("lote", "lote__producto", "bodega")
        .order_by("lote__fecha_vencimiento", "lote__producto__nombre")
    )

    if request.method == "POST":
        existencia = existencias.filter(id=request.POST.get("existencia")).first()
        cantidad_raw = request.POST.get("cantidad", "").strip()
        try:
            cantidad = Decimal(cantidad_raw)
        except InvalidOperation:
            cantidad = Decimal("0.00")

        if not existencia or cantidad <= 0:
            messages.error(request, "Selecciona un lote disponible y una cantidad mayor que cero.")
        elif cantidad > existencia.cantidad:
            messages.error(request, "La cantidad supera lo disponible en la bodega origen.")
        else:
            referencia = f"Rapido {ruta['origen'].nombre} -> {ruta['destino'].nombre}"
            observacion = request.POST.get("observacion", "").strip() or "Traslado rapido de inventario farmaceutico."
            with transaction.atomic():
                _registrar_movimiento_lote_bodega(
                    empresa=empresa,
                    bodega=ruta["origen"],
                    lote=existencia.lote,
                    tipo="traslado_salida",
                    cantidad=cantidad,
                    referencia=referencia,
                    observacion=observacion,
                )
                _registrar_movimiento_lote_bodega(
                    empresa=empresa,
                    bodega=ruta["destino"],
                    lote=existencia.lote,
                    tipo="traslado_entrada",
                    cantidad=cantidad,
                    referencia=referencia,
                    observacion=observacion,
                )
            messages.success(
                request,
                f"Traslado rapido realizado: {cantidad:.2f} unidad(es) de {existencia.lote.producto.nombre} hacia {ruta['destino'].nombre}."
            )
            return redirect(f"{request.path}?ruta={ruta_key}")

    return render(request, "facturacion/traslado_rapido_farmaceutico.html", {
        "empresa": empresa,
        "rutas": rutas,
        "ruta_key": ruta_key,
        "ruta": ruta,
        "existencias": existencias,
    })


@login_required
def ajustar_inventario(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)

    if request.method == "POST":
        form = AjusteInventarioForm(request.POST, empresa=empresa)
        if form.is_valid():
            producto = form.cleaned_data['producto']
            tipo_ajuste = form.cleaned_data['tipo_ajuste']
            cantidad = form.cleaned_data['cantidad']
            observacion = form.cleaned_data['observacion']
            stock_minimo = form.cleaned_data['stock_minimo']

            inventario = _obtener_inventario_producto(producto)
            if stock_minimo is not None:
                inventario.stock_minimo = stock_minimo
                inventario.save(update_fields=['stock_minimo', 'fecha_actualizacion'])

            _registrar_movimiento_inventario(
                empresa=empresa,
                producto=producto,
                tipo=tipo_ajuste,
                cantidad=cantidad,
                referencia=f"Ajuste manual {timezone.now().strftime('%Y-%m-%d')}",
                observacion=observacion or 'Ajuste manual de inventario.',
            )
            messages.success(request, "Inventario ajustado correctamente.")
            return redirect("inventario_facturacion", empresa_slug=empresa.slug)
    else:
        form = AjusteInventarioForm(empresa=empresa)

    return render(request, "facturacion/ajustar_inventario.html", {
        "empresa": empresa,
        "form": form,
    })


@login_required
def entrada_inventario(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    if empresa.slug in {"hospital_mia", "medical_spa"}:
        _asegurar_bodegas_farmaceuticas(empresa)

    if request.method == "POST":
        form = EntradaInventarioForm(request.POST, empresa=empresa)
        if form.is_valid():
            producto = form.cleaned_data['producto']
            bodega = form.cleaned_data.get('bodega')
            cantidad = form.cleaned_data['cantidad']
            referencia = form.cleaned_data['referencia']
            observacion = form.cleaned_data['observacion']
            stock_minimo = form.cleaned_data['stock_minimo']

            with transaction.atomic():
                inventario = _obtener_inventario_producto(producto)
                if stock_minimo is not None:
                    inventario.stock_minimo = stock_minimo
                    inventario.save(update_fields=['stock_minimo', 'fecha_actualizacion'])

                _registrar_movimiento_inventario(
                    empresa=empresa,
                    producto=producto,
                    bodega=bodega,
                    tipo='entrada',
                    cantidad=cantidad,
                    referencia=referencia,
                    observacion=observacion or 'Entrada formal de inventario.',
                )
                if bodega:
                    _registrar_movimiento_lote_bodega(
                        empresa=empresa,
                        bodega=bodega,
                        lote=_obtener_lote_generico(producto),
                        tipo='entrada',
                        cantidad=cantidad,
                        referencia=referencia,
                        observacion=observacion or 'Entrada formal de inventario.',
                    )
            destino = f" en {bodega.nombre}" if bodega else ""
            messages.success(request, f"Entrada de inventario registrada correctamente{destino}.")
            return redirect("inventario_facturacion", empresa_slug=empresa.slug)
    else:
        form = EntradaInventarioForm(empresa=empresa)

    return render(request, "facturacion/entrada_inventario.html", {
        "empresa": empresa,
        "form": form,
    })


@login_required
def entradas_inventario_dashboard(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    entradas = EntradaInventarioDocumento.objects.filter(empresa=empresa).prefetch_related('lineas').order_by('-fecha_documento', '-id')
    resumen = {
        "total": entradas.count(),
        "aplicadas": entradas.filter(estado='aplicada').count(),
        "borradores": entradas.filter(estado='borrador').count(),
        "unidades": sum(
            (sum((linea.cantidad for linea in entrada.lineas.all()), Decimal('0.00')) for entrada in entradas),
            Decimal('0.00')
        ),
    }

    return render(request, "facturacion/entradas_inventario_premium.html", {
        "empresa": empresa,
        "entradas": entradas,
        "resumen": resumen,
    })


@login_required
def crear_entrada_inventario_documento(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)

    EntradaForm = modelform_factory(
        EntradaInventarioDocumento,
        fields=['referencia', 'fecha_documento', 'observacion', 'estado']
    )
    LineaFormSet = inlineformset_factory(
        EntradaInventarioDocumento,
        LineaEntradaInventario,
        fields=['producto', 'cantidad', 'comentario'],
        extra=1,
        can_delete=True
    )

    productos_qs = Producto.objects.filter(
        empresa=empresa,
        activo=True,
        controla_inventario=True
    ).order_by('nombre')

    if request.method == "POST":
        form = EntradaForm(request.POST)
        entrada_temp = EntradaInventarioDocumento(empresa=empresa)
        formset = LineaFormSet(request.POST, instance=entrada_temp, prefix='lineas_entrada')

        for f in formset.forms:
            f.fields['producto'].queryset = productos_qs

        lineas_validas = []
        for f in formset.forms:
            prefix = f.prefix
            producto_raw = (request.POST.get(f"{prefix}-producto") or "").strip()
            cantidad_raw = (request.POST.get(f"{prefix}-cantidad") or "").strip()
            comentario_raw = (request.POST.get(f"{prefix}-comentario") or "").strip()
            delete_raw = request.POST.get(f"{prefix}-DELETE")
            fila_vacia = not producto_raw and not cantidad_raw and not comentario_raw
            if delete_raw or fila_vacia:
                continue
            if f.is_valid():
                lineas_validas.append(f)

        if form.is_valid() and lineas_validas:
            with transaction.atomic():
                estado_destino = form.cleaned_data['estado']
                entrada = form.save(commit=False)
                entrada.empresa = empresa
                entrada.estado = 'borrador'
                entrada.save()

                for f in lineas_validas:
                    linea = f.save(commit=False)
                    linea.entrada = entrada
                    linea.save()

                if estado_destino == 'aplicada':
                    _aplicar_entrada_documento(entrada)
                    entrada.estado = 'aplicada'
                    entrada.save(update_fields=['estado'])

            messages.success(request, "Entrada de inventario guardada correctamente.")
            return redirect("ver_entrada_inventario", empresa_slug=empresa.slug, entrada_id=entrada.id)
        elif form.is_valid():
            messages.error(request, "Debe agregar al menos una linea valida en la entrada.")
    else:
        form = EntradaForm()
        formset = LineaFormSet(prefix='lineas_entrada')
        for f in formset:
            f.fields['producto'].queryset = productos_qs

    return render(request, "facturacion/crear_entrada_inventario_premium.html", {
        "empresa": empresa,
        "form": form,
        "formset": formset,
        "productos": productos_qs,
    })


@login_required
def ver_entrada_inventario(request, empresa_slug, entrada_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    entrada = get_object_or_404(
        EntradaInventarioDocumento.objects.prefetch_related('lineas__producto'),
        id=entrada_id,
        empresa=empresa,
    )
    total_unidades = sum((linea.cantidad for linea in entrada.lineas.all()), Decimal('0.00'))
    movimientos = MovimientoInventario.objects.filter(entrada_documento=entrada).select_related('producto')

    return render(request, "facturacion/ver_entrada_inventario_premium.html", {
        "empresa": empresa,
        "entrada": entrada,
        "total_unidades": total_unidades,
        "movimientos": movimientos,
    })


@login_required
def compras_dashboard(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    q = request.GET.get("q", "").strip()
    estado_filtro = request.GET.get("estado", "").strip()
    pago_filtro = request.GET.get("pago", "").strip()
    compras = (
        CompraInventario.objects.filter(empresa=empresa)
        .select_related('proveedor')
        .prefetch_related('lineas', 'pagos_compra')
        .order_by('-fecha_documento', '-id')
    )

    if q:
        compras = compras.filter(
            Q(numero_compra__icontains=q) |
            Q(proveedor_nombre__icontains=q) |
            Q(referencia_documento__icontains=q)
        )
    if estado_filtro in {"borrador", "aplicada", "anulada"}:
        compras = compras.filter(estado=estado_filtro)

    compras_lista = list(compras)
    if pago_filtro in {"pendiente", "parcial", "pagado"}:
        compras_lista = [compra for compra in compras_lista if compra.estado_pago == pago_filtro]

    resumen = {
        "total": len(compras_lista),
        "aplicadas": sum(1 for compra in compras_lista if compra.estado == 'aplicada'),
        "borradores": sum(1 for compra in compras_lista if compra.estado == 'borrador'),
        "unidades": sum((compra.total_unidades for compra in compras_lista), Decimal('0.00')),
        "monto": sum((compra.total_documento for compra in compras_lista), Decimal('0.00')),
        "saldo_total": sum((compra.saldo_pendiente for compra in compras_lista), Decimal('0.00')),
        "por_pagar": sum(1 for compra in compras_lista if compra.saldo_pendiente > 0),
    }

    return render(request, "facturacion/compras_cxp_premium.html", {
        "empresa": empresa,
        "compras": compras_lista,
        "resumen": resumen,
        "q": q,
        "estado_filtro": estado_filtro,
        "pago_filtro": pago_filtro,
    })


@login_required
def libro_compras_fiscal(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    registros_activos = RegistroCompraFiscal.objects.filter(empresa=empresa).exclude(estado="anulada")
    periodos = (
        registros_activos.values("periodo_anio", "periodo_mes")
        .annotate(
            documentos=Count("id"),
            subtotal=Sum("subtotal"),
            base_15=Sum("base_15"),
            isv_15=Sum("isv_15"),
            base_18=Sum("base_18"),
            isv_18=Sum("isv_18"),
            exento=Sum("exento"),
            total=Sum("total"),
        )
        .order_by("-periodo_anio", "-periodo_mes")
    )
    resumen = registros_activos.aggregate(
        subtotal=Sum("subtotal"),
        base_15=Sum("base_15"),
        isv_15=Sum("isv_15"),
        base_18=Sum("base_18"),
        isv_18=Sum("isv_18"),
        exento=Sum("exento"),
        exonerado=Sum("exonerado"),
        total=Sum("total"),
    )
    resumen = {clave: valor or Decimal("0.00") for clave, valor in resumen.items()}
    resumen["documentos"] = registros_activos.count()
    resumen["isv_total"] = resumen["isv_15"] + resumen["isv_18"]

    importar_form = ImportarLibroComprasForm(initial={
        "periodo_anio": timezone.now().year,
        "periodo_mes": timezone.now().month,
    })
    return render(request, "facturacion/libro_compras_meses_premium.html", {
        "empresa": empresa,
        "periodos": periodos,
        "resumen": resumen,
        "importar_form": importar_form,
    })


@login_required
def libro_compras_fiscal_detalle(request, empresa_slug, anio, mes):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    q = request.GET.get("q", "").strip()
    registros = RegistroCompraFiscal.objects.filter(
        empresa=empresa,
        periodo_anio=anio,
        periodo_mes=mes,
    ).select_related("proveedor", "clasificacion_contable").order_by("fecha_documento", "id")
    if q:
        registros = registros.filter(
            Q(proveedor_nombre__icontains=q) |
            Q(proveedor_rtn__icontains=q) |
            Q(numero_factura__icontains=q) |
            Q(cai__icontains=q)
        )
    resumen_base = registros.exclude(estado="anulada")
    resumen = resumen_base.aggregate(
        subtotal=Sum("subtotal"),
        base_15=Sum("base_15"),
        isv_15=Sum("isv_15"),
        base_18=Sum("base_18"),
        isv_18=Sum("isv_18"),
        exento=Sum("exento"),
        exonerado=Sum("exonerado"),
        total=Sum("total"),
    )
    resumen = {clave: valor or Decimal("0.00") for clave, valor in resumen.items()}
    resumen["documentos"] = resumen_base.count()
    resumen["isv_total"] = resumen["isv_15"] + resumen["isv_18"]
    return render(request, "facturacion/libro_compras_fiscal_detalle_premium.html", {
        "empresa": empresa,
        "registros": registros,
        "resumen": resumen,
        "periodo": {"anio": anio, "mes": mes},
        "q": q,
    })


@login_required
def importar_libro_compras_fiscal(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    if request.method != "POST":
        return redirect("libro_compras_fiscal", empresa_slug=empresa.slug)

    form = ImportarLibroComprasForm(request.POST, request.FILES)
    if form.is_valid():
        try:
            resultado = importar_libro_compras_desde_excel(
                empresa,
                form.cleaned_data["archivo"],
                form.cleaned_data["periodo_anio"],
                form.cleaned_data["periodo_mes"],
            )
            messages.success(request, f"Libro importado: {resultado['creadas']} compras fiscales creadas.")
            if resultado["duplicadas"]:
                messages.warning(request, f"Se omitieron {len(resultado['duplicadas'])} posibles facturas duplicadas.")
            if resultado["omitidas"]:
                messages.warning(request, f"Se omitieron {len(resultado['omitidas'])} filas incompletas.")
        except Exception as exc:
            messages.error(request, f"No se pudo importar el libro de compras: {exc}")
    else:
        messages.error(request, "Revisa el archivo y el periodo antes de importar.")
    periodo_anio = request.POST.get("periodo_anio")
    periodo_mes = request.POST.get("periodo_mes")
    if periodo_anio and periodo_mes:
        return redirect("libro_compras_fiscal_detalle", empresa_slug=empresa.slug, anio=periodo_anio, mes=periodo_mes)
    return redirect("libro_compras_fiscal", empresa_slug=empresa.slug)


def _decimal_desde_fila(valor):
    try:
        return Decimal(str(valor or "0").replace(",", "")).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


def _centavos(valor):
    return valor.quantize(Decimal("0.01"))


def _fecha_desde_fila(valor):
    valor = (valor or "").strip()
    if not valor:
        return None
    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(valor, formato).date()
        except ValueError:
            continue
    return None


@login_required
def crear_registro_compra_fiscal(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    filas = [{
        "fecha_documento": timezone.now().date().isoformat(),
        "clasificacion_id": "",
        "proveedor_nombre": "",
        "numero_factura": "",
        "exento": "",
        "base_15": "",
        "base_18": "",
        "isv_15": "0.00",
        "isv_18": "0.00",
        "total": "0.00",
    }]
    errores = []

    if request.method == "POST":
        fechas = request.POST.getlist("fecha_documento[]")
        proveedores = request.POST.getlist("proveedor_nombre[]")
        facturas = request.POST.getlist("numero_factura[]")
        clasificaciones = request.POST.getlist("clasificacion_id[]")
        exentos = request.POST.getlist("exento[]")
        bases_15 = request.POST.getlist("base_15[]")
        bases_18 = request.POST.getlist("base_18[]")
        total_filas = max(len(fechas), len(proveedores), len(facturas), len(clasificaciones), len(exentos), len(bases_15), len(bases_18))
        filas = []
        registros = []
        claves_en_pantalla = set()

        for index in range(total_filas):
            fila = {
                "fecha_documento": fechas[index] if index < len(fechas) else "",
                "clasificacion_id": clasificaciones[index] if index < len(clasificaciones) else "",
                "proveedor_nombre": proveedores[index].strip() if index < len(proveedores) else "",
                "numero_factura": facturas[index].strip() if index < len(facturas) else "",
                "exento": exentos[index] if index < len(exentos) else "",
                "base_15": bases_15[index] if index < len(bases_15) else "",
                "base_18": bases_18[index] if index < len(bases_18) else "",
            }
            exento = _decimal_desde_fila(fila["exento"])
            base_15 = _decimal_desde_fila(fila["base_15"])
            base_18 = _decimal_desde_fila(fila["base_18"])
            isv_15 = _centavos(base_15 * Decimal("0.15"))
            isv_18 = _centavos(base_18 * Decimal("0.18"))
            total = _centavos(exento + base_15 + base_18 + isv_15 + isv_18)
            fila.update({"isv_15": isv_15, "isv_18": isv_18, "total": total})
            filas.append(fila)

            fila_vacia = not fila["fecha_documento"] and not fila["proveedor_nombre"] and not fila["numero_factura"] and total == 0
            if fila_vacia:
                continue

            fecha_documento = _fecha_desde_fila(fila["fecha_documento"])
            if not fecha_documento or not fila["proveedor_nombre"] or not fila["numero_factura"]:
                errores.append(f"Fila {index + 1}: fecha, proveedor y numero de factura son obligatorios.")
                continue
            if total <= 0:
                errores.append(f"Fila {index + 1}: el total de la factura debe ser mayor que cero.")
                continue

            registro = RegistroCompraFiscal(
                empresa=empresa,
                clasificacion_contable_id=fila["clasificacion_id"] or None,
                proveedor_nombre=fila["proveedor_nombre"],
                numero_factura=fila["numero_factura"],
                fecha_documento=fecha_documento,
                periodo_anio=fecha_documento.year,
                periodo_mes=fecha_documento.month,
                subtotal=exento + base_15 + base_18,
                exento=exento,
                base_15=base_15,
                isv_15=isv_15,
                base_18=base_18,
                isv_18=isv_18,
                total=total,
            )
            duplicada = registro.buscar_duplicada()
            if duplicada:
                errores.append(
                    f"Fila {index + 1}: la factura {registro.numero_factura} ya existe en "
                    f"{duplicada.periodo_mes}/{duplicada.periodo_anio}."
                )
                continue
            clave_pantalla = (registro.proveedor_nombre.lower(), registro.numero_factura.lower())
            if clave_pantalla in claves_en_pantalla:
                errores.append(f"Fila {index + 1}: esta factura ya esta repetida en las lineas que estas guardando.")
                continue
            claves_en_pantalla.add(clave_pantalla)
            registros.append(registro)

        if not errores and registros:
            with transaction.atomic():
                for registro in registros:
                    registro.save()
            primer_registro = registros[0]
            messages.success(request, f"Libro actualizado: {len(registros)} facturas guardadas correctamente.")
            return redirect(
                "libro_compras_fiscal_detalle",
                empresa_slug=empresa.slug,
                anio=primer_registro.periodo_anio,
                mes=primer_registro.periodo_mes,
            )
        if not registros and not errores:
            errores.append("Agrega al menos una linea valida al libro de compras.")

    return render(request, "facturacion/registro_compra_fiscal_table_form.html", {
        "empresa": empresa,
        "filas": filas,
        "errores": errores,
        "clasificaciones": ClasificacionCompraFiscal.objects.filter(empresa=empresa, activa=True).select_related("cuenta_contable").order_by("nombre"),
    })


@login_required
@require_POST
def anular_registro_compra_fiscal(request, empresa_slug, registro_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    registro = get_object_or_404(RegistroCompraFiscal, id=registro_id, empresa=empresa)
    registro.estado = "anulada"
    registro.save(update_fields=["estado"])
    messages.success(request, "Registro fiscal anulado correctamente.")
    return redirect("libro_compras_fiscal", empresa_slug=empresa.slug)


@login_required
def crear_compra(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    estados_disponibles = [estado for estado in CompraInventario.ESTADOS if estado[0] != 'anulada']
    proveedores_qs = Proveedor.objects.filter(empresa=empresa, activo=True).order_by('nombre')

    CompraForm = modelform_factory(
        CompraInventario,
        fields=[
            'proveedor',
            'proveedor_nombre',
            'referencia_documento',
            'fecha_documento',
            'condicion_pago',
            'dias_credito',
            'fecha_vencimiento',
            'observacion',
            'estado',
        ]
    )
    LineaFormSet = inlineformset_factory(
        CompraInventario,
        LineaCompraInventario,
        fields=['producto', 'cantidad', 'costo_unitario', 'comentario'],
        extra=1,
        can_delete=True
    )

    productos_qs = Producto.objects.filter(
        empresa=empresa,
        activo=True,
        controla_inventario=True
    ).order_by('nombre')

    if request.method == "POST":
        post_data = request.POST.copy()
        post_data.setdefault('condicion_pago', 'contado')
        post_data.setdefault('dias_credito', '0')
        form = CompraForm(post_data)
        form.fields['estado'].choices = estados_disponibles
        form.fields['proveedor'].queryset = proveedores_qs
        form.fields['proveedor_nombre'].required = False
        form.fields['fecha_vencimiento'].required = False
        compra_temp = CompraInventario(empresa=empresa)
        formset = LineaFormSet(post_data, instance=compra_temp, prefix='lineas_compra')

        for f in formset.forms:
            f.fields['producto'].queryset = productos_qs

        lineas_validas = []
        for f in formset.forms:
            prefix = f.prefix
            producto_raw = (request.POST.get(f"{prefix}-producto") or "").strip()
            cantidad_raw = (request.POST.get(f"{prefix}-cantidad") or "").strip()
            costo_raw = (request.POST.get(f"{prefix}-costo_unitario") or "").strip()
            comentario_raw = (request.POST.get(f"{prefix}-comentario") or "").strip()
            delete_raw = request.POST.get(f"{prefix}-DELETE")
            fila_vacia = not producto_raw and not cantidad_raw and not costo_raw and not comentario_raw
            if delete_raw or fila_vacia:
                continue
            if f.is_valid():
                lineas_validas.append(f)

        if form.is_valid() and lineas_validas:
            with transaction.atomic():
                estado_destino = form.cleaned_data['estado']
                compra = form.save(commit=False)
                compra.empresa = empresa
                if compra.proveedor:
                    compra.proveedor_nombre = compra.proveedor.nombre
                    if not compra.pk:
                        compra.condicion_pago = compra.condicion_pago or compra.proveedor.condicion_pago
                compra.estado = 'borrador'
                compra.save()

                for f in lineas_validas:
                    linea = f.save(commit=False)
                    linea.compra = compra
                    linea.save()

                if estado_destino == 'aplicada':
                    _aplicar_compra_documento(compra)
                    compra.estado = 'aplicada'
                    compra.save(update_fields=['estado'])

            messages.success(request, "Compra guardada correctamente.")
            return redirect("ver_compra", empresa_slug=empresa.slug, compra_id=compra.id)
        elif form.is_valid():
            messages.error(request, "Debe agregar al menos una linea valida en la compra.")
    else:
        form = CompraForm()
        form.fields['estado'].choices = estados_disponibles
        form.fields['proveedor'].queryset = proveedores_qs
        form.fields['proveedor_nombre'].required = False
        form.fields['fecha_vencimiento'].required = False
        formset = LineaFormSet(prefix='lineas_compra')
        for f in formset:
            f.fields['producto'].queryset = productos_qs

    return render(request, "facturacion/crear_compra_premium.html", {
        "empresa": empresa,
        "form": form,
        "formset": formset,
        "productos": productos_qs,
        "proveedores": proveedores_qs,
        "proveedores_sugeridos": proveedores_qs.values_list('nombre', flat=True).distinct(),
        "proveedores_config": {
            str(proveedor.id): {
                "condicion_pago": proveedor.condicion_pago,
                "dias_credito": proveedor.dias_credito,
            }
            for proveedor in proveedores_qs
        },
        "modo_edicion": False,
    })


@login_required
def editar_compra(request, empresa_slug, compra_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    compra = get_object_or_404(CompraInventario, id=compra_id, empresa=empresa)
    estados_disponibles = [estado for estado in CompraInventario.ESTADOS if estado[0] != 'anulada']
    proveedores_qs = Proveedor.objects.filter(empresa=empresa, activo=True).order_by('nombre')

    if compra.estado == 'anulada':
        messages.error(request, "No se puede editar una compra anulada.")
        return redirect("ver_compra", empresa_slug=empresa.slug, compra_id=compra.id)

    if compra.estado == 'aplicada':
        messages.error(request, "No se puede editar una compra aplicada porque ya impactó el inventario.")
        return redirect("ver_compra", empresa_slug=empresa.slug, compra_id=compra.id)

    CompraForm = modelform_factory(
        CompraInventario,
        fields=[
            'proveedor',
            'proveedor_nombre',
            'referencia_documento',
            'fecha_documento',
            'condicion_pago',
            'dias_credito',
            'fecha_vencimiento',
            'observacion',
            'estado',
        ]
    )
    LineaFormSet = inlineformset_factory(
        CompraInventario,
        LineaCompraInventario,
        fields=['producto', 'cantidad', 'costo_unitario', 'comentario'],
        extra=0,
        can_delete=True
    )

    productos_qs = Producto.objects.filter(
        empresa=empresa,
        activo=True,
        controla_inventario=True
    ).order_by('nombre')

    if request.method == "POST":
        post_data = request.POST.copy()
        post_data.setdefault('condicion_pago', compra.condicion_pago or 'contado')
        post_data.setdefault('dias_credito', str(compra.dias_credito or 0))
        form = CompraForm(post_data, instance=compra)
        form.fields['estado'].choices = estados_disponibles
        form.fields['proveedor'].queryset = proveedores_qs
        form.fields['proveedor_nombre'].required = False
        form.fields['fecha_vencimiento'].required = False
        formset = LineaFormSet(post_data, instance=compra, prefix='lineas_compra')

        for f in formset.forms:
            f.fields['producto'].queryset = productos_qs

        lineas_validas = []
        for f in formset.forms:
            prefix = f.prefix
            producto_raw = (request.POST.get(f"{prefix}-producto") or "").strip()
            cantidad_raw = (request.POST.get(f"{prefix}-cantidad") or "").strip()
            costo_raw = (request.POST.get(f"{prefix}-costo_unitario") or "").strip()
            comentario_raw = (request.POST.get(f"{prefix}-comentario") or "").strip()
            delete_raw = request.POST.get(f"{prefix}-DELETE")
            fila_vacia = not producto_raw and not cantidad_raw and not costo_raw and not comentario_raw
            if delete_raw or fila_vacia:
                continue
            if f.is_valid():
                lineas_validas.append(f)

        if form.is_valid() and formset.is_valid() and lineas_validas:
            with transaction.atomic():
                compra = form.save(commit=False)
                if compra.proveedor:
                    compra.proveedor_nombre = compra.proveedor.nombre
                compra.estado = 'borrador'
                compra.save()
                formset.save()
            messages.success(request, "Compra actualizada correctamente.")
            return redirect("ver_compra", empresa_slug=empresa.slug, compra_id=compra.id)
        elif form.is_valid() and formset.is_valid():
            messages.error(request, "Debe agregar al menos una linea valida en la compra.")
    else:
        form = CompraForm(instance=compra)
        form.fields['estado'].choices = estados_disponibles
        form.fields['proveedor'].queryset = proveedores_qs
        form.fields['proveedor_nombre'].required = False
        form.fields['fecha_vencimiento'].required = False
        formset = LineaFormSet(instance=compra, prefix='lineas_compra')
        for f in formset:
            f.fields['producto'].queryset = productos_qs

    return render(request, "facturacion/crear_compra_premium.html", {
        "empresa": empresa,
        "form": form,
        "formset": formset,
        "productos": productos_qs,
        "proveedores": proveedores_qs,
        "proveedores_sugeridos": proveedores_qs.values_list('nombre', flat=True).distinct(),
        "proveedores_config": {
            str(proveedor.id): {
                "condicion_pago": proveedor.condicion_pago,
                "dias_credito": proveedor.dias_credito,
            }
            for proveedor in proveedores_qs
        },
        "modo_edicion": True,
        "compra": compra,
    })


@login_required
@require_POST
def aplicar_compra(request, empresa_slug, compra_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    compra = get_object_or_404(CompraInventario, id=compra_id, empresa=empresa)

    if compra.estado == 'anulada':
        messages.error(request, "No se puede aplicar una compra anulada.")
        return redirect("ver_compra", empresa_slug=empresa.slug, compra_id=compra.id)

    if compra.estado == 'aplicada':
        messages.info(request, "La compra ya estaba aplicada.")
        return redirect("ver_compra", empresa_slug=empresa.slug, compra_id=compra.id)

    if not compra.lineas.exists():
        messages.error(request, "No se puede aplicar una compra sin lineas.")
        return redirect("ver_compra", empresa_slug=empresa.slug, compra_id=compra.id)

    with transaction.atomic():
        _aplicar_compra_documento(compra)
        compra.estado = 'aplicada'
        compra.save(update_fields=['estado'])
        registrar_asiento_compra_aplicada(compra)

    messages.success(request, "Compra aplicada correctamente e inventario actualizado.")
    return redirect("ver_compra", empresa_slug=empresa.slug, compra_id=compra.id)


@login_required
@require_POST
def anular_compra(request, empresa_slug, compra_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    compra = get_object_or_404(CompraInventario, id=compra_id, empresa=empresa)

    if compra.estado == 'anulada':
        messages.info(request, "La compra ya estaba anulada.")
        return redirect("ver_compra", empresa_slug=empresa.slug, compra_id=compra.id)

    with transaction.atomic():
        if compra.estado == 'aplicada':
            _revertir_compra_documento(compra)
            registrar_reversion_documento(
                empresa=compra.empresa,
                documento_tipo='compra',
                documento_id=compra.id,
                evento_origen='aplicacion',
                evento_reversion='anulacion',
                fecha=timezone.now().date(),
                descripcion=f"Reversion compra {compra.numero_compra or compra.id}",
                referencia=compra.numero_compra or compra.referencia_documento or str(compra.id),
                origen_modulo='compras',
            )
        compra.estado = 'anulada'
        compra.save(update_fields=['estado'])

    messages.success(request, "Compra anulada correctamente.")
    return redirect("ver_compra", empresa_slug=empresa.slug, compra_id=compra.id)


@login_required
def registrar_pago_compra(request, empresa_slug, compra_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    _cuentas_financieras_activas_para_pago(empresa)
    compra = get_object_or_404(
        CompraInventario.objects.select_related('proveedor').prefetch_related('pagos_compra'),
        id=compra_id,
        empresa=empresa,
    )

    if not compra.puede_registrar_pago:
        messages.error(request, compra.motivo_bloqueo_pago or "No se puede registrar un pago para esta compra.")
        return redirect("ver_compra", empresa_slug=empresa.slug, compra_id=compra.id)

    if request.method == "POST":
        form = PagoCompraForm(request.POST, empresa=empresa)
        form.instance.compra = compra
        if form.is_valid():
            try:
                with transaction.atomic():
                    pago = form.save(commit=False)
                    pago.compra = compra
                    pago.save()
                    registrar_asiento_pago_proveedor(pago)
                messages.success(request, "Pago de compra registrado correctamente.")
                return redirect("ver_compra", empresa_slug=empresa.slug, compra_id=compra.id)
            except ValidationError as exc:
                if hasattr(exc, "message_dict"):
                    for campo, errores in exc.message_dict.items():
                        for error in errores:
                            form.add_error(campo if campo != "__all__" else None, error)
                else:
                    form.add_error(None, exc.messages[0] if exc.messages else "No se pudo registrar el pago.")
        for campo, errores in form.errors.items():
            etiqueta = form.fields[campo].label if campo in form.fields else "Formulario"
            for error in errores:
                messages.error(request, f"{etiqueta}: {error}")
        if not form.errors:
            messages.error(request, "Revisa la informacion del pago antes de continuar.")
    else:
        form = PagoCompraForm(initial={"fecha": timezone.now().date()}, empresa=empresa)

    historial = compra.pagos_compra.all()
    return render(request, "facturacion/registrar_pago_compra.html", {
        "empresa": empresa,
        "compra": compra,
        "form": form,
        "historial": historial,
    })


@login_required
@require_POST
def revertir_pago_compra(request, empresa_slug, compra_id, pago_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    compra = get_object_or_404(CompraInventario, id=compra_id, empresa=empresa)
    pago = get_object_or_404(PagoCompra, id=pago_id, compra=compra)

    with transaction.atomic():
        registrar_reversion_documento(
            empresa=compra.empresa,
            documento_tipo='pago_compra',
            documento_id=pago.id,
            evento_origen='egreso',
            evento_reversion='reversion',
            fecha=timezone.now().date(),
            descripcion=f"Reversion pago compra {compra.numero_compra or compra.id}",
            referencia=pago.referencia or (compra.numero_compra or str(compra.id)),
            origen_modulo='compras',
        )
        pago.delete()

    messages.success(request, "Pago de compra revertido correctamente.")
    return redirect("ver_compra", empresa_slug=empresa.slug, compra_id=compra.id)


@login_required
def ver_compra(request, empresa_slug, compra_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    compra = get_object_or_404(
        CompraInventario.objects.select_related('proveedor').prefetch_related('lineas__producto', 'pagos_compra'),
        id=compra_id,
        empresa=empresa,
    )
    movimientos = MovimientoInventario.objects.filter(compra_documento=compra).select_related('producto')
    pagos = compra.pagos_compra.select_related('comprobante').all()

    return render(request, "facturacion/ver_compra_cxp_premium.html", {
        "empresa": empresa,
        "compra": compra,
        "movimientos": movimientos,
        "pagos": pagos,
    })


@login_required
def kardex_inventario(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    producto_id = request.GET.get("producto")
    tipo = request.GET.get("tipo", "").strip()
    fecha_desde = request.GET.get("fecha_desde", "").strip()
    fecha_hasta = request.GET.get("fecha_hasta", "").strip()

    productos = Producto.objects.filter(
        empresa=empresa,
        controla_inventario=True
    ).order_by('nombre')

    movimientos = MovimientoInventario.objects.filter(
        empresa=empresa
    ).select_related('producto', 'bodega', 'factura', 'nota_credito', 'compra_documento')

    if producto_id:
        try:
            movimientos = movimientos.filter(producto_id=int(producto_id))
        except (TypeError, ValueError):
            producto_id = ""

    if tipo:
        movimientos = movimientos.filter(tipo=tipo)

    if fecha_desde:
        movimientos = movimientos.filter(fecha__date__gte=fecha_desde)

    if fecha_hasta:
        movimientos = movimientos.filter(fecha__date__lte=fecha_hasta)

    resumen = {
        "total_movimientos": movimientos.count(),
        "entradas": movimientos.filter(
            tipo__in=['entrada', 'entrada_compra', 'ajuste_entrada', 'devolucion_nota_credito', 'reversion_factura']
        ).count(),
        "salidas": movimientos.filter(
            tipo__in=['salida_factura', 'ajuste_salida', 'reversion_nota_credito']
        ).count(),
    }

    return render(request, "facturacion/kardex_premium.html", {
        "empresa": empresa,
        "productos": productos,
        "movimientos": movimientos[:150],
        "resumen": resumen,
        "producto_id": str(producto_id or ""),
        "tipo": tipo,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "tipos_movimiento": MovimientoInventario.TIPOS,
    })


@login_required
def cai_facturacion(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    config_avanzada = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)
    uso_documento = request.GET.get("uso_documento", "").strip()
    cais = CAI.objects.filter(empresa=empresa).order_by('-fecha_creacion')
    hoy = timezone.now().date()
    fecha_alerta = hoy + timezone.timedelta(days=15)

    if uso_documento in {"factura", "nota_credito"}:
        cais = cais.filter(uso_documento=uso_documento)

    resumen_base = CAI.objects.filter(empresa=empresa)
    resumen = {
        "total": resumen_base.count(),
        "activos": resumen_base.filter(activo=True).count(),
        "vencidos": resumen_base.filter(fecha_limite__lt=hoy).count(),
        "vigentes": resumen_base.filter(
            fecha_activacion__lte=hoy,
            fecha_limite__gte=hoy,
            activo=True,
        ).count(),
        "por_vencer": resumen_base.filter(
            activo=True,
            fecha_activacion__lte=hoy,
            fecha_limite__gte=hoy,
            fecha_limite__lte=fecha_alerta
        ).count(),
        "agotados": resumen_base.filter(correlativo_actual__gte=F('rango_final')).count(),
    }

    return render(request, "facturacion/cai_premium.html", {
        "empresa": empresa,
        "cais": cais,
        "resumen": resumen,
        "hoy": hoy,
        "fecha_alerta": fecha_alerta,
        "uso_documento": uso_documento,
        "permite_gestion_fiscal_historica": config_avanzada.permite_gestion_fiscal_historica,
    })


@login_required
def crear_cai(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)

    if request.method == "POST":
        form = CAIForm(request.POST)
        if form.is_valid():
            cai = form.save(commit=False)
            cai.empresa = empresa
            cai.save()
            messages.success(request, "CAI creado correctamente.")
            return redirect("cai_facturacion", empresa_slug=empresa.slug)
    else:
        form = CAIForm()

    return render(request, "facturacion/crear_cai.html", {
        "empresa": empresa,
        "form": form,
        "titulo": "Nuevo CAI",
        "texto_boton": "Guardar CAI",
    })


@login_required
def editar_cai(request, empresa_slug, cai_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    cai = get_object_or_404(CAI, id=cai_id, empresa=empresa)

    if request.method == "POST":
        form = CAIForm(request.POST, instance=cai)
        if form.is_valid():
            form.save()
            messages.success(request, "CAI actualizado correctamente.")
            return redirect("cai_facturacion", empresa_slug=empresa.slug)
    else:
        form = CAIForm(instance=cai)

    return render(request, "facturacion/crear_cai.html", {
        "empresa": empresa,
        "form": form,
        "titulo": "Editar CAI",
        "texto_boton": "Guardar Cambios",
    })


@login_required
@require_POST
def eliminar_cai(request, empresa_slug, cai_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    cai = get_object_or_404(CAI, id=cai_id, empresa=empresa)
    config_avanzada = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)

    if not config_avanzada.permite_gestion_fiscal_historica:
        messages.error(request, "Esta empresa no tiene habilitada la correccion fiscal historica.")
    elif cai.factura_set.exists() or cai.notacredito_set.exists():
        messages.error(request, "No se puede eliminar este CAI porque todavia tiene documentos asociados.")
    else:
        cai.delete()
        messages.success(request, "CAI eliminado correctamente.")

    return redirect("cai_facturacion", empresa_slug=empresa.slug)


@login_required
def impuestos_facturacion(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    impuestos = TipoImpuesto.objects.all().order_by('porcentaje', 'nombre')
    resumen = {
        "total": impuestos.count(),
        "activos": impuestos.filter(activo=True).count(),
        "inactivos": impuestos.filter(activo=False).count(),
        "promedio": (impuestos.aggregate(valor=Sum('porcentaje')).get('valor') or Decimal('0.00')) / impuestos.count() if impuestos.count() else Decimal('0.00'),
    }

    return render(request, "facturacion/impuestos_premium.html", {
        "empresa": empresa,
        "impuestos": impuestos,
        "resumen": resumen,
    })


@login_required
def crear_impuesto(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)

    if request.method == "POST":
        form = TipoImpuestoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Impuesto creado correctamente.")
            return redirect("impuestos_facturacion", empresa_slug=empresa.slug)
    else:
        form = TipoImpuestoForm()

    return render(request, "facturacion/crear_impuesto.html", {
        "empresa": empresa,
        "form": form,
        "titulo": "Nuevo Impuesto",
        "texto_boton": "Guardar Impuesto",
    })


@login_required
def editar_impuesto(request, empresa_slug, impuesto_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    impuesto = get_object_or_404(TipoImpuesto, id=impuesto_id)

    if request.method == "POST":
        form = TipoImpuestoForm(request.POST, instance=impuesto)
        if form.is_valid():
            form.save()
            messages.success(request, "Impuesto actualizado correctamente.")
            return redirect("impuestos_facturacion", empresa_slug=empresa.slug)
    else:
        form = TipoImpuestoForm(instance=impuesto)

    return render(request, "facturacion/crear_impuesto.html", {
        "empresa": empresa,
        "form": form,
        "titulo": "Editar Impuesto",
        "texto_boton": "Guardar Cambios",
    })


@login_required
@xframe_options_sameorigin
def crear_producto(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    quick_mode = request.GET.get("modal") == "1" or request.POST.get("quick_mode") == "1"
    config_avanzada = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)
    if config_avanzada.usa_bodegas_internas:
        _asegurar_bodegas_farmaceuticas(empresa)

    if request.method == "POST":
        form = ProductoForm(request.POST, request.FILES, empresa=empresa)
        if form.is_valid():
            try:
                with transaction.atomic():
                    producto = form.save(commit=False)
                    producto.empresa = empresa
                    producto.save()
                    if _empresa_muestra_costo_real_inventario(empresa):
                        _registrar_cambio_costo_real(producto, Decimal("0"), request.user)
                    if getattr(form, "mostrar_perfil_farmaceutico", False):
                        form.guardar_perfil_farmaceutico(producto)
                    _aplicar_distribucion_bodegas_producto(producto, form)
                if quick_mode:
                    return render(request, "facturacion/crear_producto_rapido_modal.html", {
                        "empresa": empresa,
                        "success_payload": {
                            "type": "erp-producto-creado",
                            "producto": {
                                "id": producto.id,
                                "nombre": producto.nombre,
                                "precio": str(producto.precio or "0"),
                                "foto_url": producto.foto.url if producto.foto else "",
                                "impuesto_id": str(producto.impuesto_predeterminado_id or ""),
                                "tipo_item": producto.tipo_item,
                                "unidad_medida": producto.unidad_medida,
                            },
                        },
                    })
                messages.success(request, "Producto creado correctamente.")
                return _redirect_seguro(request, "productos_facturacion", empresa_slug=empresa.slug)
            except ValidationError as exc:
                if hasattr(exc, "message_dict"):
                    for campo, errores in exc.message_dict.items():
                        destino = campo if campo in form.fields else None
                        for error in errores:
                            form.add_error(destino, error)
                else:
                    form.add_error(None, str(exc))
    else:
        form = ProductoForm(empresa=empresa)
        nombre_prefill = (request.GET.get("nombre") or "").strip()
        if nombre_prefill:
            form.fields["nombre"].initial = nombre_prefill

    template_name = "facturacion/crear_producto_rapido_modal.html" if quick_mode else "facturacion/crear_producto.html"
    return render(request, template_name, {
        "empresa": empresa,
        "form": form,
        "quick_mode": quick_mode,
        "mostrar_perfil_farmaceutico": getattr(form, "mostrar_perfil_farmaceutico", False),
        "mostrar_bodega_inicial": getattr(form, "mostrar_bodega_inicial", False),
        "lector_codigo_activo": getattr(form, "lector_codigo_activo", False),
        "distribucion_bodegas": form.distribucion_bodegas() if getattr(form, "mostrar_bodega_inicial", False) else [],
        "next": request.GET.get("next", ""),
        "titulo": "Nuevo Producto",
        "texto_boton": "Guardar Producto",
    })


@login_required
def editar_producto(request, empresa_slug, producto_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    producto = get_object_or_404(Producto, id=producto_id, empresa=empresa, eliminado=False)
    costo_real_anterior = producto.costo_real_inventario

    if request.method == "POST":
        form = ProductoForm(request.POST, request.FILES, instance=producto, empresa=empresa)
        if form.is_valid():
            try:
                with transaction.atomic():
                    producto = form.save()
                    if _empresa_muestra_costo_real_inventario(empresa):
                        _registrar_cambio_costo_real(producto, costo_real_anterior, request.user)
                    if getattr(form, "mostrar_perfil_farmaceutico", False):
                        form.guardar_perfil_farmaceutico(producto)
                    _aplicar_distribucion_bodegas_producto(producto, form)
                messages.success(request, "Producto actualizado correctamente.")
                return redirect("productos_facturacion", empresa_slug=empresa.slug)
            except ValidationError as exc:
                if hasattr(exc, "message_dict"):
                    for campo, errores in exc.message_dict.items():
                        destino = campo if campo in form.fields else None
                        for error in errores:
                            form.add_error(destino, error)
                else:
                    form.add_error(None, str(exc))
    else:
        form = ProductoForm(instance=producto, empresa=empresa)

    lotes_producto = (
        ExistenciaLoteBodega.objects.filter(
            empresa=empresa,
            lote__producto=producto,
        )
        .select_related("bodega", "lote")
        .order_by("lote__fecha_vencimiento", "bodega__tipo", "bodega__nombre")
    )

    return render(request, "facturacion/crear_producto.html", {
        "empresa": empresa,
        "form": form,
        "mostrar_perfil_farmaceutico": getattr(form, "mostrar_perfil_farmaceutico", False),
        "mostrar_bodega_inicial": getattr(form, "mostrar_bodega_inicial", False),
        "lector_codigo_activo": getattr(form, "lector_codigo_activo", False),
        "distribucion_bodegas": form.distribucion_bodegas() if getattr(form, "mostrar_bodega_inicial", False) else [],
        "lotes_producto": lotes_producto,
        "producto_actual": producto,
        "titulo": "Editar Producto",
        "texto_boton": "Guardar Cambios",
    })


@login_required
@require_POST
def eliminar_producto(request, empresa_slug, producto_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    producto = get_object_or_404(Producto, id=producto_id, empresa=empresa, eliminado=False)
    form = EliminarProductoForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Indica un motivo claro para dejar registrada la baja del producto.")
        return redirect("productos_facturacion", empresa_slug=empresa.slug)

    motivo = form.cleaned_data["motivo"].strip()
    with transaction.atomic():
        BitacoraProductoEliminado.objects.create(
            empresa=empresa,
            producto=producto,
            usuario=request.user,
            motivo=motivo,
            nombre_producto=producto.nombre,
            codigo_producto=producto.codigo,
            tipo_item=producto.tipo_item,
            precio=producto.precio,
            controlaba_inventario=producto.controla_inventario,
            stock_al_momento=producto.stock_actual,
        )
        producto.activo = False
        producto.eliminado = True
        producto.fecha_eliminacion = timezone.now()
        producto.eliminado_por = request.user
        producto.motivo_eliminacion = motivo
        producto.save(update_fields=[
            "activo",
            "eliminado",
            "fecha_eliminacion",
            "eliminado_por",
            "motivo_eliminacion",
        ])

    messages.success(request, f"Producto {producto.nombre} eliminado del catalogo operativo y registrado en bitacora.")
    return redirect("productos_facturacion", empresa_slug=empresa.slug)


@login_required
def categorias_farmaceuticas(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    if not _empresa_usa_perfil_farmaceutico(empresa):
        messages.error(request, "Las categorias farmaceuticas no estan activas para esta empresa.")
        return redirect("productos_facturacion", empresa_slug=empresa.slug)

    categoria_id = request.GET.get("editar", "").strip()
    categoria = None
    if categoria_id:
        categoria = get_object_or_404(CategoriaProductoFarmaceutico, id=categoria_id, empresa=empresa)

    form = CategoriaProductoFarmaceuticoForm(request.POST or None, instance=categoria)
    if request.method == "POST" and form.is_valid():
        categoria_guardada = form.save(commit=False)
        categoria_guardada.empresa = empresa
        categoria_guardada.save()
        messages.success(request, "Categoria farmaceutica guardada correctamente.")
        return redirect("categorias_farmaceuticas", empresa_slug=empresa.slug)

    categorias = CategoriaProductoFarmaceutico.objects.filter(empresa=empresa).order_by("nombre")
    return render(request, "facturacion/categorias_farmaceuticas.html", {
        "empresa": empresa,
        "form": form,
        "categorias": categorias,
        "categoria": categoria,
    })


@login_required
def notas_credito_dashboard(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    q = request.GET.get("q", "").strip()
    notas_credito = NotaCredito.objects.filter(empresa=empresa).select_related('cliente', 'factura_origen').order_by('-fecha_creacion')
    if q:
        notas_credito = notas_credito.filter(
            Q(numero_nota__icontains=q) |
            Q(cliente__nombre__icontains=q) |
            Q(factura_origen__numero_factura__icontains=q) |
            Q(motivo__icontains=q)
        )
    resumen = {
        "total": notas_credito.count(),
        "borradores": notas_credito.filter(estado='borrador').count(),
        "emitidas": notas_credito.filter(estado='emitida').count(),
        "monto_total": sum((nota.total for nota in notas_credito), Decimal('0.00')),
    }

    return render(request, "facturacion/notas_credito_premium.html", {
        "empresa": empresa,
        "notas_credito": notas_credito,
        "resumen": resumen,
        "q": q,
        "notas_sugeridas": list(
            NotaCredito.objects.filter(empresa=empresa)
            .select_related('cliente', 'factura_origen')
            .order_by('-fecha_creacion')
            .values_list('numero_nota', flat=True)
        ),
        "clientes_sugeridos": NotaCredito.objects.filter(empresa=empresa).select_related('cliente').values_list('cliente__nombre', flat=True).distinct(),
    })


@login_required
def recibos_dashboard(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    q = request.GET.get("q", "").strip()
    recibos = ReciboPago.objects.filter(empresa=empresa).select_related('cliente', 'factura').order_by('-fecha_creacion')
    if q:
        recibos = recibos.filter(
            Q(numero_recibo__icontains=q) |
            Q(cliente__nombre__icontains=q) |
            Q(referencia__icontains=q) |
            Q(concepto__icontains=q) |
            Q(factura__numero_factura__icontains=q)
        )
    resumen = {
        "total": recibos.count(),
        "monto_total": sum((recibo.monto for recibo in recibos), Decimal('0.00')),
        "efectivo": recibos.filter(metodo='efectivo').count(),
        "transferencia": recibos.filter(metodo='transferencia').count(),
    }

    return render(request, "facturacion/recibos_premium.html", {
        "empresa": empresa,
        "recibos": recibos,
        "resumen": resumen,
        "q": q,
        "recibos_sugeridos": ReciboPago.objects.filter(empresa=empresa).values_list('numero_recibo', flat=True),
        "clientes_sugeridos": ReciboPago.objects.filter(empresa=empresa).select_related('cliente').values_list('cliente__nombre', flat=True).distinct(),
    })


@login_required
def ver_recibo(request, empresa_slug, recibo_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    recibo = get_object_or_404(
        ReciboPago.objects.select_related('cliente', 'factura', 'pago'),
        id=recibo_id,
        empresa=empresa,
    )
    pagos_factura = list(
        recibo.factura.pagos_facturacion.select_related("cuenta_financiera", "cajero").order_by("fecha", "id")
    )
    total_pagos_previos = sum(
        (pago.total_aplicado for pago in pagos_factura if recibo.pago_id and pago.id != recibo.pago_id),
        Decimal("0.00"),
    )

    return render(request, "facturacion/ver_recibo_premium.html", {
        "empresa": empresa,
        "recibo": recibo,
        "pagos_factura": pagos_factura,
        "total_pagos_previos": total_pagos_previos,
    })


@login_required
def editar_recibo(request, empresa_slug, recibo_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    recibo = get_object_or_404(
        ReciboPago.objects.select_related('cliente', 'factura', 'pago'),
        id=recibo_id,
        empresa=empresa,
    )
    form = ReciboPagoForm(request.POST or None, instance=recibo)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, f"Recibo {recibo.numero_recibo} actualizado correctamente.")
        return redirect("ver_recibo", empresa_slug=empresa.slug, recibo_id=recibo.id)

    return render(request, "facturacion/editar_recibo_premium.html", {
        "empresa": empresa,
        "recibo": recibo,
        "form": form,
    })


@login_required
def descargar_recibo_pdf(request, empresa_slug, recibo_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    recibo = get_object_or_404(
        ReciboPago.objects.select_related('cliente', 'factura', 'pago'),
        id=recibo_id,
        empresa=empresa,
    )

    logo_url = _obtener_logo_url(empresa)

    html_string = render_to_string(
        "facturacion/recibo_pdf.html",
        {
            "empresa": empresa,
            "recibo": recibo,
            "logo_url": logo_url,
        }
    )

    html = HTML(string=html_string, base_url=str(settings.BASE_DIR))
    pdf_file = html.write_pdf()

    nombre_archivo = _nombre_archivo_pdf(
        "Recibo",
        recibo.numero_recibo or f"recibo_{recibo.id}",
        recibo.cliente.nombre,
    )

    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return response


@login_required
def egresos_dashboard(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    q = request.GET.get("q", "").strip()
    metodo = request.GET.get("metodo", "").strip()
    egresos = (
        ComprobanteEgresoCompra.objects.filter(empresa=empresa)
        .select_related('compra', 'proveedor', 'pago')
        .order_by('-fecha_creacion')
    )
    if q:
        egresos = egresos.filter(
            Q(numero_comprobante__icontains=q) |
            Q(proveedor_nombre__icontains=q) |
            Q(referencia__icontains=q) |
            Q(compra__numero_compra__icontains=q)
        )
    if metodo in {"efectivo", "transferencia", "tarjeta"}:
        egresos = egresos.filter(metodo=metodo)

    resumen = {
        "total": egresos.count(),
        "monto_total": sum((egreso.monto for egreso in egresos), Decimal('0.00')),
        "efectivo": egresos.filter(metodo='efectivo').count(),
        "transferencia": egresos.filter(metodo='transferencia').count(),
    }

    return render(request, "facturacion/egresos_premium.html", {
        "empresa": empresa,
        "egresos": egresos,
        "resumen": resumen,
        "q": q,
        "metodo": metodo,
    })


@login_required
def ver_comprobante_egreso(request, empresa_slug, comprobante_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    comprobante = get_object_or_404(
        ComprobanteEgresoCompra.objects.select_related('compra', 'proveedor', 'pago'),
        id=comprobante_id,
        empresa=empresa,
    )

    return render(request, "facturacion/ver_comprobante_egreso_premium.html", {
        "empresa": empresa,
        "comprobante": comprobante,
    })


@login_required
def descargar_comprobante_egreso_pdf(request, empresa_slug, comprobante_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    comprobante = get_object_or_404(
        ComprobanteEgresoCompra.objects.select_related('compra', 'proveedor', 'pago'),
        id=comprobante_id,
        empresa=empresa,
    )

    logo_url = _obtener_logo_url(empresa)
    html_string = render_to_string(
        "facturacion/comprobante_egreso_pdf.html",
        {
            "empresa": empresa,
            "comprobante": comprobante,
            "logo_url": logo_url,
        }
    )

    html = HTML(string=html_string, base_url=str(settings.BASE_DIR))
    pdf_file = html.write_pdf()

    nombre_archivo = _nombre_archivo_pdf(
        "Comprobante_Egreso",
        comprobante.numero_comprobante or f"egreso_{comprobante.id}",
        comprobante.proveedor_nombre,
    )

    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return response


@login_required
def detalle_factura_para_nota_credito(request, empresa_slug, factura_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    factura = get_object_or_404(
        Factura.objects.select_related('cliente').prefetch_related('lineas__producto', 'lineas__impuesto'),
        id=factura_id,
        empresa=empresa,
        estado='emitida',
    )

    lineas = []
    for linea in factura.lineas.all():
        lineas.append({
            "producto_id": linea.producto_id,
            "producto_nombre": linea.descripcion_visual,
            "cantidad": str(linea.cantidad),
            "precio_unitario": str(linea.precio_unitario),
            "descuento_porcentaje": str(linea.descuento_porcentaje or 0),
            "comentario": linea.comentario or "",
            "impuesto_id": linea.impuesto_id,
            "impuesto_nombre": linea.impuesto.nombre,
            "linea_total": str(linea.subtotal + linea.impuesto_monto),
        })

    return JsonResponse({
        "factura_id": factura.id,
        "numero_factura": factura.numero_factura or str(factura.id),
        "cliente": factura.cliente.nombre,
        "moneda": factura.moneda,
        "total": str(factura.total),
        "saldo_disponible": str(factura.total_documento_ajustado),
        "lineas": lineas,
    })


@login_required
def crear_nota_credito(request, empresa_slug):

    empresa = get_object_or_404(Empresa, slug=empresa_slug)

    NotaCreditoForm = modelform_factory(
        NotaCredito,
        fields=[
            'factura_origen',
            'fecha_emision',
            'motivo',
            'estado',
        ]
    )

    LineaFormSet = inlineformset_factory(
        NotaCredito,
        LineaNotaCredito,
        fields=[
            'producto',
            'cantidad',
            'precio_unitario',
            'descuento_porcentaje',
            'comentario',
            'impuesto'
        ],
        extra=1,
        can_delete=True
    )

    productos_qs = Producto.objects.filter(empresa=empresa, activo=True).select_related('impuesto_predeterminado')
    impuestos_qs = TipoImpuesto.objects.filter(activo=True)
    facturas_qs = Factura.objects.filter(empresa=empresa, estado='emitida').order_by('-fecha_emision')

    if request.method == "POST":
        form = NotaCreditoForm(request.POST)
        form.fields['factura_origen'].queryset = facturas_qs

        nota_temp = NotaCredito(empresa=empresa)
        formset = LineaFormSet(request.POST, instance=nota_temp)

        for f in formset.forms:
            f.fields['producto'].queryset = productos_qs
            f.fields['impuesto'].queryset = impuestos_qs

        lineas_validas = []

        for f in formset.forms:
            prefix = f.prefix
            producto_raw = (request.POST.get(f"{prefix}-producto") or "").strip()
            cantidad_raw = (request.POST.get(f"{prefix}-cantidad") or "").strip()
            precio_raw = (request.POST.get(f"{prefix}-precio_unitario") or "").strip()
            descuento_raw = (request.POST.get(f"{prefix}-descuento_porcentaje") or "").strip()
            comentario_raw = (request.POST.get(f"{prefix}-comentario") or "").strip()
            impuesto_raw = (request.POST.get(f"{prefix}-impuesto") or "").strip()
            delete_raw = request.POST.get(f"{prefix}-DELETE")

            fila_vacia = (
                not producto_raw and
                not cantidad_raw and
                not precio_raw and
                not descuento_raw and
                not comentario_raw and
                not impuesto_raw
            )

            if delete_raw or fila_vacia:
                continue

            if f.is_valid():
                lineas_validas.append(f)

        if form.is_valid() and lineas_validas:
            try:
                with transaction.atomic():
                    estado_destino = form.cleaned_data['estado']
                    nota = form.save(commit=False)
                    nota.empresa = empresa
                    nota.cliente = nota.factura_origen.cliente
                    nota.vendedor = nota.factura_origen.vendedor
                    nota.moneda = nota.factura_origen.moneda
                    nota.tipo_cambio = nota.factura_origen.tipo_cambio
                    nota.estado = 'borrador'
                    nota.save()
                    formset.instance = nota

                    for f in lineas_validas:
                        linea = f.save(commit=False)
                        linea.nota_credito = nota
                        linea.precio_incluye_impuesto = _precios_incluyen_impuesto(empresa)
                        linea.save()

                    nota.calcular_totales()
                    nota.estado = estado_destino
                    nota.full_clean()
                    nota.save(update_fields=['subtotal', 'impuesto', 'total', 'total_lempiras', 'estado'])
                    if estado_destino == 'emitida':
                        _registrar_entrada_nota_credito(nota)
                        registrar_asiento_nota_credito(nota)
                    nota.factura_origen.actualizar_estado_pago()

                messages.success(request, "Nota de crédito guardada correctamente.")
                return redirect("notas_credito_dashboard", empresa_slug=empresa.slug)
            except (ValueError, ValidationError) as exc:
                messages.error(request, str(exc))
        elif form.is_valid():
            messages.error(request, "Debe agregar al menos una línea válida en la nota de crédito.")

    else:
        form = NotaCreditoForm()
        form.fields['factura_origen'].queryset = facturas_qs
        formset = LineaFormSet()

        for f in formset:
            f.fields['producto'].queryset = productos_qs
            f.fields['impuesto'].queryset = impuestos_qs

    return render(request, "facturacion/crear_nota_credito_premium.html", {
        "empresa": empresa,
        "form": form,
        "formset": formset,
        "productos": productos_qs,
        "impuestos": impuestos_qs,
        "facturas_sugeridas": [
            {
                "id": factura.id,
                "label": f"{factura.numero_factura or factura.id} - {factura.cliente.nombre}"
            }
            for factura in facturas_qs.select_related('cliente')
        ],
        "titulo_documento": "Nota de Crédito",
    })


@login_required
def editar_nota_credito(request, empresa_slug, nota_id):

    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    nota = get_object_or_404(NotaCredito, id=nota_id, empresa=empresa)

    NotaCreditoForm = modelform_factory(
        NotaCredito,
        fields=[
            'factura_origen',
            'fecha_emision',
            'motivo',
            'estado',
        ]
    )

    LineaFormSet = inlineformset_factory(
        NotaCredito,
        LineaNotaCredito,
        fields=[
            'producto',
            'cantidad',
            'precio_unitario',
            'descuento_porcentaje',
            'comentario',
            'impuesto'
        ],
        extra=0,
        can_delete=True
    )

    productos_qs = Producto.objects.filter(empresa=empresa, activo=True).select_related('impuesto_predeterminado')
    impuestos_qs = TipoImpuesto.objects.filter(activo=True)
    facturas_qs = Factura.objects.filter(empresa=empresa, estado='emitida').order_by('-fecha_emision')

    if request.method == "POST":
        estado_original = nota.estado
        form = NotaCreditoForm(request.POST, instance=nota)
        form.fields['factura_origen'].queryset = facturas_qs
        formset = LineaFormSet(request.POST, instance=nota)

        for f in formset.forms:
            f.fields['producto'].queryset = productos_qs
            f.fields['impuesto'].queryset = impuestos_qs

        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    estado_destino = form.cleaned_data['estado']
                    nota = form.save(commit=False)
                    nota.cliente = nota.factura_origen.cliente
                    nota.vendedor = nota.factura_origen.vendedor
                    nota.moneda = nota.factura_origen.moneda
                    nota.tipo_cambio = nota.factura_origen.tipo_cambio
                    nota.estado = 'borrador' if estado_destino == 'emitida' else estado_destino
                    nota.save()
                    lineas_actualizadas = formset.save(commit=False)
                    for linea in formset.deleted_objects:
                        linea.delete()
                    for linea in lineas_actualizadas:
                        if not linea.pk:
                            linea.precio_incluye_impuesto = _precios_incluyen_impuesto(empresa)
                        linea.save()
                    formset.save_m2m()
                    nota.calcular_totales()
                    nota.estado = estado_destino
                    nota.full_clean()
                    nota.save(update_fields=['subtotal', 'impuesto', 'total', 'total_lempiras', 'estado'])
                    if estado_original != 'emitida' and estado_destino == 'emitida':
                        _registrar_entrada_nota_credito(nota)
                        registrar_asiento_nota_credito(nota)
                    elif estado_original == 'emitida' and estado_destino == 'emitida':
                        _reconstruir_nota_credito_emitida(nota)
                    elif estado_original == 'emitida' and estado_destino == 'anulada':
                        _revertir_entrada_nota_credito(nota)
                        registrar_reversion_documento(
                            empresa=nota.empresa,
                            documento_tipo='nota_credito',
                            documento_id=nota.id,
                            evento_origen='emision',
                            evento_reversion='anulacion',
                            fecha=timezone.now().date(),
                            descripcion=f"Reversion nota de credito {nota.numero_nota or nota.id}",
                            referencia=nota.numero_nota or str(nota.id),
                            origen_modulo='facturacion',
                            creado_por=nota.vendedor,
                        )
                    nota.factura_origen.actualizar_estado_pago()

                messages.success(request, "Nota de crédito actualizada correctamente.")
                return redirect("notas_credito_dashboard", empresa_slug=empresa.slug)
            except (ValueError, ValidationError) as exc:
                messages.error(request, str(exc))

    else:
        form = NotaCreditoForm(instance=nota)
        form.fields['factura_origen'].queryset = facturas_qs
        formset = LineaFormSet(instance=nota)

        for f in formset:
            f.fields['producto'].queryset = productos_qs
            f.fields['impuesto'].queryset = impuestos_qs

    return render(request, "facturacion/crear_nota_credito_premium.html", {
        "empresa": empresa,
        "form": form,
        "formset": formset,
        "productos": productos_qs,
        "impuestos": impuestos_qs,
        "facturas_sugeridas": [
            {
                "id": factura.id,
                "label": f"{factura.numero_factura or factura.id} - {factura.cliente.nombre}"
            }
            for factura in facturas_qs.select_related('cliente')
        ],
        "titulo_documento": "Editar Nota de Crédito",
        "nota_credito": nota,
        "modo_edicion": True,
    })


@login_required
def ver_nota_credito(request, empresa_slug, nota_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    nota = get_object_or_404(NotaCredito, id=nota_id, empresa=empresa)
    resumen = nota.resumen_fiscal()

    return render(request, "facturacion/ver_nota_credito_premium.html", {
        "empresa": empresa,
        "nota": nota,
        "resumen": resumen,
    })


@login_required
def descargar_nota_credito_pdf(request, empresa_slug, nota_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    nota = get_object_or_404(NotaCredito, id=nota_id, empresa=empresa)
    resumen = nota.resumen_fiscal()

    logo_url = None

    if empresa.logo:
        logo_path = os.path.join(settings.MEDIA_ROOT, empresa.logo.name)
        logo_path = logo_path.replace("\\", "/")
        logo_url = "file:///" + logo_path

    html_string = render_to_string(
        "facturacion/nota_credito_pdf.html",
        {
            "empresa": empresa,
            "nota": nota,
            "resumen": resumen,
            "logo_url": logo_url,
        }
    )

    html = HTML(string=html_string, base_url=str(settings.BASE_DIR))
    pdf_file = html.write_pdf()

    nombre_archivo = _nombre_archivo_pdf(
        "Nota_Credito",
        nota.numero_nota or f"nota_{nota.id}",
        nota.cliente.nombre,
    )

    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return response


@login_required
@require_POST
def generar_nota_credito_desde_factura(request, empresa_slug, factura_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)

    nota = NotaCredito.objects.create(
        empresa=empresa,
        factura_origen=factura,
        cliente=factura.cliente,
        vendedor=factura.vendedor,
        moneda=factura.moneda,
        tipo_cambio=factura.tipo_cambio,
        fecha_emision=timezone.now().date(),
        motivo=f"Nota de crédito generada desde factura {factura.numero_factura or factura.id}",
        estado='borrador',
    )

    for linea in factura.lineas.all():
        LineaNotaCredito.objects.create(
            nota_credito=nota,
            producto=linea.producto,
            cantidad=linea.cantidad,
            precio_unitario=linea.precio_unitario,
            precio_incluye_impuesto=linea.precio_incluye_impuesto,
            costo_unitario=linea.costo_unitario,
            descuento_porcentaje=linea.descuento_porcentaje,
            comentario=linea.comentario,
            impuesto=linea.impuesto,
        )

    nota.calcular_totales()
    nota.save(update_fields=['subtotal', 'impuesto', 'total', 'total_lempiras'])

    messages.success(request, "Nota de crédito creada como borrador a partir de la factura.")
    return redirect("editar_nota_credito", empresa_slug=empresa.slug, nota_id=nota.id)


@login_required
@require_POST
def anular_nota_credito(request, empresa_slug, nota_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    nota = get_object_or_404(NotaCredito, id=nota_id, empresa=empresa)

    if nota.estado == 'anulada':
        messages.info(request, "La nota de crédito ya estaba anulada.")
    else:
        if nota.estado == 'emitida':
            _revertir_entrada_nota_credito(nota)
            registrar_reversion_documento(
                empresa=nota.empresa,
                documento_tipo='nota_credito',
                documento_id=nota.id,
                evento_origen='emision',
                evento_reversion='anulacion',
                fecha=timezone.now().date(),
                descripcion=f"Reversion nota de credito {nota.numero_nota or nota.id}",
                referencia=nota.numero_nota or str(nota.id),
                origen_modulo='facturacion',
                creado_por=nota.vendedor,
            )
        nota.estado = 'anulada'
        nota.save(update_fields=['estado'])
        nota.factura_origen.actualizar_estado_pago()
        messages.success(request, "Nota de crédito anulada correctamente.")

    return redirect("ver_nota_credito", empresa_slug=empresa.slug, nota_id=nota.id)


# =====================================================
# CREAR FACTURA
# =====================================================

@login_required
def crear_factura(request, empresa_slug):

    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    config_avanzada = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)
    campos_factura = [
        'cliente',
        'fecha_emision',
        'fecha_vencimiento',
        'vendedor',
        'tipo_cambio',
        'moneda',
        'estado',
        'orden_compra_exenta',
        'registro_exonerado',
        'registro_sag',
    ]
    if config_avanzada.permite_gestion_fiscal_historica:
        campos_factura.insert(6, 'numero_factura')

    FacturaForm = modelform_factory(
        Factura,
        fields=campos_factura
    )

    LineaFormSet = inlineformset_factory(
        Factura,
        LineaFactura,
        fields=[
            'producto',
            'descripcion_manual',
            'cantidad',
            'precio_unitario',
            'descuento_porcentaje',
            'comentario',
            'impuesto'
        ],
        extra=1,
        can_delete=True
    )

    productos_qs = Producto.objects.filter(empresa=empresa, activo=True).select_related('impuesto_predeterminado')
    impuestos_qs = TipoImpuesto.objects.filter(activo=True)
    clientes_qs = Cliente.objects.filter(empresa=empresa)
    vendedores_qs = Usuario.objects.filter(Q(empresa=empresa) | Q(empresas_acceso=empresa)).distinct()

    def obtener_prefijo_manual(fecha_referencia=None, numero_actual=""):
        numero_actual = (numero_actual or "").strip()
        if numero_actual:
            return numero_actual[:-3], numero_actual[-3:]

        if not config_avanzada.permite_gestion_fiscal_historica:
            return "", ""

        fecha_referencia = fecha_referencia or timezone.localdate()
        factura_tmp = Factura(empresa=empresa, fecha_emision=fecha_referencia)
        cai = factura_tmp._obtener_queryset_cai_factura(
            fecha_referencia=factura_tmp._obtener_fecha_referencia_cai()
        ).first()
        if not cai:
            return "", ""
        prefijo = (
            f"{cai.establecimiento}-"
            f"{cai.punto_emision}-"
            f"{cai.tipo_documento}-"
            f"{str(cai.rango_inicial).zfill(8)[:5]}"
        )
        return prefijo, ""

    def preparar_post_factura(post_data):
        if not config_avanzada.permite_gestion_fiscal_historica:
            return post_data

        post_data = post_data.copy()
        fecha_referencia = _parsear_fecha_latam(post_data.get("fecha_emision")) or timezone.localdate()
        prefijo, _ = obtener_prefijo_manual(
            fecha_referencia=fecha_referencia,
            numero_actual=post_data.get("numero_factura"),
        )
        sufijo = (post_data.get("numero_factura_sufijo") or "").strip()
        if sufijo:
            post_data["numero_factura"] = f"{prefijo}{sufijo.zfill(3)}"
        else:
            post_data["numero_factura"] = ""
        return post_data

    def preparar_factura_form(form, numero_prefijo="", numero_sufijo=""):
        configurar_campo_fecha(form.fields['fecha_emision'])
        configurar_campo_fecha(form.fields['fecha_vencimiento'])
        form.fields['fecha_vencimiento'].required = False
        if 'tipo_cambio' in form.fields:
            form.fields['tipo_cambio'] = forms.DecimalField(
                max_digits=10,
                decimal_places=4,
                required=False,
                initial=form.initial.get('tipo_cambio', getattr(form.instance, 'tipo_cambio', Decimal("1.0000"))),
                widget=forms.NumberInput(attrs={
                    'step': '0.0001',
                    'min': '0.0001',
                    'placeholder': '1.0000',
                    'inputmode': 'decimal',
                }),
                label=form.fields['tipo_cambio'].label,
            )
        if 'estado' in form.fields:
            form.fields['estado'].choices = [
                ('borrador', 'Borrador'),
                ('emitida', 'Emitida'),
            ]
        if 'numero_factura' in form.fields:
            form.fields['numero_factura'].required = False
            if config_avanzada.permite_gestion_fiscal_historica:
                form.fields['numero_factura'].widget = forms.HiddenInput()
                form.fields['numero_factura_sufijo'] = forms.RegexField(
                    regex=r'^\d{0,3}$',
                    required=False,
                    label='Ultimos 3 digitos',
                    error_messages={"invalid": "Ingresa solo los ultimos 3 digitos."},
                    widget=forms.TextInput(attrs={
                        "placeholder": "461",
                        "maxlength": "3",
                        "autocomplete": "off",
                        "inputmode": "numeric",
                    }),
                )
                if not form.is_bound:
                    form.initial['numero_factura_sufijo'] = numero_sufijo
            else:
                form.fields['numero_factura'].widget.attrs.update({
                    'placeholder': '000-000-00-00000000',
                    'autocomplete': 'off',
                })
                form.fields['numero_factura'].help_text = (
                    "Opcional. Si lo completas, el ERP validara que pertenezca al CAI activo para la fecha de esta factura."
                )
        return form

    if request.method == "POST":

        post_data = preparar_post_factura(request.POST)
        prefijo_manual, sufijo_manual = obtener_prefijo_manual(
            fecha_referencia=_parsear_fecha_latam(post_data.get("fecha_emision")),
            numero_actual=post_data.get("numero_factura"),
        )
        form = preparar_factura_form(
            FacturaForm(post_data),
            numero_prefijo=prefijo_manual,
            numero_sufijo=sufijo_manual,
        )
        form.fields['cliente'].queryset = clientes_qs
        form.fields['vendedor'].queryset = vendedores_qs

        factura_temp = Factura(empresa=empresa)
        formset = LineaFormSet(request.POST, instance=factura_temp)

        for f in formset.forms:
            f.fields['producto'].queryset = productos_qs
            f.fields['impuesto'].queryset = impuestos_qs
            f.fields['descripcion_manual'].required = False
            f.fields['descripcion_manual'].widget = forms.HiddenInput()

        lineas_validas = []

        for f in formset.forms:
            prefix = f.prefix

            producto_raw = (request.POST.get(f"{prefix}-producto") or "").strip()
            descripcion_manual_raw = (request.POST.get(f"{prefix}-descripcion_manual") or "").strip()
            cantidad_raw = (request.POST.get(f"{prefix}-cantidad") or "").strip()
            precio_raw = (request.POST.get(f"{prefix}-precio_unitario") or "").strip()
            descuento_raw = (request.POST.get(f"{prefix}-descuento_porcentaje") or "").strip()
            comentario_raw = (request.POST.get(f"{prefix}-comentario") or "").strip()
            impuesto_raw = (request.POST.get(f"{prefix}-impuesto") or "").strip()
            delete_raw = request.POST.get(f"{prefix}-DELETE")

            fila_vacia = (
                not producto_raw and
                not descripcion_manual_raw and
                not cantidad_raw and
                not precio_raw and
                not descuento_raw and
                not comentario_raw and
                not impuesto_raw
            )

            if delete_raw:
                continue

            if fila_vacia:
                continue

            if not f.is_valid():
                continue
            else:
                lineas_validas.append(f)

        if form.is_valid() and lineas_validas:

            try:
                with transaction.atomic():
                    estado_emitida = form.cleaned_data.get('estado') == 'emitida'
                    factura = form.save(commit=False)
                    factura.empresa = empresa
                    factura.save()
                    formset.instance = factura

                    lineas_guardadas = []
                    for f in lineas_validas:
                        linea = f.save(commit=False)
                        linea.factura = factura
                        linea.precio_incluye_impuesto = _precios_incluyen_impuesto(empresa)
                        linea.save()
                        lineas_guardadas.append(linea)

                    if estado_emitida:
                        _validar_stock_disponible_para_lineas(lineas_guardadas)

                    _actualizar_totales_factura(factura)

                    if estado_emitida:
                        _registrar_salida_factura(factura)
                        registrar_asiento_factura_emitida(factura)

                messages.success(request, "Factura guardada correctamente.")
                return redirect(f"{redirect('ver_factura', empresa_slug=empresa.slug, factura_id=factura.id).url}?nueva=1")
            except ValidationError as exc:
                if hasattr(exc, "message_dict"):
                    for campo, errores in exc.message_dict.items():
                        destino = campo if campo in form.fields else None
                        for error in errores:
                            form.add_error(destino, error)
                else:
                    form.add_error(None, str(exc))
            except ValueError as exc:
                messages.error(request, str(exc))
        elif form.is_valid():
            messages.error(request, "Debe agregar al menos una línea de producto válida.")

    else:
        prefijo_manual, sufijo_manual = obtener_prefijo_manual()
        form = preparar_factura_form(
            FacturaForm(),
            numero_prefijo=prefijo_manual,
            numero_sufijo=sufijo_manual,
        )
        form.fields['cliente'].queryset = clientes_qs
        form.fields['vendedor'].queryset = vendedores_qs

        formset = LineaFormSet()

        for f in formset:
            f.fields['producto'].queryset = productos_qs
            f.fields['impuesto'].queryset = impuestos_qs
            f.fields['descripcion_manual'].required = False
            f.fields['descripcion_manual'].widget = forms.HiddenInput()

    return render(request, "facturacion/crear_factura_premium.html", {
        "empresa": empresa,
        "form": form,
        "formset": formset,
        "productos": productos_qs,
        "impuestos": impuestos_qs,
        "clientes_sugeridos": clientes_qs.order_by('nombre').values_list('nombre', flat=True).distinct(),
        "clientes_preview": list(
            clientes_qs.order_by('nombre').values(
                'id',
                'nombre',
                'rtn',
                'direccion',
            )
        ),
        "permite_gestion_fiscal_historica": config_avanzada.permite_gestion_fiscal_historica,
        "numero_factura_prefijo_manual": prefijo_manual,
        "prefijo_factura_manual_url": reverse("prefijo_factura_manual", args=[empresa.slug]),
        "precios_incluyen_impuesto": _precios_incluyen_impuesto(empresa),
    })

# =====================================================
# EDITAR FACTURA
# =====================================================

@login_required
def corregir_numero_factura(request, empresa_slug, factura_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)
    config_avanzada = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)

    if not config_avanzada.permite_gestion_fiscal_historica:
        messages.error(request, "La correccion fiscal historica no esta habilitada para esta empresa.")
        return redirect("ver_factura", empresa_slug=empresa.slug, factura_id=factura.id)

    if factura.estado != "emitida":
        messages.info(request, "Los borradores se corrigen desde la edicion normal de la factura.")
        return redirect("editar_factura", empresa_slug=empresa.slug, factura_id=factura.id)

    form = CorreccionNumeroFacturaForm(
        request.POST or None,
        initial={"numero_factura": factura.numero_factura},
    )

    if request.method == "POST" and form.is_valid():
        try:
            with transaction.atomic():
                factura_bloqueada = Factura.objects.select_for_update().get(
                    id=factura.id,
                    empresa=empresa,
                )
                numero_anterior = (factura_bloqueada.numero_factura or "").strip()
                numero_nuevo = form.cleaned_data["numero_factura"]

                if numero_nuevo == numero_anterior:
                    form.add_error("numero_factura", "El nuevo numero es igual al numero actual.")
                else:
                    cai_anterior = factura_bloqueada.cai_numero_historico
                    factura_bloqueada.numero_factura = numero_nuevo
                    factura_bloqueada.save(update_fields=[
                        "numero_factura",
                        "cai",
                        "cai_numero",
                        "cai_establecimiento",
                        "cai_punto_emision",
                        "cai_tipo_documento",
                        "cai_rango_inicial",
                        "cai_rango_final",
                        "cai_fecha_limite",
                    ])

                    AsientoContable.objects.filter(
                        empresa=empresa,
                        documento_tipo="factura",
                        documento_id=factura_bloqueada.id,
                        evento="emision",
                    ).update(
                        referencia=numero_nuevo,
                        descripcion=f"Emision factura {numero_nuevo}",
                    )
                    MovimientoInventario.objects.filter(
                        empresa=empresa,
                        factura=factura_bloqueada,
                    ).update(referencia=numero_nuevo)

                    CorreccionNumeroFactura.objects.create(
                        empresa=empresa,
                        factura=factura_bloqueada,
                        numero_anterior=numero_anterior,
                        numero_nuevo=numero_nuevo,
                        cai_anterior=cai_anterior,
                        cai_nuevo=factura_bloqueada.cai_numero_historico,
                        motivo=form.cleaned_data["motivo"],
                        realizado_por=request.user,
                    )

                    messages.success(
                        request,
                        f"Numero fiscal corregido de {numero_anterior} a {numero_nuevo}. "
                        "Los pagos, totales y recibos no fueron modificados.",
                    )
                    return redirect(
                        "ver_factura",
                        empresa_slug=empresa.slug,
                        factura_id=factura_bloqueada.id,
                    )
        except ValidationError as exc:
            if hasattr(exc, "message_dict"):
                for error in exc.message_dict.get("numero_factura", []):
                    form.add_error("numero_factura", error)
                for campo, errores in exc.message_dict.items():
                    if campo != "numero_factura":
                        for error in errores:
                            form.add_error(None, error)
            else:
                form.add_error("numero_factura", str(exc))
        except ValueError as exc:
            form.add_error("numero_factura", str(exc))

    return render(request, "facturacion/corregir_numero_factura.html", {
        "empresa": empresa,
        "factura": factura,
        "form": form,
    })


@login_required
def editar_factura(request, empresa_slug, factura_id):

    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)
    config_avanzada = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)

    if _factura_bloqueada_para_edicion(factura):
        messages.error(
            request,
            f"No se puede editar esta factura emitida. {factura.motivo_bloqueo_edicion}"
        )
        return redirect("ver_factura", empresa_slug=empresa.slug, factura_id=factura.id)

    campos_factura = [
        'cliente',
        'fecha_emision',
        'fecha_vencimiento',
        'vendedor',
        'tipo_cambio',
        'moneda',
        'estado',
        'orden_compra_exenta',
        'registro_exonerado',
        'registro_sag',
    ]
    if config_avanzada.permite_gestion_fiscal_historica:
        campos_factura.insert(6, 'numero_factura')

    FacturaForm = modelform_factory(Factura, fields=campos_factura)

    LineaFormSet = inlineformset_factory(
        Factura,
        LineaFactura,
        fields=[
            'producto',
            'descripcion_manual',
            'cantidad',
            'precio_unitario',
            'descuento_porcentaje',
            'comentario',
            'impuesto'
        ],
        extra=0,
        can_delete=True
    )

    productos_qs = Producto.objects.filter(empresa=empresa, activo=True).select_related('impuesto_predeterminado')
    impuestos_qs = TipoImpuesto.objects.filter(activo=True)
    clientes_qs = Cliente.objects.filter(empresa=empresa)
    vendedores_qs = Usuario.objects.filter(Q(empresa=empresa) | Q(empresas_acceso=empresa)).distinct()

    def obtener_prefijo_manual(fecha_referencia=None, numero_actual=""):
        numero_actual = (numero_actual or "").strip()
        if numero_actual:
            return numero_actual[:-3], numero_actual[-3:]

        if not config_avanzada.permite_gestion_fiscal_historica:
            return "", ""

        fecha_referencia = fecha_referencia or factura.fecha_emision or timezone.localdate()
        factura_tmp = Factura(empresa=empresa, fecha_emision=fecha_referencia)
        cai = factura_tmp._obtener_queryset_cai_factura(
            fecha_referencia=factura_tmp._obtener_fecha_referencia_cai()
        ).first()
        if not cai:
            return "", ""
        prefijo = (
            f"{cai.establecimiento}-"
            f"{cai.punto_emision}-"
            f"{cai.tipo_documento}-"
            f"{str(cai.rango_inicial).zfill(8)[:5]}"
        )
        return prefijo, ""

    def preparar_post_factura(post_data):
        if not config_avanzada.permite_gestion_fiscal_historica:
            return post_data

        post_data = post_data.copy()
        fecha_referencia = _parsear_fecha_latam(post_data.get("fecha_emision")) or factura.fecha_emision or timezone.localdate()
        prefijo, _ = obtener_prefijo_manual(
            fecha_referencia=fecha_referencia,
            numero_actual=post_data.get("numero_factura"),
        )
        sufijo = (post_data.get("numero_factura_sufijo") or "").strip()
        if sufijo:
            post_data["numero_factura"] = f"{prefijo}{sufijo.zfill(3)}"
        else:
            post_data["numero_factura"] = ""
        return post_data

    def preparar_factura_form(form, numero_prefijo="", numero_sufijo=""):
        configurar_campo_fecha(form.fields['fecha_emision'])
        configurar_campo_fecha(form.fields['fecha_vencimiento'])
        form.fields['fecha_vencimiento'].required = False
        if 'tipo_cambio' in form.fields:
            form.fields['tipo_cambio'] = forms.DecimalField(
                max_digits=10,
                decimal_places=4,
                required=False,
                initial=form.initial.get('tipo_cambio', getattr(form.instance, 'tipo_cambio', Decimal("1.0000"))),
                widget=forms.NumberInput(attrs={
                    'step': '0.0001',
                    'min': '0.0001',
                    'placeholder': '1.0000',
                    'inputmode': 'decimal',
                }),
                label=form.fields['tipo_cambio'].label,
            )
        if 'estado' in form.fields:
            form.fields['estado'].choices = [
                ('borrador', 'Borrador'),
                ('emitida', 'Emitida'),
            ]
        if 'numero_factura' in form.fields:
            form.fields['numero_factura'].required = False
            if config_avanzada.permite_gestion_fiscal_historica:
                form.fields['numero_factura'].widget = forms.HiddenInput()
                form.fields['numero_factura_sufijo'] = forms.RegexField(
                    regex=r'^\d{0,3}$',
                    required=False,
                    label='Ultimos 3 digitos',
                    error_messages={"invalid": "Ingresa solo los ultimos 3 digitos."},
                    widget=forms.TextInput(attrs={
                        "placeholder": "461",
                        "maxlength": "3",
                        "autocomplete": "off",
                        "inputmode": "numeric",
                    }),
                )
                if not form.is_bound:
                    form.initial['numero_factura_sufijo'] = numero_sufijo
            else:
                form.fields['numero_factura'].widget.attrs.update({
                    'placeholder': '000-000-00-00000000',
                    'autocomplete': 'off',
                })
                form.fields['numero_factura'].help_text = (
                    "Opcional. Si lo completas, el ERP validara que pertenezca al CAI activo para la fecha de esta factura."
                )
        form.fields['motivo_auditoria'] = forms.CharField(
            required=True,
            min_length=8,
            label='Motivo de la modificacion',
            help_text='Explica brevemente por que se modifica esta factura. Quedara en la bitacora permanente.',
            widget=forms.Textarea(attrs={
                'rows': 2,
                'placeholder': 'Ejemplo: Correccion solicitada por el cliente.',
            }),
        )
        return form

    if request.method == "POST":
        estado_original = factura.estado

        post_data = preparar_post_factura(request.POST)
        prefijo_manual, sufijo_manual = obtener_prefijo_manual(
            fecha_referencia=_parsear_fecha_latam(post_data.get("fecha_emision")) or factura.fecha_emision,
            numero_actual=post_data.get("numero_factura"),
        )
        form = preparar_factura_form(
            FacturaForm(post_data, instance=factura),
            numero_prefijo=prefijo_manual,
            numero_sufijo=sufijo_manual,
        )
        form.fields['cliente'].queryset = clientes_qs
        form.fields['vendedor'].queryset = vendedores_qs

        formset = LineaFormSet(request.POST, instance=factura)

        for f in formset.forms:
            f.fields['producto'].queryset = productos_qs
            f.fields['impuesto'].queryset = impuestos_qs
            f.fields['descripcion_manual'].required = False
            f.fields['descripcion_manual'].widget = forms.HiddenInput()

        if form.is_valid() and formset.is_valid():

            try:
                with transaction.atomic():
                    factura = form.save()
                    lineas_actualizadas = formset.save(commit=False)
                    for linea in formset.deleted_objects:
                        linea.delete()
                    for linea in lineas_actualizadas:
                        if not linea.pk:
                            linea.precio_incluye_impuesto = _precios_incluyen_impuesto(empresa)
                        linea.save()
                    formset.save_m2m()

                    if estado_original != 'emitida' and factura.estado == 'emitida':
                        _validar_stock_disponible_para_lineas(
                            factura.lineas.select_related('producto').all()
                        )

                    _actualizar_totales_factura(factura)

                    if estado_original != 'emitida' and factura.estado == 'emitida':
                        _registrar_salida_factura(factura)
                        registrar_asiento_factura_emitida(factura)
                    elif estado_original == 'emitida' and factura.estado == 'anulada':
                        _revertir_salida_factura(factura)
                        registrar_reversion_documento(
                            empresa=factura.empresa,
                            documento_tipo='factura',
                            documento_id=factura.id,
                            evento_origen='emision',
                            evento_reversion='anulacion',
                            fecha=timezone.now().date(),
                            descripcion=f"Reversion factura {factura.numero_factura or factura.id}",
                            referencia=factura.numero_factura or str(factura.id),
                            origen_modulo='facturacion',
                            creado_por=factura.vendedor,
                        )

                messages.success(request, "Factura actualizada correctamente.")
                return redirect("facturas_dashboard", empresa_slug=empresa.slug)
            except ValidationError as exc:
                if hasattr(exc, "message_dict"):
                    for campo, errores in exc.message_dict.items():
                        destino = campo if campo in form.fields else None
                        for error in errores:
                            form.add_error(destino, error)
                else:
                    form.add_error(None, str(exc))
            except ValueError as exc:
                messages.error(request, str(exc))

    else:
        prefijo_manual, sufijo_manual = obtener_prefijo_manual(numero_actual=factura.numero_factura)
        form = preparar_factura_form(
            FacturaForm(instance=factura),
            numero_prefijo=prefijo_manual,
            numero_sufijo=sufijo_manual,
        )
        form.fields['cliente'].queryset = clientes_qs
        form.fields['vendedor'].queryset = vendedores_qs

        formset = LineaFormSet(instance=factura)

        for f in formset:
            f.fields['producto'].queryset = productos_qs
            f.fields['impuesto'].queryset = impuestos_qs
            f.fields['descripcion_manual'].required = False
            f.fields['descripcion_manual'].widget = forms.HiddenInput()

    return render(request, "facturacion/crear_factura_premium.html", {
        "empresa": empresa,
        "form": form,
        "formset": formset,
        "productos": productos_qs,
        "impuestos": impuestos_qs,
        "clientes_sugeridos": clientes_qs.order_by('nombre').values_list('nombre', flat=True).distinct(),
        "clientes_preview": list(
            clientes_qs.order_by('nombre').values(
                'id',
                'nombre',
                'rtn',
                'direccion',
            )
        ),
        "modo_edicion": True,
        "factura": factura,
        "permite_gestion_fiscal_historica": config_avanzada.permite_gestion_fiscal_historica,
        "numero_factura_prefijo_manual": prefijo_manual,
        "prefijo_factura_manual_url": reverse("prefijo_factura_manual", args=[empresa.slug]),
        "precios_incluyen_impuesto": _precios_incluyen_impuesto(empresa),
    })

# =====================================================
# REGISTRAR PAGO
# =====================================================

@login_required
@xframe_options_sameorigin
def registrar_pago(request, empresa_slug, factura_id):

    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)
    es_modal = request.GET.get("modal") == "1"
    template_name = "facturacion/registrar_pago_modal.html" if es_modal else "facturacion/registrar_pago.html"
    config_avanzada = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)
    cuentas_financieras = _cuentas_financieras_activas_para_pago(empresa)
    cajas = cuentas_financieras.filter(tipo='caja')
    bancos = cuentas_financieras.filter(tipo='banco')
    tarjetas = cuentas_financieras.filter(tipo='tarjeta_credito')
    cuentas_tarjeta = tarjetas if tarjetas.exists() else cuentas_financieras

    def _contexto_pago(form_data=None):
        return {
            "factura": factura,
            "empresa": empresa,
            "cuentas_financieras": cuentas_financieras,
            "cajas": cajas,
            "bancos": bancos,
            "cuentas_tarjeta": cuentas_tarjeta,
            "usa_pagos_mixtos": config_avanzada.usa_pagos_mixtos,
            "subtotal_pendiente": factura.subtotal_pendiente_cobro,
            "impuesto_pendiente": factura.impuesto_pendiente_cobro,
            "today": timezone.now().date(),
            "form_data": form_data or _form_data_pago_por_defecto(cuentas_financieras),
            "es_modal": es_modal,
        }

    def _respuesta_modal_exito(mensaje):
        if not es_modal:
            return None
        mensaje_seguro = json.dumps(mensaje)
        html = f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="utf-8">
    <title>Pago registrado</title>
    <style>
        body {{
            margin: 0;
            font-family: "Segoe UI", Roboto, Arial, sans-serif;
            background: #f5f8fc;
            color: #102338;
            display: flex;
            align-items: center;
            justify-content: center;
            min-height: 100vh;
        }}
        .card {{
            max-width: 460px;
            margin: 24px;
            padding: 28px;
            border-radius: 18px;
            background: #ffffff;
            border: 1px solid #d9e3f0;
            box-shadow: 0 16px 50px rgba(12, 28, 48, 0.12);
            text-align: center;
        }}
        h2 {{ margin: 0 0 10px; font-size: 24px; }}
        p {{ margin: 0; line-height: 1.6; color: #51657d; }}
    </style>
</head>
<body>
    <div class="card">
        <h2>Pago registrado</h2>
        <p>{mensaje}</p>
    </div>
    <script>
        if (window.parent && window.parent !== window) {{
            window.parent.postMessage({{
                type: "erp-pago-guardado",
                facturaId: "{factura.id}",
                mensaje: {mensaje_seguro}
            }}, "*");
        }}
    </script>
</body>
</html>
        """.strip()
        return HttpResponse(html)

    if factura.estado == 'anulada':
        messages.error(request, "No se pueden registrar pagos en una factura anulada.")
        return redirect("ver_factura", empresa_slug=empresa.slug, factura_id=factura.id)

    if factura.saldo_pendiente <= 0:
        messages.info(request, "Esta factura no tiene saldo pendiente.")
        return redirect("ver_factura", empresa_slug=empresa.slug, factura_id=factura.id)

    if request.method == "POST":

        monto = request.POST.get("monto", "").strip()
        metodo = request.POST.get("metodo")
        referencia = request.POST.get("referencia", "").strip()
        fecha_pago = request.POST.get("fecha")
        cuenta_financiera_id = request.POST.get("cuenta_financiera")
        retencion_isr_raw = request.POST.get("retencion_isr", "").strip()
        retencion_isv_raw = request.POST.get("retencion_isv", "").strip()
        separar_isv = request.POST.get("separar_isv") == "on"
        cuenta_financiera_impuesto_id = request.POST.get("cuenta_financiera_impuesto")
        error = None

        def _parsear_decimal(valor, etiqueta):
            if not valor:
                return Decimal('0.00')
            try:
                return Decimal(valor)
            except InvalidOperation as exc:
                raise ValidationError(f"El valor de {etiqueta} no es valido.") from exc

        if config_avanzada.usa_pagos_mixtos:
            try:
                fecha_convertida = datetime.strptime(fecha_pago, "%Y-%m-%d").date() if fecha_pago else timezone.now().date()
            except ValueError:
                fecha_convertida = timezone.now().date()
                error = "Ingrese una fecha valida."

            pagos_solicitados = [
                ("efectivo", "Efectivo", request.POST.get("monto_efectivo", "").strip(), request.POST.get("cuenta_efectivo"), request.POST.get("referencia_efectivo", "").strip()),
                ("tarjeta", "Tarjeta", request.POST.get("monto_tarjeta", "").strip(), request.POST.get("cuenta_tarjeta"), request.POST.get("referencia_tarjeta", "").strip()),
                ("transferencia", "Transferencia", request.POST.get("monto_transferencia", "").strip(), request.POST.get("cuenta_transferencia"), request.POST.get("referencia_transferencia", "").strip()),
            ]
            pagos_validos = []
            total_pago = Decimal('0.00')
            retencion_isr = Decimal('0.00')
            retencion_isv = Decimal('0.00')

            if not error:
                try:
                    retencion_isr = _parsear_decimal(retencion_isr_raw, "retencion ISR")
                    retencion_isv = _parsear_decimal(retencion_isv_raw, "retencion ISV")
                except ValidationError as exc:
                    error = "; ".join(exc.messages)

            if not error:
                for metodo_mixto, etiqueta, monto_raw, cuenta_id, referencia_mixta in pagos_solicitados:
                    if not monto_raw:
                        continue
                    try:
                        monto_decimal = Decimal(monto_raw)
                    except InvalidOperation:
                        error = f"El monto de {etiqueta} no es valido."
                        break
                    if monto_decimal <= 0:
                        error = f"El monto de {etiqueta} debe ser mayor que cero."
                        break
                    cuenta_mixta = cuentas_financieras.filter(id=cuenta_id).first() or _cuenta_financiera_por_defecto(cuentas_financieras, metodo_mixto)
                    if not cuenta_mixta:
                        error = f"Selecciona la cuenta financiera para {etiqueta}."
                        break
                    total_pago += monto_decimal
                    pagos_validos.append((metodo_mixto, monto_decimal, referencia_mixta, cuenta_mixta))

            if not error:
                if not pagos_validos and retencion_isr <= 0 and retencion_isv <= 0:
                    error = "Ingresa al menos un monto o una retencion para registrar el pago."
                elif not pagos_validos:
                    error = "Con pagos mixtos debes registrar al menos un cobro recibido. Si solo deseas aplicar retenciones, usa monto 0.00 en el flujo simple."
                elif retencion_isr < 0 or retencion_isv < 0:
                    error = "Las retenciones no pueden ser negativas."
                elif total_pago + retencion_isr + retencion_isv > factura.saldo_pendiente:
                    error = "La suma aplicada no puede ser mayor que el saldo pendiente."
                else:
                    recibos = []
                    try:
                        with transaction.atomic():
                            cuenta_financiera_impuesto = cuentas_financieras.filter(id=cuenta_financiera_impuesto_id).first() if cuenta_financiera_impuesto_id else None
                            for indice, (metodo_mixto, monto_decimal, referencia_mixta, cuenta_mixta) in enumerate(pagos_validos, start=1):
                                pago = PagoFactura.objects.create(
                                    factura=factura,
                                    monto=monto_decimal,
                                    retencion_isr=retencion_isr if indice == 1 else Decimal('0.00'),
                                    retencion_isv=retencion_isv if indice == 1 else Decimal('0.00'),
                                    separar_isv=separar_isv,
                                    cuenta_financiera_impuesto=cuenta_financiera_impuesto,
                                    metodo=metodo_mixto,
                                    referencia=referencia_mixta,
                                    cuenta_financiera=cuenta_mixta,
                                    cajero=request.user,
                                    fecha=fecha_convertida
                                )
                                registrar_asiento_pago_cliente(pago)
                                if hasattr(pago, 'recibo'):
                                    recibos.append(pago.recibo.numero_recibo)
                        messages.success(
                            request,
                            f"Pago mixto registrado correctamente. Cobro recibido: L. {total_pago:.2f}. Aplicado total: L. {(total_pago + retencion_isr + retencion_isv):.2f}. Recibos: {', '.join(recibos) or 'sin numero'}."
                        )
                        respuesta_modal = _respuesta_modal_exito(
                            f"Cobro recibido: L. {total_pago:.2f}. Aplicado total: L. {(total_pago + retencion_isr + retencion_isv):.2f}."
                        )
                        if respuesta_modal:
                            return respuesta_modal
                        return redirect("ver_factura", empresa_slug=empresa.slug, factura_id=factura.id)
                    except ValidationError as exc:
                        error = "; ".join(exc.messages)
                    except Exception as exc:
                        logger.exception("Error registrando pago mixto en factura %s", factura.id)
                        error = f"No se pudo completar el cobro: {exc}"

            messages.error(request, error)
            return render(request, template_name, _contexto_pago(request.POST))

        try:
            monto_decimal = _parsear_decimal(monto, "monto recibido")
            retencion_isr = _parsear_decimal(retencion_isr_raw, "retencion ISR")
            retencion_isv = _parsear_decimal(retencion_isv_raw, "retencion ISV")

            if monto_decimal < 0:
                error = "El monto recibido no puede ser negativo."
            elif retencion_isr < 0 or retencion_isv < 0:
                error = "Las retenciones no pueden ser negativas."
            elif monto_decimal + retencion_isr + retencion_isv <= 0:
                error = "Debes registrar un monto o una retencion mayor que cero."
            elif monto_decimal + retencion_isr + retencion_isv > factura.saldo_pendiente:
                error = "El pago aplicado no puede ser mayor que el saldo pendiente."
            else:
                cuenta_financiera = cuentas_financieras.filter(id=cuenta_financiera_id).first() or _cuenta_financiera_por_defecto(cuentas_financieras, metodo)
                cuenta_financiera_impuesto = cuentas_financieras.filter(id=cuenta_financiera_impuesto_id).first() if cuenta_financiera_impuesto_id else None
                pago_borrador = PagoFactura(
                    factura=factura,
                    monto=monto_decimal,
                    retencion_isr=retencion_isr,
                    retencion_isv=retencion_isv,
                    separar_isv=separar_isv,
                    cuenta_financiera=cuenta_financiera,
                    cuenta_financiera_impuesto=cuenta_financiera_impuesto,
                    metodo=metodo,
                    referencia=referencia,
                    cajero=request.user,
                )
                pago_borrador.recalcular_componentes_pago()
                requiere_cuenta_principal = pago_borrador.subtotal_recibido > 0 or (
                    pago_borrador.impuesto_recibido > 0 and not cuenta_financiera_impuesto
                )
                if requiere_cuenta_principal and not cuenta_financiera:
                    raise ValidationError("Selecciona la cuenta bancaria o caja donde entro el dinero recibido.")

                fecha_convertida = datetime.strptime(fecha_pago, "%Y-%m-%d").date() if fecha_pago else timezone.now().date()

                with transaction.atomic():
                    pago = PagoFactura.objects.create(
                        factura=factura,
                        monto=monto_decimal,
                        retencion_isr=retencion_isr,
                        retencion_isv=retencion_isv,
                        separar_isv=separar_isv,
                        cuenta_financiera_impuesto=cuenta_financiera_impuesto,
                        metodo=metodo,
                        referencia=referencia,
                        cuenta_financiera=cuenta_financiera,
                        cajero=request.user,
                        fecha=fecha_convertida
                    )
                    registrar_asiento_pago_cliente(pago)
                recibo_numero = pago.recibo.numero_recibo if hasattr(pago, 'recibo') else None
                if recibo_numero:
                    messages.success(request, f"Pago registrado correctamente. Cobro recibido: L. {monto_decimal:.2f}. Aplicado total: L. {pago.total_aplicado:.2f}. Recibo generado: {recibo_numero}.")
                else:
                    messages.success(request, f"Pago registrado correctamente. Aplicado total: L. {pago.total_aplicado:.2f}.")
                respuesta_modal = _respuesta_modal_exito(
                    f"Cobro recibido: L. {monto_decimal:.2f}. Aplicado total: L. {pago.total_aplicado:.2f}."
                )
                if respuesta_modal:
                    return respuesta_modal
                return redirect("ver_factura", empresa_slug=empresa.slug, factura_id=factura.id)
        except (InvalidOperation, ValueError):
            error = "Ingrese un monto y una fecha validos."
        except ValidationError as exc:
            error = "; ".join(exc.messages)
        except Exception as exc:
            logger.exception("Error registrando pago en factura %s", factura.id)
            error = f"No se pudo completar el cobro: {exc}"

        messages.error(request, error)

    return render(request, template_name, _contexto_pago(request.POST if request.method == "POST" else {}))


def _recalcular_y_recontabilizar_pagos_factura(factura):
    pagos = list(
        factura.pagos_facturacion.select_related(
            "cuenta_financiera",
            "cuenta_financiera_impuesto",
            "factura__cliente",
            "factura__vendedor",
        ).order_by("fecha", "id")
    )
    if not pagos:
        factura.actualizar_estado_pago()
        return

    subtotal_pendiente = Decimal(factura.subtotal_documento_ajustado or 0).quantize(Decimal("0.01"))
    impuesto_pendiente = Decimal(factura.impuesto_documento_ajustado or 0).quantize(Decimal("0.01"))
    ids_pagos = [p.id for p in pagos]

    AsientoContable.objects.filter(
        empresa=factura.empresa,
        documento_tipo="pago_factura",
        documento_id__in=ids_pagos,
        evento="cobro",
    ).delete()

    for pago in pagos:
        total_aplicado = Decimal(pago.total_aplicado or 0).quantize(Decimal("0.01"))
        total_pendiente = (subtotal_pendiente + impuesto_pendiente).quantize(Decimal("0.01"))

        if total_aplicado <= 0 or total_pendiente <= 0:
            subtotal_aplicado = Decimal("0.00")
            impuesto_aplicado = Decimal("0.00")
        elif subtotal_pendiente <= 0:
            subtotal_aplicado = Decimal("0.00")
            impuesto_aplicado = min(total_aplicado, impuesto_pendiente).quantize(Decimal("0.01"))
        elif impuesto_pendiente <= 0:
            subtotal_aplicado = min(total_aplicado, subtotal_pendiente).quantize(Decimal("0.01"))
            impuesto_aplicado = Decimal("0.00")
        else:
            proporcion_impuesto = impuesto_pendiente / total_pendiente
            impuesto_aplicado = min((total_aplicado * proporcion_impuesto).quantize(Decimal("0.01")), impuesto_pendiente)
            subtotal_aplicado = (total_aplicado - impuesto_aplicado).quantize(Decimal("0.01"))

            if subtotal_aplicado > subtotal_pendiente:
                excedente = subtotal_aplicado - subtotal_pendiente
                subtotal_aplicado = subtotal_pendiente
                impuesto_aplicado = (impuesto_aplicado + excedente).quantize(Decimal("0.01"))

            if impuesto_aplicado > impuesto_pendiente:
                excedente = impuesto_aplicado - impuesto_pendiente
                impuesto_aplicado = impuesto_pendiente
                subtotal_aplicado = (subtotal_aplicado + excedente).quantize(Decimal("0.01"))

        PagoFactura.objects.filter(pk=pago.pk).update(
            subtotal_aplicado=subtotal_aplicado,
            impuesto_aplicado=impuesto_aplicado,
        )
        pago.subtotal_aplicado = subtotal_aplicado
        pago.impuesto_aplicado = impuesto_aplicado
        registrar_asiento_pago_cliente(pago)

        subtotal_pendiente = max((subtotal_pendiente - subtotal_aplicado).quantize(Decimal("0.01")), Decimal("0.00"))
        impuesto_pendiente = max((impuesto_pendiente - impuesto_aplicado).quantize(Decimal("0.01")), Decimal("0.00"))

    factura.actualizar_estado_pago()


@login_required
@xframe_options_sameorigin
def editar_pago_factura(request, empresa_slug, factura_id, pago_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)
    pago = get_object_or_404(
        PagoFactura.objects.select_related("cuenta_financiera", "cuenta_financiera_impuesto"),
        id=pago_id,
        factura=factura,
    )
    es_modal = request.GET.get("modal") == "1"
    template_name = "facturacion/registrar_pago_modal.html" if es_modal else "facturacion/registrar_pago.html"

    if factura.estado == "anulada":
        messages.error(request, "No se pueden editar pagos de una factura anulada.")
        return redirect("ver_factura", empresa_slug=empresa.slug, factura_id=factura.id)

    cuentas_financieras = _cuentas_financieras_activas_para_pago(empresa)

    def _form_data_base():
        return {
            "fecha": pago.fecha.isoformat() if pago.fecha else timezone.now().date().isoformat(),
            "monto": f"{Decimal(pago.monto or 0):.2f}",
            "metodo": pago.metodo,
            "cuenta_financiera": str(pago.cuenta_financiera_id or ""),
            "referencia": pago.referencia or "",
            "retencion_isr": f"{Decimal(pago.retencion_isr or 0):.2f}",
            "retencion_isv": f"{Decimal(pago.retencion_isv or 0):.2f}",
            "separar_isv": pago.separar_isv,
            "cuenta_financiera_impuesto": str(pago.cuenta_financiera_impuesto_id or ""),
        }

    def _contexto_pago(form_data=None):
        tarjetas = cuentas_financieras.filter(tipo='tarjeta_credito')
        return {
            "factura": factura,
            "empresa": empresa,
            "cuentas_financieras": cuentas_financieras,
            "cajas": cuentas_financieras.filter(tipo='caja'),
            "bancos": cuentas_financieras.filter(tipo='banco'),
            "cuentas_tarjeta": tarjetas if tarjetas.exists() else cuentas_financieras,
            "usa_pagos_mixtos": False,
            "subtotal_pendiente": factura.subtotal_pendiente_cobro,
            "impuesto_pendiente": factura.impuesto_pendiente_cobro,
            "today": timezone.now().date(),
            "form_data": form_data or _form_data_base(),
            "es_modal": es_modal,
            "modo_edicion": True,
            "pago_edicion": pago,
            "modal_heading": "Editar pago",
            "modal_description": "Ajusta el monto, la cuenta o la referencia de este cobro. El sistema reconstruira recibo y contabilidad automaticamente.",
        }

    def _respuesta_modal_exito(mensaje):
        if not es_modal:
            return None
        mensaje_seguro = json.dumps(mensaje)
        html = f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="utf-8">
    <title>Pago actualizado</title>
    <style>
        body {{
            margin: 0;
            font-family: "Segoe UI", Roboto, Arial, sans-serif;
            background: #f5f8fc;
            color: #102338;
            display: flex;
            align-items: center;
            justify-content: center;
            min-height: 100vh;
        }}
        .card {{
            max-width: 460px;
            margin: 24px;
            padding: 28px;
            border-radius: 18px;
            background: #ffffff;
            border: 1px solid #d9e3f0;
            box-shadow: 0 16px 50px rgba(12, 28, 48, 0.12);
            text-align: center;
        }}
        h2 {{ margin: 0 0 10px; font-size: 24px; }}
        p {{ margin: 0; line-height: 1.6; color: #51657d; }}
    </style>
</head>
<body>
    <div class="card">
        <h2>Pago actualizado</h2>
        <p>{mensaje}</p>
    </div>
    <script>
        if (window.parent && window.parent !== window) {{
            window.parent.postMessage({{
                type: "erp-pago-guardado",
                facturaId: "{factura.id}",
                mensaje: {mensaje_seguro}
            }}, "*");
        }}
    </script>
</body>
</html>
        """.strip()
        return HttpResponse(html)

    if request.method == "POST":
        monto = request.POST.get("monto", "").strip()
        metodo = request.POST.get("metodo")
        referencia = request.POST.get("referencia", "").strip()
        fecha_pago = request.POST.get("fecha")
        cuenta_financiera_id = request.POST.get("cuenta_financiera")
        retencion_isr_raw = request.POST.get("retencion_isr", "").strip()
        retencion_isv_raw = request.POST.get("retencion_isv", "").strip()
        separar_isv = request.POST.get("separar_isv") == "on"
        cuenta_financiera_impuesto_id = request.POST.get("cuenta_financiera_impuesto")
        error = None

        def _parsear_decimal(valor, etiqueta):
            if not valor:
                return Decimal('0.00')
            try:
                return Decimal(valor)
            except InvalidOperation as exc:
                raise ValidationError(f"El valor de {etiqueta} no es valido.") from exc

        try:
            monto_decimal = _parsear_decimal(monto, "monto recibido")
            retencion_isr = _parsear_decimal(retencion_isr_raw, "retencion ISR")
            retencion_isv = _parsear_decimal(retencion_isv_raw, "retencion ISV")
            cuenta_financiera = cuentas_financieras.filter(id=cuenta_financiera_id).first() or _cuenta_financiera_por_defecto(cuentas_financieras, metodo)
            cuenta_financiera_impuesto = cuentas_financieras.filter(id=cuenta_financiera_impuesto_id).first() if cuenta_financiera_impuesto_id else None

            pago_borrador = PagoFactura(
                pk=pago.pk,
                factura=factura,
                monto=monto_decimal,
                retencion_isr=retencion_isr,
                retencion_isv=retencion_isv,
                separar_isv=separar_isv,
                cuenta_financiera=cuenta_financiera,
                cuenta_financiera_impuesto=cuenta_financiera_impuesto,
                metodo=metodo,
                referencia=referencia,
                cajero=pago.cajero,
                fecha=pago.fecha,
            )
            pago_borrador.recalcular_componentes_pago()
            requiere_cuenta_principal = pago_borrador.subtotal_recibido > 0 or (
                pago_borrador.impuesto_recibido > 0 and not cuenta_financiera_impuesto
            )
            if requiere_cuenta_principal and not cuenta_financiera:
                raise ValidationError("Selecciona la cuenta bancaria o caja donde entro el dinero recibido.")

            fecha_convertida = datetime.strptime(fecha_pago, "%Y-%m-%d").date() if fecha_pago else timezone.now().date()

            with transaction.atomic():
                pago.fecha = fecha_convertida
                pago.monto = monto_decimal
                pago.retencion_isr = retencion_isr
                pago.retencion_isv = retencion_isv
                pago.separar_isv = separar_isv
                pago.cuenta_financiera = cuenta_financiera
                pago.cuenta_financiera_impuesto = cuenta_financiera_impuesto
                pago.metodo = metodo
                pago.referencia = referencia
                pago.save()
                _recalcular_y_recontabilizar_pagos_factura(factura)

            messages.success(request, f"Pago actualizado correctamente. Aplicado total: L. {pago.total_aplicado:.2f}.")
            respuesta_modal = _respuesta_modal_exito(
                f"Cobro actualizado por L. {monto_decimal:.2f}. Aplicado total: L. {pago.total_aplicado:.2f}."
            )
            if respuesta_modal:
                return respuesta_modal
            return redirect("ver_factura", empresa_slug=empresa.slug, factura_id=factura.id)
        except (InvalidOperation, ValueError):
            error = "Ingrese un monto y una fecha validos."
        except ValidationError as exc:
            error = "; ".join(exc.messages)
        except Exception as exc:
            logger.exception("Error editando pago %s de factura %s", pago.id, factura.id)
            error = f"No se pudo actualizar el cobro: {exc}"

        messages.error(request, error)
        return render(request, template_name, _contexto_pago(request.POST))

    return render(request, template_name, _contexto_pago())


@login_required
@require_POST
def anular_factura(request, empresa_slug, factura_id):

    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)
    motivo = (request.POST.get("motivo") or "").strip()

    if factura.estado == 'anulada':
        messages.info(request, "La factura ya estaba anulada.")
        return redirect("ver_factura", empresa_slug=empresa.slug, factura_id=factura.id)

    if len(motivo) < 5:
        messages.error(request, "Explica el motivo de la anulacion con al menos 5 caracteres.")
        return redirect("ver_factura", empresa_slug=empresa.slug, factura_id=factura.id)

    with transaction.atomic():
        factura = Factura.objects.select_for_update().get(id=factura.id, empresa=empresa)
        if factura.estado == 'emitida':
            _revertir_salida_factura(factura)
            for pago in factura.pagos_facturacion.all():
                registrar_reversion_documento(
                    empresa=factura.empresa,
                    documento_tipo="pago_factura",
                    documento_id=pago.id,
                    evento_origen="cobro",
                    evento_reversion="anulacion",
                    fecha=timezone.now().date(),
                    descripcion=(
                        f"Reversion cobro de factura {factura.numero_factura or factura.id}. "
                        f"Motivo: {motivo}"
                    ),
                    referencia=pago.referencia or factura.numero_factura or str(factura.id),
                    origen_modulo="facturacion",
                    creado_por=request.user,
                )
            registrar_reversion_documento(
                empresa=factura.empresa,
                documento_tipo='factura',
                documento_id=factura.id,
                evento_origen='emision',
                evento_reversion='anulacion',
                fecha=timezone.now().date(),
                descripcion=(
                    f"Reversion factura {factura.numero_factura or factura.id}. "
                    f"Motivo: {motivo}"
                ),
                referencia=factura.numero_factura or str(factura.id),
                origen_modulo='facturacion',
                creado_por=request.user,
            )
        factura.estado = 'anulada'
        factura.estado_pago = 'pagado'
        factura.subtotal = Decimal('0.00')
        factura.impuesto = Decimal('0.00')
        factura.total = Decimal('0.00')
        factura.total_lempiras = Decimal('0.00')
        factura.save(update_fields=['estado', 'estado_pago', 'subtotal', 'impuesto', 'total', 'total_lempiras'])
    messages.success(request, "Factura anulada correctamente y registrada en la bitacora.")

    return redirect("ver_factura", empresa_slug=empresa.slug, factura_id=factura.id)


@login_required
@require_POST
def eliminar_factura(request, empresa_slug, factura_id):

    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)
    config_avanzada = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)

    if factura.estado == 'borrador' and not factura.numero_factura:
        factura.delete()
        messages.success(request, "Factura borrador eliminada correctamente.")
    else:
        messages.error(request, "Solo se pueden eliminar facturas en borrador y sin número.")

    return redirect("facturas_dashboard", empresa_slug=empresa.slug)


@login_required
@require_POST
def eliminar_factura_historica(request, empresa_slug, factura_id):

    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)
    config_avanzada = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)

    if factura.estado == 'borrador' and not factura.numero_factura:
        factura.delete()
        messages.success(request, "Factura borrador eliminada correctamente.")
        return redirect("facturas_dashboard", empresa_slug=empresa.slug)

    if not config_avanzada.permite_gestion_fiscal_historica:
        messages.error(request, "Esta empresa no tiene habilitada la correccion fiscal historica.")
        return redirect("facturas_dashboard", empresa_slug=empresa.slug)

    if factura.tiene_pagos_registrados or factura.recibos_pago.exists():
        messages.error(request, "No se puede eliminar esta factura porque ya tiene pagos o recibos registrados.")
        return redirect("facturas_dashboard", empresa_slug=empresa.slug)

    if factura.tiene_notas_credito_activas:
        messages.error(request, "No se puede eliminar esta factura porque tiene notas de credito relacionadas.")
        return redirect("facturas_dashboard", empresa_slug=empresa.slug)

    cai_id = factura.cai_id

    with transaction.atomic():
        if factura.estado == 'emitida':
            _revertir_salida_factura(factura)
            registrar_reversion_documento(
                empresa=factura.empresa,
                documento_tipo='factura',
                documento_id=factura.id,
                evento_origen='emision',
                evento_reversion='eliminacion',
                fecha=timezone.now().date(),
                descripcion=f"Eliminacion historica factura {factura.numero_factura or factura.id}",
                referencia=factura.numero_factura or str(factura.id),
                origen_modulo='facturacion',
                creado_por=factura.vendedor,
            )

        factura.delete()

        if cai_id:
            _recalcular_correlativo_cai_factura(cai_id)

    messages.success(request, "Factura eliminada correctamente para correccion historica.")
    return redirect("facturas_dashboard", empresa_slug=empresa.slug)


@login_required
@require_POST
def duplicar_factura(request, empresa_slug, factura_id):

    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    factura_original = get_object_or_404(Factura, id=factura_id, empresa=empresa)

    factura_nueva = Factura.objects.create(
        empresa=empresa,
        cliente=factura_original.cliente,
        vendedor=factura_original.vendedor,
        moneda=factura_original.moneda,
        tipo_cambio=factura_original.tipo_cambio,
        fecha_emision=timezone.now().date(),
        fecha_vencimiento=None,
        estado='borrador',
        orden_compra_exenta=factura_original.orden_compra_exenta,
        registro_exonerado=factura_original.registro_exonerado,
        registro_sag=factura_original.registro_sag,
    )

    for linea in factura_original.lineas.all():
        LineaFactura.objects.create(
            factura=factura_nueva,
            producto=linea.producto,
            cantidad=linea.cantidad,
            precio_unitario=linea.precio_unitario,
            precio_incluye_impuesto=linea.precio_incluye_impuesto,
            descuento_porcentaje=linea.descuento_porcentaje,
            comentario=linea.comentario,
            impuesto=linea.impuesto,
        )

    factura_nueva.calcular_totales()
    factura_nueva.save(update_fields=[
        'subtotal',
        'impuesto',
        'total',
        'total_lempiras'
    ])

    messages.success(request, "Factura duplicada como borrador. Revise los datos antes de guardarla o emitirla.")
    return redirect("editar_factura", empresa_slug=empresa.slug, factura_id=factura_nueva.id)


# =====================================================
# VER FACTURA
# =====================================================

@login_required
def ver_factura(request, empresa_slug, factura_id):

    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    config_avanzada = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)

    resumen = factura.resumen_fiscal()
    resumen_detallado = _resumen_detallado(factura.subtotal, resumen)
    ultimo_pago_factura_id = factura.pagos_facturacion.order_by("fecha", "id").values_list("id", flat=True).last()
    auditoria_factura = RegistroAuditoria.objects.filter(
        empresa=empresa,
        app_label="facturacion",
        modelo="factura",
        objeto_id=str(factura.id),
    )
    solicitudes_factura = auditoria_factura.values_list("identificador_solicitud", flat=True)
    historial_auditoria = RegistroAuditoria.objects.filter(empresa=empresa).filter(
        Q(app_label="facturacion", modelo="factura", objeto_id=str(factura.id))
        | Q(app_label="facturacion", identificador_solicitud__in=solicitudes_factura)
    ).select_related("usuario").distinct()[:30]

    return render(request, "facturacion/ver_factura_premium.html", {
        "empresa": empresa,
        "factura": factura,
        "resumen": resumen,
        "resumen_detallado": resumen_detallado,
        "ultimo_pago_factura_id": ultimo_pago_factura_id,
        "permite_gestion_fiscal_historica": config_avanzada.permite_gestion_fiscal_historica,
        "permite_plantilla_notas_extensas": _empresa_permite_plantilla_notas_extensas(empresa),
        "permite_plantilla_independiente": _empresa_permite_plantilla_independiente(empresa),
        "configuracion_facturacion": ConfiguracionFacturacionEmpresa.objects.get_or_create(empresa=empresa)[0],
        "historial_auditoria": historial_auditoria,
    })


@login_required
@require_POST
def validar_factura(request, empresa_slug, factura_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)

    try:
        with transaction.atomic():
            _emitir_factura_desde_borrador(factura)
        messages.success(request, "Factura validada correctamente.")
    except ValidationError as exc:
        messages.error(request, exc.message if hasattr(exc, "message") else str(exc))
    except ValueError as exc:
        messages.error(request, str(exc))

    return redirect("ver_factura", empresa_slug=empresa.slug, factura_id=factura.id)


# =====================================================
# PDF
# =====================================================

@login_required
def descargar_factura_pdf(request, empresa_slug, factura_id):

    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)
    configuracion, _ = ConfiguracionFacturacionEmpresa.objects.get_or_create(empresa=empresa)
    plantilla = _resolver_plantilla_factura(configuracion, empresa)
    return _render_factura_pdf_response(
        empresa=empresa,
        factura=factura,
        plantilla=plantilla,
        inline=False,
        prefijo_archivo="Factura",
    )


@login_required
def descargar_factura_pdf_alternativo(request, empresa_slug, factura_id):

    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)
    return _render_factura_pdf_response(
        empresa=empresa,
        factura=factura,
        plantilla="facturacion/factura_pdf_alternativa.html",
        inline=False,
        prefijo_archivo="Factura_Alternativa",
    )


@login_required
def descargar_factura_pdf_notas_extensas(request, empresa_slug, factura_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    if not _empresa_permite_plantilla_notas_extensas(empresa):
        raise Http404("Esta plantilla no esta disponible para esta empresa.")
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)
    return _render_factura_pdf_response(
        empresa=empresa,
        factura=factura,
        plantilla="facturacion/factura_pdf_notas_extensas.html",
        inline=False,
        prefijo_archivo="Factura_Notas",
    )


@login_required
def descargar_factura_pdf_independiente(request, empresa_slug, factura_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    if not _empresa_permite_plantilla_independiente(empresa):
        raise Http404("Esta plantilla no esta disponible para esta empresa.")
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)
    return _render_factura_pdf_response(
        empresa=empresa,
        factura=factura,
        plantilla="facturacion/factura_pdf_independiente.html",
        inline=False,
        prefijo_archivo="Factura_Independiente",
    )


@login_required
@xframe_options_sameorigin
def vista_previa_factura_pdf(request, empresa_slug, factura_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)
    configuracion, _ = ConfiguracionFacturacionEmpresa.objects.get_or_create(empresa=empresa)
    plantilla = _resolver_plantilla_factura(configuracion, empresa)
    return _render_factura_pdf_response(
        empresa=empresa,
        factura=factura,
        plantilla=plantilla,
        inline=True,
        prefijo_archivo="Factura",
    )


@login_required
def imprimir_factura_pos(request, empresa_slug, factura_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)
    if empresa.slug not in {"hospital_mia", "medical_spa", "demo_1"}:
        return redirect("vista_previa_factura_pdf", empresa_slug=empresa.slug, factura_id=factura.id)
    return render(request, "facturacion/factura_pos_auto_print.html", {
        "empresa": empresa,
        "factura": factura,
        "pdf_url": reverse("vista_previa_factura_pdf", args=[empresa.slug, factura.id]),
    })


@login_required
@xframe_options_sameorigin
def vista_previa_factura_pdf_alternativo(request, empresa_slug, factura_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)
    return _render_factura_pdf_response(
        empresa=empresa,
        factura=factura,
        plantilla="facturacion/factura_pdf_alternativa.html",
        inline=True,
        prefijo_archivo="Factura_Alternativa",
    )


@login_required
@xframe_options_sameorigin
def vista_previa_factura_pdf_notas_extensas(request, empresa_slug, factura_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    if not _empresa_permite_plantilla_notas_extensas(empresa):
        raise Http404("Esta plantilla no esta disponible para esta empresa.")
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)
    return _render_factura_pdf_response(
        empresa=empresa,
        factura=factura,
        plantilla="facturacion/factura_pdf_notas_extensas.html",
        inline=True,
        prefijo_archivo="Factura_Notas",
    )


@login_required
@xframe_options_sameorigin
def vista_previa_factura_pdf_independiente(request, empresa_slug, factura_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    if not _empresa_permite_plantilla_independiente(empresa):
        raise Http404("Esta plantilla no esta disponible para esta empresa.")
    factura = get_object_or_404(Factura, id=factura_id, empresa=empresa)
    return _render_factura_pdf_response(
        empresa=empresa,
        factura=factura,
        plantilla="facturacion/factura_pdf_independiente.html",
        inline=True,
        prefijo_archivo="Factura_Independiente",
    )


# =====================================================
# REPORTES
# =====================================================

@login_required
def reportes_facturacion(request, empresa_slug):

    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    facturas = _filtrar_facturas_reporte(empresa, request.GET)
    configuracion_power_bi, _ = ConfiguracionPowerBIEmpresa.objects.get_or_create(empresa=empresa)

    totales = {
        "subtotal": sum((f.subtotal_documento_ajustado for f in facturas), Decimal('0.00')),
        "impuesto": sum((f.impuesto_documento_ajustado for f in facturas), Decimal('0.00')),
        "total": sum((f.total_documento_ajustado for f in facturas), Decimal('0.00')),
    }

    total_saldo = sum((f.saldo_pendiente for f in facturas), Decimal('0.00'))

    total_base_15 = Decimal('0.00')
    total_isv_15 = Decimal('0.00')
    total_base_18 = Decimal('0.00')
    total_isv_18 = Decimal('0.00')
    total_exento = Decimal('0.00')
    total_exonerado = Decimal('0.00')
    total_descuento = Decimal('0.00')

    for f in facturas:
        r = f.resumen_fiscal()

        total_base_15 += Decimal(str(r["base_15"]))
        total_isv_15 += Decimal(str(r["isv_15"]))
        total_base_18 += Decimal(str(r["base_18"]))
        total_isv_18 += Decimal(str(r["isv_18"]))
        total_exento += Decimal(str(r["base_exento"]))
        total_exonerado += Decimal(str(r["base_exonerado"]))

        # 🔥 CALCULO CORRECTO DEL DESCUENTO
        total_descuento += Decimal(str(r.get("descuento_total", 0) or 0))

    clientes = Cliente.objects.filter(empresa=empresa)
    bi_interno = _construir_bi_interno_facturacion(facturas)

    return render(request, "facturacion/reportes_premium.html", {
        "empresa": empresa,
        "clientes": clientes,
        "facturas": facturas,
        "totales": totales,
        "total_saldo": total_saldo,
        "total_base_15": total_base_15,
        "total_isv_15": total_isv_15,
        "total_base_18": total_base_18,
        "total_isv_18": total_isv_18,
        "total_exento": total_exento,
        "total_exonerado": total_exonerado,
        "total_descuento": total_descuento,
        "bi_interno": bi_interno,
        "configuracion_power_bi": configuracion_power_bi,
        "puede_configurar_power_bi": _puede_configurar_power_bi(request.user),
    })


def _rango_reporte_retenciones(periodo, fecha_referencia, fecha_desde="", fecha_hasta=""):
    periodo = periodo if periodo in {"dia", "semana", "mes", "trimestre", "semestre", "anio", "personalizado"} else "mes"
    if periodo == "personalizado":
        try:
            inicio = date.fromisoformat(fecha_desde)
            fin = date.fromisoformat(fecha_hasta)
        except (TypeError, ValueError):
            inicio = date(fecha_referencia.year, fecha_referencia.month, 1)
            fin = fecha_referencia
            periodo = "mes"
        if inicio > fin:
            inicio, fin = fin, inicio
        return periodo, inicio, fin

    if periodo == "dia":
        return periodo, fecha_referencia, fecha_referencia
    if periodo == "semana":
        inicio = fecha_referencia - timedelta(days=fecha_referencia.weekday())
        return periodo, inicio, inicio + timedelta(days=6)
    if periodo == "mes":
        inicio = date(fecha_referencia.year, fecha_referencia.month, 1)
        fin = date(fecha_referencia.year, fecha_referencia.month, calendar.monthrange(fecha_referencia.year, fecha_referencia.month)[1])
        return periodo, inicio, fin
    if periodo == "trimestre":
        mes_inicio = ((fecha_referencia.month - 1) // 3) * 3 + 1
        inicio = date(fecha_referencia.year, mes_inicio, 1)
        mes_fin = mes_inicio + 2
        fin = date(fecha_referencia.year, mes_fin, calendar.monthrange(fecha_referencia.year, mes_fin)[1])
        return periodo, inicio, fin
    if periodo == "semestre":
        mes_inicio = 1 if fecha_referencia.month <= 6 else 7
        inicio = date(fecha_referencia.year, mes_inicio, 1)
        mes_fin = mes_inicio + 5
        fin = date(fecha_referencia.year, mes_fin, calendar.monthrange(fecha_referencia.year, mes_fin)[1])
        return periodo, inicio, fin
    return periodo, date(fecha_referencia.year, 1, 1), date(fecha_referencia.year, 12, 31)


@login_required
def reporte_retenciones_pagos(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    if empresa.slug != "digital_planning":
        raise Http404("Este reporte solo esta habilitado para Digital Planning.")

    hoy = timezone.localdate()
    try:
        fecha_referencia = date.fromisoformat(request.GET.get("fecha", ""))
    except (TypeError, ValueError):
        fecha_referencia = hoy

    periodo, fecha_desde, fecha_hasta = _rango_reporte_retenciones(
        request.GET.get("periodo", "mes"),
        fecha_referencia,
        request.GET.get("fecha_desde", "").strip(),
        request.GET.get("fecha_hasta", "").strip(),
    )
    pagos = (
        PagoFactura.objects.filter(
            factura__empresa=empresa,
            fecha__range=(fecha_desde, fecha_hasta),
        )
        .exclude(factura__estado="anulada")
        .filter(Q(retencion_isr__gt=0) | Q(retencion_isv__gt=0))
        .select_related("factura", "factura__cliente", "cuenta_financiera", "cajero")
        .order_by("-fecha", "-id")
    )
    agregados = pagos.aggregate(
        cobrado=Sum("monto"),
        retencion_isr=Sum("retencion_isr"),
        retencion_isv=Sum("retencion_isv"),
    )
    resumen = {
        "cobrado": agregados["cobrado"] or Decimal("0.00"),
        "retencion_isr": agregados["retencion_isr"] or Decimal("0.00"),
        "retencion_isv": agregados["retencion_isv"] or Decimal("0.00"),
        "cantidad": pagos.count(),
    }
    resumen["total_retenciones"] = resumen["retencion_isr"] + resumen["retencion_isv"]
    resumen["total_aplicado"] = resumen["cobrado"] + resumen["total_retenciones"]

    resumen_diario = []
    for item in (
        pagos.values("fecha")
        .annotate(
            cobrado=Sum("monto"),
            retencion_isr=Sum("retencion_isr"),
            retencion_isv=Sum("retencion_isv"),
            operaciones=Count("id"),
        )
        .order_by("fecha")
    ):
        item["total_retenciones"] = (item["retencion_isr"] or Decimal("0.00")) + (item["retencion_isv"] or Decimal("0.00"))
        resumen_diario.append(item)

    return render(request, "facturacion/reporte_retenciones_pagos.html", {
        "empresa": empresa,
        "pagos": pagos[:500],
        "resumen": resumen,
        "resumen_diario": resumen_diario,
        "periodo": periodo,
        "fecha_referencia": fecha_referencia,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
    })


@login_required
def dashboard_bi_facturacion(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    facturas = _filtrar_facturas_reporte(empresa, request.GET)
    bi_interno = _construir_bi_interno_facturacion(facturas)

    hoy = timezone.localdate()
    inicio_mes_actual = date(hoy.year, hoy.month, 1)
    if hoy.month == 1:
        inicio_mes_anterior = date(hoy.year - 1, 12, 1)
    else:
        inicio_mes_anterior = date(hoy.year, hoy.month - 1, 1)

    facturas_emitidas = facturas.exclude(estado="anulada")
    facturas_mes_actual = facturas_emitidas.filter(fecha_emision__gte=inicio_mes_actual)
    facturas_mes_anterior = facturas_emitidas.filter(
        fecha_emision__year=inicio_mes_anterior.year,
        fecha_emision__month=inicio_mes_anterior.month,
    )

    total_actual = facturas_mes_actual.aggregate(total=Sum("total"))["total"] or Decimal("0.00")
    total_anterior = facturas_mes_anterior.aggregate(total=Sum("total"))["total"] or Decimal("0.00")
    docs_actual = facturas_mes_actual.count()
    docs_anterior = facturas_mes_anterior.count()
    saldo_total = sum((factura.saldo_pendiente for factura in facturas_emitidas), Decimal("0.00"))
    cobrado_total = sum(
        ((factura.total or Decimal("0.00")) - factura.saldo_pendiente for factura in facturas_emitidas),
        Decimal("0.00"),
    )
    ticket_promedio = (total_actual / docs_actual) if docs_actual else Decimal("0.00")
    tasa_cobro = (cobrado_total / (cobrado_total + saldo_total) * Decimal("100.00")) if (cobrado_total + saldo_total) else Decimal("0.00")

    docs_pagados_actual = facturas_mes_actual.filter(estado_pago="pagado").count()
    docs_pagados_anterior = facturas_mes_anterior.filter(estado_pago="pagado").count()
    pagos_visibles = PagoFactura.objects.filter(factura__in=facturas_emitidas).select_related(
        "cuenta_financiera",
        "factura__cliente",
    )

    cobradores_qs = (
        pagos_visibles
        .values("cajero__username")
        .annotate(total=Sum("monto"), operaciones=Count("id"))
        .order_by("-total")[:6]
    )
    cobranza_por_cajero = []
    max_cajero = Decimal("0.00")
    for item in cobradores_qs:
        total_cajero = item["total"] or Decimal("0.00")
        max_cajero = max(max_cajero, total_cajero)
        cobranza_por_cajero.append(
            {
                "nombre": item["cajero__username"] or "Sin cajero",
                "total": total_cajero,
                "operaciones": item["operaciones"],
            }
        )
    for item in cobranza_por_cajero:
        item["ancho"] = float((item["total"] / max_cajero) * 100) if max_cajero else 0

    ventas_por_vendedor_qs = (
        facturas_emitidas.values("vendedor__username")
        .annotate(total=Sum("total"), documentos=Count("id"))
        .order_by("-total")[:6]
    )
    ventas_por_vendedor = []
    max_vendedor = Decimal("0.00")
    for item in ventas_por_vendedor_qs:
        total_vendedor = item["total"] or Decimal("0.00")
        max_vendedor = max(max_vendedor, total_vendedor)
        ventas_por_vendedor.append(
            {
                "nombre": item["vendedor__username"] or "Sin vendedor",
                "total": total_vendedor,
                "documentos": item["documentos"],
            }
        )
    for item in ventas_por_vendedor:
        item["ancho"] = float((item["total"] / max_vendedor) * 100) if max_vendedor else 0

    ingresos_por_banco_qs = (
        pagos_visibles.filter(cuenta_financiera__isnull=False)
        .values("cuenta_financiera__nombre", "cuenta_financiera__tipo")
        .annotate(total=Sum("monto"), operaciones=Count("id"))
        .order_by("-total")[:6]
    )
    ingresos_por_banco = []
    max_banco = Decimal("0.00")
    for item in ingresos_por_banco_qs:
        total_banco = item["total"] or Decimal("0.00")
        max_banco = max(max_banco, total_banco)
        ingresos_por_banco.append(
            {
                "nombre": item["cuenta_financiera__nombre"] or "Sin cuenta",
                "tipo": item["cuenta_financiera__tipo"] or "sin_tipo",
                "total": total_banco,
                "operaciones": item["operaciones"],
            }
        )
    for item in ingresos_por_banco:
        item["ancho"] = float((item["total"] / max_banco) * 100) if max_banco else 0

    clientes_con_saldo_map = {}
    hoy_control = timezone.localdate()
    for factura in facturas_emitidas:
        saldo = factura.saldo_pendiente
        fecha_control = factura.fecha_vencimiento or factura.fecha_emision
        if saldo <= 0 or not fecha_control or fecha_control >= hoy_control:
            continue
        nombre = factura.cliente.nombre if factura.cliente_id else "Cliente sin nombre"
        bucket = clientes_con_saldo_map.setdefault(
            nombre,
            {
                "nombre": nombre,
                "saldo": Decimal("0.00"),
                "documentos": 0,
                "fecha_mas_antigua": fecha_control,
            },
        )
        bucket["saldo"] += saldo
        bucket["documentos"] += 1
        if fecha_control < bucket["fecha_mas_antigua"]:
            bucket["fecha_mas_antigua"] = fecha_control
    clientes_con_saldo_vencido = sorted(
        clientes_con_saldo_map.values(),
        key=lambda item: item["saldo"],
        reverse=True,
    )[:6]
    max_vencido = max((item["saldo"] for item in clientes_con_saldo_vencido), default=Decimal("0.00"))
    for item in clientes_con_saldo_vencido:
        item["dias_vencido"] = (hoy_control - item["fecha_mas_antigua"]).days
        item["ancho"] = float((item["saldo"] / max_vencido) * 100) if max_vencido else 0

    inicio_serie = date(hoy.year, hoy.month, 1) - timedelta(days=150)
    facturado_por_mes = {
        item["periodo"].date() if hasattr(item["periodo"], "date") else item["periodo"]: item["total"]
        for item in (
            facturas_emitidas.filter(fecha_emision__gte=inicio_serie)
            .annotate(periodo=TruncMonth("fecha_emision"))
            .values("periodo")
            .annotate(total=Sum("total"))
            .order_by("periodo")
        )
    }
    cobrado_por_mes = {
        item["periodo"].date() if hasattr(item["periodo"], "date") else item["periodo"]: item["total"]
        for item in (
            pagos_visibles.filter(fecha__gte=inicio_serie)
            .annotate(periodo=TruncMonth("fecha"))
            .values("periodo")
            .annotate(total=Sum("monto"))
            .order_by("periodo")
        )
    }
    periodos_mensuales = sorted(set(facturado_por_mes.keys()) | set(cobrado_por_mes.keys()))
    comparativo_cobro = []
    max_comparativo = Decimal("0.00")
    for periodo in periodos_mensuales:
        facturado = facturado_por_mes.get(periodo) or Decimal("0.00")
        cobrado = cobrado_por_mes.get(periodo) or Decimal("0.00")
        max_comparativo = max(max_comparativo, facturado, cobrado)
        comparativo_cobro.append(
            {
                "periodo": periodo,
                "facturado": facturado,
                "cobrado": cobrado,
            }
        )
    for item in comparativo_cobro:
        item["ancho_facturado"] = float((item["facturado"] / max_comparativo) * 100) if max_comparativo else 0
        item["ancho_cobrado"] = float((item["cobrado"] / max_comparativo) * 100) if max_comparativo else 0

    total_base_15 = Decimal("0.00")
    total_base_18 = Decimal("0.00")
    total_exento = Decimal("0.00")
    total_exonerado = Decimal("0.00")
    total_isv_15 = Decimal("0.00")
    total_isv_18 = Decimal("0.00")
    for factura in facturas_emitidas:
        resumen = factura.resumen_fiscal()
        total_base_15 += Decimal(str(resumen.get("base_15", 0) or 0))
        total_base_18 += Decimal(str(resumen.get("base_18", 0) or 0))
        total_exento += Decimal(str(resumen.get("base_exento", 0) or 0))
        total_exonerado += Decimal(str(resumen.get("base_exonerado", 0) or 0))
        total_isv_15 += Decimal(str(resumen.get("isv_15", 0) or 0))
        total_isv_18 += Decimal(str(resumen.get("isv_18", 0) or 0))
    mezcla_impuestos = [
        {"etiqueta": "Base 15%", "total": total_base_15, "color": "#2968f2"},
        {"etiqueta": "Base 18%", "total": total_base_18, "color": "#14b8a6"},
        {"etiqueta": "Exento", "total": total_exento, "color": "#8b5cf6"},
        {"etiqueta": "Exonerado", "total": total_exonerado, "color": "#f59e0b"},
        {"etiqueta": "ISV 15%", "total": total_isv_15, "color": "#ef4444"},
        {"etiqueta": "ISV 18%", "total": total_isv_18, "color": "#0f766e"},
    ]
    total_impuestos_visual = sum((item["total"] for item in mezcla_impuestos), Decimal("0.00"))
    for item in mezcla_impuestos:
        item["porcentaje"] = float((item["total"] / total_impuestos_visual) * 100) if total_impuestos_visual else 0

    color_map = {
        "Pagado": "#22c55e",
        "Parcial": "#f59e0b",
        "Pendiente": "#ef4444",
    }
    total_estado = sum((item["total"] for item in bi_interno["estado_cobro"]), Decimal("0.00"))
    acumulado = 0
    segmentos = []
    for item in bi_interno["estado_cobro"]:
        porcentaje = float((item["total"] / total_estado) * 100) if total_estado else 0
        inicio = acumulado
        acumulado += porcentaje
        segmentos.append(
            {
                "etiqueta": item["etiqueta"],
                "color": color_map.get(item["etiqueta"], "#2968f2"),
                "inicio": inicio,
                "fin": acumulado,
                "porcentaje": porcentaje,
                "total": item["total"],
                "cantidad": item["cantidad"],
            }
        )
    donut_background = "conic-gradient(" + ", ".join(
        f"{segmento['color']} {segmento['inicio']:.2f}% {segmento['fin']:.2f}%"
        for segmento in segmentos
    ) + ")" if segmentos else "conic-gradient(#dbe7f8 0% 100%)"

    contexto = {
        "empresa": empresa,
        "facturas": facturas,
        "clientes": Cliente.objects.filter(empresa=empresa),
        "bi_interno": bi_interno,
        "kpis_bi": {
            "total_actual": total_actual,
            "total_anterior": total_anterior,
            "variacion_total": _calcular_variacion_porcentual(total_actual, total_anterior),
            "docs_actual": docs_actual,
            "docs_anterior": docs_anterior,
            "variacion_docs": _calcular_variacion_porcentual(Decimal(docs_actual), Decimal(docs_anterior)),
            "docs_pagados_actual": docs_pagados_actual,
            "docs_pagados_anterior": docs_pagados_anterior,
            "variacion_pagados": _calcular_variacion_porcentual(Decimal(docs_pagados_actual), Decimal(docs_pagados_anterior)),
            "ticket_promedio": ticket_promedio,
            "saldo_total": saldo_total,
            "cobrado_total": cobrado_total,
            "tasa_cobro": tasa_cobro,
        },
        "ventas_por_vendedor": ventas_por_vendedor,
        "cobranza_por_cajero": cobranza_por_cajero,
        "ingresos_por_banco": ingresos_por_banco,
        "clientes_con_saldo_vencido": clientes_con_saldo_vencido,
        "comparativo_cobro": comparativo_cobro,
        "mezcla_impuestos": mezcla_impuestos,
        "total_impuestos_visual": total_impuestos_visual,
        "estado_cobro_segmentos": segmentos,
        "estado_cobro_donut": donut_background,
        "inicio_mes_actual": inicio_mes_actual,
        "inicio_mes_anterior": inicio_mes_anterior,
    }
    return render(request, "facturacion/dashboard_bi_facturacion.html", contexto)


@login_required
def configuracion_power_bi_reportes(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)

    if not request.user.is_superuser and not request.user.puede_acceder_empresa(empresa):
        messages.error(request, "No puedes configurar Power BI para otra empresa.")
        return redirect("reportes_facturacion", empresa_slug=empresa.slug)

    if not _puede_configurar_power_bi(request.user):
        messages.error(request, "Solo un administrador puede configurar el dashboard BI de esta empresa.")
        return redirect("reportes_facturacion", empresa_slug=empresa.slug)

    configuracion, _ = ConfiguracionPowerBIEmpresa.objects.get_or_create(empresa=empresa)
    form = ConfiguracionPowerBIForm(request.POST or None, instance=configuracion)

    if request.method == "POST" and form.is_valid():
        configuracion = form.save(commit=False)
        configuracion.empresa = empresa
        configuracion.save()
        messages.success(request, "Configuracion Power BI actualizada correctamente.")
        return redirect("reportes_facturacion", empresa_slug=empresa.slug)

    return render(request, "facturacion/configuracion_power_bi.html", {
        "empresa": empresa,
        "form": form,
        "configuracion_power_bi": configuracion,
    })


@login_required
def reporte_ingresos_bancos(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    config_avanzada = ConfiguracionAvanzadaEmpresa.para_empresa(empresa)
    if not config_avanzada.usa_reporte_bancos:
        messages.error(request, "El reporte bancario no esta activo para esta empresa.")
        return redirect("facturacion_dashboard", empresa_slug=empresa.slug)

    fecha_desde = request.GET.get("fecha_desde", "").strip()
    fecha_hasta = request.GET.get("fecha_hasta", "").strip()
    cuenta_id = request.GET.get("cuenta", "").strip()
    metodo = request.GET.get("metodo", "").strip()

    pagos = (
        PagoFactura.objects.filter(
            factura__empresa=empresa,
            cuenta_financiera__isnull=False,
        )
        .select_related("factura", "factura__cliente", "cuenta_financiera")
        .order_by("-fecha", "-id")
    )

    if fecha_desde:
        pagos = pagos.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        pagos = pagos.filter(fecha__lte=fecha_hasta)
    if cuenta_id:
        pagos = pagos.filter(cuenta_financiera_id=cuenta_id)
    if metodo:
        pagos = pagos.filter(metodo=metodo)

    cuentas_financieras = CuentaFinanciera.objects.filter(
        empresa=empresa,
        activa=True
    ).order_by("tipo", "nombre")
    resumen_cuentas = []
    for cuenta in cuentas_financieras:
        pagos_cuenta = pagos.filter(cuenta_financiera=cuenta)
        total_cuenta = pagos_cuenta.aggregate(total=Sum("monto"))["total"] or Decimal("0.00")
        if total_cuenta:
            resumen_cuentas.append({
                "cuenta": cuenta,
                "total": total_cuenta,
                "cantidad": pagos_cuenta.count(),
            })

    resumen = {
        "total_ingresado": pagos.aggregate(total=Sum("monto"))["total"] or Decimal("0.00"),
        "cantidad_pagos": pagos.count(),
        "transferencias": pagos.filter(metodo="transferencia").aggregate(total=Sum("monto"))["total"] or Decimal("0.00"),
        "tarjetas": pagos.filter(metodo="tarjeta").aggregate(total=Sum("monto"))["total"] or Decimal("0.00"),
    }

    return render(request, "facturacion/reporte_ingresos_bancos.html", {
        "empresa": empresa,
        "pagos": pagos[:250],
        "cuentas_financieras": cuentas_financieras,
        "resumen_cuentas": resumen_cuentas,
        "resumen": resumen,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "cuenta_id": cuenta_id,
        "metodo": metodo,
        "metodos_pago": PagoFactura.METODOS,
    })


@login_required
def cierres_caja(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    if not _empresa_usa_cierre_caja(empresa):
        messages.error(request, "El cierre de caja no esta activo para esta empresa.")
        return redirect("facturacion_dashboard", empresa_slug=empresa.slug)

    fecha = _fecha_caja_desde_parametro(request.POST.get("fecha") or request.GET.get("fecha") or timezone.localdate())
    turno = request.POST.get("turno") or request.GET.get("turno") or "general"

    pagos_usuario = _preparar_pagos_caja(
        PagoFactura.objects.filter(
            factura__empresa=empresa,
            cajero=request.user,
            fecha=fecha,
        ).select_related("factura", "factura__cliente", "cuenta_financiera"),
        empresa=empresa,
    )

    desglose_caja = {
        metodo: _desglose_metodo_caja(pagos_usuario, metodo)
        for metodo, _etiqueta in PagoFactura.METODOS
    }

    efectivo_sistema = desglose_caja["efectivo"]["neto"]
    tarjeta_sistema = desglose_caja["tarjeta"]["neto"]
    transferencia_sistema = desglose_caja["transferencia"]["neto"]
    aperturas_caja = sum(1 for pago in pagos_usuario if pago.metodo == "efectivo" and not pago.anulacion_no_suma_caja)

    if request.method == "POST":
        def leer_decimal(nombre, valor_sistema):
            valor = request.POST.get(nombre, "").strip()
            if not valor:
                return valor_sistema
            return Decimal(valor)

        try:
            efectivo_reportado = leer_decimal("efectivo_reportado", efectivo_sistema)
            tarjeta_reportado = leer_decimal("tarjeta_reportado", tarjeta_sistema)
            transferencia_reportado = leer_decimal("transferencia_reportado", transferencia_sistema)
        except InvalidOperation:
            messages.error(request, "Revisa los montos reportados antes de cerrar caja.")
        else:
            CierreCaja.objects.update_or_create(
                empresa=empresa,
                cajero=request.user,
                fecha=fecha,
                turno=turno,
                defaults={
                    "efectivo_sistema": efectivo_sistema,
                    "tarjeta_sistema": tarjeta_sistema,
                    "transferencia_sistema": transferencia_sistema,
                    "efectivo_reportado": efectivo_reportado,
                    "tarjeta_reportado": tarjeta_reportado,
                    "transferencia_reportado": transferencia_reportado,
                    "observacion": request.POST.get("observacion", "").strip(),
                    "estado": "cerrado",
                },
            )
            messages.success(request, "Cierre de caja registrado correctamente.")
            return redirect(f"{request.path}?fecha={fecha.isoformat()}&turno={turno}")

    puede_ver_historial = _puede_ver_historial_cierres(request, empresa)
    cierres = (
        CierreCaja.objects.filter(empresa=empresa).select_related("cajero")[:60]
        if puede_ver_historial
        else CierreCaja.objects.none()
    )
    resumen = {
        "efectivo_sistema": efectivo_sistema,
        "tarjeta_sistema": tarjeta_sistema,
        "transferencia_sistema": transferencia_sistema,
        "total_sistema": efectivo_sistema + tarjeta_sistema + transferencia_sistema,
        "pagos": len(pagos_usuario),
        "anulaciones": sum(1 for pago in pagos_usuario if pago.anulacion_no_suma_caja),
        "aperturas_caja": aperturas_caja,
        "desglose_caja": desglose_caja,
    }

    return render(request, "facturacion/cierres_caja.html", {
        "empresa": empresa,
        "fecha": fecha.isoformat(),
        "turno": turno,
        "turnos": CierreCaja.TURNOS,
        "pagos_usuario": pagos_usuario[:120],
        "cierres": cierres,
        "resumen": resumen,
        "puede_ver_historial": puede_ver_historial,
    })


@login_required
def ver_cierre_caja(request, empresa_slug, cierre_id):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    if not _empresa_usa_cierre_caja(empresa):
        messages.error(request, "El cierre de caja no esta activo para esta empresa.")
        return redirect("facturacion_dashboard", empresa_slug=empresa.slug)
    if not _puede_ver_historial_cierres(request, empresa):
        messages.error(request, "Solo los administradores pueden consultar el historial de cierres.")
        return redirect("cierres_caja", empresa_slug=empresa.slug)

    cierre = get_object_or_404(
        CierreCaja.objects.select_related("cajero", "empresa"),
        id=cierre_id,
        empresa=empresa,
    )
    pagos = _preparar_pagos_caja(
        PagoFactura.objects.filter(
            factura__empresa=empresa,
            cajero=cierre.cajero,
            fecha=cierre.fecha,
        )
        .select_related("factura", "factura__cliente", "cuenta_financiera")
        .order_by("metodo", "factura__numero_factura", "id"),
        empresa=empresa,
    )

    resumen_metodos = []
    for metodo, etiqueta in PagoFactura.METODOS:
        pagos_metodo = [pago for pago in pagos if pago.metodo == metodo]
        resumen_metodos.append({
            "metodo": etiqueta,
            "cantidad": len(pagos_metodo),
            "total": _total_neto_caja(pagos_metodo),
        })
    aperturas_caja = sum(1 for pago in pagos if pago.metodo == "efectivo" and not pago.anulacion_no_suma_caja)

    return render(request, "facturacion/ver_cierre_caja.html", {
        "empresa": empresa,
        "cierre": cierre,
        "pagos": pagos,
        "resumen_metodos": resumen_metodos,
        "aperturas_caja": aperturas_caja,
    })


@login_required
def resumen_diario_caja(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    if not _empresa_usa_cierre_caja(empresa):
        messages.error(request, "El resumen diario de caja no esta activo para esta empresa.")
        return redirect("facturacion_dashboard", empresa_slug=empresa.slug)
    if not _puede_ver_historial_cierres(request, empresa):
        messages.error(request, "Solo los administradores pueden consultar el resumen e historial de cierres.")
        return redirect("cierres_caja", empresa_slug=empresa.slug)

    fecha = _fecha_caja_desde_parametro(request.GET.get("fecha") or timezone.localdate())
    pagos = _preparar_pagos_caja(
        PagoFactura.objects.filter(factura__empresa=empresa, fecha=fecha)
        .select_related("factura", "factura__cliente", "cuenta_financiera", "cajero")
        .order_by("cajero__username", "metodo", "factura__numero_factura", "id"),
        empresa=empresa,
    )
    cierres = (
        CierreCaja.objects.filter(empresa=empresa, fecha=fecha)
        .select_related("cajero")
        .order_by("cajero__username", "turno")
    )

    resumen_metodos = []
    for metodo, etiqueta in PagoFactura.METODOS:
        pagos_metodo = [pago for pago in pagos if pago.metodo == metodo]
        resumen_metodos.append({
            "metodo": etiqueta,
            "cantidad": len(pagos_metodo),
            "total": _total_neto_caja(pagos_metodo),
        })

    resumen_cajeros = []
    for cajero_id in sorted({pago.cajero_id for pago in pagos}, key=lambda value: value or 0):
        pagos_cajero = [pago for pago in pagos if pago.cajero_id == cajero_id]
        cajero = pagos_cajero[0].cajero if pagos_cajero else None
        resumen_cajeros.append({
            "cajero": cajero,
            "cantidad": len(pagos_cajero),
            "aperturas_caja": sum(1 for pago in pagos_cajero if pago.metodo == "efectivo" and not pago.anulacion_no_suma_caja),
            "total": _total_neto_caja(pagos_cajero),
            "cierres": cierres.filter(cajero_id=cajero_id),
        })

    resumen_cuentas = []
    for cuenta_id in sorted({pago.cuenta_financiera_id for pago in pagos if pago.cuenta_financiera_id}):
        pagos_cuenta = [pago for pago in pagos if pago.cuenta_financiera_id == cuenta_id]
        cuenta = pagos_cuenta[0].cuenta_financiera if pagos_cuenta else None
        resumen_cuentas.append({
            "cuenta": cuenta,
            "cantidad": len(pagos_cuenta),
            "total": _total_neto_caja(pagos_cuenta),
        })

    resumen = {
        "total": _total_neto_caja(pagos),
        "pagos": len(pagos),
        "anulaciones": sum(1 for pago in pagos if pago.anulacion_no_suma_caja),
        "facturas": len({pago.factura_id for pago in pagos}),
        "aperturas_caja": sum(1 for pago in pagos if pago.metodo == "efectivo" and not pago.anulacion_no_suma_caja),
        "cierres": cierres.count(),
        "total_cerrado": sum((cierre.total_reportado for cierre in cierres), Decimal("0.00")),
        "diferencia_cierres": sum((cierre.diferencia for cierre in cierres), Decimal("0.00")),
    }

    return render(request, "facturacion/resumen_diario_caja.html", {
        "empresa": empresa,
        "fecha": fecha.isoformat(),
        "pagos": pagos[:300],
        "cierres": cierres,
        "resumen": resumen,
        "resumen_metodos": resumen_metodos,
        "resumen_cajeros": resumen_cajeros,
        "resumen_cuentas": resumen_cuentas,
    })


# =====================================================
# REPORTES CXC
# =====================================================

@login_required
def reporte_cxc(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    contexto = _construir_reporte_cxc(empresa, request.GET)
    contexto["empresa"] = empresa
    return render(request, "facturacion/reporte_cxc_premium.html", contexto)


@login_required
def reporte_cxp(request, empresa_slug):
    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    compras = (
        CompraInventario.objects.filter(empresa=empresa, estado='aplicada')
        .select_related('proveedor')
        .prefetch_related('pagos_compra', 'lineas')
        .order_by('fecha_documento', 'id')
    )

    hoy = date.today()
    proveedor_id = request.GET.get("proveedor")
    proveedor_key = request.GET.get("proveedor_key", "").strip()
    q = request.GET.get("q", "").strip()

    data = {}
    compras_pendientes_proveedor = None
    proveedor_seleccionado = None
    proveedor_resumen = None
    proveedor_nombre_seleccionado = ""
    total_cxp = Decimal('0.00')

    for compra in compras:
        saldo = compra.saldo_pendiente
        if saldo <= 0:
            continue

        total_cxp += saldo
        proveedor = compra.proveedor
        nombre = (proveedor.nombre if proveedor else (compra.proveedor_nombre or "Proveedor sin nombre")).strip()
        llave = f"proveedor-{proveedor.id}" if proveedor else f"manual-{slugify(nombre)}"
        dias = (hoy - compra.fecha_control_cxp).days

        if llave not in data:
            data[llave] = {
                "llave": llave,
                "proveedor": proveedor,
                "nombre": nombre,
                "0_30": Decimal('0.00'),
                "31_60": Decimal('0.00'),
                "61_90": Decimal('0.00'),
                "90_mas": Decimal('0.00'),
                "total": Decimal('0.00'),
            }

        if dias <= 30:
            data[llave]["0_30"] += saldo
        elif dias <= 60:
            data[llave]["31_60"] += saldo
        elif dias <= 90:
            data[llave]["61_90"] += saldo
        else:
            data[llave]["90_mas"] += saldo

        data[llave]["total"] += saldo

    if proveedor_id:
        try:
            proveedor_id_int = int(proveedor_id)
            proveedor_seleccionado = get_object_or_404(Proveedor, id=proveedor_id_int, empresa=empresa)
            proveedor_nombre_seleccionado = proveedor_seleccionado.nombre
            compras_pendientes_proveedor = [
                compra for compra in compras.filter(proveedor_id=proveedor_id_int)
                if compra.saldo_pendiente > 0
            ]
            if compras_pendientes_proveedor:
                proveedor_resumen = {
                    "compras": len(compras_pendientes_proveedor),
                    "saldo_total": sum((compra.saldo_pendiente for compra in compras_pendientes_proveedor), Decimal('0.00')),
                    "vencidas": sum(1 for compra in compras_pendientes_proveedor if compra.fecha_control_cxp < hoy),
                    "proxima_fecha": min((compra.fecha_control_cxp for compra in compras_pendientes_proveedor)),
                }
        except (TypeError, ValueError):
            proveedor_seleccionado = None
            compras_pendientes_proveedor = None
            proveedor_resumen = None
            proveedor_nombre_seleccionado = ""
    elif proveedor_key.startswith("manual-"):
        proveedor_nombre_seleccionado = proveedor_key.replace("manual-", "").replace("-", " ").strip()
        compras_pendientes_proveedor = [
            compra for compra in compras
            if compra.saldo_pendiente > 0
            and not compra.proveedor_id
            and slugify(compra.proveedor_nombre or "Proveedor sin nombre") == proveedor_key.replace("manual-", "")
        ]
        if compras_pendientes_proveedor:
            proveedor_nombre_seleccionado = compras_pendientes_proveedor[0].proveedor_nombre or proveedor_nombre_seleccionado
            proveedor_resumen = {
                "compras": len(compras_pendientes_proveedor),
                "saldo_total": sum((compra.saldo_pendiente for compra in compras_pendientes_proveedor), Decimal('0.00')),
                "vencidas": sum(1 for compra in compras_pendientes_proveedor if compra.fecha_control_cxp < hoy),
                "proxima_fecha": min((compra.fecha_control_cxp for compra in compras_pendientes_proveedor)),
            }

    data_lista = list(data.values())
    if q:
        data_lista = [
            item for item in data_lista
            if q.lower() in item["nombre"].lower()
        ]

    resumen = {
        "proveedores_con_saldo": len(data_lista),
        "total_cxp": total_cxp,
        "compras_pendientes": sum(1 for compra in compras if compra.saldo_pendiente > 0),
    }

    return render(request, "facturacion/reporte_cxp_premium.html", {
        "empresa": empresa,
        "data": data_lista,
        "resumen": resumen,
        "proveedor_seleccionado": proveedor_seleccionado,
        "proveedor_nombre_seleccionado": proveedor_nombre_seleccionado,
        "proveedor_resumen": proveedor_resumen,
        "compras_pendientes_proveedor": compras_pendientes_proveedor,
        "q": q,
        "proveedores_sugeridos": Proveedor.objects.filter(empresa=empresa).values_list('nombre', flat=True).distinct(),
    })


# =====================================================
# EXPORTAR EXCEL
# =====================================================

@login_required
def exportar_excel_reportes(request, empresa_slug):

    empresa = get_object_or_404(Empresa, slug=empresa_slug)
    if request.GET.get("reporte") == "cxc":
        contexto_cxc = _construir_reporte_cxc(empresa, request.GET)
        wb = Workbook()

        ws_resumen = wb.active
        ws_resumen.title = "Resumen CxC"
        ws_resumen["A1"] = f"EMPRESA: {empresa.nombre}"
        ws_resumen["A3"] = "REPORTE DE CUENTAS POR COBRAR"
        ws_resumen["A5"] = "Clientes con saldo"
        ws_resumen["B5"] = contexto_cxc["resumen"]["clientes_con_saldo"]
        ws_resumen["A6"] = "Facturas pendientes"
        ws_resumen["B6"] = contexto_cxc["resumen"]["facturas_pendientes"]
        ws_resumen["A7"] = "Total cartera"
        ws_resumen["B7"] = float(contexto_cxc["resumen"]["total_cartera"])

        ws_cxc = wb.create_sheet("Antiguedad Cartera")
        ws_cxc.append(["Cliente", "0-30", "31-60", "61-90", "90+", "Total"])
        for item in contexto_cxc["data"]:
            ws_cxc.append([
                item["cliente"].nombre,
                float(item["0_30"]),
                float(item["31_60"]),
                float(item["61_90"]),
                float(item["90_mas"]),
                float(item["total"]),
            ])

        ws_detalle = wb.create_sheet("Detalle Facturas")
        ws_detalle.append(["Cliente", "Factura", "Fecha", "Vencimiento", "Estado", "Total", "Saldo"])
        for factura in contexto_cxc["facturas_con_saldo"]:
            ws_detalle.append([
                factura.cliente.nombre,
                factura.numero_factura or str(factura.id),
                str(factura.fecha_emision),
                str(factura.fecha_vencimiento or ""),
                factura.estado_pago,
                float(factura.total_documento_ajustado),
                float(factura.saldo_pendiente),
            ])

        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value not in (None, ""):
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                sheet.column_dimensions[col_letter].width = max_length + 3

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="Reporte_CxC.xlsx"'
        wb.save(response)
        return response

    facturas = _filtrar_facturas_reporte(empresa, request.GET)

    wb = Workbook()

    ws_resumen = wb.active
    ws_resumen.title = "Resumen Ejecutivo"

    total_facturado = sum((f.total_documento_ajustado for f in facturas), Decimal('0.00'))
    total_saldo = sum((f.saldo_pendiente for f in facturas), Decimal('0.00'))
    total_cobrado = total_facturado - total_saldo
    porcentaje = (total_cobrado / total_facturado * 100) if total_facturado > 0 else 0

    ws_resumen["A1"] = f"EMPRESA: {empresa.nombre}"
    ws_resumen["A3"] = "INDICADORES CLAVE"

    data = [
        ("Total Facturado", total_facturado),
        ("Total Cobrado", total_cobrado),
        ("Total Pendiente", total_saldo),
        ("% Recuperación", round(porcentaje, 2)),
    ]

    row = 5
    for label, value in data:
        ws_resumen.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_resumen.cell(row=row, column=2, value=float(value))
        row += 1

    ws_detalle = wb.create_sheet("Detalle Facturas")

    headers = [
        "Cliente", "Fecha", "Número",
        "Base 15%", "ISV 15%",
        "Base 18%", "ISV 18%",
        "Exento", "Exonerado",
        'Descuento',
        "Total", "Saldo", "Estado"
    ]

    for col, h in enumerate(headers, 1):
        cell = ws_detalle.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1e2a38", fill_type="solid")

    row = 2

    for f in facturas:
        r = f.resumen_fiscal()


        descuento_total = Decimal(str(r.get("descuento_total", 0) or 0))

        ws_detalle.cell(row=row, column=1, value=f.cliente.nombre)
        ws_detalle.cell(row=row, column=2, value=str(f.fecha_emision))
        ws_detalle.cell(row=row, column=3, value=f.numero_factura)

        ws_detalle.cell(row=row, column=4, value=float(r["base_15"]))
        ws_detalle.cell(row=row, column=5, value=float(r["isv_15"]))
        ws_detalle.cell(row=row, column=6, value=float(r["base_18"]))
        ws_detalle.cell(row=row, column=7, value=float(r["isv_18"]))
        ws_detalle.cell(row=row, column=8, value=float(r["base_exento"]))
        ws_detalle.cell(row=row, column=9, value=float(r["base_exonerado"]))
        ws_detalle.cell(row=row, column=10, value=float(descuento_total))
        ws_detalle.cell(row=row, column=11, value=float(f.total_documento_ajustado))
        ws_detalle.cell(row=row, column=12, value=float(f.saldo_pendiente))
        ws_detalle.cell(row=row, column=13, value=f.estado if f.estado == "anulada" else f.estado_pago)

        row += 1

    ws_cliente = wb.create_sheet("Clientes")
    ws_cliente.append(["Cliente", "Total", "Saldo"])

    data_clientes = {}

    for f in facturas:
        c = f.cliente.nombre
        if c not in data_clientes:
            data_clientes[c] = {"total": 0, "saldo": 0}

        data_clientes[c]["total"] += f.total_documento_ajustado
        data_clientes[c]["saldo"] += f.saldo_pendiente

    for c, v in data_clientes.items():
        ws_cliente.append([c, float(v["total"]), float(v["saldo"])])

    ws_cxc = wb.create_sheet("Cuentas por Cobrar")
    ws_cxc.append(["Cliente", "0-30", "31-60", "61-90", "90+", "Total"])

    hoy = date.today()
    aging = {}

    for f in facturas:
        saldo = f.saldo_pendiente
        if saldo <= 0:
            continue

        cliente = f.cliente.nombre
        dias = (hoy - f.fecha_emision).days

        if cliente not in aging:
            aging[cliente] = {"0": 0, "30": 0, "60": 0, "90": 0, "total": 0}

        if dias <= 30:
            aging[cliente]["0"] += saldo
        elif dias <= 60:
            aging[cliente]["30"] += saldo
        elif dias <= 90:
            aging[cliente]["60"] += saldo
        else:
            aging[cliente]["90"] += saldo

        aging[cliente]["total"] += saldo

    for c, v in aging.items():
        ws_cxc.append([
            c,
            float(v["0"]),
            float(v["30"]),
            float(v["60"]),
            float(v["90"]),
            float(v["total"])
        ])

    ws_chart = wb.create_sheet("Gráficos")
    ws_chart["A1"] = "Facturación por Cliente"

    chart = BarChart()
    data_ref = Reference(ws_cliente, min_col=2, min_row=1, max_row=len(data_clientes) + 1)
    cats = Reference(ws_cliente, min_col=1, min_row=2, max_row=len(data_clientes) + 1)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)

    ws_chart.add_chart(chart, "A3")

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            sheet.column_dimensions[col_letter].width = max_length + 3

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=Reporte_Financiero.xlsx"

    wb.save(response)
    return response
