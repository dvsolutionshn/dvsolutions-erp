from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from openpyxl import load_workbook

from .models import Proveedor, RegistroCompraFiscal


MONTO_CERO = Decimal("0.00")


def _normalizar_texto(valor):
    return str(valor or "").strip()


def _normalizar_header(valor):
    texto = _normalizar_texto(valor).lower()
    return (
        texto.replace(".", "")
        .replace(" ", "_")
        .replace("no_de", "numero")
        .replace("n_de", "numero")
        .replace("nro", "numero")
    )


def _a_decimal(valor):
    if valor in (None, ""):
        return MONTO_CERO
    try:
        return Decimal(str(valor)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError, TypeError):
        return MONTO_CERO


def _a_fecha(valor):
    if hasattr(valor, "date"):
        return valor.date()
    if isinstance(valor, str):
        texto = valor.strip()
        for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(texto, formato).date()
            except ValueError:
                continue
    return None


def _detectar_encabezado(sheet):
    for row in range(1, min(sheet.max_row, 40) + 1):
        valores = [sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)]
        normalizados = [_normalizar_header(valor) for valor in valores]
        if "fecha" in normalizados and any("beneficiario" in valor for valor in normalizados):
            return row, normalizados
    raise ValueError("No se encontro la fila de encabezados del libro de compras.")


def _indice(headers, *opciones):
    for opcion in opciones:
        if opcion in headers:
            return headers.index(opcion) + 1
    for index, header in enumerate(headers, start=1):
        if any(opcion in header for opcion in opciones):
            return index
    return None


def importar_libro_compras_desde_excel(empresa, archivo, periodo_anio, periodo_mes):
    workbook = load_workbook(archivo, data_only=True, read_only=True)
    sheet = workbook["COMPRAS"] if "COMPRAS" in workbook.sheetnames else workbook.worksheets[0]
    header_row, headers = _detectar_encabezado(sheet)
    columnas = {
        "fecha": _indice(headers, "fecha"),
        "proveedor": _indice(headers, "beneficiario", "proveedor"),
        "numero": _indice(headers, "numero_factura", "numero_de_factura", "factura"),
        "subtotal": _indice(headers, "subtotal", "sub_total"),
        "isv_15": _indice(headers, "isv_15%", "isv_15", "15%"),
        "isv_18": _indice(headers, "018", "18%", "isv_18"),
        "total": _indice(headers, "total"),
    }

    requeridas = ["fecha", "proveedor", "numero", "subtotal"]
    faltantes = [campo for campo in requeridas if not columnas[campo]]
    if faltantes:
        raise ValueError(f"Faltan columnas requeridas: {', '.join(faltantes)}.")

    resultado = {"creadas": 0, "duplicadas": [], "omitidas": []}
    for row in range(header_row + 1, sheet.max_row + 1):
        fecha = _a_fecha(sheet.cell(row, columnas["fecha"]).value)
        proveedor_nombre = _normalizar_texto(sheet.cell(row, columnas["proveedor"]).value)
        numero_factura = _normalizar_texto(sheet.cell(row, columnas["numero"]).value)
        subtotal = _a_decimal(sheet.cell(row, columnas["subtotal"]).value)
        isv_15 = _a_decimal(sheet.cell(row, columnas["isv_15"]).value) if columnas["isv_15"] else MONTO_CERO
        isv_18 = _a_decimal(sheet.cell(row, columnas["isv_18"]).value) if columnas["isv_18"] else MONTO_CERO
        total = _a_decimal(sheet.cell(row, columnas["total"]).value) if columnas["total"] else MONTO_CERO

        if not fecha and not proveedor_nombre and not numero_factura and subtotal == MONTO_CERO:
            continue
        if not fecha or not proveedor_nombre or not numero_factura:
            resultado["omitidas"].append({"fila": row, "motivo": "Fila incompleta"})
            continue

        if total == MONTO_CERO:
            total = subtotal + isv_15 + isv_18

        proveedor = Proveedor.objects.filter(empresa=empresa, nombre__iexact=proveedor_nombre).first()
        registro = RegistroCompraFiscal(
            empresa=empresa,
            proveedor=proveedor,
            proveedor_nombre=proveedor_nombre,
            proveedor_rtn=proveedor.rtn if proveedor else "",
            numero_factura=numero_factura,
            fecha_documento=fecha,
            periodo_anio=periodo_anio,
            periodo_mes=periodo_mes,
            subtotal=subtotal,
            base_15=subtotal if isv_15 > 0 else MONTO_CERO,
            isv_15=isv_15,
            base_18=(isv_18 / Decimal("0.18")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) if isv_18 > 0 else MONTO_CERO,
            isv_18=isv_18,
            exento=subtotal if isv_15 == 0 and isv_18 == 0 else MONTO_CERO,
            total=total,
            origen_importacion=getattr(archivo, "name", "Libro de compras"),
        )
        duplicada = registro.buscar_duplicada()
        if duplicada:
            resultado["duplicadas"].append({
                "fila": row,
                "numero_factura": numero_factura,
                "proveedor": proveedor_nombre,
                "periodo": f"{duplicada.periodo_mes}/{duplicada.periodo_anio}",
            })
            continue
        registro.save()
        resultado["creadas"] += 1

    return resultado
