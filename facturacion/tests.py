from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
import json
from unittest.mock import patch

from pypdf import PdfReader
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db.models import Sum
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from core.models import ConfiguracionAvanzadaEmpresa, ConfiguracionPowerBIEmpresa, Empresa, EmpresaModulo, Modulo, RegistroAuditoria, RolSistema
from clinica.models import Paciente
from contabilidad.models import AsientoContable, ClasificacionCompraFiscal, CuentaContable, CuentaFinanciera
from contabilidad.services import registrar_asiento_pago_cliente
from .forms import ConfiguracionFacturacionEmpresaForm, ProductoForm
from .models import CAI, BodegaInventario, Cliente, CierreCaja, ComprobanteEgresoCompra, CompraInventario, ConfiguracionFacturacionEmpresa, CorreccionNumeroFactura, EntradaInventarioDocumento, ExistenciaLoteBodega, Factura, HistorialCostoRealProducto, InventarioProducto, LineaCompraInventario, LineaFactura, LineaNotaCredito, LoteInventario, MovimientoInventario, MovimientoLoteBodega, NotaCredito, PagoCompra, PagoFactura, Producto, Proveedor, ReciboPago, RegistroCompraFiscal, TipoImpuesto
from .views import _registrar_entrada_nota_credito


class FacturacionTests(TestCase):
    def setUp(self):
        self.empresa = Empresa.objects.create(
            nombre="Empresa Demo",
            slug="demo",
            rtn="08011999123456",
        )
        self.modulo_facturacion, _ = Modulo.objects.get_or_create(nombre="Facturacion", codigo="facturacion")
        EmpresaModulo.objects.create(empresa=self.empresa, modulo=self.modulo_facturacion, activo=True)
        self.modulo_contabilidad, _ = Modulo.objects.get_or_create(nombre="Contabilidad", codigo="contabilidad")
        EmpresaModulo.objects.create(empresa=self.empresa, modulo=self.modulo_contabilidad, activo=True)
        self.rol_total = RolSistema.objects.create(
            nombre="Administrador Operativo",
            codigo="admin-operativo",
            activo=True,
            puede_punto_venta=True,
            puede_configuracion_facturacion=True,
            puede_cierres_caja=True,
            puede_facturas=True,
            puede_clientes=True,
            puede_productos=True,
            puede_proveedores=True,
            puede_inventario=True,
            puede_compras=True,
            puede_cai=True,
            puede_impuestos=True,
            puede_notas_credito=True,
            puede_recibos=True,
            puede_egresos=True,
            puede_reportes=True,
            puede_cxc=True,
            puede_cxp=True,
            puede_crear_facturas=True,
            puede_editar_facturas=True,
            puede_anular_facturas=True,
            puede_eliminar_borradores=True,
            puede_eliminar_facturas=True,
            puede_registrar_pagos_clientes=True,
            puede_crear_clientes=True,
            puede_editar_clientes=True,
            puede_crear_productos=True,
            puede_editar_productos=True,
            puede_crear_proveedores=True,
            puede_editar_proveedores=True,
            puede_ajustar_inventario=True,
            puede_crear_compras=True,
            puede_editar_compras=True,
            puede_aplicar_compras=True,
            puede_anular_compras=True,
            puede_registrar_pagos_proveedores=True,
            puede_crear_notas_credito=True,
            puede_editar_notas_credito=True,
            puede_anular_notas_credito=True,
            puede_exportar_reportes=True,
        )
        self.user = get_user_model().objects.create_user(
            username="admin",
            password="pass",
            empresa=self.empresa,
            rol_sistema=self.rol_total,
        )
        self.client.force_login(self.user)
        self.cliente = Cliente.objects.create(empresa=self.empresa, nombre="Cliente Demo")
        self.producto = Producto.objects.create(
            empresa=self.empresa,
            nombre="Producto Demo",
            precio=Decimal("100.00"),
        )
        self.impuesto = TipoImpuesto.objects.create(nombre="ISV 15", porcentaje=Decimal("15.00"))
        self.cai = CAI.objects.create(
            empresa=self.empresa,
            numero_cai="CAI-TEST",
            uso_documento="factura",
            establecimiento="001",
            punto_emision="001",
            tipo_documento="01",
            rango_inicial=1,
            rango_final=10,
            correlativo_actual=0,
            fecha_activacion=date(2026, 1, 1),
            fecha_limite=date.today() + timedelta(days=30),
        )
        self.cai_nota_credito = CAI.objects.create(
            empresa=self.empresa,
            numero_cai="CAI-NC",
            uso_documento="nota_credito",
            establecimiento="001",
            punto_emision="002",
            tipo_documento="03",
            rango_inicial=1,
            rango_final=10,
            correlativo_actual=0,
            fecha_activacion=date(2026, 1, 1),
            fecha_limite=date.today() + timedelta(days=30),
        )

    def crear_factura_con_linea(self, estado="emitida", fecha_emision=None):
        factura = Factura.objects.create(
            empresa=self.empresa,
            cliente=self.cliente,
            estado=estado,
            fecha_emision=fecha_emision or date.today(),
        )
        LineaFactura.objects.create(
            factura=factura,
            producto=self.producto,
            cantidad=Decimal("1.00"),
            precio_unitario=Decimal("100.00"),
            impuesto=self.impuesto,
        )
        factura.calcular_totales()
        factura.save(update_fields=["subtotal", "impuesto", "total", "total_lempiras"])
        return factura

    def crear_factura_para_cliente(self, cliente, *, estado="emitida", fecha_emision=None):
        factura = Factura.objects.create(
            empresa=self.empresa,
            cliente=cliente,
            estado=estado,
            fecha_emision=fecha_emision or date.today(),
        )
        LineaFactura.objects.create(
            factura=factura,
            producto=self.producto,
            cantidad=Decimal("1.00"),
            precio_unitario=Decimal("100.00"),
            impuesto=self.impuesto,
        )
        factura.calcular_totales()
        factura.save(update_fields=["subtotal", "impuesto", "total", "total_lempiras"])
        return factura

    def test_linea_factura_total_linea_conserva_centavos(self):
        factura = Factura.objects.create(
            empresa=self.empresa,
            cliente=self.cliente,
            estado="emitida",
            fecha_emision=date.today(),
        )
        linea = LineaFactura.objects.create(
            factura=factura,
            producto=self.producto,
            cantidad=Decimal("1.01"),
            precio_unitario=Decimal("46286.45"),
            impuesto=self.impuesto,
        )

        self.assertEqual(linea.subtotal, Decimal("46749.31"))
        self.assertEqual(linea.impuesto_monto, Decimal("7012.40"))
        self.assertEqual(linea.total_linea, Decimal("53761.71"))

    def test_ver_factura_muestra_total_linea_con_decimales_correctos(self):
        factura = Factura.objects.create(
            empresa=self.empresa,
            cliente=self.cliente,
            estado="emitida",
            fecha_emision=date.today(),
        )
        LineaFactura.objects.create(
            factura=factura,
            producto=self.producto,
            cantidad=Decimal("1.01"),
            precio_unitario=Decimal("46286.45"),
            impuesto=self.impuesto,
        )
        factura.calcular_totales()
        factura.save(update_fields=["subtotal", "impuesto", "total", "total_lempiras"])

        response = self.client.get(reverse("ver_factura", args=[self.empresa.slug, factura.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "53,761.71")

    def test_factura_clasica_genera_pdf_para_empresa_general(self):
        factura = self.crear_factura_con_linea()

        response = self.client.get(
            reverse("descargar_factura_pdf", args=[self.empresa.slug, factura.id])
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        pdf = PdfReader(BytesIO(response.content))
        self.assertEqual(len(pdf.pages), 1)
        texto = pdf.pages[0].extract_text()
        self.assertIn("FACTURA", texto)
        self.assertIn("Cliente Demo", texto)

    def test_factura_termica_genera_pdf_de_80_mm_para_empresa_medica(self):
        self.empresa.slug = "hospital_mia"
        self.empresa.save(update_fields=["slug"])
        configuracion, _ = ConfiguracionFacturacionEmpresa.objects.get_or_create(empresa=self.empresa)
        configuracion.plantilla_factura_pdf = "termica_80mm"
        configuracion.save(update_fields=["plantilla_factura_pdf"])
        factura = self.crear_factura_con_linea()

        response = self.client.get(
            reverse("descargar_factura_pdf", args=[self.empresa.slug, factura.id])
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        pdf = PdfReader(BytesIO(response.content))
        self.assertEqual(len(pdf.pages), 1)
        ancho_puntos = float(pdf.pages[0].mediabox.width)
        ancho_esperado = 80 / 25.4 * 72
        self.assertAlmostEqual(ancho_puntos, ancho_esperado, delta=1)
        alto_puntos = float(pdf.pages[0].mediabox.height)
        alto_esperado = 132 / 25.4 * 72
        self.assertAlmostEqual(alto_puntos, alto_esperado, delta=1)
        texto = pdf.pages[0].extract_text()
        self.assertIn("FACTURA", texto)
        self.assertIn("Sub Total Antes De Impuesto", texto)
        self.assertIn("Base 18%", texto)
        self.assertIn("Exento", texto)
        self.assertIn("Exonerado", texto)
        self.assertIn("ISV 18%", texto)
        self.assertIn("DATOS FISCALES", texto)
        self.assertIn("Documento generado por DV Solutions ERP", texto)

    def test_plantilla_termica_solo_aparece_en_empresas_autorizadas(self):
        configuracion = ConfiguracionFacturacionEmpresa(empresa=self.empresa)
        form = ConfiguracionFacturacionEmpresaForm(instance=configuracion)
        opciones = dict(form.fields["plantilla_factura_pdf"].choices)
        self.assertNotIn("termica_80mm", opciones)

        self.empresa.slug = "medical_spa"
        self.empresa.save(update_fields=["slug"])
        form_medico = ConfiguracionFacturacionEmpresaForm(instance=configuracion)
        opciones_medicas = dict(form_medico.fields["plantilla_factura_pdf"].choices)
        self.assertEqual(opciones_medicas["termica_80mm"], "Factura termica 80 mm")

    def test_demo_1_puede_usar_plantilla_termica_y_autoimpresion_pos(self):
        self.empresa.slug = "demo_1"
        self.empresa.save(update_fields=["slug"])
        configuracion, _ = ConfiguracionFacturacionEmpresa.objects.get_or_create(empresa=self.empresa)
        configuracion.plantilla_factura_pdf = "termica_80mm"
        configuracion.save(update_fields=["plantilla_factura_pdf"])
        factura = self.crear_factura_con_linea()

        form = ConfiguracionFacturacionEmpresaForm(instance=configuracion)
        opciones = dict(form.fields["plantilla_factura_pdf"].choices)
        self.assertEqual(opciones["termica_80mm"], "Factura termica 80 mm")

        response = self.client.get(
            reverse("imprimir_factura_pos", args=[self.empresa.slug, factura.id])
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "frame.contentWindow.print()")
        self.assertContains(
            response,
            reverse("vista_previa_factura_pdf", args=[self.empresa.slug, factura.id]),
        )

    def test_punto_venta_crea_factura_pago_recibo_y_asientos(self):
        modulo_pos, _ = Modulo.objects.get_or_create(nombre="Punto de Venta", codigo="punto_venta")
        EmpresaModulo.objects.create(empresa=self.empresa, modulo=modulo_pos, activo=True)
        self.producto.impuesto_predeterminado = self.impuesto
        self.producto.controla_inventario = False
        self.producto.save()

        response = self.client.post(
            reverse("punto_venta", args=[self.empresa.slug]),
            {
                "payload": json.dumps({
                    "metodo": "efectivo",
                    "referencia": "Caja 1",
                    "items": [
                        {
                            "producto_id": self.producto.id,
                            "cantidad": "2",
                            "precio_unitario": "100.00",
                        }
                    ],
                })
            },
        )

        factura = Factura.objects.get(cliente__nombre="Consumidor Final")
        self.assertRedirects(response, reverse("ver_factura", args=[self.empresa.slug, factura.id]))
        self.assertEqual(factura.estado, "emitida")
        self.assertEqual(factura.subtotal, Decimal("200.00"))
        self.assertEqual(factura.impuesto, Decimal("30.00"))
        self.assertEqual(factura.total, Decimal("230.00"))
        self.assertEqual(factura.estado_pago, "pagado")
        linea = factura.lineas.get()
        self.assertFalse(linea.precio_incluye_impuesto)
        pago = PagoFactura.objects.get(factura=factura)
        self.assertEqual(pago.monto, Decimal("230.00"))
        self.assertTrue(ReciboPago.objects.filter(pago=pago, monto=Decimal("230.00")).exists())
        self.assertTrue(
            AsientoContable.objects.filter(
                empresa=self.empresa,
                documento_tipo="factura",
                documento_id=factura.id,
                estado="contabilizado",
            ).exists()
        )
        self.assertTrue(
            AsientoContable.objects.filter(
                empresa=self.empresa,
                documento_tipo="pago_factura",
                documento_id=pago.id,
                estado="contabilizado",
            ).exists()
        )

    def test_precio_con_impuesto_incluido_conserva_total_exacto(self):
        factura = Factura.objects.create(
            empresa=self.empresa,
            cliente=self.cliente,
            estado="borrador",
            fecha_emision=date.today(),
        )
        linea = LineaFactura.objects.create(
            factura=factura,
            producto=self.producto,
            cantidad=Decimal("1.00"),
            precio_unitario=Decimal("2500.00"),
            precio_incluye_impuesto=True,
            impuesto=self.impuesto,
        )
        factura.calcular_totales()

        self.assertEqual(linea.subtotal, Decimal("2173.91"))
        self.assertEqual(linea.impuesto_monto, Decimal("326.09"))
        self.assertEqual(linea.total_linea, Decimal("2500.00"))
        self.assertEqual(factura.total, Decimal("2500.00"))

    def test_pos_medical_spa_interpreta_precio_catalogo_como_total_final(self):
        self.empresa.slug = "medical_spa"
        self.empresa.save(update_fields=["slug"])
        ConfiguracionFacturacionEmpresa.objects.update_or_create(
            empresa=self.empresa,
            defaults={"precios_incluyen_impuesto": True},
        )
        modulo_pos, _ = Modulo.objects.get_or_create(nombre="Punto de Venta", codigo="punto_venta")
        EmpresaModulo.objects.create(empresa=self.empresa, modulo=modulo_pos, activo=True)
        self.producto.precio = Decimal("2500.00")
        self.producto.impuesto_predeterminado = self.impuesto
        self.producto.controla_inventario = False
        self.producto.save()

        response = self.client.post(
            reverse("punto_venta", args=[self.empresa.slug]),
            {
                "payload": json.dumps({
                    "metodo": "efectivo",
                    "cliente_id": self.cliente.id,
                    "monto_recibido": "2500.00",
                    "items": [{
                        "producto_id": self.producto.id,
                        "cantidad": "1",
                        "precio_unitario": "2500.00",
                    }],
                })
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        factura = Factura.objects.get(pk=response.json()["factura_id"])
        linea = factura.lineas.get()
        self.assertTrue(linea.precio_incluye_impuesto)
        self.assertEqual(linea.subtotal, Decimal("2173.91"))
        self.assertEqual(linea.impuesto_monto, Decimal("326.09"))
        self.assertEqual(factura.total, Decimal("2500.00"))

    def test_punto_venta_ajax_cobra_y_devuelve_ticket_para_imprimir(self):
        self.empresa.slug = "hospital_mia"
        self.empresa.save(update_fields=["slug"])
        modulo_pos, _ = Modulo.objects.get_or_create(nombre="Punto de Venta", codigo="punto_venta")
        EmpresaModulo.objects.create(empresa=self.empresa, modulo=modulo_pos, activo=True)
        self.producto.impuesto_predeterminado = self.impuesto
        self.producto.controla_inventario = False
        self.producto.save()

        response = self.client.post(
            reverse("punto_venta", args=[self.empresa.slug]),
            {
                "payload": json.dumps({
                    "metodo": "efectivo",
                    "cliente_id": self.cliente.id,
                    "monto_recibido": "250.00",
                    "items": [
                        {
                            "producto_id": self.producto.id,
                            "cantidad": "2",
                            "precio_unitario": "100.00",
                        }
                    ],
                })
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        resultado = response.json()
        self.assertTrue(resultado["ok"])
        self.assertEqual(resultado["total"], "200.00")
        self.assertEqual(resultado["cambio"], "50.00")
        factura = Factura.objects.get(pk=resultado["factura_id"])
        pago = PagoFactura.objects.get(factura=factura)
        self.assertIn("Recibido L. 250.00", pago.referencia)
        self.assertEqual(
            resultado["ticket_url"],
            reverse("imprimir_factura_pos", args=[self.empresa.slug, factura.id]),
        )
        impresion = self.client.get(resultado["ticket_url"])
        self.assertEqual(impresion.status_code, 200)
        self.assertContains(impresion, "frame.contentWindow.print()")
        self.assertContains(
            impresion,
            reverse("vista_previa_factura_pdf", args=[self.empresa.slug, factura.id]),
        )

    def test_punto_venta_usa_token_csrf_vigente_y_maneja_respuesta_no_json(self):
        modulo_pos, _ = Modulo.objects.get_or_create(nombre="Punto de Venta", codigo="punto_venta")
        EmpresaModulo.objects.create(empresa=self.empresa, modulo=modulo_pos, activo=True)

        response = self.client.get(reverse("punto_venta", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "currentCsrfToken")
        self.assertContains(response, '"X-CSRFToken": csrfToken')
        self.assertContains(response, "El servidor no pudo completar la venta")

    def test_punto_venta_medico_requiere_cliente_seleccionado(self):
        self.empresa.slug = "medical_spa"
        self.empresa.save(update_fields=["slug"])
        modulo_pos, _ = Modulo.objects.get_or_create(nombre="Punto de Venta", codigo="punto_venta")
        EmpresaModulo.objects.create(empresa=self.empresa, modulo=modulo_pos, activo=True)
        self.producto.impuesto_predeterminado = self.impuesto
        self.producto.controla_inventario = False
        self.producto.save()

        response = self.client.post(
            reverse("punto_venta", args=[self.empresa.slug]),
            {
                "payload": json.dumps({
                    "metodo": "efectivo",
                    "monto_recibido": "250.00",
                    "items": [
                        {
                            "producto_id": self.producto.id,
                            "cantidad": "2",
                            "precio_unitario": "100.00",
                        }
                    ],
                })
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("cliente", response.json()["error"].lower())
        self.assertFalse(Factura.objects.filter(empresa=self.empresa).exists())

    def test_punto_venta_tarjeta_selecciona_banco_por_defecto(self):
        self.empresa.slug = "hospital_mia"
        self.empresa.save(update_fields=["slug"])
        modulo_pos, _ = Modulo.objects.get_or_create(nombre="Punto de Venta", codigo="punto_venta")
        EmpresaModulo.objects.create(empresa=self.empresa, modulo=modulo_pos, activo=True)
        self.producto.impuesto_predeterminado = self.impuesto
        self.producto.controla_inventario = False
        self.producto.save()

        response = self.client.post(
            reverse("punto_venta", args=[self.empresa.slug]),
            {
                "payload": json.dumps({
                    "metodo": "tarjeta",
                    "cliente_id": self.cliente.id,
                    "items": [
                        {
                            "producto_id": self.producto.id,
                            "cantidad": "1",
                            "precio_unitario": "100.00",
                        }
                    ],
                })
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        pago = PagoFactura.objects.get(factura_id=response.json()["factura_id"])
        self.assertEqual(pago.cuenta_financiera.tipo, "banco")

    def test_cierres_caja_respeta_permiso_del_rol(self):
        modulo_pos, _ = Modulo.objects.get_or_create(nombre="Punto de Venta", codigo="punto_venta")
        EmpresaModulo.objects.create(empresa=self.empresa, modulo=modulo_pos, activo=True)
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        self.assertFalse(configuracion.usa_cierre_caja)
        self.rol_total.puede_cierres_caja = False
        self.rol_total.save(update_fields=["puede_cierres_caja"])

        bloqueado = self.client.get(reverse("cierres_caja", args=[self.empresa.slug]))
        self.assertRedirects(
            bloqueado,
            reverse("dashboard", args=[self.empresa.slug]),
            fetch_redirect_response=False,
        )

        self.rol_total.puede_cierres_caja = True
        self.rol_total.save(update_fields=["puede_cierres_caja"])
        permitido = self.client.get(reverse("cierres_caja", args=[self.empresa.slug]))
        self.assertEqual(permitido.status_code, 200)

    def test_hospital_mia_solo_admin_ve_historial_y_resumen_de_cierres(self):
        self.empresa.slug = "hospital_mia"
        self.empresa.save(update_fields=["slug"])
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.usa_cierre_caja = True
        configuracion.save(update_fields=["usa_cierre_caja"])
        cierre = CierreCaja.objects.create(
            empresa=self.empresa,
            cajero=self.user,
            fecha=timezone.localdate(),
            turno="general",
        )

        response = self.client.get(reverse("cierres_caja", args=[self.empresa.slug]))
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["puede_ver_historial"])
        self.assertNotContains(response, "Historial de cierres")

        detalle = self.client.get(reverse("ver_cierre_caja", args=[self.empresa.slug, cierre.id]))
        resumen = self.client.get(reverse("resumen_diario_caja", args=[self.empresa.slug]))
        self.assertRedirects(detalle, reverse("cierres_caja", args=[self.empresa.slug]))
        self.assertRedirects(resumen, reverse("cierres_caja", args=[self.empresa.slug]))

        self.user.es_administrador_empresa = True
        self.user.save(update_fields=["es_administrador_empresa"])
        response = self.client.get(reverse("cierres_caja", args=[self.empresa.slug]))
        self.assertTrue(response.context["puede_ver_historial"])
        self.assertContains(response, "Historial de cierres")
        self.assertEqual(
            self.client.get(reverse("ver_cierre_caja", args=[self.empresa.slug, cierre.id])).status_code,
            200,
        )

    def test_resumen_caja_cuenta_aperturas_solo_por_efectivo(self):
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.usa_cierre_caja = True
        configuracion.save(update_fields=["usa_cierre_caja"])
        factura_efectivo = self.crear_factura_con_linea()
        factura_tarjeta = self.crear_factura_con_linea()
        PagoFactura.objects.create(
            factura=factura_efectivo,
            fecha=timezone.localdate(),
            monto=Decimal("115.00"),
            metodo="efectivo",
            cajero=self.user,
        )
        PagoFactura.objects.create(
            factura=factura_tarjeta,
            fecha=timezone.localdate(),
            monto=Decimal("115.00"),
            metodo="tarjeta",
            cajero=self.user,
        )

        response = self.client.get(reverse("resumen_diario_caja", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Aperturas caja")
        self.assertEqual(response.context["resumen"]["aperturas_caja"], 1)
        self.assertEqual(response.context["resumen_cajeros"][0]["aperturas_caja"], 1)

    def test_cierre_caja_excluye_pagos_anulados_en_empresas_medicas(self):
        self.empresa.slug = "hospital_mia"
        self.empresa.save(update_fields=["slug"])
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.usa_cierre_caja = True
        configuracion.save(update_fields=["usa_cierre_caja"])
        fecha_caja = timezone.localdate()
        factura_activa = self.crear_factura_con_linea()
        factura_anulada = self.crear_factura_con_linea()
        PagoFactura.objects.create(
            factura=factura_activa,
            fecha=fecha_caja,
            monto=Decimal("115.00"),
            metodo="efectivo",
            cajero=self.user,
        )
        PagoFactura.objects.create(
            factura=factura_anulada,
            fecha=fecha_caja,
            monto=Decimal("115.00"),
            metodo="efectivo",
            cajero=self.user,
        )
        factura_anulada.estado = "anulada"
        factura_anulada.save(update_fields=["estado"])

        response = self.client.get(reverse("cierres_caja", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["resumen"]["efectivo_sistema"], Decimal("115.00"))
        self.assertEqual(response.context["resumen"]["total_sistema"], Decimal("115.00"))
        self.assertEqual(response.context["resumen"]["anulaciones"], 1)

        response = self.client.post(
            reverse("cierres_caja", args=[self.empresa.slug]),
            {"fecha": fecha_caja.isoformat(), "turno": "general"},
        )

        self.assertEqual(response.status_code, 302)
        cierre = CierreCaja.objects.get(empresa=self.empresa, cajero=self.user, fecha=fecha_caja)
        self.assertEqual(cierre.efectivo_sistema, Decimal("115.00"))
        self.assertEqual(cierre.total_sistema, Decimal("115.00"))

        self.user.es_administrador_empresa = True
        self.user.save(update_fields=["es_administrador_empresa"])
        resumen = self.client.get(reverse("resumen_diario_caja", args=[self.empresa.slug]))

        self.assertEqual(resumen.status_code, 200)
        self.assertEqual(resumen.context["resumen"]["total"], Decimal("115.00"))
        self.assertEqual(resumen.context["resumen"]["anulaciones"], 1)

    def test_pos_crea_cliente_rapido_para_empresa_medica(self):
        self.empresa.slug = "hospital_mia"
        self.empresa.save(update_fields=["slug"])

        response = self.client.post(
            reverse("pos_crear_cliente_rapido", args=[self.empresa.slug]),
            data=json.dumps({
                "nombre": "Paciente POS",
                "rtn": "0801199911111",
                "telefono": "99990000",
                "correo": "paciente@example.com",
                "ciudad": "San Pedro Sula",
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        resultado = response.json()
        self.assertTrue(resultado["ok"])
        cliente = Cliente.objects.get(pk=resultado["cliente"]["id"])
        self.assertEqual(cliente.rtn, "0801199911111")
        self.assertIsNotNone(cliente.cuenta_contable)

    def test_pos_empresas_medicas_permiten_cliente_sin_correo(self):
        for slug in ("hospital_mia", "medical_spa"):
            with self.subTest(slug=slug):
                self.empresa.slug = slug
                self.empresa.save(update_fields=["slug"])
                response = self.client.post(
                    reverse("pos_crear_cliente_rapido", args=[slug]),
                    data=json.dumps({
                        "nombre": f"Paciente sin correo {slug}",
                        "rtn": f"08011999{1 if slug == 'hospital_mia' else 2:05d}",
                        "telefono": "99990000",
                        "ciudad": "Tegucigalpa",
                    }),
                    content_type="application/json",
                )

                self.assertEqual(response.status_code, 200)
                cliente = Cliente.objects.get(pk=response.json()["cliente"]["id"])
                self.assertFalse(cliente.correo)

    def test_pos_empresas_medicas_mantienen_nombre_identidad_y_telefono_obligatorios(self):
        self.empresa.slug = "hospital_mia"
        self.empresa.save(update_fields=["slug"])

        response = self.client.post(
            reverse("pos_crear_cliente_rapido", args=[self.empresa.slug]),
            data=json.dumps({"correo": "opcional@example.com"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("nombre", response.json()["error"])
        self.assertIn("identidad", response.json()["error"])
        self.assertIn("telefono", response.json()["error"])

    def test_pos_medical_spa_crea_cliente_con_datos_reales_del_modal(self):
        self.empresa.slug = "medical_spa"
        self.empresa.save(update_fields=["slug"])

        response = self.client.post(
            reverse("pos_crear_cliente_rapido", args=[self.empresa.slug]),
            data=json.dumps({
                "nombre": "Osman Ivan Maldonado",
                "rtn": "1706197900050",
                "telefono": "94584821",
                "correo": "osmannaples@icloud.com",
                "telefono_whatsapp": "94584821",
                "fecha_nacimiento": "1979-01-28",
                "ciudad": "Tegucigalpa",
                "canal_preferido": "whatsapp",
                "direccion": "",
                "acepta_promociones": True,
            }),
            content_type="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        cliente = Cliente.objects.get(pk=response.json()["cliente"]["id"])
        self.assertEqual(cliente.fecha_nacimiento, date(1979, 1, 28))
        self.assertEqual(cliente.telefono_whatsapp, "94584821")
        self.assertIsNotNone(cliente.cuenta_contable_id)

    @patch("facturacion.views.asegurar_cuenta_contable_cliente", side_effect=RuntimeError("fallo contable"))
    def test_pos_cliente_rapido_siempre_responde_json_si_falla_contabilidad(self, _mock_cuenta):
        self.empresa.slug = "medical_spa"
        self.empresa.save(update_fields=["slug"])

        response = self.client.post(
            reverse("pos_crear_cliente_rapido", args=[self.empresa.slug]),
            data=json.dumps({
                "nombre": "Cliente Error Controlado",
                "rtn": "0801199911777",
                "telefono": "99990001",
                "correo": "error@example.com",
            }),
            content_type="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response["Content-Type"], "application/json")
        self.assertFalse(response.json()["ok"])
        self.assertFalse(Cliente.objects.filter(nombre="Cliente Error Controlado").exists())

    def test_pos_busca_clientes_en_base_completa_y_documento_sin_guiones(self):
        self.empresa.slug = "hospital_mia"
        self.empresa.save(update_fields=["slug"])
        for indice in range(260):
            Cliente.objects.create(
                empresa=self.empresa,
                nombre=f"Cliente Masivo {indice:03d}",
                rtn=f"08011994{indice:05d}",
                activo=True,
            )
        cliente = Cliente.objects.create(
            empresa=self.empresa,
            nombre="Tatiana Prueba POS",
            rtn="0801-1994-13996",
            telefono="99998888",
            correo="tatiana@example.com",
            activo=True,
        )

        response = self.client.get(
            reverse("pos_buscar_clientes", args=[self.empresa.slug]),
            {"q": "0801199413996"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        resultados = response.json()["clientes"]
        self.assertTrue(any(item["id"] == cliente.id for item in resultados))

    def test_pos_busca_paciente_clinico_y_lo_enlaza_como_cliente(self):
        self.empresa.slug = "medical_spa"
        self.empresa.save(update_fields=["slug"])
        paciente = Paciente.objects.create(
            empresa=self.empresa,
            expediente_codigo="MIA-0999",
            primer_nombre="Mariana",
            primer_apellido="Lopez",
            nombre="Mariana Lopez",
            identidad="0801199413996",
            whatsapp="99991111",
            correo="mariana@example.com",
            activo=True,
        )

        response = self.client.get(
            reverse("pos_buscar_clientes", args=[self.empresa.slug]),
            {"q": "Mariana"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        paciente.refresh_from_db()
        self.assertIsNotNone(paciente.cliente)
        resultados = response.json()["clientes"]
        self.assertTrue(any(item["id"] == paciente.cliente_id for item in resultados))
        self.assertTrue(Cliente.objects.filter(empresa=self.empresa, rtn="0801199413996").exists())

    def test_pos_crea_producto_rapido_con_distribucion_bodega(self):
        self.empresa.slug = "medical_spa"
        self.empresa.save(update_fields=["slug"])
        bodega = BodegaInventario.objects.create(empresa=self.empresa, nombre="Vitrina", tipo="vitrina")

        response = self.client.post(
            reverse("pos_crear_producto_rapido", args=[self.empresa.slug]),
            data=json.dumps({
                "nombre": "Crema POS",
                "codigo": "POS-001",
                "precio": "350.00",
                "lote_inicial": "L-001",
                "vencimiento_lote": "2027-08-31",
                "bodegas": {str(bodega.id): "4"},
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        producto = Producto.objects.get(codigo="POS-001", empresa=self.empresa)
        self.assertEqual(producto.stock_actual, Decimal("4.00"))
        existencia = ExistenciaLoteBodega.objects.get(lote__producto=producto, bodega=bodega)
        self.assertEqual(existencia.cantidad, Decimal("4.00"))

    def test_punto_venta_ajax_rechaza_efectivo_insuficiente_sin_crear_factura(self):
        modulo_pos, _ = Modulo.objects.get_or_create(nombre="Punto de Venta", codigo="punto_venta")
        EmpresaModulo.objects.create(empresa=self.empresa, modulo=modulo_pos, activo=True)
        self.producto.impuesto_predeterminado = self.impuesto
        self.producto.controla_inventario = False
        self.producto.save()

        response = self.client.post(
            reverse("punto_venta", args=[self.empresa.slug]),
            {
                "payload": json.dumps({
                    "metodo": "efectivo",
                    "monto_recibido": "199.99",
                    "items": [
                        {
                            "producto_id": self.producto.id,
                            "cantidad": "2",
                            "precio_unitario": "100.00",
                        }
                    ],
                })
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.json()["ok"])
        self.assertEqual(Factura.objects.filter(empresa=self.empresa).count(), 0)

    def test_punto_venta_expone_codigo_para_lector_de_barras(self):
        modulo_pos, _ = Modulo.objects.get_or_create(nombre="Punto de Venta", codigo="punto_venta")
        EmpresaModulo.objects.create(empresa=self.empresa, modulo=modulo_pos, activo=True)
        self.producto.codigo = "7501234567890"
        self.producto.controla_inventario = False
        self.producto.save()

        response = self.client.get(reverse("punto_venta", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Lector listo")
        self.assertContains(response, "7501234567890")
        self.assertContains(response, "scanCode")
        self.assertContains(response, 'replace(/[^a-z0-9]/g,"")')
        self.assertContains(response, "scannerLastKeyAt > 160")
        self.assertContains(response, "productCode.replace(/^0+/,")

    def test_formulario_producto_prepara_campo_para_codigo_de_barras(self):
        form = ProductoForm(empresa=self.empresa)

        self.assertEqual(form.fields["codigo"].label, "Codigo de barras / SKU")
        self.assertEqual(form.fields["codigo"].widget.attrs["data-barcode-input"], "true")

    def test_formulario_producto_costo_real_solo_empresas_medicas(self):
        form_general = ProductoForm(empresa=self.empresa)
        self.assertNotIn("costo_real_inventario", form_general.fields)

        for slug in ["hospital_mia", "medical_spa"]:
            self.empresa.slug = slug
            self.empresa.save(update_fields=["slug"])
            form_medico = ProductoForm(empresa=self.empresa)
            self.assertIn("costo_real_inventario", form_medico.fields)
            self.assertEqual(form_medico.fields["costo_real_inventario"].label, "Costo real")
            self.assertIn("No afecta el costo promedio contable", form_medico.fields["costo_real_inventario"].help_text)

    def test_precio_final_con_impuesto_solo_aplica_a_empresas_medicas_definidas(self):
        form_general = ProductoForm(empresa=self.empresa)
        self.assertEqual(form_general.fields["precio"].label, "Precio")
        self.assertIn("Precio base", form_general.fields["precio"].help_text)

        for slug in ["hospital_mia", "medical_spa"]:
            self.empresa.slug = slug
            self.empresa.save(update_fields=["slug"])
            form_medico = ProductoForm(empresa=self.empresa)
            self.assertEqual(form_medico.fields["precio"].label, "Precio final (impuesto incluido)")
            self.assertIn("total que pagara", form_medico.fields["precio"].help_text)

    def test_configuracion_no_permite_activar_precio_final_en_erp_general(self):
        configuracion = ConfiguracionFacturacionEmpresa.objects.create(
            empresa=self.empresa,
            precios_incluyen_impuesto=True,
        )

        response = self.client.get(reverse("configuracion_facturacion", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        configuracion.refresh_from_db()
        self.assertFalse(configuracion.precios_incluyen_impuesto)
        self.assertContains(response, "Precio base más impuesto")
        self.assertNotContains(response, 'name="precios_incluyen_impuesto"', html=False)

    def test_nuevo_producto_muestra_estacion_lector_en_empresas_pos(self):
        for slug in ["demo_1", "hospital_mia", "medical_spa"]:
            self.empresa.slug = slug
            self.empresa.save(update_fields=["slug"])
            response = self.client.get(
                reverse("crear_producto_facturacion", args=[self.empresa.slug])
            )
            self.assertEqual(response.status_code, 200)
            self.assertContains(response, "Lector listo")
            self.assertContains(response, "Activar lector")
            self.assertContains(response, 'data-barcode-enabled="true"')

    def test_nuevo_producto_no_duplica_campo_codigo_con_estacion_lector(self):
        self.empresa.slug = "demo_1"
        self.empresa.save(update_fields=["slug"])

        response = self.client.get(
            reverse("crear_producto_facturacion", args=[self.empresa.slug])
        )

        self.assertEqual(response.content.decode().count('id="id_codigo"'), 1)

    def test_pantalla_asigna_codigo_a_producto_existente(self):
        response = self.client.get(
            reverse("codigos_barras_productos", args=[self.empresa.slug])
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Asignar codigos de barras")
        self.assertContains(response, self.producto.nombre)

        response = self.client.post(
            reverse("codigos_barras_productos", args=[self.empresa.slug]),
            {
                "producto_id": self.producto.id,
                "codigo": "7421234567890",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertJSONEqual(
            response.content,
            {
                "ok": True,
                "producto_id": self.producto.id,
                "producto": self.producto.nombre,
                "codigo": "7421234567890",
            },
        )
        self.producto.refresh_from_db()
        self.assertEqual(self.producto.codigo, "7421234567890")

    def test_asignacion_rapida_rechaza_codigo_duplicado(self):
        otro_producto = Producto.objects.create(
            empresa=self.empresa,
            nombre="Producto con barra",
            codigo="7420000000001",
            precio=Decimal("50.00"),
        )

        response = self.client.post(
            reverse("codigos_barras_productos", args=[self.empresa.slug]),
            {
                "producto_id": self.producto.id,
                "codigo": otro_producto.codigo,
            },
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.json()["ok"])
        self.assertIn("Ya existe", response.json()["error"])
        self.producto.refresh_from_db()
        self.assertFalse(self.producto.codigo)

    def test_libro_compras_fiscal_evita_duplicados_entre_meses(self):
        RegistroCompraFiscal.objects.create(
            empresa=self.empresa,
            proveedor_nombre="Proveedor Fiscal",
            numero_factura="001-001-01-00000001",
            fecha_documento=date(2026, 1, 10),
            periodo_anio=2026,
            periodo_mes=1,
            subtotal=Decimal("100.00"),
            base_15=Decimal("100.00"),
            isv_15=Decimal("15.00"),
            total=Decimal("115.00"),
        )
        duplicada = RegistroCompraFiscal(
            empresa=self.empresa,
            proveedor_nombre="Proveedor Fiscal",
            numero_factura="001-001-01-00000001",
            fecha_documento=date(2026, 3, 10),
            periodo_anio=2026,
            periodo_mes=3,
            subtotal=Decimal("100.00"),
            base_15=Decimal("100.00"),
            isv_15=Decimal("15.00"),
            total=Decimal("115.00"),
        )
        with self.assertRaises(ValidationError):
            duplicada.full_clean()

    def test_importar_libro_compras_fiscal_omite_duplicadas(self):
        RegistroCompraFiscal.objects.create(
            empresa=self.empresa,
            proveedor_nombre="Proveedor Fiscal",
            numero_factura="001-001-01-00000001",
            fecha_documento=date(2026, 1, 10),
            periodo_anio=2026,
            periodo_mes=1,
            subtotal=Decimal("100.00"),
            base_15=Decimal("100.00"),
            isv_15=Decimal("15.00"),
            total=Decimal("115.00"),
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "COMPRAS"
        sheet.append([])
        sheet.append(["Fecha", "beneficiario", "No. De factura", "Subtotal", "ISV 15%", "Total"])
        sheet.append([date(2026, 3, 10), "Proveedor Fiscal", "001-001-01-00000001", 100, 15, 115])
        sheet.append([date(2026, 3, 11), "Proveedor Nuevo", "001-001-01-00000002", 200, 30, 230])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        archivo = SimpleUploadedFile(
            "libro_compras.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(
            reverse("importar_libro_compras_fiscal", args=[self.empresa.slug]),
            {"archivo": archivo, "periodo_anio": "2026", "periodo_mes": "3"},
        )
        self.assertRedirects(response, reverse("libro_compras_fiscal_detalle", args=[self.empresa.slug, 2026, 3]))
        self.assertTrue(RegistroCompraFiscal.objects.filter(numero_factura="001-001-01-00000002").exists())
        self.assertEqual(RegistroCompraFiscal.objects.filter(numero_factura="001-001-01-00000001").count(), 1)

    def test_libro_compras_fiscal_muestra_periodos_y_detalle(self):
        RegistroCompraFiscal.objects.create(
            empresa=self.empresa,
            proveedor_nombre="Proveedor Fiscal",
            numero_factura="001-001-01-00000003",
            fecha_documento=date(2026, 3, 10),
            periodo_anio=2026,
            periodo_mes=3,
            subtotal=Decimal("100.00"),
            base_15=Decimal("100.00"),
            isv_15=Decimal("15.00"),
            total=Decimal("115.00"),
        )
        response = self.client.get(reverse("libro_compras_fiscal", args=[self.empresa.slug]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Libros por mes")
        self.assertContains(response, "3/2026")

        response = self.client.get(reverse("libro_compras_fiscal_detalle", args=[self.empresa.slug, 2026, 3]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "001-001-01-00000003")

    def test_crear_libro_compras_guarda_clasificacion_contable(self):
        cuenta_gasto = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="6101",
            nombre="Combustible",
            tipo="gasto",
        )
        clasificacion = ClasificacionCompraFiscal.objects.create(
            empresa=self.empresa,
            nombre="Combustible",
            cuenta_contable=cuenta_gasto,
        )
        response = self.client.post(
            reverse("crear_registro_compra_fiscal", args=[self.empresa.slug]),
            {
                "fecha_documento[]": ["2026-03-10"],
                "clasificacion_id[]": [str(clasificacion.id)],
                "proveedor_nombre[]": ["Proveedor Fiscal"],
                "numero_factura[]": ["001-001-01-00000004"],
                "exento[]": ["0.00"],
                "base_15[]": ["100.00"],
                "base_18[]": ["0.00"],
            },
        )
        self.assertRedirects(response, reverse("libro_compras_fiscal_detalle", args=[self.empresa.slug, 2026, 3]))
        registro = RegistroCompraFiscal.objects.get(numero_factura="001-001-01-00000004")
        self.assertEqual(registro.clasificacion_contable, clasificacion)

    def crear_compra_con_linea(self, estado="aplicada", condicion_pago="contado", dias_credito=0):
        proveedor = Proveedor.objects.create(
            empresa=self.empresa,
            nombre="Proveedor Demo",
            condicion_pago=condicion_pago,
            dias_credito=dias_credito,
        )
        compra = CompraInventario.objects.create(
            empresa=self.empresa,
            proveedor=proveedor,
            proveedor_nombre=proveedor.nombre,
            fecha_documento=date.today(),
            condicion_pago=condicion_pago,
            dias_credito=dias_credito,
            estado=estado,
        )
        LineaCompraInventario.objects.create(
            compra=compra,
            producto=self.producto,
            cantidad=Decimal("3.00"),
            costo_unitario=Decimal("40.00"),
        )
        return compra

    def test_pago_no_puede_superar_saldo(self):
        factura = self.crear_factura_con_linea()

        with self.assertRaises(ValidationError):
            PagoFactura.objects.create(
                factura=factura,
                monto=factura.total + Decimal("1.00"),
                metodo="efectivo",
                fecha=date.today(),
            )

    def test_dashboard_facturacion_muestra_menu_modular(self):
        response = self.client.get(reverse("facturacion_dashboard", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Facturación")
        self.assertContains(response, "Configuracion")
        self.assertContains(response, "Clientes")
        self.assertContains(response, "Productos")
        self.assertContains(response, "Inventario")
        self.assertContains(response, "Notas de Crédito")
        self.assertContains(response, "Recibos")
        self.assertContains(response, "Cuentas por Cobrar")

    def test_reporte_cxc_filtra_por_busqueda_cliente(self):
        cliente_extra = Cliente.objects.create(empresa=self.empresa, nombre="Alchemia Digital Lab")
        self.crear_factura_para_cliente(self.cliente)
        self.crear_factura_para_cliente(cliente_extra)

        response = self.client.get(
            reverse("reporte_cxc", args=[self.empresa.slug]),
            {"q": "Alchemia"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Alchemia Digital Lab")
        self.assertNotContains(response, f'?cliente={self.cliente.id}')
        self.assertContains(response, 'id="cxc-sugerencias"', html=False)

    def test_exportar_excel_cxc_respeta_busqueda_actual(self):
        cliente_extra = Cliente.objects.create(empresa=self.empresa, nombre="Alchemia Digital Lab")
        self.crear_factura_para_cliente(self.cliente)
        self.crear_factura_para_cliente(cliente_extra)

        response = self.client.get(
            reverse("exportar_excel", args=[self.empresa.slug]),
            {"reporte": "cxc", "q": "Alchemia"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        ws = workbook["Antiguedad Cartera"]
        nombres = [ws.cell(row=row, column=1).value for row in range(2, ws.max_row + 1)]
        self.assertEqual(nombres, ["Alchemia Digital Lab"])

    def test_exportar_excel_cxc_respeta_cliente_seleccionado(self):
        cliente_extra = Cliente.objects.create(empresa=self.empresa, nombre="Cliente Secundario")
        self.crear_factura_para_cliente(self.cliente)
        self.crear_factura_para_cliente(cliente_extra)

        response = self.client.get(
            reverse("exportar_excel", args=[self.empresa.slug]),
            {"reporte": "cxc", "cliente": str(cliente_extra.id)},
        )

        workbook = load_workbook(BytesIO(response.content))
        ws = workbook["Detalle Facturas"]
        clientes = [ws.cell(row=row, column=1).value for row in range(2, ws.max_row + 1)]
        self.assertEqual(clientes, ["Cliente Secundario"])

    def test_configuracion_facturacion_guarda_preferencias_por_empresa(self):
        response = self.client.post(
            reverse("configuracion_facturacion", args=[self.empresa.slug]),
            {
                "plantilla_factura_pdf": "alternativa",
                "nombre_comercial_documentos": "Marca Premium Demo",
                "color_primario": "#123456",
                "color_secundario": "#abcdef",
                "logo_ancho_pdf": "145",
                "logo_alto_pdf": "72",
                "mostrar_vendedor": "on",
                "mostrar_descuentos": "on",
                "leyenda_factura": "Gracias por su compra.",
                "pie_factura": "Documento generado por DV Solutions ERP.",
            },
        )

        self.assertRedirects(response, reverse("configuracion_facturacion", args=[self.empresa.slug]))
        configuracion = ConfiguracionFacturacionEmpresa.objects.get(empresa=self.empresa)
        self.assertEqual(configuracion.plantilla_factura_pdf, "alternativa")
        self.assertEqual(configuracion.nombre_comercial_documentos, "Marca Premium Demo")
        self.assertEqual(configuracion.color_primario, "#123456")
        self.assertEqual(configuracion.logo_ancho_pdf, 145)
        self.assertEqual(configuracion.logo_alto_pdf, 72)
        self.assertTrue(configuracion.mostrar_vendedor)
        self.assertFalse(configuracion.mostrar_notas_linea)

    def test_form_configuracion_oculta_plantilla_notas_extensas_si_empresa_no_aplica(self):
        configuracion = ConfiguracionFacturacionEmpresa.objects.create(empresa=self.empresa)
        form = ConfiguracionFacturacionEmpresaForm(
            instance=configuracion,
            permite_plantilla_notas_extensas=False,
        )
        self.assertNotIn(
            ("notas_extensas", "Factura notas extensas"),
            list(form.fields["plantilla_factura_pdf"].choices),
        )

    def test_form_configuracion_muestra_plantilla_notas_extensas_si_empresa_aplica(self):
        configuracion = ConfiguracionFacturacionEmpresa.objects.create(
            empresa=Empresa.objects.create(
                nombre="Digital Planning",
                slug="digital-planning",
                rtn="08011999123457",
            )
        )
        form = ConfiguracionFacturacionEmpresaForm(
            instance=configuracion,
            permite_plantilla_notas_extensas=True,
        )
        self.assertIn(
            ("notas_extensas", "Factura notas extensas"),
            list(form.fields["plantilla_factura_pdf"].choices),
        )

    def test_form_configuracion_oculta_plantilla_independiente_si_empresa_no_aplica(self):
        configuracion = ConfiguracionFacturacionEmpresa.objects.create(empresa=self.empresa)
        form = ConfiguracionFacturacionEmpresaForm(
            instance=configuracion,
            permite_plantilla_independiente=False,
        )
        self.assertNotIn(
            ("independiente", "Factura independiente"),
            list(form.fields["plantilla_factura_pdf"].choices),
        )

    def test_form_configuracion_muestra_plantilla_independiente_si_empresa_aplica(self):
        configuracion = ConfiguracionFacturacionEmpresa.objects.create(empresa=self.empresa)
        form = ConfiguracionFacturacionEmpresaForm(
            instance=configuracion,
            permite_plantilla_independiente=True,
        )
        self.assertIn(
            ("independiente", "Factura independiente"),
            list(form.fields["plantilla_factura_pdf"].choices),
        )

    def test_facturas_dashboard_muestra_listado_facturas(self):
        factura = self.crear_factura_con_linea()

        response = self.client.get(reverse("facturas_dashboard", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, factura.cliente.nombre)

    def test_rol_limitado_bloquea_clientes(self):
        rol_limitado = RolSistema.objects.create(
            nombre="Solo Facturas",
            codigo="solo-facturas",
            puede_facturas=True,
        )
        self.user.rol_sistema = rol_limitado
        self.user.es_administrador_empresa = False
        self.user.save(update_fields=["rol_sistema", "es_administrador_empresa"])

        response = self.client.get(reverse("clientes_facturacion", args=[self.empresa.slug]))

        self.assertRedirects(response, reverse("dashboard", args=[self.empresa.slug]))

    def test_rol_limitado_bloquea_crear_factura_sin_permiso_especifico(self):
        rol_limitado = RolSistema.objects.create(
            nombre="Consulta Facturas",
            codigo="consulta-facturas",
            puede_facturas=True,
        )
        self.user.rol_sistema = rol_limitado
        self.user.es_administrador_empresa = False
        self.user.save(update_fields=["rol_sistema", "es_administrador_empresa"])

        response = self.client.get(reverse("crear_factura", args=[self.empresa.slug]))

        self.assertRedirects(response, reverse("dashboard", args=[self.empresa.slug]))

    def test_rol_limitado_bloquea_crear_clientes_sin_permiso_especifico(self):
        rol_limitado = RolSistema.objects.create(
            nombre="Solo Facturas Activas",
            codigo="solo-facturas-activo",
            puede_facturas=True,
            puede_crear_facturas=True,
        )
        self.user.rol_sistema = rol_limitado
        self.user.es_administrador_empresa = False
        self.user.save(update_fields=["rol_sistema", "es_administrador_empresa"])

        response = self.client.get(reverse("crear_cliente_facturacion", args=[self.empresa.slug]))

        self.assertRedirects(response, reverse("dashboard", args=[self.empresa.slug]))

    def test_facturas_dashboard_filtra_por_cliente_con_sugerencias(self):
        factura = self.crear_factura_con_linea()
        otro_cliente = Cliente.objects.create(empresa=self.empresa, nombre="Zeta Retail")
        Factura.objects.create(
            empresa=self.empresa,
            cliente=otro_cliente,
            estado="borrador",
            fecha_emision=date.today(),
        )

        response = self.client.get(
            reverse("facturas_dashboard", args=[self.empresa.slug]),
            {"q": "Cliente Demo"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="facturas-sugerencias"', html=False)
        self.assertContains(response, "Cliente Demo")
        self.assertEqual(len(response.context["facturas"]), 1)
        self.assertEqual(response.context["facturas"][0].cliente.nombre, "Cliente Demo")

    def test_vista_pago_rechaza_monto_mayor_al_saldo(self):
        factura = self.crear_factura_con_linea()

        response = self.client.post(
            reverse("registrar_pago", args=[self.empresa.slug, factura.id]),
            {"monto": str(factura.total + Decimal("1.00")), "metodo": "efectivo", "fecha": date.today()},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(PagoFactura.objects.filter(factura=factura).count(), 0)

    def test_pago_genera_recibo_automaticamente(self):
        factura = self.crear_factura_con_linea()

        pago = PagoFactura.objects.create(
            factura=factura,
            monto=Decimal("50.00"),
            metodo="efectivo",
            fecha=date.today(),
        )

        self.assertTrue(ReciboPago.objects.filter(pago=pago).exists())
        self.assertEqual(pago.recibo.factura, factura)
        self.assertEqual(pago.recibo.cliente, factura.cliente)
        self.assertEqual(pago.recibo.monto, Decimal("50.00"))
        self.assertEqual(
            pago.recibo.concepto,
            f"Pago aplicado a factura {factura.numero_factura or factura.id}",
        )

    def test_pago_cliente_usa_cuenta_financiera_en_asiento(self):
        modulo_contabilidad, _ = Modulo.objects.get_or_create(
            codigo="contabilidad",
            defaults={"nombre": "Contabilidad"},
        )
        EmpresaModulo.objects.update_or_create(empresa=self.empresa, modulo=modulo_contabilidad, defaults={"activo": True})
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1102",
            nombre="Banco Principal",
            tipo="activo",
        )
        cuenta_financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Principal Lempiras",
            tipo="banco",
            cuenta_contable=cuenta_banco,
        )
        factura = self.crear_factura_con_linea()

        response = self.client.post(
            reverse("registrar_pago", args=[self.empresa.slug, factura.id]),
            {
                "fecha": str(date.today()),
                "monto": "50.00",
                "metodo": "transferencia",
                "cuenta_financiera": str(cuenta_financiera.id),
                "referencia": "DEP-001",
            },
        )

        self.assertRedirects(response, reverse("ver_factura", args=[self.empresa.slug, factura.id]))
        pago = PagoFactura.objects.get(factura=factura, referencia="DEP-001")
        self.assertEqual(pago.cuenta_financiera, cuenta_financiera)
        asiento = AsientoContable.objects.get(documento_tipo="pago_factura", documento_id=pago.id, evento="cobro")
        self.assertTrue(asiento.lineas.filter(cuenta=cuenta_banco, debe=Decimal("50.00")).exists())

    def test_pago_con_retenciones_aplica_total_y_separa_componentes(self):
        modulo_contabilidad, _ = Modulo.objects.get_or_create(
            codigo="contabilidad",
            defaults={"nombre": "Contabilidad"},
        )
        EmpresaModulo.objects.update_or_create(empresa=self.empresa, modulo=modulo_contabilidad, defaults={"activo": True})
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1102",
            nombre="Banco Principal",
            tipo="activo",
        )
        cuenta_financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Principal Lempiras",
            tipo="banco",
            cuenta_contable=cuenta_banco,
        )
        factura = self.crear_factura_con_linea()

        response = self.client.post(
            reverse("registrar_pago", args=[self.empresa.slug, factura.id]),
            {
                "fecha": str(date.today()),
                "monto": "100.00",
                "retencion_isr": "10.00",
                "retencion_isv": "5.00",
                "metodo": "transferencia",
                "cuenta_financiera": str(cuenta_financiera.id),
                "referencia": "DEP-RET-001",
            },
        )

        self.assertRedirects(response, reverse("ver_factura", args=[self.empresa.slug, factura.id]))
        pago = PagoFactura.objects.get(factura=factura, referencia="DEP-RET-001")
        self.assertEqual(pago.total_aplicado, Decimal("115.00"))
        self.assertEqual(pago.total_retenciones, Decimal("15.00"))
        self.assertEqual(pago.impuesto_aplicado, Decimal("15.00"))
        self.assertEqual(pago.subtotal_aplicado, Decimal("100.00"))
        factura.refresh_from_db()
        self.assertEqual(factura.saldo_pendiente, Decimal("0.00"))
        asiento = AsientoContable.objects.get(documento_tipo="pago_factura", documento_id=pago.id, evento="cobro")
        self.assertTrue(asiento.lineas.filter(cuenta__codigo="113003", debe=Decimal("10.00")).exists())
        self.assertTrue(asiento.lineas.filter(cuenta__codigo="113005", debe=Decimal("5.00")).exists())

    def test_pago_total_separa_isv_a_otra_cuenta(self):
        modulo_contabilidad, _ = Modulo.objects.get_or_create(
            codigo="contabilidad",
            defaults={"nombre": "Contabilidad"},
        )
        EmpresaModulo.objects.update_or_create(empresa=self.empresa, modulo=modulo_contabilidad, defaults={"activo": True})
        cuenta_operativa = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1102",
            nombre="Banco Operativo",
            tipo="activo",
        )
        cuenta_isv = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1103",
            nombre="Banco Impuestos",
            tipo="activo",
        )
        cuenta_financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Operativo HNL",
            tipo="banco",
            cuenta_contable=cuenta_operativa,
        )
        cuenta_financiera_impuesto = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Fiscal HNL",
            tipo="banco",
            cuenta_contable=cuenta_isv,
        )
        factura = self.crear_factura_con_linea()

        response = self.client.post(
            reverse("registrar_pago", args=[self.empresa.slug, factura.id]),
            {
                "fecha": str(date.today()),
                "monto": "115.00",
                "retencion_isr": "0.00",
                "retencion_isv": "0.00",
                "separar_isv": "on",
                "cuenta_financiera_impuesto": str(cuenta_financiera_impuesto.id),
                "metodo": "transferencia",
                "cuenta_financiera": str(cuenta_financiera.id),
                "referencia": "DEP-ISV-001",
            },
        )

        self.assertRedirects(response, reverse("ver_factura", args=[self.empresa.slug, factura.id]))
        pago = PagoFactura.objects.get(factura=factura, referencia="DEP-ISV-001")
        self.assertTrue(pago.separar_isv)
        self.assertEqual(pago.subtotal_recibido, Decimal("100.00"))
        self.assertEqual(pago.impuesto_recibido, Decimal("15.00"))
        asiento = AsientoContable.objects.get(documento_tipo="pago_factura", documento_id=pago.id, evento="cobro")
        self.assertTrue(asiento.lineas.filter(cuenta=cuenta_operativa, debe=Decimal("100.00")).exists())
        self.assertTrue(asiento.lineas.filter(cuenta=cuenta_isv, debe=Decimal("15.00")).exists())

    def test_pago_solo_isv_pendiente_permite_aplicar_a_cuenta_fiscal(self):
        modulo_contabilidad, _ = Modulo.objects.get_or_create(
            codigo="contabilidad",
            defaults={"nombre": "Contabilidad"},
        )
        EmpresaModulo.objects.update_or_create(empresa=self.empresa, modulo=modulo_contabilidad, defaults={"activo": True})
        cuenta_isv = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1103",
            nombre="Banco Fiscal",
            tipo="activo",
        )
        cuenta_financiera_impuesto = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Fiscal HNL",
            tipo="banco",
            cuenta_contable=cuenta_isv,
        )
        factura = self.crear_factura_con_linea()
        pago_base = PagoFactura.objects.create(
            factura=factura,
            monto=Decimal("100.00"),
            metodo="transferencia",
            referencia="BASE-001",
            fecha=date.today(),
        )
        PagoFactura.objects.filter(pk=pago_base.pk).update(
            subtotal_aplicado=Decimal("100.00"),
            impuesto_aplicado=Decimal("0.00"),
        )
        factura.refresh_from_db()

        response = self.client.post(
            reverse("registrar_pago", args=[self.empresa.slug, factura.id]),
            {
                "fecha": str(date.today()),
                "monto": "15.00",
                "retencion_isr": "0.00",
                "retencion_isv": "0.00",
                "separar_isv": "on",
                "cuenta_financiera_impuesto": str(cuenta_financiera_impuesto.id),
                "metodo": "transferencia",
                "referencia": "ISV-ONLY-001",
            },
        )

        self.assertRedirects(response, reverse("ver_factura", args=[self.empresa.slug, factura.id]))
        pago = PagoFactura.objects.get(factura=factura, referencia="ISV-ONLY-001")
        self.assertEqual(pago.subtotal_aplicado, Decimal("0.00"))
        self.assertEqual(pago.impuesto_aplicado, Decimal("15.00"))
        self.assertEqual(pago.impuesto_recibido, Decimal("15.00"))
        asiento = AsientoContable.objects.get(documento_tipo="pago_factura", documento_id=pago.id, evento="cobro")
        self.assertTrue(asiento.lineas.filter(cuenta=cuenta_isv, debe=Decimal("15.00")).exists())
        self.assertFalse(asiento.lineas.filter(cuenta__codigo="1101", debe=Decimal("15.00")).exists())
        self.assertFalse(asiento.lineas.filter(cuenta__codigo="1102", debe=Decimal("15.00")).exists())

    def test_pago_separar_isv_sin_cuenta_fiscal_usa_misma_cuenta_del_cobro(self):
        modulo_contabilidad, _ = Modulo.objects.get_or_create(
            codigo="contabilidad",
            defaults={"nombre": "Contabilidad"},
        )
        EmpresaModulo.objects.update_or_create(empresa=self.empresa, modulo=modulo_contabilidad, defaults={"activo": True})
        cuenta_operativa = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1102",
            nombre="Banco Operativo",
            tipo="activo",
        )
        cuenta_financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Operativo HNL",
            tipo="banco",
            cuenta_contable=cuenta_operativa,
        )
        factura = self.crear_factura_con_linea()

        response = self.client.post(
            reverse("registrar_pago", args=[self.empresa.slug, factura.id]),
            {
                "fecha": str(date.today()),
                "monto": "115.00",
                "retencion_isr": "0.00",
                "retencion_isv": "0.00",
                "separar_isv": "on",
                "metodo": "transferencia",
                "cuenta_financiera": str(cuenta_financiera.id),
                "referencia": "DEP-ISV-AUTO-001",
            },
        )

        self.assertRedirects(response, reverse("ver_factura", args=[self.empresa.slug, factura.id]))
        pago = PagoFactura.objects.get(factura=factura, referencia="DEP-ISV-AUTO-001")
        self.assertTrue(pago.separar_isv)
        self.assertIsNone(pago.cuenta_financiera_impuesto)
        asiento = AsientoContable.objects.get(documento_tipo="pago_factura", documento_id=pago.id, evento="cobro")
        self.assertTrue(asiento.lineas.filter(cuenta=cuenta_operativa, debe=Decimal("15.00")).exists())

    def test_pago_fallido_no_deja_registro_guardado(self):
        modulo_contabilidad, _ = Modulo.objects.get_or_create(
            codigo="contabilidad",
            defaults={"nombre": "Contabilidad"},
        )
        EmpresaModulo.objects.update_or_create(empresa=self.empresa, modulo=modulo_contabilidad, defaults={"activo": True})
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1102",
            nombre="Banco Principal",
            tipo="activo",
        )
        cuenta_financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Principal Lempiras",
            tipo="banco",
            cuenta_contable=cuenta_banco,
        )
        factura = self.crear_factura_con_linea()

        with patch("facturacion.views.registrar_asiento_pago_cliente", side_effect=RuntimeError("fallo contable")):
            response = self.client.post(
                reverse("registrar_pago", args=[self.empresa.slug, factura.id]),
                {
                    "fecha": str(date.today()),
                    "monto": "50.00",
                    "metodo": "transferencia",
                    "cuenta_financiera": str(cuenta_financiera.id),
                    "referencia": "DEP-ERR-001",
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(PagoFactura.objects.filter(factura=factura, referencia="DEP-ERR-001").exists())

    def test_registrar_pago_modal_responde_con_postmessage_al_guardar(self):
        modulo_contabilidad, _ = Modulo.objects.get_or_create(
            codigo="contabilidad",
            defaults={"nombre": "Contabilidad"},
        )
        EmpresaModulo.objects.update_or_create(empresa=self.empresa, modulo=modulo_contabilidad, defaults={"activo": True})
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1102",
            nombre="Banco Principal",
            tipo="activo",
        )
        cuenta_financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Principal Lempiras",
            tipo="banco",
            cuenta_contable=cuenta_banco,
        )
        factura = self.crear_factura_con_linea()

        response = self.client.post(
            reverse("registrar_pago", args=[self.empresa.slug, factura.id]) + "?modal=1",
            {
                "fecha": str(date.today()),
                "monto": "50.00",
                "metodo": "transferencia",
                "cuenta_financiera": str(cuenta_financiera.id),
                "referencia": "DEP-MODAL-001",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "erp-pago-guardado")
        self.assertTrue(PagoFactura.objects.filter(factura=factura, referencia="DEP-MODAL-001").exists())

    def test_registrar_pago_modal_no_renderiza_shell_completo_del_erp(self):
        factura = self.crear_factura_con_linea()

        response = self.client.get(
            reverse("registrar_pago", args=[self.empresa.slug, factura.id]) + "?modal=1"
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Pago rapido")
        self.assertNotContains(response, "Panel Principal")
        self.assertNotContains(response, "DV Solutions ERP")

    def test_editar_ultimo_pago_actualiza_recibo_y_asiento(self):
        modulo_contabilidad, _ = Modulo.objects.get_or_create(
            codigo="contabilidad",
            defaults={"nombre": "Contabilidad"},
        )
        EmpresaModulo.objects.update_or_create(empresa=self.empresa, modulo=modulo_contabilidad, defaults={"activo": True})
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110201",
            nombre="Banco Moneda Nacional",
            tipo="activo",
            acepta_movimientos=True,
        )
        cuenta_clientes = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1110",
            nombre="Cuentas por Cobrar Clientes",
            tipo="activo",
            acepta_movimientos=True,
        )
        cuenta_isv = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="2102",
            nombre="Impuestos por Pagar",
            tipo="pasivo",
            acepta_movimientos=True,
        )
        banco = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Principal HNL",
            tipo="banco",
            cuenta_contable=cuenta_banco,
            activa=True,
        )
        factura = self.crear_factura_con_linea()
        pago = PagoFactura.objects.create(
            factura=factura,
            fecha=date.today(),
            monto=Decimal("50.00"),
            metodo="transferencia",
            referencia="DEP-EDIT-001",
            cuenta_financiera=banco,
            cajero=self.user,
        )
        registrar_asiento_pago_cliente(pago)

        response = self.client.post(
            reverse("editar_pago_factura", args=[self.empresa.slug, factura.id, pago.id]) + "?modal=1",
            {
                "fecha": str(date.today()),
                "monto": "60.00",
                "metodo": "transferencia",
                "cuenta_financiera": str(banco.id),
                "referencia": "DEP-EDIT-002",
                "retencion_isr": "0.00",
                "retencion_isv": "0.00",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "erp-pago-guardado")

        pago.refresh_from_db()
        self.assertEqual(pago.monto, Decimal("60.00"))
        self.assertEqual(pago.referencia, "DEP-EDIT-002")
        self.assertEqual(pago.recibo.monto, Decimal("60.00"))
        self.assertEqual(pago.recibo.referencia, "DEP-EDIT-002")

        asientos = AsientoContable.objects.filter(
            empresa=self.empresa,
            documento_tipo="pago_factura",
            documento_id=pago.id,
            evento="cobro",
        )
        self.assertEqual(asientos.count(), 1)
        asiento = asientos.first()
        factura.cliente.refresh_from_db()
        self.assertTrue(asiento.lineas.filter(cuenta=cuenta_banco, debe=Decimal("60.00")).exists())
        self.assertTrue(asiento.lineas.filter(cuenta=factura.cliente.cuenta_contable, haber=Decimal("60.00")).exists())

    def test_editar_pago_anterior_recontabiliza_todos_los_pagos(self):
        modulo_contabilidad, _ = Modulo.objects.get_or_create(
            codigo="contabilidad",
            defaults={"nombre": "Contabilidad"},
        )
        EmpresaModulo.objects.update_or_create(empresa=self.empresa, modulo=modulo_contabilidad, defaults={"activo": True})
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110201",
            nombre="Banco Moneda Nacional",
            tipo="activo",
            acepta_movimientos=True,
        )
        cuenta_clientes = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1110",
            nombre="Cuentas por Cobrar Clientes",
            tipo="activo",
            acepta_movimientos=True,
        )
        CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="2102",
            nombre="Impuestos por Pagar",
            tipo="pasivo",
            acepta_movimientos=True,
        )
        banco = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Principal HNL",
            tipo="banco",
            cuenta_contable=cuenta_banco,
            activa=True,
        )
        factura = self.crear_factura_con_linea()
        pago_1 = PagoFactura.objects.create(
            factura=factura,
            fecha=date.today(),
            monto=Decimal("40.00"),
            metodo="transferencia",
            referencia="DEP-EDIT-A",
            cuenta_financiera=banco,
            cajero=self.user,
        )
        registrar_asiento_pago_cliente(pago_1)
        pago_2 = PagoFactura.objects.create(
            factura=factura,
            fecha=date.today(),
            monto=Decimal("50.00"),
            metodo="transferencia",
            referencia="DEP-EDIT-B",
            cuenta_financiera=banco,
            cajero=self.user,
        )
        registrar_asiento_pago_cliente(pago_2)

        response = self.client.post(
            reverse("editar_pago_factura", args=[self.empresa.slug, factura.id, pago_1.id]) + "?modal=1",
            {
                "fecha": str(date.today()),
                "monto": "45.00",
                "metodo": "transferencia",
                "cuenta_financiera": str(banco.id),
                "referencia": "DEP-EDIT-A2",
                "retencion_isr": "0.00",
                "retencion_isv": "0.00",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "erp-pago-guardado")

        pago_1.refresh_from_db()
        pago_2.refresh_from_db()
        self.assertEqual(pago_1.monto, Decimal("45.00"))
        self.assertEqual(pago_1.recibo.referencia, "DEP-EDIT-A2")

        asientos = AsientoContable.objects.filter(
            empresa=self.empresa,
            documento_tipo="pago_factura",
            documento_id__in=[pago_1.id, pago_2.id],
            evento="cobro",
        )
        self.assertEqual(asientos.count(), 2)
        self.assertTrue(
            AsientoContable.objects.filter(
                empresa=self.empresa,
                documento_tipo="pago_factura",
                documento_id=pago_1.id,
                evento="cobro",
                lineas__cuenta=cuenta_banco,
                lineas__debe=Decimal("45.00"),
            ).exists()
        )
        factura.cliente.refresh_from_db()
        self.assertTrue(
            AsientoContable.objects.filter(
                empresa=self.empresa,
                documento_tipo="pago_factura",
                documento_id=pago_2.id,
                evento="cobro",
                lineas__cuenta=factura.cliente.cuenta_contable,
                lineas__haber=Decimal("50.00"),
            ).exists()
        )

    def test_registrar_pago_prepara_cuentas_financieras_base(self):
        CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1101",
            nombre="Caja General",
            tipo="activo",
            acepta_movimientos=True,
        )
        CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110201",
            nombre="Banco Moneda Nacional",
            tipo="activo",
            acepta_movimientos=True,
        )
        factura = self.crear_factura_con_linea()

        response = self.client.get(reverse("registrar_pago", args=[self.empresa.slug, factura.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Caja General - 1101 Caja General")
        self.assertContains(response, "Banco Principal HNL - 110201 Banco Moneda Nacional")
        self.assertContains(response, "Misma cuenta del cobro")
        self.assertTrue(CuentaFinanciera.objects.filter(empresa=self.empresa, nombre="Caja General", tipo="caja").exists())

    def test_registrar_pago_carga_catalogo_base_si_no_hay_cuentas(self):
        factura = self.crear_factura_con_linea()

        response = self.client.get(reverse("registrar_pago", args=[self.empresa.slug, factura.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Caja General - 1101 Caja General")
        self.assertContains(response, "Banco Principal HNL - 110201 Banco Moneda Nacional")
        self.assertTrue(CuentaContable.objects.filter(empresa=self.empresa, codigo="110201").exists())
        self.assertTrue(CuentaFinanciera.objects.filter(empresa=self.empresa, nombre="Banco Principal HNL", tipo="banco").exists())

    def test_registrar_pago_repara_banco_principal_hnl_a_110201(self):
        cuenta_padre_bancos = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1102",
            nombre="Bancos",
            tipo="activo",
            acepta_movimientos=True,
        )
        cuenta_banco_correcta = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110201",
            nombre="Banco Moneda Nacional",
            tipo="activo",
            cuenta_padre=cuenta_padre_bancos,
            acepta_movimientos=True,
        )
        CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Principal HNL",
            tipo="banco",
            cuenta_contable=cuenta_padre_bancos,
        )
        factura = self.crear_factura_con_linea()

        self.client.get(reverse("registrar_pago", args=[self.empresa.slug, factura.id]))

        banco = CuentaFinanciera.objects.get(empresa=self.empresa, nombre="Banco Principal HNL")
        self.assertEqual(banco.cuenta_contable, cuenta_banco_correcta)

    def test_recibos_dashboard_muestra_recibo(self):
        factura = self.crear_factura_con_linea()
        pago = PagoFactura.objects.create(
            factura=factura,
            monto=Decimal("50.00"),
            metodo="efectivo",
            fecha=date.today(),
        )
        pago.recibo.concepto = "Abono parcial por transferencia"
        pago.recibo.save(update_fields=["concepto"])

        response = self.client.get(reverse("recibos_dashboard", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, pago.recibo.numero_recibo)
        self.assertContains(response, "Abono parcial por transferencia")

    def test_editar_recibo_actualiza_concepto_documental(self):
        factura = self.crear_factura_con_linea()
        pago = PagoFactura.objects.create(
            factura=factura,
            monto=Decimal("50.00"),
            metodo="efectivo",
            fecha=date.today(),
            referencia="REF-001",
        )

        response = self.client.post(
            reverse("editar_recibo", args=[self.empresa.slug, pago.recibo.id]),
            {
                "fecha": "2026-05-30",
                "referencia": "REF-EDITADA",
                "concepto": "Cancelacion parcial de factura segun acuerdo",
            },
        )

        self.assertRedirects(response, reverse("ver_recibo", args=[self.empresa.slug, pago.recibo.id]))
        pago.recibo.refresh_from_db()
        self.assertEqual(str(pago.recibo.fecha), "2026-05-30")
        self.assertEqual(pago.recibo.referencia, "REF-EDITADA")
        self.assertEqual(pago.recibo.concepto, "Cancelacion parcial de factura segun acuerdo")
        pago.refresh_from_db()
        self.assertEqual(pago.monto, Decimal("50.00"))

    def test_recibo_pago_usa_consecutivo_global_y_no_repite_entre_empresas(self):
        otra_empresa = Empresa.objects.create(nombre="Otra empresa", slug="otra-empresa")
        ConfiguracionAvanzadaEmpresa.objects.create(empresa=otra_empresa)
        cliente_otro = Cliente.objects.create(
            empresa=otra_empresa,
            nombre="Cliente Global",
            rtn="08011999123456",
            telefono="9999-9999",
            correo="global@example.com",
            direccion="Centro",
        )
        factura_otra = Factura.objects.create(
            empresa=otra_empresa,
            cliente=cliente_otro,
            subtotal=Decimal("100.00"),
            impuesto=Decimal("15.00"),
            total=Decimal("115.00"),
            total_lempiras=Decimal("115.00"),
            estado="borrador",
        )
        pago_otro = PagoFactura.objects.create(
            factura=factura_otra,
            monto=Decimal("50.00"),
            metodo="efectivo",
            fecha=date.today(),
        )
        factura_local = self.crear_factura_con_linea()
        pago_local = PagoFactura.objects.create(
            factura=factura_local,
            monto=Decimal("50.00"),
            metodo="efectivo",
            fecha=date.today(),
        )

        self.assertEqual(pago_otro.recibo.numero_recibo, "REC-00000001")
        self.assertEqual(pago_local.recibo.numero_recibo, "REC-00000002")

    def test_anular_factura_cambia_estado(self):
        factura = self.crear_factura_con_linea()

        response = self.client.post(
            reverse("anular_factura", args=[self.empresa.slug, factura.id]),
            {"motivo": "Documento emitido por error"},
        )
        factura.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(factura.estado, "anulada")

    def test_anular_factura_exige_motivo_y_lo_registra_en_bitacora(self):
        factura = self.crear_factura_con_linea()
        url = reverse("anular_factura", args=[self.empresa.slug, factura.id])

        response = self.client.post(url, {"motivo": "no"})
        factura.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(factura.estado, "emitida")

        motivo = "Tarjeta del cliente denegada"
        response = self.client.post(url, {"motivo": motivo})
        factura.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(factura.estado, "anulada")
        evento = RegistroAuditoria.objects.filter(
            empresa=self.empresa,
            app_label="facturacion",
            modelo="factura",
            objeto_id=str(factura.id),
            accion=RegistroAuditoria.ACCION_MODIFICAR,
        ).latest("fecha")
        self.assertEqual(evento.usuario, self.user)
        self.assertEqual(evento.motivo, motivo)
        self.assertEqual(evento.cambios["estado"]["anterior"], "emitida")
        self.assertEqual(evento.cambios["estado"]["nuevo"], "anulada")

    def test_anular_factura_pos_revierte_asiento_del_cobro(self):
        factura = self.crear_factura_con_linea()
        cuenta = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco tarjeta",
            tipo="banco",
            cuenta_contable=CuentaContable.objects.create(
                empresa=self.empresa,
                codigo="1102.0099",
                nombre="Banco tarjeta",
                tipo="activo",
            ),
        )
        pago = PagoFactura.objects.create(
            factura=factura,
            fecha=date.today(),
            monto=factura.total,
            metodo="tarjeta",
            cuenta_financiera=cuenta,
            cajero=self.user,
        )
        registrar_asiento_pago_cliente(pago)

        response = self.client.post(
            reverse("anular_factura", args=[self.empresa.slug, factura.id]),
            {"motivo": "Tarjeta denegada por el banco"},
        )

        self.assertEqual(response.status_code, 302)
        factura.refresh_from_db()
        self.assertEqual(factura.estado, "anulada")
        self.assertTrue(
            AsientoContable.objects.filter(
                empresa=self.empresa,
                documento_tipo="pago_factura",
                documento_id=pago.id,
                evento="anulacion",
                estado="contabilizado",
            ).exists()
        )

    def test_borrador_sin_cai_se_puede_ver(self):
        factura = self.crear_factura_con_linea(estado="borrador")
        factura.cai = None
        factura.numero_factura = None
        factura.save(update_fields=["cai", "numero_factura"])

        response = self.client.get(reverse("ver_factura", args=[self.empresa.slug, factura.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sin CAI asignado")

    def test_eliminar_factura_borrador_sin_numero(self):
        factura = self.crear_factura_con_linea(estado="borrador")
        factura.cai = None
        factura.numero_factura = None
        factura.save(update_fields=["cai", "numero_factura"])

        response = self.client.post(reverse("eliminar_factura_borrador", args=[self.empresa.slug, factura.id]))

        self.assertEqual(response.status_code, 302)
        self.assertFalse(Factura.objects.filter(id=factura.id).exists())

    def test_rol_facturacion_sin_permisos_criticos_no_ve_ni_ejecuta_anular_o_eliminar(self):
        rol_facturacion = RolSistema.objects.create(
            nombre="Facturacion Segura",
            codigo="facturacion-segura",
            puede_facturas=True,
            puede_crear_facturas=True,
            puede_editar_facturas=True,
        )
        self.user.rol_sistema = rol_facturacion
        self.user.es_administrador_empresa = False
        self.user.save(update_fields=["rol_sistema", "es_administrador_empresa"])
        factura = self.crear_factura_con_linea(estado="emitida")

        detalle = self.client.get(reverse("ver_factura", args=[self.empresa.slug, factura.id]))

        self.assertEqual(detalle.status_code, 200)
        self.assertNotContains(detalle, "Anular factura")
        self.assertNotContains(detalle, "Eliminar factura")

        anulacion = self.client.post(
            reverse("anular_factura", args=[self.empresa.slug, factura.id]),
            {"motivo": "Intento no autorizado"},
        )
        eliminacion = self.client.post(
            reverse("eliminar_factura", args=[self.empresa.slug, factura.id]),
        )

        self.assertRedirects(
            anulacion,
            reverse("dashboard", args=[self.empresa.slug]),
            fetch_redirect_response=False,
        )
        self.assertRedirects(
            eliminacion,
            reverse("dashboard", args=[self.empresa.slug]),
            fetch_redirect_response=False,
        )
        factura.refresh_from_db()
        self.assertEqual(factura.estado, "emitida")

    def test_no_elimina_factura_emitida(self):
        factura = self.crear_factura_con_linea(estado="emitida")

        response = self.client.post(reverse("eliminar_factura", args=[self.empresa.slug, factura.id]))

        self.assertEqual(response.status_code, 302)
        self.assertTrue(Factura.objects.filter(id=factura.id).exists())

    def test_empresa_especial_puede_eliminar_factura_emitida_historica(self):
        self.empresa.nombre = "AMKT Digital"
        self.empresa.slug = "amkt-digital"
        self.empresa.save(update_fields=["nombre", "slug"])
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.permite_cai_historico = True
        configuracion.save(update_fields=["permite_cai_historico"])
        factura = self.crear_factura_con_linea(estado="emitida")

        response = self.client.post(reverse("eliminar_factura", args=[self.empresa.slug, factura.id]))

        self.assertEqual(response.status_code, 302)
        self.assertFalse(Factura.objects.filter(id=factura.id).exists())

    def test_eliminar_factura_historica_recalcula_correlativo_cai(self):
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.permite_cai_historico = True
        configuracion.save(update_fields=["permite_cai_historico"])

        factura_1 = self.crear_factura_con_linea(estado="emitida")
        factura_2 = self.crear_factura_con_linea(estado="emitida")

        self.assertEqual(factura_1.numero_factura, "001-001-01-00000001")
        self.assertEqual(factura_2.numero_factura, "001-001-01-00000002")

        response = self.client.post(reverse("eliminar_factura", args=[self.empresa.slug, factura_2.id]))

        self.assertEqual(response.status_code, 302)
        self.assertFalse(Factura.objects.filter(id=factura_2.id).exists())

        self.cai.refresh_from_db()
        self.assertEqual(self.cai.correlativo_actual, 1)

        factura_3 = self.crear_factura_con_linea(estado="emitida")
        self.assertEqual(factura_3.numero_factura, "001-001-01-00000002")

    def test_recalculo_correlativo_ignora_borradores_y_anuladas(self):
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.permite_cai_historico = True
        configuracion.save(update_fields=["permite_cai_historico"])

        factura_emitida = self.crear_factura_con_linea(estado="emitida")
        self.assertEqual(factura_emitida.numero_factura, "001-001-01-00000001")

        factura_anulada = self.crear_factura_con_linea(estado="borrador")
        factura_anulada.fecha_emision = date.today()
        factura_anulada.numero_factura = "001-001-01-00000002"
        factura_anulada.estado = "anulada"
        factura_anulada.save(update_fields=["fecha_emision", "numero_factura", "estado"])

        factura_borrador = self.crear_factura_con_linea(estado="borrador")
        factura_borrador.fecha_emision = date.today()
        factura_borrador.numero_factura = "001-001-01-00000003"
        factura_borrador.save(update_fields=["fecha_emision", "numero_factura"])

        self.client.post(reverse("eliminar_factura", args=[self.empresa.slug, factura_emitida.id]))

        self.cai.refresh_from_db()
        self.assertEqual(self.cai.correlativo_actual, 0)

        nueva = self.crear_factura_con_linea(estado="emitida")
        self.assertEqual(nueva.numero_factura, "001-001-01-00000001")

    def test_empresa_especial_puede_editar_cai_usado(self):
        self.empresa.nombre = "INTEGRATED SALES AND SERVICES S. DE R.L."
        self.empresa.slug = "integrated-sales-and-services-s-de-r-l"
        self.empresa.save(update_fields=["nombre", "slug"])
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.permite_cai_historico = True
        configuracion.save(update_fields=["permite_cai_historico"])
        factura = self.crear_factura_con_linea(estado="emitida")

        self.cai.numero_cai = "CAI-AJUSTADO"
        self.cai.save()

        self.cai.refresh_from_db()
        factura.refresh_from_db()
        self.assertEqual(self.cai.numero_cai, "CAI-AJUSTADO")
        self.assertEqual(factura.cai_numero_historico, "CAI-TEST")

    def test_empresa_especial_puede_eliminar_cai_sin_documentos_asociados(self):
        self.empresa.nombre = "AMKT Digital"
        self.empresa.slug = "amkt-digital"
        self.empresa.save(update_fields=["nombre", "slug"])
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.permite_cai_historico = True
        configuracion.save(update_fields=["permite_cai_historico"])
        cai = CAI.objects.create(
            empresa=self.empresa,
            numero_cai="CAI-BORRAR",
            uso_documento="factura",
            establecimiento="009",
            punto_emision="001",
            tipo_documento="01",
            rango_inicial=91,
            rango_final=100,
            correlativo_actual=90,
            fecha_activacion=date(2026, 1, 1),
            fecha_limite=date(2026, 12, 31),
            activo=True,
        )

        response = self.client.post(reverse("eliminar_cai_facturacion", args=[self.empresa.slug, cai.id]))

        self.assertEqual(response.status_code, 302)
        self.assertFalse(CAI.objects.filter(id=cai.id).exists())

    def test_duplicar_factura_crea_borrador_sin_numero(self):
        factura = self.crear_factura_con_linea(estado="emitida")

        response = self.client.post(reverse("duplicar_factura", args=[self.empresa.slug, factura.id]))
        factura_nueva = Factura.objects.exclude(id=factura.id).get()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(factura_nueva.estado, "borrador")
        self.assertIsNone(factura_nueva.numero_factura)
        self.assertIsNone(factura_nueva.cai)
        self.assertEqual(factura_nueva.fecha_emision, timezone.now().date())
        self.assertEqual(factura_nueva.lineas.count(), factura.lineas.count())
        self.assertEqual(factura_nueva.total, factura.total)

    def test_crear_cliente_desde_facturacion(self):
        response = self.client.post(
            reverse("crear_cliente_facturacion", args=[self.empresa.slug]),
            {
                "nombre": "Cliente Nuevo",
                "rtn": "08011999111111",
                "direccion": "Tegucigalpa",
                "ciudad": "Tegucigalpa",
                "activo": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(Cliente.objects.filter(empresa=self.empresa, nombre="Cliente Nuevo").exists())

    def test_facturacion_empresas_medicas_exige_datos_generales_pero_no_correo(self):
        self.empresa.slug = "medical_spa"
        self.empresa.save(update_fields=["slug"])
        url = reverse("crear_cliente_facturacion", args=[self.empresa.slug])

        response = self.client.post(
            url,
            {
                "nombre": "Cliente sin correo",
                "rtn": "0801199912345",
                "telefono": "99990000",
                "direccion": "Tegucigalpa",
                "ciudad": "Tegucigalpa",
                "activo": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        cliente = Cliente.objects.get(empresa=self.empresa, nombre="Cliente sin correo")
        self.assertFalse(cliente.correo)

        response = self.client.post(
            url,
            {
                "nombre": "Cliente incompleto",
                "correo": "opcional@example.com",
                "activo": "on",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertFormError(response.context["form"], "rtn", "Este campo es obligatorio.")
        self.assertFormError(response.context["form"], "telefono", "Este campo es obligatorio.")
        self.assertFalse(Cliente.objects.filter(empresa=self.empresa, nombre="Cliente incompleto").exists())

    def test_cliente_medico_no_se_puede_crear_sin_identidad_desde_modelo(self):
        for slug in ("hospital_mia", "medical_spa"):
            with self.subTest(slug=slug):
                self.empresa.slug = slug
                self.empresa.save(update_fields=["slug"])

                with self.assertRaisesMessage(ValidationError, "identidad/RTN es obligatoria"):
                    Cliente.objects.create(
                        empresa=self.empresa,
                        nombre=f"Cliente sin identidad {slug}",
                        telefono="99990000",
                    )

    def test_cliente_se_comparte_entre_luque_hospital_y_medical_spa(self):
        self.empresa.slug = "hospital_mia"
        self.empresa.save(update_fields=["slug"])
        luque = Empresa.objects.create(
            nombre="Luque Aestetic",
            slug="luque_aestetic",
            rtn="08011999100001",
        )
        medical_spa = Empresa.objects.create(
            nombre="Medical Spa",
            slug="medical_spa",
            rtn="08011999100002",
        )

        response = self.client.post(
            reverse("crear_cliente_facturacion", args=[self.empresa.slug]),
            {
                "nombre": "Paciente Compartido",
                "rtn": "0801199912777",
                "telefono": "99887766",
                "direccion": "Colonia Centro",
                "ciudad": "Tegucigalpa",
                "activo": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        fichas = Cliente.objects.filter(
            empresa__in=[self.empresa, luque, medical_spa],
            rtn="0801199912777",
        ).select_related("empresa")
        self.assertEqual(fichas.count(), 3)
        self.assertEqual(len({ficha.perfil_compartido_id for ficha in fichas}), 1)
        self.assertEqual(len({ficha.empresa_id for ficha in fichas}), 3)
        paciente = Paciente.objects.get(empresa=self.empresa, identidad="0801199912777")
        self.assertEqual(paciente.cliente, fichas.get(empresa=self.empresa))
        self.assertEqual(paciente.telefono, "99887766")

        origen = fichas.get(empresa=self.empresa)
        origen.telefono = "99991111"
        origen.correo = "compartido@example.com"
        origen.save(update_fields=["telefono", "correo"])

        self.assertFalse(
            fichas.exclude(empresa=self.empresa).exclude(
                telefono="99991111",
                correo="compartido@example.com",
            ).exists()
        )
        self.assertIsNotNone(origen.cuenta_contable_id)
        self.assertTrue(
            Cliente.objects.filter(
                empresa__in=[luque, medical_spa],
                rtn="0801199912777",
                cuenta_contable__isnull=True,
            ).exists()
        )

    def test_crear_cliente_genera_cuenta_contable(self):
        response = self.client.post(
            reverse("crear_cliente_facturacion", args=[self.empresa.slug]),
            {
                "nombre": "Cliente con Cuenta",
                "rtn": "08011999111112",
                "direccion": "Tegucigalpa",
                "ciudad": "Tegucigalpa",
                "activo": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        cliente = Cliente.objects.get(empresa=self.empresa, nombre="Cliente con Cuenta")
        self.assertIsNotNone(cliente.cuenta_contable_id)
        self.assertEqual(cliente.cuenta_contable.tipo, "activo")
        self.assertTrue(cliente.cuenta_contable.acepta_movimientos)
        self.assertTrue(cliente.cuenta_contable.codigo.startswith("111001."))

    def test_crear_cliente_rapido_desde_factura_retorna_payload(self):
        response = self.client.post(
            f"{reverse('crear_cliente_facturacion', args=[self.empresa.slug])}?modal=1",
            {
                "quick_mode": "1",
                "nombre": "Cliente Modal",
                "rtn": "08011999112233",
                "direccion": "Tegucigalpa",
                "ciudad": "Tegucigalpa",
                "activo": "on",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "erp-cliente-creado")
        self.assertEqual(response.headers.get("X-Frame-Options"), "SAMEORIGIN")
        self.assertTrue(Cliente.objects.filter(empresa=self.empresa, nombre="Cliente Modal").exists())

    def test_no_permite_cliente_duplicado_por_nombre(self):
        Cliente.objects.create(empresa=self.empresa, nombre="Cliente Unico")

        response = self.client.post(
            reverse("crear_cliente_facturacion", args=[self.empresa.slug]),
            {
                "nombre": "cliente unico",
                "rtn": "",
                "direccion": "Tegucigalpa",
                "ciudad": "Tegucigalpa",
                "activo": "on",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ya existe un cliente con este nombre en la empresa.")

    def test_no_permite_cliente_duplicado_por_rtn(self):
        Cliente.objects.create(empresa=self.empresa, nombre="Cliente RTN", rtn="08011999111111")

        response = self.client.post(
            reverse("crear_cliente_facturacion", args=[self.empresa.slug]),
            {
                "nombre": "Otro Cliente",
                "rtn": "08011999111111",
                "direccion": "Tegucigalpa",
                "ciudad": "Tegucigalpa",
                "activo": "on",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ya existe un cliente con este RTN en la empresa.")

    def test_listado_clientes_facturacion(self):
        response = self.client.get(reverse("clientes_facturacion", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.cliente.nombre)

    def test_listado_clientes_tiene_buscador_con_sugerencias(self):
        response = self.client.get(reverse("clientes_facturacion", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="clientes-sugerencias"', html=False)

    def test_editar_cliente_facturacion(self):
        response = self.client.post(
            reverse("editar_cliente_facturacion", args=[self.empresa.slug, self.cliente.id]),
            {
                "nombre": "Cliente Editado",
                "rtn": self.cliente.rtn or "",
                "direccion": self.cliente.direccion or "",
                "ciudad": "San Pedro Sula",
                "activo": "on",
            },
        )
        self.cliente.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.cliente.nombre, "Cliente Editado")

    def test_puede_eliminar_cliente_sin_facturas(self):
        cliente = Cliente.objects.create(empresa=self.empresa, nombre="Cliente Temporal")

        response = self.client.post(
            reverse("eliminar_cliente_facturacion", args=[self.empresa.slug, cliente.id]),
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Cliente.objects.filter(id=cliente.id).exists())
        self.assertContains(response, "Cliente eliminado correctamente.")

    def test_no_elimina_cliente_con_facturas_registradas(self):
        factura = self.crear_factura_con_linea(estado="emitida")

        response = self.client.post(
            reverse("eliminar_cliente_facturacion", args=[self.empresa.slug, factura.cliente.id]),
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(Cliente.objects.filter(id=factura.cliente.id).exists())
        self.assertContains(response, "No se puede eliminar este cliente porque tiene facturas registradas.")

    def test_crear_producto_desde_facturacion(self):
        response = self.client.post(
            reverse("crear_producto_facturacion", args=[self.empresa.slug]),
            {
                "nombre": "Producto Nuevo",
                "codigo": "PRD-002",
                "tipo_item": "producto",
                "unidad_medida": "unidad",
                "descripcion": "Servicio mensual",
                "precio": "250.00",
                "impuesto_predeterminado": str(self.impuesto.id),
                "activo": "on",
                "controla_inventario": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(Producto.objects.filter(empresa=self.empresa, nombre="Producto Nuevo").exists())

    def test_crear_y_editar_producto_guarda_costo_real_en_empresas_medicas(self):
        self.empresa.slug = "hospital_mia"
        self.empresa.save(update_fields=["slug"])

        response = self.client.get(reverse("crear_producto_facturacion", args=[self.empresa.slug]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="costo_real_inventario"', html=False)
        self.assertContains(response, "Costo real")

        response = self.client.post(
            reverse("crear_producto_facturacion", args=[self.empresa.slug]),
            {
                "nombre": "Producto Margen",
                "codigo": "MRG-001",
                "tipo_item": "producto",
                "unidad_medida": "unidad",
                "descripcion": "Producto con costo real",
                "precio": "300.00",
                "costo_real_inventario": "180.2500",
                "impuesto_predeterminado": str(self.impuesto.id),
                "activo": "on",
                "controla_inventario": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        producto = Producto.objects.get(empresa=self.empresa, codigo="MRG-001")
        self.assertEqual(producto.costo_real_inventario, Decimal("180.2500"))
        self.assertEqual(producto.venta_real_inventario, Decimal("119.75"))
        self.assertEqual(producto.porcentaje_venta_real, Decimal("39.92"))
        primer_cambio = HistorialCostoRealProducto.objects.get(producto=producto)
        self.assertEqual(primer_cambio.costo_anterior, Decimal("0.0000"))
        self.assertEqual(primer_cambio.costo_nuevo, Decimal("180.2500"))

        response = self.client.post(
            reverse("editar_producto_facturacion", args=[self.empresa.slug, producto.id]),
            {
                "nombre": producto.nombre,
                "codigo": producto.codigo,
                "tipo_item": producto.tipo_item,
                "unidad_medida": producto.unidad_medida,
                "descripcion": producto.descripcion,
                "precio": "300.00",
                "costo_real_inventario": "150.0000",
                "impuesto_predeterminado": str(self.impuesto.id),
                "activo": "on",
                "controla_inventario": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        producto.refresh_from_db()
        self.assertEqual(producto.costo_real_inventario, Decimal("150.0000"))
        cambios = HistorialCostoRealProducto.objects.filter(producto=producto)
        self.assertEqual(cambios.count(), 2)
        self.assertEqual(cambios.first().costo_anterior, Decimal("180.2500"))
        self.assertEqual(cambios.first().costo_nuevo, Decimal("150.0000"))

    def test_crear_producto_registra_existencia_distribuida_por_bodega(self):
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.usa_bodegas_internas = True
        configuracion.save(update_fields=["usa_bodegas_internas"])
        bodega_general = BodegaInventario.objects.create(
            empresa=self.empresa,
            nombre="Bodega General",
            tipo="principal",
        )
        bodega_hospital = BodegaInventario.objects.create(
            empresa=self.empresa,
            nombre="Bodega Hospital",
            tipo="provisional",
        )
        vitrina = BodegaInventario.objects.create(
            empresa=self.empresa,
            nombre="Vitrina",
            tipo="vitrina",
        )

        response = self.client.post(
            reverse("crear_producto_facturacion", args=[self.empresa.slug]),
            {
                "nombre": "Jeringa 10ml",
                "codigo": "JER-10",
                "tipo_item": "producto",
                "unidad_medida": "unidad",
                "descripcion": "Insumo medico",
                "precio": "12.50",
                "impuesto_predeterminado": str(self.impuesto.id),
                "activo": "on",
                "controla_inventario": "on",
                f"stock_bodega_{bodega_general.id}": "2.00",
                f"stock_bodega_{vitrina.id}": "3.00",
                f"stock_bodega_{bodega_hospital.id}": "1.00",
                "lote_inicial": "L-001",
            },
        )

        self.assertEqual(response.status_code, 302)
        producto = Producto.objects.get(empresa=self.empresa, codigo="JER-10")
        self.assertEqual(producto.stock_actual, Decimal("6.00"))
        self.assertEqual(
            ExistenciaLoteBodega.objects.filter(
                empresa=self.empresa,
                lote__producto=producto,
            ).aggregate(total=Sum("cantidad"))["total"],
            Decimal("6.00"),
        )
        self.assertEqual(
            ExistenciaLoteBodega.objects.get(empresa=self.empresa, bodega=vitrina, lote__producto=producto).cantidad,
            Decimal("3.00"),
        )
        self.assertEqual(
            MovimientoLoteBodega.objects.filter(empresa=self.empresa, lote__producto=producto, tipo="ajuste").count(),
            3,
        )

    def test_editar_producto_actualiza_distribucion_por_bodega(self):
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.usa_bodegas_internas = True
        configuracion.save(update_fields=["usa_bodegas_internas"])
        bodega_general = BodegaInventario.objects.create(empresa=self.empresa, nombre="Bodega General", tipo="principal")
        bodega_hospital = BodegaInventario.objects.create(empresa=self.empresa, nombre="Bodega Hospital", tipo="provisional")
        vitrina = BodegaInventario.objects.create(empresa=self.empresa, nombre="Vitrina", tipo="vitrina")
        producto = Producto.objects.create(
            empresa=self.empresa,
            nombre="Guantes",
            codigo="GUA-001",
            tipo_item="producto",
            unidad_medida="unidad",
            precio=Decimal("1.50"),
            impuesto_predeterminado=self.impuesto,
            controla_inventario=True,
        )

        response = self.client.post(
            reverse("editar_producto_facturacion", args=[self.empresa.slug, producto.id]),
            {
                "nombre": producto.nombre,
                "codigo": producto.codigo,
                "tipo_item": producto.tipo_item,
                "unidad_medida": producto.unidad_medida,
                "descripcion": "",
                "precio": "1.50",
                "impuesto_predeterminado": str(self.impuesto.id),
                "activo": "on",
                "controla_inventario": "on",
                f"stock_bodega_{bodega_general.id}": "2.00",
                f"stock_bodega_{vitrina.id}": "3.00",
                f"stock_bodega_{bodega_hospital.id}": "1.00",
                "lote_inicial": "AJ-001",
            },
        )

        self.assertEqual(response.status_code, 302)
        producto.refresh_from_db()
        self.assertEqual(producto.stock_actual, Decimal("6.00"))
        self.assertEqual(
            ExistenciaLoteBodega.objects.get(empresa=self.empresa, bodega=bodega_hospital, lote__producto=producto).cantidad,
            Decimal("1.00"),
        )

        response = self.client.get(reverse("bodegas_dashboard", args=[self.empresa.slug]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Bodega General")

        response = self.client.get(reverse("ver_bodega_inventario", args=[self.empresa.slug, vitrina.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Guantes")

    def test_crear_lotes_del_mismo_producto_con_vencimientos_rapidos(self):
        modulo_clinica, _ = Modulo.objects.get_or_create(nombre="Clinica Medica", codigo="clinica_medica")
        EmpresaModulo.objects.update_or_create(empresa=self.empresa, modulo=modulo_clinica, defaults={"activo": True})
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.usa_bodegas_internas = True
        configuracion.usa_inventario_farmaceutico = True
        configuracion.save(update_fields=["usa_bodegas_internas", "usa_inventario_farmaceutico"])
        bodega_general = BodegaInventario.objects.create(empresa=self.empresa, nombre="Bodega General", tipo="principal")
        vitrina = BodegaInventario.objects.create(empresa=self.empresa, nombre="Vitrina", tipo="vitrina")
        producto = Producto.objects.create(
            empresa=self.empresa,
            nombre="Brightening Cream Dia",
            codigo="BCD-001",
            tipo_item="producto",
            unidad_medida="unidad",
            precio=Decimal("850.00"),
            impuesto_predeterminado=self.impuesto,
            controla_inventario=True,
        )

        response = self.client.post(
            reverse("crear_lote_inventario", args=[self.empresa.slug]),
            {
                "producto": str(producto.id),
                "numero_lote": "BRIGHT-AGO27",
                "fecha_vencimiento_rapida": "agos-27",
                "bodega": str(vitrina.id),
                "cantidad": "3.00",
            },
        )
        self.assertEqual(response.status_code, 302)

        response = self.client.post(
            reverse("crear_lote_inventario", args=[self.empresa.slug]),
            {
                "producto": str(producto.id),
                "numero_lote": "BRIGHT-DIC26",
                "fecha_vencimiento_rapida": "dic-26",
                "bodega": str(bodega_general.id),
                "cantidad": "2.00",
            },
        )

        self.assertEqual(response.status_code, 302)
        producto.refresh_from_db()
        self.assertEqual(producto.stock_actual, Decimal("5.00"))
        self.assertTrue(
            LoteInventario.objects.filter(
                empresa=self.empresa,
                producto=producto,
                numero_lote="BRIGHT-AGO27",
                fecha_vencimiento=date(2027, 8, 31),
            ).exists()
        )
        self.assertTrue(
            LoteInventario.objects.filter(
                empresa=self.empresa,
                producto=producto,
                numero_lote="BRIGHT-DIC26",
                fecha_vencimiento=date(2026, 12, 31),
            ).exists()
        )
        response = self.client.get(reverse("editar_producto_facturacion", args=[self.empresa.slug, producto.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "BRIGHT-AGO27")
        self.assertContains(response, "BRIGHT-DIC26")

    def test_crear_producto_rapido_desde_factura_retorna_payload(self):
        response = self.client.post(
            f"{reverse('crear_producto_facturacion', args=[self.empresa.slug])}?modal=1",
            {
                "quick_mode": "1",
                "nombre": "Producto Modal",
                "codigo": "PM-001",
                "tipo_item": "producto",
                "unidad_medida": "unidad",
                "precio": "325.50",
                "impuesto_predeterminado": str(self.impuesto.id),
                "activo": "on",
                "controla_inventario": "on",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "erp-producto-creado")
        self.assertEqual(response.headers.get("X-Frame-Options"), "SAMEORIGIN")
        self.assertTrue(Producto.objects.filter(empresa=self.empresa, nombre="Producto Modal").exists())

    def test_crear_producto_oculta_perfil_farmaceutico_en_empresa_base(self):
        response = self.client.get(reverse("crear_producto_facturacion", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Perfil farmaceutico")
        self.assertNotContains(response, "Principio activo")

    def test_crear_producto_muestra_perfil_farmaceutico_si_empresa_lo_activa(self):
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.usa_inventario_farmaceutico = True
        configuracion.save(update_fields=["usa_inventario_farmaceutico"])

        response = self.client.get(reverse("crear_producto_facturacion", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Perfil farmaceutico")
        self.assertContains(response, "Principio activo")

    def test_crear_proveedor_desde_facturacion(self):
        response = self.client.post(
            reverse("crear_proveedor_facturacion", args=[self.empresa.slug]),
            {
                "nombre": "Proveedor Nuevo",
                "rtn": "08011999111111",
                "contacto": "Ana López",
                "telefono": "9999-1111",
                "correo": "proveedor@example.com",
                "direccion": "Tegucigalpa",
                "ciudad": "Tegucigalpa",
                "activo": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(Proveedor.objects.filter(empresa=self.empresa, nombre="Proveedor Nuevo").exists())

    def test_listado_proveedores_facturacion(self):
        proveedor = Proveedor.objects.create(
            empresa=self.empresa,
            nombre="Proveedor Demo",
            activo=True,
        )

        response = self.client.get(reverse("proveedores_facturacion", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, proveedor.nombre)

    def test_ver_proveedor_muestra_historial_compras(self):
        proveedor = Proveedor.objects.create(
            empresa=self.empresa,
            nombre="Proveedor Historial",
            activo=True,
        )
        compra = CompraInventario.objects.create(
            empresa=self.empresa,
            proveedor=proveedor,
            proveedor_nombre=proveedor.nombre,
            referencia_documento="REF-HIST-001",
            fecha_documento=date.today(),
            estado="borrador",
        )
        LineaCompraInventario.objects.create(
            compra=compra,
            producto=self.producto,
            cantidad=Decimal("2.00"),
            costo_unitario=Decimal("90.00"),
        )

        response = self.client.get(
            reverse("ver_proveedor_facturacion", args=[self.empresa.slug, proveedor.id]),
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, proveedor.nombre)
        self.assertContains(response, compra.numero_compra)
        self.assertContains(response, "Compras y saldo abierto")

    def test_compra_guarda_proveedor_relacionado(self):
        proveedor = Proveedor.objects.create(
            empresa=self.empresa,
            nombre="Proveedor Compra",
            activo=True,
        )

        response = self.client.post(
            reverse("crear_compra", args=[self.empresa.slug]),
            {
                "proveedor": str(proveedor.id),
                "proveedor_nombre": "",
                "referencia_documento": "FAC-PROV-010",
                "fecha_documento": str(date.today()),
                "observacion": "Compra con proveedor",
                "estado": "borrador",
                "lineas_compra-TOTAL_FORMS": "1",
                "lineas_compra-INITIAL_FORMS": "0",
                "lineas_compra-MIN_NUM_FORMS": "0",
                "lineas_compra-MAX_NUM_FORMS": "1000",
                "lineas_compra-0-producto": str(self.producto.id),
                "lineas_compra-0-cantidad": "1.00",
                "lineas_compra-0-costo_unitario": "90.00",
                "lineas_compra-0-comentario": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        compra = CompraInventario.objects.get(referencia_documento="FAC-PROV-010")
        self.assertEqual(compra.proveedor, proveedor)
        self.assertEqual(compra.proveedor_nombre, proveedor.nombre)

    def test_listado_productos_facturacion(self):
        response = self.client.get(reverse("productos_facturacion", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.producto.nombre)

    def test_listado_productos_tiene_buscador_con_sugerencias(self):
        response = self.client.get(reverse("productos_facturacion", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="productos-sugerencias"', html=False)

    def test_inventario_dashboard_muestra_producto(self):
        response = self.client.get(reverse("inventario_facturacion", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.producto.nombre)
        self.assertNotContains(response, "Costo real")
        producto_url = reverse("crear_producto_facturacion", args=[self.empresa.slug])
        inventario_url = reverse("inventario_facturacion", args=[self.empresa.slug])
        self.assertContains(response, "Nuevo Producto")
        self.assertContains(response, f'href="{producto_url}?next={inventario_url}"', html=False)

    def test_costo_real_inventario_solo_hospital_mia_y_medical_spa(self):
        inventario_url = reverse("inventario_facturacion", args=[self.empresa.slug])
        response = self.client.post(
            inventario_url,
            {
                "accion": "actualizar_costo_real",
                "producto_id": str(self.producto.id),
                "costo_real_inventario": "55.25",
            },
        )
        self.assertRedirects(response, inventario_url)
        self.producto.refresh_from_db()
        self.assertEqual(self.producto.costo_real_inventario, Decimal("0.0000"))

        self.empresa.slug = "hospital_mia"
        self.empresa.save(update_fields=["slug"])
        inventario_url = reverse("inventario_facturacion", args=[self.empresa.slug])
        response = self.client.get(inventario_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Costo real")
        self.assertContains(response, "Contable:")
        self.assertContains(response, "Editar costo")

        self.producto.costo_promedio = Decimal("40.0000")
        self.producto.save(update_fields=["costo_promedio"])
        response = self.client.post(
            inventario_url,
            {
                "accion": "actualizar_costo_real",
                "producto_id": str(self.producto.id),
                "costo_real_inventario": "55.25",
            },
        )
        self.assertRedirects(response, f"{inventario_url}?producto={self.producto.id}")
        self.producto.refresh_from_db()
        self.assertEqual(self.producto.costo_real_inventario, Decimal("55.2500"))
        self.assertEqual(self.producto.costo_promedio, Decimal("40.0000"))
        self.assertEqual(self.producto.venta_real_inventario, Decimal("44.75"))
        self.assertEqual(self.producto.porcentaje_venta_real, Decimal("44.75"))
        cambio = HistorialCostoRealProducto.objects.get(producto=self.producto)
        self.assertEqual(cambio.costo_anterior, Decimal("0.0000"))
        self.assertEqual(cambio.costo_nuevo, Decimal("55.2500"))
        self.assertEqual(cambio.usuario, self.user)

        response = self.client.get(inventario_url)
        self.assertContains(response, "Historial de costos")
        self.assertContains(response, "55.2500")

    def test_inventario_farmaceutico_tiene_acceso_a_crear_producto(self):
        modulo_clinica, _ = Modulo.objects.get_or_create(nombre="Clinica Medica", codigo="clinica_medica")
        EmpresaModulo.objects.create(empresa=self.empresa, modulo=modulo_clinica, activo=True)
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.usa_inventario_farmaceutico = True
        configuracion.save(update_fields=["usa_inventario_farmaceutico"])

        response = self.client.get(reverse("inventario_farmaceutico", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        producto_url = reverse("crear_producto_facturacion", args=[self.empresa.slug])
        inventario_farmaceutico_url = reverse("inventario_farmaceutico", args=[self.empresa.slug])
        self.assertContains(response, "Nuevo Producto")
        self.assertContains(response, f'href="{producto_url}?next={inventario_farmaceutico_url}"', html=False)

    def test_kardex_inventario_muestra_movimientos(self):
        InventarioProducto.objects.create(
            empresa=self.empresa,
            producto=self.producto,
            existencias=Decimal("10.00"),
            stock_minimo=Decimal("2.00"),
        )
        MovimientoInventario.objects.create(
            empresa=self.empresa,
            producto=self.producto,
            tipo="ajuste_entrada",
            cantidad=Decimal("10.00"),
            existencia_anterior=Decimal("0.00"),
            existencia_resultante=Decimal("10.00"),
            referencia="Carga inicial",
        )

        response = self.client.get(reverse("kardex_inventario", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.producto.nombre)
        self.assertContains(response, "Carga inicial")

    def test_ajuste_manual_inventario_actualiza_existencias(self):
        response = self.client.post(
            reverse("ajustar_inventario", args=[self.empresa.slug]),
            {
                "producto": str(self.producto.id),
                "tipo_ajuste": "ajuste_entrada",
                "cantidad": "10.00",
                "observacion": "Carga inicial",
                "stock_minimo": "2.00",
            },
        )

        self.assertEqual(response.status_code, 302)
        inventario = InventarioProducto.objects.get(producto=self.producto)
        self.assertEqual(inventario.existencias, Decimal("10.00"))
        self.assertEqual(inventario.stock_minimo, Decimal("2.00"))
        self.assertTrue(MovimientoInventario.objects.filter(producto=self.producto, tipo="ajuste_entrada").exists())

    def test_entrada_inventario_actualiza_existencias(self):
        response = self.client.post(
            reverse("entrada_inventario", args=[self.empresa.slug]),
            {
                "producto": str(self.producto.id),
                "cantidad": "15.00",
                "referencia": "Carga inicial abril",
                "observacion": "Ingreso formal de stock",
                "stock_minimo": "3.00",
            },
        )

        self.assertEqual(response.status_code, 302)
        inventario = InventarioProducto.objects.get(producto=self.producto)
        self.assertEqual(inventario.existencias, Decimal("15.00"))
        self.assertEqual(inventario.stock_minimo, Decimal("3.00"))
        self.assertTrue(MovimientoInventario.objects.filter(producto=self.producto, tipo="entrada", referencia="Carga inicial abril").exists())

    def test_entrada_medical_spa_exige_bodega_y_carga_existencia_en_destino(self):
        self.empresa.slug = "medical_spa"
        self.empresa.save(update_fields=["slug"])
        bodega = BodegaInventario.objects.create(
            empresa=self.empresa,
            nombre="Bodega Tratamientos",
            tipo="principal",
            activa=True,
        )

        formulario = self.client.get(reverse("entrada_inventario", args=[self.empresa.slug]))
        self.assertContains(formulario, "Bodega de destino")
        self.assertContains(formulario, "Bodega Tratamientos")

        response = self.client.post(
            reverse("entrada_inventario", args=[self.empresa.slug]),
            {
                "producto": str(self.producto.id),
                "bodega": str(bodega.id),
                "cantidad": "6.00",
                "referencia": "Ingreso Medical Spa",
                "observacion": "Producto recibido en bodega",
                "stock_minimo": "2.00",
            },
        )

        self.assertEqual(response.status_code, 302)
        movimiento = MovimientoInventario.objects.get(
            producto=self.producto,
            tipo="entrada",
            referencia="Ingreso Medical Spa",
        )
        self.assertEqual(movimiento.bodega, bodega)
        existencia = ExistenciaLoteBodega.objects.get(
            empresa=self.empresa,
            bodega=bodega,
            lote__producto=self.producto,
        )
        self.assertEqual(existencia.cantidad, Decimal("6.00"))
        self.assertTrue(
            MovimientoLoteBodega.objects.filter(
                empresa=self.empresa,
                bodega=bodega,
                lote__producto=self.producto,
                tipo="entrada",
                cantidad=Decimal("6.00"),
            ).exists()
        )

    def test_crear_documento_entrada_inventario_aplicada(self):
        response = self.client.post(
            reverse("crear_entrada_inventario_documento", args=[self.empresa.slug]),
            {
                "referencia": "DOC-ENT-001",
                "fecha_documento": str(date.today()),
                "observacion": "Carga por documento",
                "estado": "aplicada",
                "lineas_entrada-TOTAL_FORMS": "1",
                "lineas_entrada-INITIAL_FORMS": "0",
                "lineas_entrada-MIN_NUM_FORMS": "0",
                "lineas_entrada-MAX_NUM_FORMS": "1000",
                "lineas_entrada-0-producto": str(self.producto.id),
                "lineas_entrada-0-cantidad": "7.00",
                "lineas_entrada-0-comentario": "Lote inicial",
            },
        )

        self.assertEqual(response.status_code, 302)
        entrada = EntradaInventarioDocumento.objects.get(referencia="DOC-ENT-001")
        inventario = InventarioProducto.objects.get(producto=self.producto)
        self.assertEqual(entrada.estado, "aplicada")
        self.assertEqual(inventario.existencias, Decimal("7.00"))
        self.assertTrue(MovimientoInventario.objects.filter(entrada_documento=entrada, tipo="entrada").exists())

    def test_compras_dashboard_muestra_modulo(self):
        response = self.client.get(reverse("compras_dashboard", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Compras")

    def test_crear_compra_tiene_buscador_con_sugerencias_de_proveedor(self):
        Proveedor.objects.create(empresa=self.empresa, nombre="Proveedor Sugerido")

        response = self.client.get(reverse("crear_compra", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="proveedores-sugerencias"', html=False)

    def test_crear_compra_aplicada_ingresa_inventario(self):
        response = self.client.post(
            reverse("crear_compra", args=[self.empresa.slug]),
            {
                "proveedor_nombre": "Distribuidora Central",
                "referencia_documento": "FAC-PROV-001",
                "fecha_documento": str(date.today()),
                "observacion": "Compra inicial",
                "estado": "aplicada",
                "lineas_compra-TOTAL_FORMS": "1",
                "lineas_compra-INITIAL_FORMS": "0",
                "lineas_compra-MIN_NUM_FORMS": "0",
                "lineas_compra-MAX_NUM_FORMS": "1000",
                "lineas_compra-0-producto": str(self.producto.id),
                "lineas_compra-0-cantidad": "5.00",
                "lineas_compra-0-costo_unitario": "80.00",
                "lineas_compra-0-comentario": "Lote abril",
            },
        )

        self.assertEqual(response.status_code, 302)
        compra = CompraInventario.objects.get(proveedor_nombre="Distribuidora Central")
        inventario = InventarioProducto.objects.get(producto=self.producto)
        self.assertEqual(compra.estado, "aplicada")
        self.assertEqual(compra.total_unidades, Decimal("5.00"))
        self.assertEqual(compra.total_documento, Decimal("400.00"))
        self.assertEqual(inventario.existencias, Decimal("5.00"))
        self.assertTrue(MovimientoInventario.objects.filter(compra_documento=compra, tipo="entrada_compra").exists())

    def test_compra_borrador_no_mueve_inventario(self):
        response = self.client.post(
            reverse("crear_compra", args=[self.empresa.slug]),
            {
                "proveedor_nombre": "Distribuidora Norte",
                "referencia_documento": "FAC-PROV-002",
                "fecha_documento": str(date.today()),
                "observacion": "",
                "estado": "borrador",
                "lineas_compra-TOTAL_FORMS": "1",
                "lineas_compra-INITIAL_FORMS": "0",
                "lineas_compra-MIN_NUM_FORMS": "0",
                "lineas_compra-MAX_NUM_FORMS": "1000",
                "lineas_compra-0-producto": str(self.producto.id),
                "lineas_compra-0-cantidad": "3.00",
                "lineas_compra-0-costo_unitario": "75.00",
                "lineas_compra-0-comentario": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        compra = CompraInventario.objects.get(proveedor_nombre="Distribuidora Norte")
        self.assertEqual(compra.estado, "borrador")
        self.assertFalse(InventarioProducto.objects.filter(producto=self.producto).exists())
        self.assertFalse(MovimientoInventario.objects.filter(compra_documento=compra).exists())

    def test_editar_compra_borrador_permite_corregir_cantidad(self):
        compra = CompraInventario.objects.create(
            empresa=self.empresa,
            proveedor_nombre="Proveedor Demo",
            referencia_documento="REF-001",
            fecha_documento=date.today(),
            estado="borrador",
        )
        linea = LineaCompraInventario.objects.create(
            compra=compra,
            producto=self.producto,
            cantidad=Decimal("3.00"),
            costo_unitario=Decimal("70.00"),
        )

        response = self.client.post(
            reverse("editar_compra", args=[self.empresa.slug, compra.id]),
            {
                "proveedor_nombre": "Proveedor Demo",
                "referencia_documento": "REF-001-A",
                "fecha_documento": str(date.today()),
                "observacion": "Compra corregida",
                "estado": "borrador",
                "lineas_compra-TOTAL_FORMS": "1",
                "lineas_compra-INITIAL_FORMS": "1",
                "lineas_compra-MIN_NUM_FORMS": "0",
                "lineas_compra-MAX_NUM_FORMS": "1000",
                "lineas_compra-0-id": str(linea.id),
                "lineas_compra-0-producto": str(self.producto.id),
                "lineas_compra-0-cantidad": "8.00",
                "lineas_compra-0-costo_unitario": "75.00",
                "lineas_compra-0-comentario": "Cantidad corregida",
            },
        )

        self.assertEqual(response.status_code, 302)
        compra.refresh_from_db()
        linea.refresh_from_db()
        self.assertEqual(compra.referencia_documento, "REF-001-A")
        self.assertEqual(linea.cantidad, Decimal("8.00"))
        self.assertEqual(linea.costo_unitario, Decimal("75.00"))
        self.assertFalse(MovimientoInventario.objects.filter(compra_documento=compra).exists())

    def test_no_permite_editar_compra_aplicada(self):
        compra = CompraInventario.objects.create(
            empresa=self.empresa,
            proveedor_nombre="Proveedor Aplicado",
            referencia_documento="REF-AP-001",
            fecha_documento=date.today(),
            estado="aplicada",
        )

        response = self.client.get(
            reverse("editar_compra", args=[self.empresa.slug, compra.id]),
            follow=True,
        )

        self.assertRedirects(
            response,
            reverse("ver_compra", args=[self.empresa.slug, compra.id]),
        )
        self.assertContains(response, "No se puede editar una compra aplicada")

    def test_puede_aplicar_compra_borrador_despues(self):
        compra = CompraInventario.objects.create(
            empresa=self.empresa,
            proveedor_nombre="Proveedor Aplicar",
            referencia_documento="REF-APL-001",
            fecha_documento=date.today(),
            estado="borrador",
        )
        LineaCompraInventario.objects.create(
            compra=compra,
            producto=self.producto,
            cantidad=Decimal("4.00"),
            costo_unitario=Decimal("85.00"),
        )

        response = self.client.post(
            reverse("aplicar_compra", args=[self.empresa.slug, compra.id]),
        )

        self.assertEqual(response.status_code, 302)
        compra.refresh_from_db()
        inventario = InventarioProducto.objects.get(producto=self.producto)
        self.assertEqual(compra.estado, "aplicada")
        self.assertEqual(inventario.existencias, Decimal("4.00"))
        self.assertTrue(MovimientoInventario.objects.filter(compra_documento=compra, tipo="entrada_compra").exists())

    def test_aplicar_compra_ya_aplicada_no_duplica_movimiento(self):
        compra = CompraInventario.objects.create(
            empresa=self.empresa,
            proveedor_nombre="Proveedor Ya Aplicado",
            referencia_documento="REF-APL-002",
            fecha_documento=date.today(),
            estado="borrador",
        )
        LineaCompraInventario.objects.create(
            compra=compra,
            producto=self.producto,
            cantidad=Decimal("2.00"),
            costo_unitario=Decimal("90.00"),
        )

        self.client.post(reverse("aplicar_compra", args=[self.empresa.slug, compra.id]))
        response = self.client.post(
            reverse("aplicar_compra", args=[self.empresa.slug, compra.id]),
            follow=True,
        )

        compra.refresh_from_db()
        inventario = InventarioProducto.objects.get(producto=self.producto)
        self.assertRedirects(
            response,
            reverse("ver_compra", args=[self.empresa.slug, compra.id]),
        )
        self.assertContains(response, "La compra ya estaba aplicada")
        self.assertEqual(compra.estado, "aplicada")
        self.assertEqual(inventario.existencias, Decimal("2.00"))
        self.assertEqual(MovimientoInventario.objects.filter(compra_documento=compra, tipo="entrada_compra").count(), 1)

    def test_anular_compra_borrador_cambia_estado_sin_mover_inventario(self):
        compra = CompraInventario.objects.create(
            empresa=self.empresa,
            proveedor_nombre="Proveedor Borrador",
            referencia_documento="REF-AN-001",
            fecha_documento=date.today(),
            estado="borrador",
        )
        LineaCompraInventario.objects.create(
            compra=compra,
            producto=self.producto,
            cantidad=Decimal("3.00"),
            costo_unitario=Decimal("50.00"),
        )

        response = self.client.post(
            reverse("anular_compra", args=[self.empresa.slug, compra.id]),
        )

        self.assertEqual(response.status_code, 302)
        compra.refresh_from_db()
        self.assertEqual(compra.estado, "anulada")
        self.assertFalse(InventarioProducto.objects.filter(producto=self.producto).exists())
        self.assertFalse(MovimientoInventario.objects.filter(compra_documento=compra).exists())

    def test_anular_compra_aplicada_revierte_inventario(self):
        compra = CompraInventario.objects.create(
            empresa=self.empresa,
            proveedor_nombre="Proveedor Revertir",
            referencia_documento="REF-AN-002",
            fecha_documento=date.today(),
            estado="borrador",
        )
        LineaCompraInventario.objects.create(
            compra=compra,
            producto=self.producto,
            cantidad=Decimal("6.00"),
            costo_unitario=Decimal("65.00"),
        )
        self.client.post(reverse("aplicar_compra", args=[self.empresa.slug, compra.id]))

        response = self.client.post(
            reverse("anular_compra", args=[self.empresa.slug, compra.id]),
        )

        self.assertEqual(response.status_code, 302)
        compra.refresh_from_db()
        inventario = InventarioProducto.objects.get(producto=self.producto)
        self.assertEqual(compra.estado, "anulada")
        self.assertEqual(inventario.existencias, Decimal("0.00"))
        self.assertEqual(MovimientoInventario.objects.filter(compra_documento=compra, tipo="entrada_compra").count(), 1)
        self.assertEqual(MovimientoInventario.objects.filter(compra_documento=compra, tipo="reversion_compra").count(), 1)

    def test_anular_compra_ya_anulada_no_duplica_reversion(self):
        compra = CompraInventario.objects.create(
            empresa=self.empresa,
            proveedor_nombre="Proveedor Ya Anulado",
            referencia_documento="REF-AN-003",
            fecha_documento=date.today(),
            estado="borrador",
        )
        LineaCompraInventario.objects.create(
            compra=compra,
            producto=self.producto,
            cantidad=Decimal("2.00"),
            costo_unitario=Decimal("70.00"),
        )
        self.client.post(reverse("aplicar_compra", args=[self.empresa.slug, compra.id]))
        self.client.post(reverse("anular_compra", args=[self.empresa.slug, compra.id]))

        response = self.client.post(
            reverse("anular_compra", args=[self.empresa.slug, compra.id]),
            follow=True,
        )

        compra.refresh_from_db()
        inventario = InventarioProducto.objects.get(producto=self.producto)
        self.assertRedirects(
            response,
            reverse("ver_compra", args=[self.empresa.slug, compra.id]),
        )
        self.assertContains(response, "La compra ya estaba anulada")
        self.assertEqual(compra.estado, "anulada")
        self.assertEqual(inventario.existencias, Decimal("0.00"))
        self.assertEqual(MovimientoInventario.objects.filter(compra_documento=compra, tipo="reversion_compra").count(), 1)

    def test_inventario_dashboard_muestra_alerta_stock_minimo(self):
        InventarioProducto.objects.create(
            empresa=self.empresa,
            producto=self.producto,
            existencias=Decimal("2.00"),
            stock_minimo=Decimal("2.00"),
        )

        response = self.client.get(reverse("inventario_facturacion", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Alertas de stock minimo")
        self.assertContains(response, self.producto.nombre)

    def test_editar_producto_facturacion(self):
        response = self.client.post(
            reverse("editar_producto_facturacion", args=[self.empresa.slug, self.producto.id]),
            {
                "nombre": "Producto Editado",
                "codigo": "PRD-001",
                "tipo_item": "producto",
                "unidad_medida": "caja",
                "descripcion": self.producto.descripcion or "",
                "precio": "300.00",
                "impuesto_predeterminado": str(self.impuesto.id),
                "activo": "on",
                "controla_inventario": "on",
            },
        )
        self.producto.refresh_from_db()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.producto.nombre, "Producto Editado")
        self.assertEqual(self.producto.codigo, "PRD-001")
        self.assertEqual(self.producto.unidad_medida, "caja")
        self.assertEqual(self.producto.impuesto_predeterminado, self.impuesto)
        self.assertEqual(self.producto.precio, Decimal("300.00"))

    def test_producto_servicio_no_puede_controlar_inventario(self):
        producto = Producto(
            empresa=self.empresa,
            nombre="Consultoria",
            tipo_item="servicio",
            unidad_medida="hora",
            precio=Decimal("500.00"),
            controla_inventario=True,
        )

        with self.assertRaises(ValidationError):
            producto.save()

    def test_producto_no_repite_codigo_en_empresa(self):
        Producto.objects.create(
            empresa=self.empresa,
            nombre="Producto Codigo",
            codigo="COD-001",
            tipo_item="producto",
            unidad_medida="unidad",
            precio=Decimal("10.00"),
            controla_inventario=True,
        )
        producto = Producto(
            empresa=self.empresa,
            nombre="Producto Duplicado",
            codigo="cod-001",
            tipo_item="producto",
            unidad_medida="unidad",
            precio=Decimal("12.00"),
            controla_inventario=True,
        )

        with self.assertRaises(ValidationError):
            producto.save()

    def test_listado_cai_facturacion(self):
        response = self.client.get(reverse("cai_facturacion", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.cai.numero_cai)
        self.assertContains(response, self.cai_nota_credito.numero_cai)

    def test_listado_cai_facturacion_filtra_por_uso(self):
        response = self.client.get(
            reverse("cai_facturacion", args=[self.empresa.slug]),
            {"uso_documento": "nota_credito"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.cai_nota_credito.numero_cai)
        self.assertNotContains(response, self.cai.numero_cai)

    def test_crear_cai_desde_facturacion(self):
        response = self.client.post(
            reverse("crear_cai_facturacion", args=[self.empresa.slug]),
            {
                "numero_cai": "CAI-NUEVO",
                "uso_documento": "factura",
                "establecimiento": "002",
                "punto_emision": "001",
                "tipo_documento": "01",
                "rango_inicial": "11",
                "rango_final": "20",
                "correlativo_actual": "10",
                "fecha_activacion": str(date.today()),
                "fecha_limite": str(date.today() + timedelta(days=60)),
                "activo": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(CAI.objects.filter(empresa=self.empresa, numero_cai="CAI-NUEVO").exists())

    def test_crear_cai_desde_facturacion_acepta_formato_fecha_latino(self):
        response = self.client.post(
            reverse("crear_cai_facturacion", args=[self.empresa.slug]),
            {
                "numero_cai": "CAI-FECHA-LATAM",
                "uso_documento": "factura",
                "establecimiento": "007",
                "punto_emision": "001",
                "tipo_documento": "01",
                "rango_inicial": "71",
                "rango_final": "80",
                "correlativo_actual": "70",
                "fecha_activacion": "12/05/2022",
                "fecha_limite": "12/05/2023",
                "activo": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        cai = CAI.objects.get(empresa=self.empresa, numero_cai="CAI-FECHA-LATAM")
        self.assertEqual(cai.fecha_activacion, date(2022, 5, 12))
        self.assertEqual(cai.fecha_limite, date(2023, 5, 12))

    def test_factura_historica_sin_permiso_sigue_usando_cai_vigente_actual(self):
        cai_historico = CAI.objects.create(
            empresa=self.empresa,
            numero_cai="CAI-2024",
            uso_documento="factura",
            establecimiento="003",
            punto_emision="001",
            tipo_documento="01",
            rango_inicial=30,
            rango_final=40,
            correlativo_actual=29,
            fecha_activacion=date(2024, 1, 1),
            fecha_limite=date(2024, 12, 31),
            activo=True,
        )

        factura = Factura.objects.create(
            empresa=self.empresa,
            cliente=self.cliente,
            estado="borrador",
            fecha_emision=date(2024, 1, 15),
        )
        LineaFactura.objects.create(
            factura=factura,
            producto=self.producto,
            cantidad=Decimal("1.00"),
            precio_unitario=Decimal("100.00"),
            impuesto=self.impuesto,
        )
        factura.calcular_totales()
        factura.estado = "emitida"

        factura.save()

        self.assertEqual(factura.cai, self.cai)
        self.assertEqual(factura.cai_numero_historico, self.cai.numero_cai)
        cai_historico.refresh_from_db()
        self.assertEqual(cai_historico.correlativo_actual, 29)

    def test_factura_historica_puede_usarse_con_cai_vencido_si_empresa_tiene_permiso(self):
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.permite_cai_historico = True
        configuracion.save(update_fields=["permite_cai_historico"])

        cai_historico = CAI.objects.create(
            empresa=self.empresa,
            numero_cai="CAI-2024",
            uso_documento="factura",
            establecimiento="003",
            punto_emision="001",
            tipo_documento="01",
            rango_inicial=30,
            rango_final=40,
            correlativo_actual=29,
            fecha_activacion=date(2024, 1, 1),
            fecha_limite=date(2024, 12, 31),
            activo=True,
        )

        cai_2025 = CAI.objects.create(
            empresa=self.empresa,
            numero_cai="CAI-2025",
            uso_documento="factura",
            establecimiento="004",
            punto_emision="001",
            tipo_documento="01",
            rango_inicial=41,
            rango_final=50,
            correlativo_actual=40,
            fecha_activacion=date(2025, 1, 1),
            fecha_limite=date(2025, 12, 31),
            activo=True,
        )

        factura = self.crear_factura_con_linea(
            estado="borrador",
            fecha_emision=date(2024, 1, 15),
        )
        factura.estado = "emitida"
        factura.save()

        self.assertEqual(factura.cai, cai_historico)
        self.assertEqual(factura.numero_factura, "003-001-01-00000030")
        self.assertEqual(factura.cai_fecha_limite_historico, date(2024, 12, 31))

        factura_2025 = self.crear_factura_con_linea(
            estado="borrador",
            fecha_emision=date(2025, 7, 10),
        )
        factura_2025.estado = "emitida"
        factura_2025.save()

        self.assertEqual(factura_2025.cai, cai_2025)
        self.assertEqual(factura_2025.numero_factura, "004-001-01-00000041")

    def test_factura_historica_no_usa_cai_antes_de_su_fecha_activacion(self):
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.permite_cai_historico = True
        configuracion.save(update_fields=["permite_cai_historico"])

        cai_2025_vigente = CAI.objects.create(
            empresa=self.empresa,
            numero_cai="CAI-2025-VIG",
            uso_documento="factura",
            establecimiento="006",
            punto_emision="001",
            tipo_documento="01",
            rango_inicial=61,
            rango_final=70,
            correlativo_actual=60,
            fecha_activacion=date(2025, 1, 1),
            fecha_limite=date(2025, 5, 31),
            activo=True,
        )

        cai_programado = CAI.objects.create(
            empresa=self.empresa,
            numero_cai="CAI-PROG-2025",
            uso_documento="factura",
            establecimiento="005",
            punto_emision="001",
            tipo_documento="01",
            rango_inicial=51,
            rango_final=60,
            correlativo_actual=50,
            fecha_activacion=date(2025, 6, 1),
            fecha_limite=date(2025, 12, 31),
            activo=True,
        )

        factura = self.crear_factura_con_linea(
            estado="borrador",
            fecha_emision=date(2025, 3, 10),
        )
        factura.estado = "emitida"
        factura.save()

        self.assertNotEqual(factura.cai, cai_programado)
        self.assertEqual(factura.cai, cai_2025_vigente)

    def test_factura_conserva_snapshot_del_cai_historico(self):
        factura_antigua = self.crear_factura_con_linea(estado="emitida")
        self.cai.activo = False
        self.cai.save(update_fields=["activo"])

        cai_nuevo = CAI.objects.create(
            empresa=self.empresa,
            numero_cai="CAI-2026",
            uso_documento="factura",
            establecimiento="002",
            punto_emision="001",
            tipo_documento="01",
            rango_inicial=11,
            rango_final=20,
            correlativo_actual=10,
            fecha_limite=date.today() + timedelta(days=90),
            activo=True,
        )

        factura_nueva = self.crear_factura_con_linea(estado="emitida")

        self.assertEqual(factura_antigua.cai_numero_historico, "CAI-TEST")
        self.assertEqual(factura_antigua.cai_establecimiento_historico, "001")
        self.assertEqual(factura_nueva.cai_numero_historico, cai_nuevo.numero_cai)
        self.assertEqual(factura_antigua.cai_numero_historico, factura_antigua.cai_numero)

    def test_no_permite_editar_datos_sensibles_de_cai_usado(self):
        self.crear_factura_con_linea(estado="emitida")
        self.cai.numero_cai = "CAI-CAMBIADO"

        with self.assertRaises(ValidationError):
            self.cai.save()

    def test_no_permite_editar_cai_usado_en_nota_credito_emitida(self):
        factura = self.crear_factura_con_linea(estado="emitida")
        nota = NotaCredito.objects.create(
            empresa=self.empresa,
            factura_origen=factura,
            cliente=factura.cliente,
            moneda=factura.moneda,
            tipo_cambio=factura.tipo_cambio,
            fecha_emision=date.today(),
            estado="borrador",
        )
        LineaNotaCredito.objects.create(
            nota_credito=nota,
            producto=self.producto,
            cantidad=Decimal("1.00"),
            precio_unitario=Decimal("100.00"),
            impuesto=self.impuesto,
        )
        nota.calcular_totales()
        nota.estado = "emitida"
        nota.save(update_fields=["subtotal", "impuesto", "total", "total_lempiras", "estado"])
        self.cai_nota_credito.numero_cai = "CAI-NC-CAMBIADO"

        with self.assertRaises(ValidationError):
            self.cai_nota_credito.save()

    def test_nota_credito_emitida_usa_cai_propio(self):
        factura = self.crear_factura_con_linea(estado="emitida")
        nota = NotaCredito.objects.create(
            empresa=self.empresa,
            factura_origen=factura,
            cliente=factura.cliente,
            moneda=factura.moneda,
            tipo_cambio=factura.tipo_cambio,
            fecha_emision=date.today(),
            estado="borrador",
        )
        LineaNotaCredito.objects.create(
            nota_credito=nota,
            producto=self.producto,
            cantidad=Decimal("1.00"),
            precio_unitario=Decimal("100.00"),
            impuesto=self.impuesto,
        )
        nota.calcular_totales()
        nota.estado = "emitida"
        nota.save(update_fields=["subtotal", "impuesto", "total", "total_lempiras", "estado"])

        self.assertEqual(nota.cai, self.cai_nota_credito)
        self.assertEqual(nota.cai_numero_historico, "CAI-NC")

    def test_listado_impuestos_facturacion(self):
        response = self.client.get(reverse("impuestos_facturacion", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.impuesto.nombre)

    def test_crear_impuesto_desde_facturacion(self):
        response = self.client.post(
            reverse("crear_impuesto_facturacion", args=[self.empresa.slug]),
            {
                "nombre": "ISV 18",
                "porcentaje": "18.00",
                "activo": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(TipoImpuesto.objects.filter(nombre="ISV 18").exists())

    def test_generar_nota_credito_desde_factura_copia_lineas(self):
        factura = self.crear_factura_con_linea(estado="emitida")

        response = self.client.post(
            reverse("generar_nota_credito_desde_factura", args=[self.empresa.slug, factura.id])
        )

        nota = NotaCredito.objects.get(factura_origen=factura)

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("editar_nota_credito", args=[self.empresa.slug, nota.id]))
        self.assertEqual(nota.estado, "borrador")
        self.assertEqual(nota.lineas.count(), factura.lineas.count())
        self.assertEqual(nota.total, factura.total)

    def test_detalle_factura_para_nota_credito_devuelve_lineas(self):
        factura = self.crear_factura_con_linea(estado="emitida")

        response = self.client.get(
            reverse("detalle_factura_para_nota_credito", args=[self.empresa.slug, factura.id])
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["factura_id"], factura.id)
        self.assertEqual(data["cliente"], factura.cliente.nombre)
        self.assertEqual(len(data["lineas"]), 1)
        self.assertEqual(data["lineas"][0]["producto_id"], self.producto.id)

    def test_nota_credito_emitida_reduce_saldo_y_limita_pago(self):
        factura = self.crear_factura_con_linea(estado="emitida")
        nota = NotaCredito.objects.create(
            empresa=self.empresa,
            factura_origen=factura,
            cliente=factura.cliente,
            moneda=factura.moneda,
            tipo_cambio=factura.tipo_cambio,
            fecha_emision=date.today(),
            estado="borrador",
        )
        LineaNotaCredito.objects.create(
            nota_credito=nota,
            producto=self.producto,
            cantidad=Decimal("0.50"),
            precio_unitario=Decimal("100.00"),
            impuesto=self.impuesto,
        )
        nota.calcular_totales()
        nota.estado = "emitida"
        nota.save(update_fields=["subtotal", "impuesto", "total", "total_lempiras", "estado"])
        factura.refresh_from_db()
        factura.actualizar_estado_pago()
        factura.refresh_from_db()

        self.assertEqual(factura.total_notas_credito, Decimal("57.50"))
        self.assertEqual(factura.total_documento_ajustado, Decimal("57.50"))
        self.assertEqual(factura.saldo_pendiente, Decimal("57.50"))

        with self.assertRaises(ValidationError):
            PagoFactura.objects.create(
                factura=factura,
                monto=Decimal("58.00"),
                metodo="efectivo",
                fecha=date.today(),
            )

    def test_editar_nota_credito_emitida_reconstruye_asiento_contable(self):
        factura = self.crear_factura_con_linea(estado="emitida")
        nota = NotaCredito.objects.create(
            empresa=self.empresa,
            factura_origen=factura,
            cliente=factura.cliente,
            vendedor=self.user,
            moneda=factura.moneda,
            tipo_cambio=factura.tipo_cambio,
            fecha_emision=date.today(),
            motivo="Ajuste inicial",
            estado="borrador",
        )
        linea = LineaNotaCredito.objects.create(
            nota_credito=nota,
            producto=self.producto,
            cantidad=Decimal("1.00"),
            precio_unitario=Decimal("100.00"),
            impuesto=self.impuesto,
        )
        nota.calcular_totales()
        nota.estado = "emitida"
        nota.save(update_fields=["subtotal", "impuesto", "total", "total_lempiras", "estado"])

        response = self.client.post(
            reverse("editar_nota_credito", args=[self.empresa.slug, nota.id]),
            {
                "factura_origen": str(factura.id),
                "fecha_emision": str(date.today()),
                "motivo": "Ajuste parcial",
                "estado": "emitida",
                "lineas-TOTAL_FORMS": "1",
                "lineas-INITIAL_FORMS": "1",
                "lineas-MIN_NUM_FORMS": "0",
                "lineas-MAX_NUM_FORMS": "1000",
                "lineas-0-id": str(linea.id),
                "lineas-0-producto": str(self.producto.id),
                "lineas-0-cantidad": "0.50",
                "lineas-0-precio_unitario": "100.00",
                "lineas-0-descuento_porcentaje": "0",
                "lineas-0-comentario": "",
                "lineas-0-impuesto": str(self.impuesto.id),
            },
        )

        self.assertEqual(response.status_code, 302)
        nota.refresh_from_db()
        factura.refresh_from_db()
        self.assertEqual(nota.total, Decimal("57.50"))
        self.assertEqual(factura.saldo_pendiente, Decimal("57.50"))

        asientos = AsientoContable.objects.filter(
            empresa=self.empresa,
            documento_tipo="nota_credito",
            documento_id=nota.id,
            evento="emision",
        )
        self.assertEqual(asientos.count(), 1)
        asiento = asientos.get()
        nota.cliente.refresh_from_db()
        self.assertTrue(asiento.lineas.filter(cuenta=nota.cliente.cuenta_contable, haber=Decimal("57.50")).exists())

    def test_crear_factura_emitida_descuenta_inventario(self):
        InventarioProducto.objects.create(
            empresa=self.empresa,
            producto=self.producto,
            existencias=Decimal("10.00"),
            stock_minimo=Decimal("1.00"),
        )

        response = self.client.post(
            reverse("crear_factura", args=[self.empresa.slug]),
            {
                "cliente": str(self.cliente.id),
                "fecha_emision": str(date.today()),
                "fecha_vencimiento": "",
                "vendedor": "",
                "tipo_cambio": "1.0000",
                "moneda": "HNL",
                "estado": "emitida",
                "orden_compra_exenta": "",
                "registro_exonerado": "",
                "registro_sag": "",
                "lineas-TOTAL_FORMS": "1",
                "lineas-INITIAL_FORMS": "0",
                "lineas-MIN_NUM_FORMS": "0",
                "lineas-MAX_NUM_FORMS": "1000",
                "lineas-0-producto": str(self.producto.id),
                "lineas-0-cantidad": "2.00",
                "lineas-0-precio_unitario": "100.00",
                "lineas-0-descuento_porcentaje": "0",
                "lineas-0-comentario": "",
                "lineas-0-impuesto": str(self.impuesto.id),
            },
        )

        self.assertEqual(response.status_code, 302)
        inventario = InventarioProducto.objects.get(producto=self.producto)
        self.assertEqual(inventario.existencias, Decimal("8.00"))
        self.assertTrue(MovimientoInventario.objects.filter(producto=self.producto, tipo="salida_factura").exists())

    def test_crear_factura_muestra_buscadores_de_cliente_y_producto(self):
        response = self.client.get(reverse("crear_factura", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Buscar cliente por nombre")
        self.assertContains(response, 'id="clientes-sugerencias"', html=False)
        self.assertContains(response, "Buscar producto")
        self.assertContains(response, 'id="inline-create-modal"', html=False)

    def test_crear_factura_usd_acepta_tipo_cambio_con_cuatro_decimales(self):
        response = self.client.post(
            reverse("crear_factura", args=[self.empresa.slug]),
            {
                "cliente": str(self.cliente.id),
                "fecha_emision": str(date.today()),
                "fecha_vencimiento": "",
                "vendedor": "",
                "tipo_cambio": "26.7348",
                "moneda": "USD",
                "estado": "borrador",
                "orden_compra_exenta": "",
                "registro_exonerado": "",
                "registro_sag": "",
                "lineas-TOTAL_FORMS": "1",
                "lineas-INITIAL_FORMS": "0",
                "lineas-MIN_NUM_FORMS": "0",
                "lineas-MAX_NUM_FORMS": "1000",
                "lineas-0-producto": str(self.producto.id),
                "lineas-0-cantidad": "1.00",
                "lineas-0-precio_unitario": "100.00",
                "lineas-0-descuento_porcentaje": "0",
                "lineas-0-comentario": "",
                "lineas-0-impuesto": str(self.impuesto.id),
            },
        )

        self.assertEqual(response.status_code, 302)
        factura = Factura.objects.latest("id")
        self.assertEqual(factura.moneda, "USD")
        self.assertEqual(factura.tipo_cambio, Decimal("26.7348"))

    def test_total_en_letras_conserva_centavos_en_usd(self):
        factura = Factura.objects.create(
            empresa=self.empresa,
            cliente=self.cliente,
            moneda="USD",
            tipo_cambio=Decimal("26.7348"),
            estado="borrador",
        )
        factura.total = Decimal("4004.80")

        self.assertEqual(
            factura.total_en_letras(),
            "SON: CUATRO MIL CUATRO DOLARES CON 80/100",
        )

    def test_crear_factura_normal_no_muestra_numero_manual(self):
        response = self.client.get(reverse("crear_factura", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertNotIn("numero_factura", response.context["form"].fields)

    def test_empresa_historica_muestra_y_acepta_numero_factura_manual(self):
        self.empresa.nombre = "AMKT Digital"
        self.empresa.slug = "amkt-digital"
        self.empresa.save(update_fields=["nombre", "slug"])
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.permite_cai_historico = True
        configuracion.save(update_fields=["permite_cai_historico"])
        InventarioProducto.objects.create(
            empresa=self.empresa,
            producto=self.producto,
            existencias=Decimal("10.00"),
            stock_minimo=Decimal("1.00"),
        )
        self.cai.fecha_activacion = date(2026, 1, 1)
        self.cai.fecha_limite = date(2026, 12, 31)
        self.cai.correlativo_actual = 0
        self.cai.save(update_fields=["fecha_activacion", "fecha_limite", "correlativo_actual"])

        response = self.client.get(reverse("crear_factura", args=[self.empresa.slug]))
        self.assertIn("numero_factura", response.context["form"].fields)
        self.assertIn("numero_factura_sufijo", response.context["form"].fields)
        self.assertContains(response, "001-001-01-00000")

        response = self.client.post(
            reverse("crear_factura", args=[self.empresa.slug]),
            {
                "cliente": str(self.cliente.id),
                "fecha_emision": "22/04/2026",
                "fecha_vencimiento": "",
                "vendedor": "",
                "tipo_cambio": "1.0000",
                "moneda": "HNL",
                "estado": "emitida",
                "numero_factura_sufijo": "005",
                "orden_compra_exenta": "",
                "registro_exonerado": "",
                "registro_sag": "",
                "lineas-TOTAL_FORMS": "1",
                "lineas-INITIAL_FORMS": "0",
                "lineas-MIN_NUM_FORMS": "0",
                "lineas-MAX_NUM_FORMS": "1000",
                "lineas-0-producto": str(self.producto.id),
                "lineas-0-cantidad": "2.00",
                "lineas-0-precio_unitario": "100.00",
                "lineas-0-descuento_porcentaje": "0",
                "lineas-0-comentario": "",
                "lineas-0-impuesto": str(self.impuesto.id),
            },
        )

        self.assertEqual(response.status_code, 302)
        factura = Factura.objects.latest("id")
        self.cai.refresh_from_db()
        self.assertEqual(factura.numero_factura, "001-001-01-00000005")
        self.assertEqual(factura.cai_id, self.cai.id)
        self.assertEqual(self.cai.correlativo_actual, 0)

        factura_auto = self.crear_factura_con_linea(estado="emitida", fecha_emision=date(2026, 4, 23))
        self.assertEqual(factura_auto.numero_factura, "001-001-01-00000001")

    def test_factura_historica_manual_no_empuja_correlativo_actual(self):
        self.empresa.nombre = "Digital Planning"
        self.empresa.slug = "digital_planning"
        self.empresa.save(update_fields=["nombre", "slug"])
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.permite_cai_historico = True
        configuracion.save(update_fields=["permite_cai_historico"])
        self.cai.rango_inicial = 276
        self.cai.rango_final = 1275
        self.cai.correlativo_actual = 330
        self.cai.save(update_fields=["rango_inicial", "rango_final", "correlativo_actual"])

        factura = self.crear_factura_con_linea(estado="borrador", fecha_emision=date(2026, 4, 22))
        factura.estado = "emitida"
        factura.numero_factura = "001-001-01-00000380"
        factura.save()

        self.cai.refresh_from_db()
        self.assertEqual(self.cai.correlativo_actual, 330)

        factura_nueva = self.crear_factura_con_linea(estado="emitida", fecha_emision=date(2026, 5, 16))
        self.assertEqual(factura_nueva.numero_factura, "001-001-01-00000331")

    def test_factura_automatica_salta_numeros_historicos_ya_usados(self):
        self.empresa.nombre = "Digital Planning"
        self.empresa.slug = "digital_planning"
        self.empresa.save(update_fields=["nombre", "slug"])
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.permite_cai_historico = True
        configuracion.save(update_fields=["permite_cai_historico"])
        self.cai.rango_inicial = 276
        self.cai.rango_final = 1275
        self.cai.correlativo_actual = 379
        self.cai.save(update_fields=["rango_inicial", "rango_final", "correlativo_actual"])

        factura_historica = self.crear_factura_con_linea(estado="borrador", fecha_emision=date(2026, 4, 22))
        factura_historica.estado = "emitida"
        factura_historica.numero_factura = "001-001-01-00000381"
        factura_historica.save()

        factura_380 = self.crear_factura_con_linea(estado="emitida", fecha_emision=date(2026, 5, 16))
        self.assertEqual(factura_380.numero_factura, "001-001-01-00000380")

        factura_382 = self.crear_factura_con_linea(estado="emitida", fecha_emision=date(2026, 5, 17))
        self.assertEqual(factura_382.numero_factura, "001-001-01-00000382")

    def test_empresa_historica_bloquea_numero_manual_fuera_de_cai(self):
        self.empresa.nombre = "INTEGRATED SALES AND SERVICES S. DE R.L."
        self.empresa.slug = "integrated-sales-and-services"
        self.empresa.save(update_fields=["nombre", "slug"])
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.permite_cai_historico = True
        configuracion.save(update_fields=["permite_cai_historico"])
        InventarioProducto.objects.create(
            empresa=self.empresa,
            producto=self.producto,
            existencias=Decimal("10.00"),
            stock_minimo=Decimal("1.00"),
        )
        self.cai.fecha_activacion = date(2026, 1, 1)
        self.cai.fecha_limite = date(2026, 12, 31)
        self.cai.rango_final = 10
        self.cai.save(update_fields=["fecha_activacion", "fecha_limite", "rango_final"])

        response = self.client.post(
            reverse("crear_factura", args=[self.empresa.slug]),
            {
                "cliente": str(self.cliente.id),
                "fecha_emision": "22/04/2026",
                "fecha_vencimiento": "",
                "vendedor": "",
                "tipo_cambio": "1.0000",
                "moneda": "HNL",
                "estado": "emitida",
                "numero_factura_sufijo": "025",
                "orden_compra_exenta": "",
                "registro_exonerado": "",
                "registro_sag": "",
                "lineas-TOTAL_FORMS": "1",
                "lineas-INITIAL_FORMS": "0",
                "lineas-MIN_NUM_FORMS": "0",
                "lineas-MAX_NUM_FORMS": "1000",
                "lineas-0-producto": str(self.producto.id),
                "lineas-0-cantidad": "1.00",
                "lineas-0-precio_unitario": "100.00",
                "lineas-0-descuento_porcentaje": "0",
                "lineas-0-comentario": "",
                "lineas-0-impuesto": str(self.impuesto.id),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No existe un CAI que cubra este numero para la fecha de la factura.")
        self.assertFalse(Factura.objects.exists())

    def test_puede_crear_factura_con_descripcion_manual_sin_producto(self):
        response = self.client.post(
            reverse("crear_factura", args=[self.empresa.slug]),
            {
                "cliente": str(self.cliente.id),
                "fecha_emision": "22/04/2026",
                "fecha_vencimiento": "",
                "vendedor": "",
                "tipo_cambio": "1.0000",
                "moneda": "HNL",
                "estado": "emitida",
                "orden_compra_exenta": "",
                "registro_exonerado": "",
                "registro_sag": "",
                "lineas-TOTAL_FORMS": "1",
                "lineas-INITIAL_FORMS": "0",
                "lineas-MIN_NUM_FORMS": "0",
                "lineas-MAX_NUM_FORMS": "1000",
                "lineas-0-producto": "",
                "lineas-0-descripcion_manual": "4 reels, 6 stories, 7 carrusel, 7 post estaticos",
                "lineas-0-cantidad": "1.00",
                "lineas-0-precio_unitario": "5000.00",
                "lineas-0-descuento_porcentaje": "0",
                "lineas-0-comentario": "",
                "lineas-0-impuesto": str(self.impuesto.id),
            },
        )

        self.assertEqual(response.status_code, 302)
        factura = Factura.objects.latest("id")
        linea = factura.lineas.get()
        self.assertIsNone(linea.producto)
        self.assertEqual(linea.descripcion_manual, "4 reels, 6 stories, 7 carrusel, 7 post estaticos")
        self.assertEqual(linea.descripcion_visual, "4 reels, 6 stories, 7 carrusel, 7 post estaticos")

    def test_no_permite_emitir_factura_si_no_hay_stock_suficiente(self):
        InventarioProducto.objects.create(
            empresa=self.empresa,
            producto=self.producto,
            existencias=Decimal("1.00"),
            stock_minimo=Decimal("1.00"),
        )

        response = self.client.post(
            reverse("crear_factura", args=[self.empresa.slug]),
            {
                "cliente": str(self.cliente.id),
                "fecha_emision": str(date.today()),
                "fecha_vencimiento": "",
                "vendedor": "",
                "tipo_cambio": "1.0000",
                "moneda": "HNL",
                "estado": "emitida",
                "orden_compra_exenta": "",
                "registro_exonerado": "",
                "registro_sag": "",
                "lineas-TOTAL_FORMS": "1",
                "lineas-INITIAL_FORMS": "0",
                "lineas-MIN_NUM_FORMS": "0",
                "lineas-MAX_NUM_FORMS": "1000",
                "lineas-0-producto": str(self.producto.id),
                "lineas-0-cantidad": "2.00",
                "lineas-0-precio_unitario": "100.00",
                "lineas-0-descuento_porcentaje": "0",
                "lineas-0-comentario": "",
                "lineas-0-impuesto": str(self.impuesto.id),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Stock insuficiente para emitir la factura")
        self.assertEqual(Factura.objects.count(), 0)
        inventario = InventarioProducto.objects.get(producto=self.producto)
        self.assertEqual(inventario.existencias, Decimal("1.00"))
        self.assertFalse(MovimientoInventario.objects.filter(producto=self.producto, tipo="salida_factura").exists())

    def test_editar_factura_exige_motivo_y_lo_guarda_en_auditoria(self):
        factura = self.crear_factura_con_linea(estado="borrador")
        linea = factura.lineas.get()
        datos = {
            "cliente": str(self.cliente.id),
            "fecha_emision": str(factura.fecha_emision),
            "fecha_vencimiento": str(factura.fecha_emision + timedelta(days=15)),
            "vendedor": "",
            "tipo_cambio": "1.0000",
            "moneda": "HNL",
            "estado": "borrador",
            "orden_compra_exenta": "",
            "registro_exonerado": "",
            "registro_sag": "",
            "lineas-TOTAL_FORMS": "1",
            "lineas-INITIAL_FORMS": "1",
            "lineas-MIN_NUM_FORMS": "0",
            "lineas-MAX_NUM_FORMS": "1000",
            "lineas-0-id": str(linea.id),
            "lineas-0-producto": str(self.producto.id),
            "lineas-0-cantidad": "1.00",
            "lineas-0-precio_unitario": "100.00",
            "lineas-0-descuento_porcentaje": "0",
            "lineas-0-comentario": "",
            "lineas-0-impuesto": str(self.impuesto.id),
        }

        sin_motivo = self.client.post(
            reverse("editar_factura", args=[self.empresa.slug, factura.id]), datos
        )
        self.assertEqual(sin_motivo.status_code, 200)
        self.assertIn("motivo_auditoria", sin_motivo.context["form"].errors)

        datos["motivo_auditoria"] = "Correccion de fecha solicitada por el cliente"
        con_motivo = self.client.post(
            reverse("editar_factura", args=[self.empresa.slug, factura.id]), datos
        )
        self.assertRedirects(
            con_motivo, reverse("facturas_dashboard", args=[self.empresa.slug])
        )
        evento = RegistroAuditoria.objects.filter(
            empresa=self.empresa,
            app_label="facturacion",
            modelo="factura",
            objeto_id=str(factura.id),
            accion="modificar",
            cambios__has_key="fecha_vencimiento",
        ).latest("fecha")
        self.assertEqual(evento.usuario, self.user)
        self.assertEqual(evento.motivo, "Correccion de fecha solicitada por el cliente")

    def test_si_permite_guardar_factura_borrador_aunque_no_haya_stock(self):
        InventarioProducto.objects.create(
            empresa=self.empresa,
            producto=self.producto,
            existencias=Decimal("1.00"),
            stock_minimo=Decimal("1.00"),
        )

        response = self.client.post(
            reverse("crear_factura", args=[self.empresa.slug]),
            {
                "cliente": str(self.cliente.id),
                "fecha_emision": str(date.today()),
                "fecha_vencimiento": "",
                "vendedor": "",
                "tipo_cambio": "1.0000",
                "moneda": "HNL",
                "estado": "borrador",
                "orden_compra_exenta": "",
                "registro_exonerado": "",
                "registro_sag": "",
                "lineas-TOTAL_FORMS": "1",
                "lineas-INITIAL_FORMS": "0",
                "lineas-MIN_NUM_FORMS": "0",
                "lineas-MAX_NUM_FORMS": "1000",
                "lineas-0-producto": str(self.producto.id),
                "lineas-0-cantidad": "2.00",
                "lineas-0-precio_unitario": "100.00",
                "lineas-0-descuento_porcentaje": "0",
                "lineas-0-comentario": "",
                "lineas-0-impuesto": str(self.impuesto.id),
            },
        )

        self.assertEqual(response.status_code, 302)
        factura = Factura.objects.get()
        self.assertEqual(factura.estado, "borrador")
        inventario = InventarioProducto.objects.get(producto=self.producto)
        self.assertEqual(inventario.existencias, Decimal("1.00"))
        self.assertFalse(MovimientoInventario.objects.filter(producto=self.producto, tipo="salida_factura").exists())

    def test_no_permite_cambiar_borrador_a_emitida_si_no_hay_stock(self):
        InventarioProducto.objects.create(
            empresa=self.empresa,
            producto=self.producto,
            existencias=Decimal("1.00"),
            stock_minimo=Decimal("1.00"),
        )
        factura = self.crear_factura_con_linea(estado="borrador")
        factura.cai = None
        factura.numero_factura = None
        factura.save(update_fields=["cai", "numero_factura"])

        response = self.client.post(
            reverse("editar_factura", args=[self.empresa.slug, factura.id]),
            {
                "cliente": str(self.cliente.id),
                "fecha_emision": str(factura.fecha_emision),
                "fecha_vencimiento": "",
                "vendedor": "",
                "tipo_cambio": "1.0000",
                "moneda": "HNL",
                "estado": "emitida",
                "orden_compra_exenta": "",
                "registro_exonerado": "",
                "registro_sag": "",
                "motivo_auditoria": "Emision autorizada por administracion",
                "lineas-TOTAL_FORMS": "1",
                "lineas-INITIAL_FORMS": "1",
                "lineas-MIN_NUM_FORMS": "0",
                "lineas-MAX_NUM_FORMS": "1000",
                "lineas-0-id": str(factura.lineas.first().id),
                "lineas-0-producto": str(self.producto.id),
                "lineas-0-cantidad": "2.00",
                "lineas-0-precio_unitario": "100.00",
                "lineas-0-descuento_porcentaje": "0",
                "lineas-0-comentario": "",
                "lineas-0-impuesto": str(self.impuesto.id),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Stock insuficiente para emitir la factura")
        factura.refresh_from_db()
        self.assertEqual(factura.estado, "borrador")
        self.assertIsNone(factura.numero_factura)
        inventario = InventarioProducto.objects.get(producto=self.producto)
        self.assertEqual(inventario.existencias, Decimal("1.00"))
        self.assertFalse(MovimientoInventario.objects.filter(producto=self.producto, tipo="salida_factura").exists())

    def test_no_permite_editar_factura_emitida_con_pagos(self):
        factura = self.crear_factura_con_linea(estado="emitida")
        PagoFactura.objects.create(
            factura=factura,
            monto=Decimal("25.00"),
            metodo="efectivo",
            fecha=date.today(),
        )

        response = self.client.get(
            reverse("editar_factura", args=[self.empresa.slug, factura.id]),
            follow=True,
        )

        self.assertRedirects(
            response,
            reverse("ver_factura", args=[self.empresa.slug, factura.id]),
        )
        self.assertContains(response, "No se puede editar esta factura emitida")
        self.assertContains(response, "pagos registrados")

    def test_permite_corregir_solo_numero_fiscal_en_factura_pagada(self):
        factura = self.crear_factura_con_linea(estado="emitida")
        pago = PagoFactura.objects.create(
            factura=factura,
            monto=Decimal("25.00"),
            metodo="efectivo",
            fecha=date.today(),
        )
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.permite_cai_historico = True
        configuracion.save(update_fields=["permite_cai_historico"])
        total_original = factura.total
        linea_original = factura.lineas.get()

        response = self.client.post(
            reverse("corregir_numero_factura", args=[self.empresa.slug, factura.id]),
            {
                "numero_factura": "001-001-01-00000002",
                "motivo": "Correccion del correlativo fiscal digitado.",
            },
        )

        self.assertRedirects(
            response,
            reverse("ver_factura", args=[self.empresa.slug, factura.id]),
        )
        factura.refresh_from_db()
        pago.refresh_from_db()
        linea_original.refresh_from_db()
        self.assertEqual(factura.numero_factura, "001-001-01-00000002")
        self.assertEqual(factura.total, total_original)
        self.assertEqual(linea_original.cantidad, Decimal("1.00"))
        self.assertEqual(pago.monto, Decimal("25.00"))
        self.assertTrue(PagoFactura.objects.filter(id=pago.id, factura=factura).exists())

        correccion = CorreccionNumeroFactura.objects.get(factura=factura)
        self.assertEqual(correccion.numero_anterior, "001-001-01-00000001")
        self.assertEqual(correccion.numero_nuevo, "001-001-01-00000002")
        self.assertEqual(correccion.realizado_por, self.user)

    def test_no_permite_correccion_numero_sin_configuracion_historica(self):
        factura = self.crear_factura_con_linea(estado="emitida")

        response = self.client.get(
            reverse("corregir_numero_factura", args=[self.empresa.slug, factura.id]),
            follow=True,
        )

        self.assertRedirects(
            response,
            reverse("ver_factura", args=[self.empresa.slug, factura.id]),
        )
        self.assertContains(response, "correccion fiscal historica no esta habilitada")
        self.assertFalse(CorreccionNumeroFactura.objects.filter(factura=factura).exists())

    def test_correccion_numero_rechaza_correlativo_fuera_del_cai(self):
        factura = self.crear_factura_con_linea(estado="emitida")
        configuracion = ConfiguracionAvanzadaEmpresa.para_empresa(self.empresa)
        configuracion.permite_cai_historico = True
        configuracion.save(update_fields=["permite_cai_historico"])

        response = self.client.post(
            reverse("corregir_numero_factura", args=[self.empresa.slug, factura.id]),
            {
                "numero_factura": "001-001-01-00000099",
                "motivo": "Prueba de numero fuera de rango.",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No existe un CAI que cubra este numero")
        factura.refresh_from_db()
        self.assertEqual(factura.numero_factura, "001-001-01-00000001")
        self.assertFalse(CorreccionNumeroFactura.objects.filter(factura=factura).exists())

    def test_no_permite_editar_factura_emitida_con_nota_credito(self):
        factura = self.crear_factura_con_linea(estado="emitida")
        nota = NotaCredito.objects.create(
            empresa=self.empresa,
            factura_origen=factura,
            cliente=factura.cliente,
            moneda=factura.moneda,
            tipo_cambio=factura.tipo_cambio,
            fecha_emision=date.today(),
            estado="borrador",
        )
        LineaNotaCredito.objects.create(
            nota_credito=nota,
            producto=self.producto,
            cantidad=Decimal("1.00"),
            precio_unitario=Decimal("100.00"),
            impuesto=self.impuesto,
        )

        response = self.client.get(
            reverse("editar_factura", args=[self.empresa.slug, factura.id]),
            follow=True,
        )

        self.assertRedirects(
            response,
            reverse("ver_factura", args=[self.empresa.slug, factura.id]),
        )
        self.assertContains(response, "No se puede editar esta factura emitida")
        self.assertContains(response, "notas de crédito")

    def test_nota_credito_emitida_devuelve_inventario(self):
        InventarioProducto.objects.create(
            empresa=self.empresa,
            producto=self.producto,
            existencias=Decimal("10.00"),
            stock_minimo=Decimal("1.00"),
        )
        factura = self.crear_factura_con_linea(estado="emitida")
        inventario = InventarioProducto.objects.get(producto=self.producto)
        inventario.existencias = Decimal("9.00")
        inventario.save(update_fields=["existencias", "fecha_actualizacion"])
        MovimientoInventario.objects.create(
            empresa=self.empresa,
            producto=self.producto,
            tipo="salida_factura",
            cantidad=Decimal("1.00"),
            existencia_anterior=Decimal("10.00"),
            existencia_resultante=Decimal("9.00"),
            referencia=factura.numero_factura or f"Factura {factura.id}",
            factura=factura,
        )
        nota = NotaCredito.objects.create(
            empresa=self.empresa,
            factura_origen=factura,
            cliente=factura.cliente,
            moneda=factura.moneda,
            tipo_cambio=factura.tipo_cambio,
            fecha_emision=date.today(),
            estado="borrador",
        )
        LineaNotaCredito.objects.create(
            nota_credito=nota,
            producto=self.producto,
            cantidad=Decimal("1.00"),
            precio_unitario=Decimal("100.00"),
            impuesto=self.impuesto,
        )
        nota.calcular_totales()
        nota.estado = "emitida"
        nota.save(update_fields=["subtotal", "impuesto", "total", "total_lempiras", "estado"])
        _registrar_entrada_nota_credito(nota)

        inventario.refresh_from_db()
        self.assertEqual(inventario.existencias, Decimal("10.00"))
        self.assertTrue(MovimientoInventario.objects.filter(producto=self.producto, tipo="devolucion_nota_credito").exists())

    def test_crear_nota_credito_tiene_buscador_con_sugerencias_de_factura(self):
        factura = self.crear_factura_con_linea(estado="emitida")

        response = self.client.get(reverse("crear_nota_credito", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="facturas-origen-sugeridas"', html=False)
        self.assertContains(response, factura.numero_factura or str(factura.id))

    def test_nota_credito_emitida_no_puede_exceder_total_factura(self):
        factura = self.crear_factura_con_linea(estado="emitida")
        nota = NotaCredito.objects.create(
            empresa=self.empresa,
            factura_origen=factura,
            cliente=factura.cliente,
            moneda=factura.moneda,
            tipo_cambio=factura.tipo_cambio,
            fecha_emision=date.today(),
            estado="borrador",
        )
        LineaNotaCredito.objects.create(
            nota_credito=nota,
            producto=self.producto,
            cantidad=Decimal("2.00"),
            precio_unitario=Decimal("100.00"),
            impuesto=self.impuesto,
        )
        nota.calcular_totales()
        nota.estado = "emitida"

        with self.assertRaises(ValidationError):
            nota.save(update_fields=["subtotal", "impuesto", "total", "total_lempiras", "estado"])

    def test_crear_cliente_respeta_next_seguro(self):
        next_url = reverse("crear_factura", args=[self.empresa.slug])
        response = self.client.post(
            f"{reverse('crear_cliente_facturacion', args=[self.empresa.slug])}?next={next_url}",
            {
                "nombre": "Cliente Con Next",
                "rtn": "08011999222222",
                "direccion": "Tegucigalpa",
                "ciudad": "Tegucigalpa",
                "activo": "on",
            },
        )

        self.assertRedirects(response, next_url)

    def test_pago_compra_no_puede_superar_saldo(self):
        compra = self.crear_compra_con_linea()

        with self.assertRaises(ValidationError):
            PagoCompra.objects.create(
                compra=compra,
                fecha=date.today(),
                monto=compra.total_documento + Decimal("1.00"),
                metodo="transferencia",
            )

    def test_no_permite_pago_en_compra_borrador(self):
        compra = self.crear_compra_con_linea(estado="borrador")

        response = self.client.get(
            reverse("registrar_pago_compra", args=[self.empresa.slug, compra.id]),
            follow=True,
        )

        self.assertRedirects(
            response,
            reverse("ver_compra", args=[self.empresa.slug, compra.id]),
        )
        self.assertContains(response, "aun esta en borrador")

    def test_pago_compra_reduce_saldo_pendiente(self):
        compra = self.crear_compra_con_linea()
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110299",
            nombre="Banco Pagos Proveedores",
            tipo="activo",
        )
        cuenta_financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Pagos Proveedores",
            tipo="banco",
            cuenta_contable=cuenta_banco,
        )

        response = self.client.post(
            reverse("registrar_pago_compra", args=[self.empresa.slug, compra.id]),
            {
                "fecha": str(date.today()),
                "monto": "60.00",
                "metodo": "transferencia",
                "cuenta_financiera": str(cuenta_financiera.id),
                "referencia": "TRX-001",
                "observacion": "Primer abono",
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        compra.refresh_from_db()
        self.assertEqual(compra.total_pagado, Decimal("60.00"))
        self.assertEqual(compra.saldo_pendiente, compra.total_documento - Decimal("60.00"))
        self.assertContains(response, "Pago de compra registrado correctamente")
        self.assertTrue(ComprobanteEgresoCompra.objects.filter(pago__compra=compra).exists())

    def test_reporte_cxp_muestra_proveedor_y_compras_pendientes(self):
        compra = self.crear_compra_con_linea()

        response = self.client.get(reverse("reporte_cxp", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, compra.proveedor_nombre)
        self.assertContains(response, "Cuentas por Pagar")

        detalle = self.client.get(
            reverse("reporte_cxp", args=[self.empresa.slug]) + f"?proveedor={compra.proveedor.id}"
        )
        self.assertEqual(detalle.status_code, 200)
        self.assertContains(detalle, compra.numero_compra)
        self.assertContains(detalle, "Registrar Pago")

    def test_reporte_cxp_permita_detalle_por_nombre_si_no_hay_fk_proveedor(self):
        compra = CompraInventario.objects.create(
            empresa=self.empresa,
            proveedor=None,
            proveedor_nombre="Proveedor Legacy",
            fecha_documento=date.today(),
            estado="aplicada",
        )
        LineaCompraInventario.objects.create(
            compra=compra,
            producto=self.producto,
            cantidad=Decimal("2.00"),
            costo_unitario=Decimal("25.00"),
        )

        response = self.client.get(
            reverse("reporte_cxp", args=[self.empresa.slug]) + "?proveedor_key=manual-proveedor-legacy"
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Proveedor Legacy")
        self.assertContains(response, compra.numero_compra)

    def test_pago_compra_genera_comprobante_egreso(self):
        compra = self.crear_compra_con_linea()
        pago = PagoCompra.objects.create(
            compra=compra,
            fecha=date.today(),
            monto=Decimal("40.00"),
            metodo="efectivo",
            referencia="CAJA-01",
            observacion="Pago parcial",
        )

        self.assertTrue(ComprobanteEgresoCompra.objects.filter(pago=pago).exists())
        comprobante = pago.comprobante
        self.assertEqual(comprobante.compra, compra)
        self.assertEqual(comprobante.proveedor_nombre, compra.proveedor_nombre)
        self.assertEqual(comprobante.monto, Decimal("40.00"))

    def test_proveedor_contado_normaliza_dias_credito(self):
        proveedor = Proveedor.objects.create(
            empresa=self.empresa,
            nombre="Proveedor Contado",
            condicion_pago="contado",
            dias_credito=30,
        )

        self.assertEqual(proveedor.dias_credito, 0)

    def test_compra_credito_calcula_fecha_vencimiento(self):
        compra = self.crear_compra_con_linea(condicion_pago="credito", dias_credito=15)

        self.assertEqual(compra.fecha_vencimiento, date.today() + timedelta(days=15))

    def test_reporte_cxp_usa_fecha_vencimiento_para_antiguedad(self):
        compra = self.crear_compra_con_linea(condicion_pago="credito", dias_credito=15)
        compra.fecha_documento = date.today() - timedelta(days=45)
        compra.fecha_vencimiento = date.today() + timedelta(days=2)
        compra.save(update_fields=["fecha_documento", "fecha_vencimiento"])

        response = self.client.get(reverse("reporte_cxp", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "L. 120.00")

    def test_egresos_dashboard_muestra_comprobante(self):
        compra = self.crear_compra_con_linea()
        pago = PagoCompra.objects.create(
            compra=compra,
            fecha=date.today(),
            monto=Decimal("30.00"),
            metodo="transferencia",
            referencia="TRX-EG-01",
        )

        response = self.client.get(reverse("egresos_dashboard", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, pago.comprobante.numero_comprobante)
        self.assertContains(response, compra.proveedor_nombre)

    def test_ver_compra_muestra_boton_pagar_cuando_hay_saldo(self):
        compra = self.crear_compra_con_linea()

        response = self.client.get(reverse("ver_compra", args=[self.empresa.slug, compra.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Registrar Pago")

    def test_revertir_pago_compra_elimina_comprobante_y_restaurar_saldo(self):
        compra = self.crear_compra_con_linea()
        pago = PagoCompra.objects.create(
            compra=compra,
            fecha=date.today(),
            monto=Decimal("30.00"),
            metodo="efectivo",
        )

        response = self.client.post(
            reverse("revertir_pago_compra", args=[self.empresa.slug, compra.id, pago.id]),
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(PagoCompra.objects.filter(id=pago.id).exists())
        self.assertFalse(ComprobanteEgresoCompra.objects.filter(compra=compra).exists())
        compra.refresh_from_db()
        self.assertEqual(compra.total_pagado, Decimal("0.00"))
        self.assertContains(response, "Pago de compra revertido correctamente")

    def test_filtro_compras_por_busqueda_estado_y_pago(self):
        compra = self.crear_compra_con_linea()
        PagoCompra.objects.create(
            compra=compra,
            fecha=date.today(),
            monto=Decimal("10.00"),
            metodo="transferencia",
        )

        response = self.client.get(
            reverse("compras_dashboard", args=[self.empresa.slug]),
            {"q": compra.numero_compra, "estado": "aplicada", "pago": "parcial"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, compra.numero_compra)

    def test_filtro_proveedores_por_busqueda_y_estado(self):
        proveedor = Proveedor.objects.create(
            empresa=self.empresa,
            nombre="Proveedor Busqueda",
            activo=False,
        )

        response = self.client.get(
            reverse("proveedores_facturacion", args=[self.empresa.slug]),
            {"q": "Busqueda", "estado": "inactivos"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, proveedor.nombre)

    def test_filtro_egresos_por_busqueda_y_metodo(self):
        compra = self.crear_compra_con_linea()
        pago = PagoCompra.objects.create(
            compra=compra,
            fecha=date.today(),
            monto=Decimal("35.00"),
            metodo="transferencia",
            referencia="TRX-FILTRO",
        )

        response = self.client.get(
            reverse("egresos_dashboard", args=[self.empresa.slug]),
            {"q": "TRX-FILTRO", "metodo": "transferencia"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, pago.comprobante.numero_comprobante)

    def test_reportes_muestra_iframe_power_bi_cuando_esta_configurado(self):
        ConfiguracionPowerBIEmpresa.objects.create(
            empresa=self.empresa,
            activo=True,
            mostrar_en_reportes=True,
            titulo_panel="Panel de Gerencia",
            descripcion_panel="Resumen ejecutivo conectado a Power BI.",
            url_embed="https://app.powerbi.com/reportEmbed?reportId=demo-report",
            alto_iframe=640,
        )

        response = self.client.get(reverse("reportes_facturacion", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Panel de Gerencia")
        self.assertContains(response, "reportEmbed?reportId=demo-report")
        self.assertContains(response, "<iframe", html=False)

    def test_administrador_empresa_puede_guardar_configuracion_power_bi(self):
        self.user.es_administrador_empresa = True
        self.user.save(update_fields=["es_administrador_empresa"])

        response = self.client.post(
            reverse("configuracion_power_bi_reportes", args=[self.empresa.slug]),
            {
                "activo": "on",
                "mostrar_en_reportes": "on",
                "titulo_panel": "Indicadores ejecutivos",
                "descripcion_panel": "Vista consolidada para direccion.",
                "url_embed": "https://app.powerbi.com/reportEmbed?reportId=panel-seguro",
                "alto_iframe": "700",
                "workspace_id": "workspace-demo",
                "report_id": "report-demo",
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Configuracion Power BI actualizada correctamente.")

        configuracion = ConfiguracionPowerBIEmpresa.objects.get(empresa=self.empresa)
        self.assertTrue(configuracion.activo)
        self.assertEqual(configuracion.titulo_panel, "Indicadores ejecutivos")
        self.assertEqual(configuracion.url_embed, "https://app.powerbi.com/reportEmbed?reportId=panel-seguro")

    def test_reportes_muestra_bi_interno_con_clientes_y_productos(self):
        self.crear_factura_con_linea(estado="emitida")

        response = self.client.get(reverse("reportes_facturacion", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "BI interno de facturacion")
        self.assertContains(response, "Clientes con mayor facturacion")
        self.assertContains(response, self.cliente.nombre)
        self.assertContains(response, self.producto.nombre)

    def test_dashboard_bi_facturacion_renderiza_panel_ejecutivo(self):
        self.crear_factura_con_linea(estado="emitida")

        response = self.client.get(reverse("dashboard_bi_facturacion", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Dashboard BI de Facturación")
        self.assertContains(response, "Lectura inteligente de la facturación en una sola vista.")
        self.assertContains(response, "Productos más vendidos")
        self.assertContains(response, self.cliente.nombre)

    def test_dashboard_bi_facturacion_muestra_bancos_y_cartera_vencida(self):
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110210",
            nombre="Banco Comercial",
            tipo="activo",
            acepta_movimientos=True,
        )
        cuenta_financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Atlántida HNL",
            tipo="banco",
            cuenta_contable=cuenta_banco,
        )
        factura_vencida = self.crear_factura_con_linea(
            estado="emitida",
            fecha_emision=date.today() - timedelta(days=45),
        )
        factura_vencida.fecha_vencimiento = date.today() - timedelta(days=15)
        factura_vencida.save(update_fields=["fecha_vencimiento"])
        PagoFactura.objects.create(
            factura=factura_vencida,
            monto=Decimal("25.00"),
            metodo="transferencia",
            fecha=date.today() - timedelta(days=3),
            cuenta_financiera=cuenta_financiera,
        )

        response = self.client.get(reverse("dashboard_bi_facturacion", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cobrado vs facturado")
        self.assertContains(response, "Ingresos por banco y caja")
        self.assertContains(response, "Clientes con saldo vencido")
        self.assertContains(response, "Banco Atlántida HNL")
        self.assertContains(response, self.cliente.nombre)
