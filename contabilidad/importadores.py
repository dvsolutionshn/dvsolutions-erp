from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import datetime
import re
import unicodedata
from zipfile import BadZipFile

from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .models import CuentaContable, MovimientoBancario


TIPOS_CUENTA = {clave: clave for clave, _ in CuentaContable.TIPO_CHOICES}
TIPOS_CUENTA.update({
    "activos": "activo",
    "assets": "activo",
    "pasivos": "pasivo",
    "liability": "pasivo",
    "liabilities": "pasivo",
    "patrimonio": "patrimonio",
    "capital": "patrimonio",
    "ingresos": "ingreso",
    "income": "ingreso",
    "revenue": "ingreso",
    "costo": "costo",
    "costos": "costo",
    "cost": "costo",
    "costs": "costo",
    "gastos": "gasto",
    "expenses": "gasto",
})

COLUMNAS = {
    "codigo": {"codigo", "code", "cuenta", "numero", "numero_cuenta"},
    "nombre": {"nombre", "name", "descripcion_cuenta", "cuenta_nombre"},
    "tipo": {"tipo", "type", "clase", "categoria"},
    "codigo_padre": {"codigo_padre", "padre", "cuenta_padre", "parent", "parent_code"},
    "acepta_movimientos": {"acepta_movimientos", "movimiento", "acepta_movimiento", "posting", "movimientos"},
    "activa": {"activa", "activo", "estado", "active"},
    "descripcion": {"descripcion", "description", "detalle"},
}


def _normalizar_texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def _normalizar_encabezado(valor):
    texto = unicodedata.normalize("NFKD", _normalizar_texto(valor))
    texto = "".join(caracter for caracter in texto if not unicodedata.combining(caracter))
    return texto.lower().replace(" ", "_").replace("-", "_")


def _normalizar_booleano(valor, por_defecto=True):
    texto = _normalizar_texto(valor).lower()
    if not texto:
        return por_defecto
    if texto in {"1", "si", "sí", "s", "true", "verdadero", "activo", "activa", "yes", "y"}:
        return True
    if texto in {"0", "no", "n", "false", "falso", "inactivo", "inactiva"}:
        return False
    return por_defecto


def _normalizar_tipo(valor):
    tipo = _normalizar_texto(valor).lower()
    return TIPOS_CUENTA.get(tipo)


def _inferir_tipo_desde_codigo(codigo):
    primer_digito = _normalizar_texto(codigo)[:1]
    return {
        "1": "activo",
        "2": "pasivo",
        "3": "patrimonio",
        "4": "ingreso",
        "5": "costo",
        "6": "gasto",
    }.get(primer_digito)


def _mapear_columnas(encabezados):
    indices = {}
    encabezados_normalizados = [_normalizar_encabezado(valor) for valor in encabezados]
    for campo, alias in COLUMNAS.items():
        for index, encabezado in enumerate(encabezados_normalizados):
            if encabezado in alias:
                indices[campo] = index
                break
    return indices


def _valor(fila, indices, campo):
    index = indices.get(campo)
    if index is None:
        return ""
    return _normalizar_texto(fila[index])


def _es_codigo_contable(valor):
    texto = _normalizar_texto(valor)
    return bool(texto) and texto.replace(".", "").replace("-", "").isdigit()


def _detectar_formato_jerarquico(sheet):
    coincidencias = 0
    for row in range(1, min(sheet.max_row, 40) + 1):
        codigo = sheet.cell(row, 8).value
        nombre = sheet.cell(row, 9).value
        if _es_codigo_contable(codigo) and _normalizar_texto(nombre):
            coincidencias += 1
    return coincidencias >= 3


def _extraer_catalogo_jerarquico(sheet):
    datos = []
    errores = []
    codigos_en_archivo = {}
    ultimo_codigo_por_nivel = {}

    for row in range(1, sheet.max_row + 1):
        codigo = _normalizar_texto(sheet.cell(row, 8).value)
        nombre = _normalizar_texto(sheet.cell(row, 9).value)
        if not codigo and not nombre:
            continue
        if not _es_codigo_contable(codigo) or not nombre:
            continue

        tipo = _inferir_tipo_desde_codigo(codigo)
        if not tipo:
            errores.append(f"Fila {row}: no se pudo inferir el tipo contable para el codigo {codigo}.")
            continue

        if codigo in codigos_en_archivo:
            fila_original, nombre_original = codigos_en_archivo[codigo]
            errores.append(
                f"Fila {row}: codigo duplicado {codigo}. Ya existe en fila {fila_original} como {nombre_original}."
            )
            continue

        nivel = sum(1 for col in range(1, 8) if _normalizar_texto(sheet.cell(row, col).value))
        nivel = nivel or 1
        codigo_padre = ultimo_codigo_por_nivel.get(nivel - 1, "")

        codigos_en_archivo[codigo] = (row, nombre)
        ultimo_codigo_por_nivel[nivel] = codigo
        for nivel_obsoleto in list(ultimo_codigo_por_nivel):
            if nivel_obsoleto > nivel:
                ultimo_codigo_por_nivel.pop(nivel_obsoleto, None)

        datos.append({
            "codigo": codigo,
            "nombre": nombre,
            "tipo": tipo,
            "codigo_padre": codigo_padre,
            "descripcion": f"Importado desde formato jerarquico. Nivel {nivel}.",
            "acepta_movimientos": True,
            "activa": True,
        })

    codigos_con_hijos = {item["codigo_padre"] for item in datos if item["codigo_padre"]}
    for item in datos:
        if item["codigo"] in codigos_con_hijos:
            item["acepta_movimientos"] = False

    if not datos and not errores:
        errores.append("No se encontraron cuentas en el formato jerarquico esperado: codigo en columna H y nombre en columna I.")

    return datos, errores


def _extraer_catalogo_con_encabezados(filas):
    indices = _mapear_columnas(filas[0])
    requeridas = {"codigo", "nombre", "tipo"}
    faltantes = sorted(requeridas - set(indices))
    if faltantes:
        raise ValidationError(f"Faltan columnas obligatorias: {', '.join(faltantes)}.")

    datos = []
    errores = []
    codigos_en_archivo = set()
    for numero_fila, fila in enumerate(filas[1:], start=2):
        codigo = _valor(fila, indices, "codigo")
        nombre = _valor(fila, indices, "nombre")
        tipo = _normalizar_tipo(_valor(fila, indices, "tipo"))
        codigo_padre = _valor(fila, indices, "codigo_padre")
        descripcion = _valor(fila, indices, "descripcion")
        acepta_movimientos = _normalizar_booleano(_valor(fila, indices, "acepta_movimientos"), por_defecto=True)
        activa = _normalizar_booleano(_valor(fila, indices, "activa"), por_defecto=True)

        if not codigo and not nombre and not tipo:
            continue
        if not codigo:
            errores.append(f"Fila {numero_fila}: falta codigo.")
            continue
        if codigo in codigos_en_archivo:
            errores.append(f"Fila {numero_fila}: codigo duplicado en el archivo ({codigo}).")
            continue
        if not nombre:
            errores.append(f"Fila {numero_fila}: falta nombre para {codigo}.")
            continue
        if not tipo:
            errores.append(f"Fila {numero_fila}: tipo invalido para {codigo}. Usa activo, pasivo, patrimonio, ingreso, costo o gasto.")
            continue

        codigos_en_archivo.add(codigo)
        datos.append({
            "codigo": codigo,
            "nombre": nombre,
            "tipo": tipo,
            "codigo_padre": codigo_padre,
            "descripcion": descripcion,
            "acepta_movimientos": acepta_movimientos,
            "activa": activa,
        })
    return datos, errores


def importar_catalogo_cuentas_desde_excel(empresa, archivo, *, actualizar_existentes=True):
    workbook = load_workbook(archivo, data_only=True)
    sheet = workbook.active
    filas = list(sheet.iter_rows(values_only=True))
    if not filas:
        raise ValidationError("El archivo esta vacio.")

    if _detectar_formato_jerarquico(sheet):
        datos, errores = _extraer_catalogo_jerarquico(sheet)
    else:
        datos, errores = _extraer_catalogo_con_encabezados(filas)

    codigos_en_archivo = {item["codigo"] for item in datos}
    codigos_existentes = set(CuentaContable.objects.filter(empresa=empresa).values_list("codigo", flat=True))
    for item in datos:
        codigo_padre = item["codigo_padre"]
        if codigo_padre and codigo_padre not in codigos_en_archivo and codigo_padre not in codigos_existentes:
            errores.append(f"Cuenta {item['codigo']}: la cuenta padre {codigo_padre} no existe.")
        if item["codigo"] in codigos_existentes and not actualizar_existentes:
            errores.append(f"Cuenta {item['codigo']}: ya existe y la actualizacion esta desactivada.")

    if errores:
        raise ValidationError(errores)

    creadas = 0
    actualizadas = 0
    with transaction.atomic():
        cuentas_por_codigo = {
            cuenta.codigo: cuenta
            for cuenta in CuentaContable.objects.filter(empresa=empresa)
        }

        for item in datos:
            cuenta = cuentas_por_codigo.get(item["codigo"])
            if cuenta:
                cuenta.nombre = item["nombre"]
                cuenta.tipo = item["tipo"]
                cuenta.descripcion = item["descripcion"]
                cuenta.acepta_movimientos = item["acepta_movimientos"]
                cuenta.activa = item["activa"]
                actualizadas += 1
            else:
                cuenta = CuentaContable(
                    empresa=empresa,
                    codigo=item["codigo"],
                    nombre=item["nombre"],
                    tipo=item["tipo"],
                    descripcion=item["descripcion"],
                    acepta_movimientos=item["acepta_movimientos"],
                    activa=item["activa"],
                )
                creadas += 1
            cuenta.cuenta_padre = None
            cuenta.full_clean()
            cuenta.save()
            cuentas_por_codigo[cuenta.codigo] = cuenta

        for item in datos:
            cuenta = cuentas_por_codigo[item["codigo"]]
            codigo_padre = item["codigo_padre"]
            cuenta.cuenta_padre = cuentas_por_codigo.get(codigo_padre) if codigo_padre else None
            cuenta.full_clean()
            cuenta.save(update_fields=["cuenta_padre"])

    return {
        "creadas": creadas,
        "actualizadas": actualizadas,
        "total": creadas + actualizadas,
    }


MOVIMIENTO_COLUMNAS = {
    "fecha": {"fecha", "date"},
    "descripcion": {"descripcion", "description", "detalle", "concepto", "transaccion", "descripcion_del_movimiento"},
    "referencia": {"referencia", "reference", "documento", "numero", "ref"},
    "debito": {"debito", "debitos", "retiro", "retiros", "cargo", "cargos", "debit", "debito_l", "debitos_l"},
    "credito": {"credito", "creditos", "deposito", "depositos", "abono", "abonos", "credit", "credito_l", "creditos_l"},
    "saldo": {"saldo", "saldos", "balance", "saldo_actual"},
}

MESES_ESTADO_CUENTA = {
    "ENE": 1,
    "FEB": 2,
    "MAR": 3,
    "ABR": 4,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AGO": 8,
    "AUG": 8,
    "SEP": 9,
    "SEPT": 9,
    "OCT": 10,
    "NOV": 11,
    "DIC": 12,
    "DEC": 12,
}


def _decimal_movimiento(valor):
    if valor in (None, ""):
        return Decimal("0.00")
    texto = str(valor).upper().replace(",", "").replace("L.", "").replace("HNL", "").strip()
    texto = texto.replace(" ", "")
    if texto in {"", "-", "--"}:
        return Decimal("0.00")
    negativo = texto.startswith("(") and texto.endswith(")")
    texto = texto.replace("(", "").replace(")", "")
    try:
        monto = Decimal(texto).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return -monto if negativo else monto
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


def _inferir_mes_estado_cuenta(sheet):
    patron = re.compile(r"\b([A-Z]{3,4})[/\-\s](\d{1,2})[/\-\s](\d{2,4})\b")
    for row in range(1, min(sheet.max_row, 25) + 1):
        for col in range(1, min(sheet.max_column, 12) + 1):
            texto = _normalizar_texto(sheet.cell(row, col).value).upper()
            coincidencia = patron.search(texto)
            if coincidencia:
                mes = MESES_ESTADO_CUENTA.get(coincidencia.group(1))
                if mes:
                    return mes
    return None


def _fecha_movimiento(valor, mes_preferido=None):
    if hasattr(valor, "date"):
        return valor.date()
    texto = _normalizar_texto(valor)

    candidatos = []
    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%y", "%d/%m/%y", "%d-%m-%y"):
        try:
            candidatos.append(datetime.strptime(texto, formato).date())
        except ValueError:
            continue
    if mes_preferido:
        for fecha in candidatos:
            if fecha.month == mes_preferido:
                return fecha
    if candidatos:
        return candidatos[0]
    return None


def _detectar_encabezado_movimientos(sheet):
    for row in range(1, min(sheet.max_row, 80) + 1):
        valores = [sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)]
        indices = _mapear_columnas_generico(valores, MOVIMIENTO_COLUMNAS)
        if {"fecha", "descripcion"}.issubset(indices) and ("debito" in indices or "credito" in indices):
            return row, indices
    raise ValidationError("No se encontro una fila de encabezados valida para el estado de cuenta.")


def _mapear_columnas_generico(encabezados, definicion):
    indices = {}
    encabezados_normalizados = [_normalizar_encabezado(valor) for valor in encabezados]
    for campo, alias in definicion.items():
        for index, encabezado in enumerate(encabezados_normalizados, start=1):
            if encabezado in alias:
                indices[campo] = index
                break
    return indices


def _crear_movimiento_bancario(empresa, cuenta_financiera, *, fecha, descripcion, referencia, debito, credito, saldo, origen_importacion):
    existe = MovimientoBancario.objects.filter(
        empresa=empresa,
        cuenta_financiera=cuenta_financiera,
        fecha=fecha,
        descripcion__iexact=descripcion,
        referencia__iexact=referencia,
        debito=debito,
        credito=credito,
    ).exists()
    if existe:
        return False

    movimiento = MovimientoBancario(
        empresa=empresa,
        cuenta_financiera=cuenta_financiera,
        fecha=fecha,
        descripcion=descripcion,
        referencia=referencia,
        debito=debito,
        credito=credito,
        saldo=saldo,
        origen_importacion=origen_importacion,
    )
    movimiento.full_clean()
    movimiento.save()
    return movimiento


def _inferir_mes_estado_cuenta_texto(texto):
    patron = re.compile(r"\b([A-Z]{3,4})[/\-\s](\d{1,2})[/\-\s](\d{2,4})\b")
    coincidencia = patron.search(texto.upper())
    if not coincidencia:
        return None
    return MESES_ESTADO_CUENTA.get(coincidencia.group(1))


def _descripcion_parece_credito(descripcion):
    texto = _normalizar_encabezado(descripcion).replace("_", " ")
    palabras_credito = {"deposito", "deposit", "abono", "credito", "credit", "cr ", "pago cliente"}
    palabras_debito = {"pago", "db ", "debito", "debit", "comision", "proveedor", "retiro", "ach", "cuota"}
    if any(palabra in f" {texto} " for palabra in palabras_credito):
        return True
    if any(palabra in f" {texto} " for palabra in palabras_debito):
        return False
    return False


def _extraer_movimientos_desde_texto_pdf(texto, mes_preferido=None):
    movimientos = []
    omitidos = []
    patron_linea = re.compile(
        r"^\s*(\d{1,2}/\d{1,2}/\d{2,4})\s+(.+?)\s+((?:-?\(?\d[\d,]*\.\d{2}\)?\s*){1,3})\s*$"
    )
    patron_monto = re.compile(r"-?\(?\d[\d,]*\.\d{2}\)?")

    for numero_linea, linea in enumerate(texto.splitlines(), start=1):
        coincidencia = patron_linea.match(linea)
        if not coincidencia:
            continue

        fecha = _fecha_movimiento(coincidencia.group(1), mes_preferido=mes_preferido)
        descripcion = _normalizar_texto(coincidencia.group(2))
        montos = [_decimal_movimiento(monto) for monto in patron_monto.findall(coincidencia.group(3))]

        if not fecha or not descripcion or len(montos) < 2:
            omitidos.append(numero_linea)
            continue

        debito = Decimal("0.00")
        credito = Decimal("0.00")
        saldo = montos[-1]

        monto_movimiento = montos[-2]
        if _descripcion_parece_credito(descripcion):
            credito = monto_movimiento
        else:
            debito = monto_movimiento

        if debito == 0 and credito == 0:
            omitidos.append(numero_linea)
            continue

        movimientos.append({
            "fecha": fecha,
            "descripcion": descripcion,
            "referencia": "",
            "debito": debito,
            "credito": credito,
            "saldo": saldo,
        })

    return movimientos, omitidos


def _importar_movimientos_bancarios_desde_pdf(empresa, cuenta_financiera, archivo):
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise ValidationError(
            "Para importar PDF instala la dependencia pypdf. Puedes hacerlo con: pip install pypdf."
        ) from exc

    try:
        archivo.seek(0)
        lector = PdfReader(archivo)
        texto = "\n".join(pagina.extract_text() or "" for pagina in lector.pages)
    except Exception as exc:
        raise ValidationError("No se pudo leer el PDF. Verifica que no este protegido, escaneado como imagen o danado.") from exc

    if not texto.strip():
        raise ValidationError("No se encontro texto legible en el PDF. Si es un PDF escaneado, necesitaremos OCR.")

    movimientos, omitidos = _extraer_movimientos_desde_texto_pdf(
        texto,
        mes_preferido=_inferir_mes_estado_cuenta_texto(texto),
    )
    if not movimientos:
        raise ValidationError("No se encontraron movimientos bancarios en el PDF con el formato esperado.")

    creados = 0
    duplicados = 0
    movimiento_ids = []
    with transaction.atomic():
        for movimiento in movimientos:
            creado = _crear_movimiento_bancario(
                empresa,
                cuenta_financiera,
                origen_importacion=getattr(archivo, "name", "Estado de cuenta PDF"),
                **movimiento,
            )
            if creado:
                creados += 1
                movimiento_ids.append(creado.id)
            else:
                duplicados += 1

    return {"creados": creados, "duplicados": duplicados, "omitidos": omitidos, "movimiento_ids": movimiento_ids}


def importar_movimientos_bancarios_desde_excel(empresa, cuenta_financiera, archivo):
    nombre = getattr(archivo, "name", "").lower()
    if nombre.endswith(".pdf"):
        return _importar_movimientos_bancarios_desde_pdf(empresa, cuenta_financiera, archivo)

    try:
        workbook = load_workbook(archivo, data_only=True)
    except (BadZipFile, InvalidFileException, OSError) as exc:
        raise ValidationError(
            "El estado de cuenta debe ser un archivo Excel valido (.xlsx o .xlsm). "
            "Si lo tienes en PDF, subelo con extension .pdf para usar el importador PDF."
        ) from exc
    sheet = workbook.active
    header_row, indices = _detectar_encabezado_movimientos(sheet)
    mes_preferido = _inferir_mes_estado_cuenta(sheet)
    creados = 0
    duplicados = 0
    movimiento_ids = []
    omitidos = []

    with transaction.atomic():
        for row in range(header_row + 1, sheet.max_row + 1):
            fecha = _fecha_movimiento(sheet.cell(row, indices["fecha"]).value, mes_preferido=mes_preferido)
            descripcion = _normalizar_texto(sheet.cell(row, indices["descripcion"]).value)
            referencia = _normalizar_texto(sheet.cell(row, indices.get("referencia")).value) if indices.get("referencia") else ""
            debito = _decimal_movimiento(sheet.cell(row, indices.get("debito")).value) if indices.get("debito") else Decimal("0.00")
            credito = _decimal_movimiento(sheet.cell(row, indices.get("credito")).value) if indices.get("credito") else Decimal("0.00")
            saldo = _decimal_movimiento(sheet.cell(row, indices.get("saldo")).value) if indices.get("saldo") else Decimal("0.00")

            if not fecha and not descripcion and debito == 0 and credito == 0:
                continue
            if not fecha or not descripcion or (debito == 0 and credito == 0):
                omitidos.append(row)
                continue

            creado = _crear_movimiento_bancario(
                empresa,
                cuenta_financiera,
                fecha=fecha,
                descripcion=descripcion,
                referencia=referencia,
                debito=debito,
                credito=credito,
                saldo=saldo,
                origen_importacion=getattr(archivo, "name", "Estado de cuenta"),
            )
            if not creado:
                duplicados += 1
                continue
            movimiento_ids.append(creado.id)
            creados += 1

    return {"creados": creados, "duplicados": duplicados, "omitidos": omitidos, "movimiento_ids": movimiento_ids}
