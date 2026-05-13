from decimal import Decimal
from datetime import date
from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook

from core.models import Empresa, Modulo, PlanComercial, PlanModulo, RolSistema, Usuario

from .importadores import _extraer_movimientos_desde_texto_pdf
from .models import AsientoContable, ClasificacionCompraFiscal, ClasificacionMovimientoBanco, ConfiguracionContableEmpresa, CuentaContable, CuentaFinanciera, LineaAsientoContable, MovimientoBancario, PeriodoContable, ReglaClasificacionBanco
from .services import registrar_asiento_factura_emitida, registrar_asiento_nota_credito
from facturacion.models import (
    CAI,
    Cliente,
    CompraInventario,
    LineaCompraInventario,
    LineaFactura,
    LineaNotaCredito,
    NotaCredito,
    PagoCompra,
    Producto,
    Proveedor,
    RegistroCompraFiscal,
    TipoImpuesto,
    Factura,
)


class ContabilidadTests(TestCase):
    def setUp(self):
        self.empresa = Empresa.objects.create(
            nombre="Empresa Contable",
            slug="empresa-contable",
            rtn="08011999111111",
            estado_licencia="activa",
        )
        self.modulo, _ = Modulo.objects.get_or_create(
            codigo="contabilidad",
            defaults={"nombre": "Contabilidad", "es_comercial": True},
        )
        self.modulo_facturacion, _ = Modulo.objects.get_or_create(
            codigo="facturacion",
            defaults={"nombre": "Facturacion", "es_comercial": True},
        )
        self.plan = PlanComercial.objects.create(nombre="Plan Conta", codigo="plan-conta", precio_mensual="199.00")
        PlanModulo.objects.create(plan=self.plan, modulo=self.modulo, activo=True)
        PlanModulo.objects.create(plan=self.plan, modulo=self.modulo_facturacion, activo=True)
        self.empresa.plan_comercial = self.plan
        self.empresa.save(update_fields=["plan_comercial"])
        self.rol = RolSistema.objects.create(
            nombre="Contador",
            codigo="contador",
            puede_contabilidad=True,
            puede_catalogo_cuentas=True,
            puede_crear_asientos=True,
            puede_contabilizar_asientos=True,
            puede_reportes_contables=True,
            puede_facturas=True,
            puede_crear_facturas=True,
            puede_registrar_pagos_clientes=True,
            puede_compras=True,
            puede_crear_compras=True,
            puede_aplicar_compras=True,
            puede_registrar_pagos_proveedores=True,
            puede_notas_credito=True,
            puede_crear_notas_credito=True,
        )
        self.usuario = Usuario.objects.create_user(
            username="contador",
            password="pass12345",
            empresa=self.empresa,
            rol_sistema=self.rol,
        )
        self.cliente = Cliente.objects.create(empresa=self.empresa, nombre="Cliente Demo")
        self.impuesto = TipoImpuesto.objects.create(nombre="ISV 15%", porcentaje=Decimal("15.00"))
        self.producto = Producto.objects.create(
            empresa=self.empresa,
            nombre="Servicio Demo",
            tipo_item="servicio",
            unidad_medida="servicio",
            precio=Decimal("100.00"),
            controla_inventario=False,
            impuesto_predeterminado=self.impuesto,
        )
        self.proveedor = Proveedor.objects.create(empresa=self.empresa, nombre="Proveedor Demo")
        self.cai_factura = CAI.objects.create(
            empresa=self.empresa,
            numero_cai="CAI-FACTURA",
            uso_documento="factura",
            establecimiento="001",
            punto_emision="001",
            tipo_documento="01",
            rango_inicial=1,
            rango_final=99999999,
            correlativo_actual=0,
            fecha_limite=self.empresa.fecha_vencimiento_plan,
            activo=True,
        )
        self.cai_nota = CAI.objects.create(
            empresa=self.empresa,
            numero_cai="CAI-NOTA",
            uso_documento="nota_credito",
            establecimiento="001",
            punto_emision="001",
            tipo_documento="02",
            rango_inicial=1,
            rango_final=99999999,
            correlativo_actual=0,
            fecha_limite=self.empresa.fecha_vencimiento_plan,
            activo=True,
        )

    def test_dashboard_contabilidad_responde(self):
        self.client.login(username="contador", password="pass12345")
        response = self.client.get(reverse("contabilidad_dashboard", args=[self.empresa.slug]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Contabilidad")

    def test_periodos_contables_responden_y_cierran_mes(self):
        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("crear_periodo_contable", args=[self.empresa.slug]),
            {"anio": "2026", "mes": "4", "estado": "abierto", "observacion": "Periodo de abril"},
        )
        self.assertRedirects(response, reverse("periodos_contables", args=[self.empresa.slug]))
        periodo = PeriodoContable.objects.get(empresa=self.empresa, anio=2026, mes=4)
        self.assertEqual(periodo.estado, "abierto")

        response = self.client.post(reverse("cerrar_periodo_contable", args=[self.empresa.slug, periodo.id]))
        self.assertRedirects(response, reverse("periodos_contables", args=[self.empresa.slug]))
        periodo.refresh_from_db()
        self.assertEqual(periodo.estado, "cerrado")
        self.assertEqual(periodo.cerrado_por, self.usuario)

    def test_cierre_periodo_bloquea_si_hay_pendientes_contables(self):
        periodo = PeriodoContable.objects.create(
            empresa=self.empresa,
            anio=2026,
            mes=4,
            estado="abierto",
        )
        AsientoContable.objects.create(
            empresa=self.empresa,
            fecha=date(2026, 4, 15),
            descripcion="Asiento pendiente de revision",
            estado="borrador",
        )
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110219",
            nombre="Banco Pendiente",
            tipo="activo",
        )
        cuenta_financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Pendiente",
            tipo="banco",
            cuenta_contable=cuenta_banco,
        )
        MovimientoBancario.objects.create(
            empresa=self.empresa,
            cuenta_financiera=cuenta_financiera,
            fecha=date(2026, 4, 20),
            descripcion="Movimiento no contabilizado",
            debito=Decimal("25.00"),
        )

        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("cerrar_periodo_contable", args=[self.empresa.slug, periodo.id]),
            follow=True,
        )

        periodo.refresh_from_db()
        self.assertEqual(periodo.estado, "abierto")
        self.assertContains(response, "No se puede cerrar")
        self.assertContains(response, "1 asientos borrador")
        self.assertContains(response, "1 movimientos bancarios")

    def test_checklist_cierre_periodo_responde_con_controles(self):
        periodo = PeriodoContable.objects.create(
            empresa=self.empresa,
            anio=2026,
            mes=4,
            estado="abierto",
        )
        self.client.login(username="contador", password="pass12345")
        response = self.client.get(reverse("checklist_cierre_periodo", args=[self.empresa.slug, periodo.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Checklist de Cierre")
        self.assertContains(response, "Asientos Borrador")
        self.assertContains(response, "Sin Conciliar")

    def test_cerrar_periodo_contable_requiere_post(self):
        periodo = PeriodoContable.objects.create(
            empresa=self.empresa,
            anio=2026,
            mes=4,
            estado="abierto",
        )
        self.client.login(username="contador", password="pass12345")
        response = self.client.get(reverse("cerrar_periodo_contable", args=[self.empresa.slug, periodo.id]))

        self.assertEqual(response.status_code, 405)
        periodo.refresh_from_db()
        self.assertEqual(periodo.estado, "abierto")

    def test_cierre_periodo_bloquea_movimientos_sin_conciliar(self):
        periodo = PeriodoContable.objects.create(
            empresa=self.empresa,
            anio=2026,
            mes=4,
            estado="abierto",
        )
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110220",
            nombre="Banco sin conciliar",
            tipo="activo",
        )
        cuenta_financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco sin conciliar",
            tipo="banco",
            cuenta_contable=cuenta_banco,
        )
        asiento = AsientoContable.objects.create(
            empresa=self.empresa,
            fecha=date(2026, 4, 10),
            descripcion="Cobro contabilizado",
            estado="contabilizado",
        )
        MovimientoBancario.objects.create(
            empresa=self.empresa,
            cuenta_financiera=cuenta_financiera,
            fecha=date(2026, 4, 10),
            descripcion="Deposito contabilizado sin conciliar",
            credito=Decimal("75.00"),
            estado="contabilizado",
            asiento=asiento,
            conciliado=False,
        )

        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("cerrar_periodo_contable", args=[self.empresa.slug, periodo.id]),
            follow=True,
        )

        periodo.refresh_from_db()
        self.assertEqual(periodo.estado, "abierto")
        self.assertContains(response, "1 movimientos sin conciliar")

    def test_duplicar_y_reversar_asiento_contable(self):
        cuenta_caja = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110221",
            nombre="Caja Reversion",
            tipo="activo",
        )
        cuenta_ingreso = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="410221",
            nombre="Ingreso Reversion",
            tipo="ingreso",
        )
        asiento = AsientoContable.objects.create(
            empresa=self.empresa,
            fecha=date(2026, 4, 10),
            descripcion="Asiento original",
            referencia="REV-01",
            estado="contabilizado",
        )
        LineaAsientoContable.objects.create(asiento=asiento, cuenta=cuenta_caja, debe=Decimal("100.00"))
        LineaAsientoContable.objects.create(asiento=asiento, cuenta=cuenta_ingreso, haber=Decimal("100.00"))

        self.client.login(username="contador", password="pass12345")
        duplicado = self.client.post(reverse("duplicar_asiento_contable", args=[self.empresa.slug, asiento.id]))
        self.assertEqual(duplicado.status_code, 302)
        copia = AsientoContable.objects.get(documento_tipo="asiento_duplicado", documento_id=asiento.id)
        self.assertEqual(copia.estado, "borrador")
        self.assertEqual(copia.lineas.count(), 2)

        reversion = self.client.post(reverse("reversar_asiento_contable", args=[self.empresa.slug, asiento.id]))
        self.assertEqual(reversion.status_code, 302)
        asiento_reversion = AsientoContable.objects.get(documento_tipo="asiento_contable", documento_id=asiento.id, evento="reversion")
        self.assertEqual(asiento_reversion.estado, "contabilizado")
        self.assertTrue(asiento_reversion.lineas.filter(cuenta=cuenta_caja, haber=Decimal("100.00")).exists())
        self.assertTrue(asiento_reversion.lineas.filter(cuenta=cuenta_ingreso, debe=Decimal("100.00")).exists())

    def test_periodo_cerrado_bloquea_asiento(self):
        PeriodoContable.objects.create(
            empresa=self.empresa,
            anio=2026,
            mes=4,
            estado="cerrado",
            cerrado_por=self.usuario,
        )
        asiento = AsientoContable(
            empresa=self.empresa,
            numero="ASI-TEST",
            fecha=date(2026, 4, 10),
            descripcion="Asiento bloqueado",
            estado="borrador",
        )
        with self.assertRaises(ValidationError):
            asiento.full_clean()

    def test_crear_clasificacion_compra_fiscal(self):
        cuenta_gasto = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="6101",
            nombre="Combustible",
            tipo="gasto",
        )
        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("crear_clasificacion_compra_fiscal", args=[self.empresa.slug]),
            {
                "nombre": "Combustible",
                "cuenta_contable": cuenta_gasto.id,
                "descripcion": "Gastos de combustible",
                "activa": "on",
            },
        )
        self.assertRedirects(response, reverse("clasificaciones_compras_fiscales", args=[self.empresa.slug]))
        self.assertTrue(
            ClasificacionCompraFiscal.objects.filter(
                empresa=self.empresa,
                nombre="Combustible",
                cuenta_contable=cuenta_gasto,
            ).exists()
        )

    def test_bancos_importa_clasifica_y_contabiliza_movimiento(self):
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1102",
            nombre="Banco Principal",
            tipo="activo",
        )
        cuenta_gasto = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="6102",
            nombre="Servicios Publicos",
            tipo="gasto",
        )
        financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Demo",
            tipo="banco",
            institucion="Banco Demo",
            cuenta_contable=cuenta_banco,
        )
        clasificacion = ClasificacionMovimientoBanco.objects.create(
            empresa=self.empresa,
            nombre="Servicios Publicos",
            tipo="egreso",
            cuenta_contable=cuenta_gasto,
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Fecha", "Descripcion", "Referencia", "Debito", "Credito", "Saldo"])
        sheet.append([date(2026, 4, 10), "Pago ENEE", "ENE-01", 3500, 0, 10000])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        archivo = SimpleUploadedFile(
            "estado.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("importar_movimientos_bancarios", args=[self.empresa.slug]),
            {"cuenta_financiera": financiera.id, "archivo": archivo},
        )
        self.assertRedirects(response, reverse("bancos_dashboard", args=[self.empresa.slug]))
        movimiento = MovimientoBancario.objects.get(empresa=self.empresa, descripcion="Pago ENEE")
        self.assertEqual(movimiento.debito, Decimal("3500.00"))

        response = self.client.post(
            reverse("clasificar_movimiento_bancario", args=[self.empresa.slug, movimiento.id]),
            {"clasificacion": clasificacion.id},
        )
        self.assertRedirects(response, reverse("movimientos_bancarios", args=[self.empresa.slug]))
        movimiento.refresh_from_db()
        self.assertEqual(movimiento.estado, "clasificado")

        response = self.client.post(reverse("contabilizar_movimiento_bancario", args=[self.empresa.slug, movimiento.id]))
        self.assertRedirects(response, reverse("movimientos_bancarios", args=[self.empresa.slug]))
        movimiento.refresh_from_db()
        self.assertEqual(movimiento.estado, "contabilizado")
        self.assertTrue(movimiento.asiento.lineas.filter(cuenta=cuenta_gasto, debe=Decimal("3500.00")).exists())
        self.assertTrue(movimiento.asiento.lineas.filter(cuenta=cuenta_banco, haber=Decimal("3500.00")).exists())

        response = self.client.get(reverse("conciliacion_bancaria", args=[self.empresa.slug]))
        self.assertEqual(response.status_code, 200)

        response = self.client.post(reverse("conciliar_movimiento_bancario", args=[self.empresa.slug, movimiento.id]))
        self.assertRedirects(response, reverse("conciliacion_bancaria", args=[self.empresa.slug]))
        movimiento.refresh_from_db()
        self.assertTrue(movimiento.conciliado)
        self.assertEqual(movimiento.conciliado_por, self.usuario)

    def test_contabilizar_movimiento_bancario_requiere_post(self):
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110299",
            nombre="Banco GET bloqueado",
            tipo="activo",
        )
        cuenta_gasto = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="610299",
            nombre="Gasto GET bloqueado",
            tipo="gasto",
        )
        financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco GET bloqueado",
            tipo="banco",
            cuenta_contable=cuenta_banco,
        )
        clasificacion = ClasificacionMovimientoBanco.objects.create(
            empresa=self.empresa,
            nombre="Clasificacion GET bloqueada",
            tipo="egreso",
            cuenta_contable=cuenta_gasto,
        )
        movimiento = MovimientoBancario.objects.create(
            empresa=self.empresa,
            cuenta_financiera=financiera,
            fecha=date(2026, 4, 10),
            descripcion="Movimiento GET bloqueado",
            debito=Decimal("50.00"),
            clasificacion=clasificacion,
            estado="clasificado",
        )

        self.client.login(username="contador", password="pass12345")
        response = self.client.get(reverse("contabilizar_movimiento_bancario", args=[self.empresa.slug, movimiento.id]))

        self.assertEqual(response.status_code, 405)
        movimiento.refresh_from_db()
        self.assertEqual(movimiento.estado, "clasificado")
        self.assertIsNone(movimiento.asiento_id)

    def test_importar_estado_cuenta_lafise_con_encabezado_y_fechas_cortas(self):
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110209",
            nombre="Banco LAFISE HNL",
            tipo="activo",
        )
        financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco LAFISE HNL",
            tipo="banco",
            institucion="Banco LAFISE",
            cuenta_contable=cuenta_banco,
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["BANCO LAFISE"])
        sheet.append(["ESTADO DE CUENTA", None, "SEP/30/25"])
        sheet.append(["NUMERO DE CUENTA", "201503001127", "MONEDA", "LEMPIRAS"])
        sheet.append([])
        sheet.append(["FECHA", "DESCRIPCION", "DEBITOS", "CREDITOS", "SALDOS"])
        sheet.append(["09/01/25", "PAGO PROMERICA CUOTA 60/60", "42,349.57", "", "4,447,999.63"])
        sheet.append(["09/01/25", "DB TSP ACH 42 349.57", "86.00", "", "4,447,913.63"])
        sheet.append(["09/02/25", "DEPOSITO CLIENTE DEMO", "", "10,000.00", "4,457,913.63"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        archivo = SimpleUploadedFile(
            "lafise_sep_2025.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("importar_movimientos_bancarios", args=[self.empresa.slug]),
            {"cuenta_financiera": financiera.id, "archivo": archivo},
        )

        self.assertRedirects(response, reverse("bancos_dashboard", args=[self.empresa.slug]))
        pago = MovimientoBancario.objects.get(empresa=self.empresa, descripcion="PAGO PROMERICA CUOTA 60/60")
        deposito = MovimientoBancario.objects.get(empresa=self.empresa, descripcion="DEPOSITO CLIENTE DEMO")
        self.assertEqual(pago.fecha, date(2025, 9, 1))
        self.assertEqual(pago.debito, Decimal("42349.57"))
        self.assertEqual(pago.saldo, Decimal("4447999.63"))
        self.assertEqual(deposito.fecha, date(2025, 9, 2))
        self.assertEqual(deposito.credito, Decimal("10000.00"))

    def test_importar_estado_cuenta_rechaza_formato_no_soportado_con_mensaje_amigable(self):
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110211",
            nombre="Banco TXT",
            tipo="activo",
        )
        financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco TXT",
            tipo="banco",
            cuenta_contable=cuenta_banco,
        )
        archivo = SimpleUploadedFile("estado.txt", b"contenido demo", content_type="text/plain")

        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("importar_movimientos_bancarios", args=[self.empresa.slug]),
            {"cuenta_financiera": financiera.id, "archivo": archivo},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "formato Excel .xlsx/.xlsm o PDF .pdf")
        self.assertFalse(MovimientoBancario.objects.filter(empresa=self.empresa, cuenta_financiera=financiera).exists())

    def test_extraer_movimientos_pdf_lafise_desde_texto(self):
        texto = """
        ESTADO DE CUENTA
        FECHA ESTADO DE CUENTA SEP/30/25
        FECHA DESCRIPCION DEBITOS CREDITOS SALDOS
        09/01/25 PAGO PROMERICA CUOTA 60/60 42,349.57 4,447,999.63
        09/01/25 DB COMISION ACH 42349.57 40.00 4,447,873.63
        09/02/25 DEPOSITO CLIENTE DEMO 10,000.00 4,457,913.63
        """

        movimientos, omitidos = _extraer_movimientos_desde_texto_pdf(texto, mes_preferido=9)

        self.assertEqual(omitidos, [])
        self.assertEqual(len(movimientos), 3)
        self.assertEqual(movimientos[0]["fecha"], date(2025, 9, 1))
        self.assertEqual(movimientos[0]["debito"], Decimal("42349.57"))
        self.assertEqual(movimientos[0]["credito"], Decimal("0.00"))
        self.assertEqual(movimientos[1]["debito"], Decimal("40.00"))
        self.assertEqual(movimientos[2]["credito"], Decimal("10000.00"))

    def test_importar_estado_cuenta_rechaza_xlsx_invalido_con_mensaje_amigable(self):
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110212",
            nombre="Banco Archivo Invalido",
            tipo="activo",
        )
        financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Archivo Invalido",
            tipo="banco",
            cuenta_contable=cuenta_banco,
        )
        archivo = SimpleUploadedFile(
            "estado.xlsx",
            b"este archivo no es un excel real",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("importar_movimientos_bancarios", args=[self.empresa.slug]),
            {"cuenta_financiera": financiera.id, "archivo": archivo},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "archivo Excel valido")
        self.assertFalse(MovimientoBancario.objects.filter(empresa=self.empresa, cuenta_financiera=financiera).exists())

    def test_clasificacion_bancaria_en_lote(self):
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110210",
            nombre="Banco Principal Lote",
            tipo="activo",
        )
        cuenta_gasto = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="610210",
            nombre="Comisiones Bancarias",
            tipo="gasto",
        )
        cuenta_ingreso = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="410210",
            nombre="Ingresos por Depositos",
            tipo="ingreso",
        )
        financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Lote",
            tipo="banco",
            institucion="Banco Demo",
            cuenta_contable=cuenta_banco,
        )
        clasificacion_gasto = ClasificacionMovimientoBanco.objects.create(
            empresa=self.empresa,
            nombre="Comisiones Bancarias",
            tipo="egreso",
            cuenta_contable=cuenta_gasto,
        )
        clasificacion_ingreso = ClasificacionMovimientoBanco.objects.create(
            empresa=self.empresa,
            nombre="Depositos Clientes",
            tipo="ingreso",
            cuenta_contable=cuenta_ingreso,
        )
        movimiento_egreso = MovimientoBancario.objects.create(
            empresa=self.empresa,
            cuenta_financiera=financiera,
            fecha=date(2026, 4, 12),
            descripcion="DB COMISION ACH",
            debito=Decimal("40.00"),
            saldo=Decimal("1000.00"),
        )
        movimiento_ingreso = MovimientoBancario.objects.create(
            empresa=self.empresa,
            cuenta_financiera=financiera,
            fecha=date(2026, 4, 12),
            descripcion="DEPOSITO CLIENTE",
            credito=Decimal("500.00"),
            saldo=Decimal("1500.00"),
        )

        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("clasificar_movimientos_bancarios_lote", args=[self.empresa.slug]),
            {
                "movimiento_ids": [str(movimiento_egreso.id), str(movimiento_ingreso.id)],
                f"clasificacion_{movimiento_egreso.id}": str(clasificacion_gasto.id),
                f"clasificacion_{movimiento_ingreso.id}": str(clasificacion_ingreso.id),
                "next": reverse("movimientos_bancarios", args=[self.empresa.slug]) + "?estado=pendiente",
            },
        )

        self.assertRedirects(response, reverse("movimientos_bancarios", args=[self.empresa.slug]) + "?estado=pendiente")
        movimiento_egreso.refresh_from_db()
        movimiento_ingreso.refresh_from_db()
        self.assertEqual(movimiento_egreso.estado, "clasificado")
        self.assertEqual(movimiento_egreso.clasificacion, clasificacion_gasto)
        self.assertEqual(movimiento_ingreso.estado, "clasificado")
        self.assertEqual(movimiento_ingreso.clasificacion, clasificacion_ingreso)

    def test_movimientos_bancarios_prepara_clasificaciones_base(self):
        self.client.login(username="contador", password="pass12345")
        self.client.post(reverse("cargar_catalogo_base_honduras", args=[self.empresa.slug]))
        financiera = CuentaFinanciera.objects.get(empresa=self.empresa, nombre="Banco Principal HNL")
        MovimientoBancario.objects.create(
            empresa=self.empresa,
            cuenta_financiera=financiera,
            fecha=date(2026, 4, 12),
            descripcion="COMISION ACH",
            debito=Decimal("40.00"),
            saldo=Decimal("1000.00"),
            origen_importacion="estado.pdf",
        )

        response = self.client.get(reverse("movimientos_bancarios", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Comisiones Bancarias")
        self.assertContains(response, "Cobro de Clientes")

    def test_regla_bancaria_clasifica_movimiento_pendiente(self):
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110215",
            nombre="Banco Reglas",
            tipo="activo",
        )
        cuenta_gasto = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="610215",
            nombre="Comisiones Reglas",
            tipo="gasto",
        )
        financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Reglas",
            tipo="banco",
            cuenta_contable=cuenta_banco,
        )
        clasificacion = ClasificacionMovimientoBanco.objects.create(
            empresa=self.empresa,
            nombre="Comisiones Reglas",
            tipo="egreso",
            cuenta_contable=cuenta_gasto,
        )
        ReglaClasificacionBanco.objects.create(
            empresa=self.empresa,
            nombre="Detectar comisiones",
            texto_busqueda="COMISION",
            tipo_movimiento="egreso",
            clasificacion=clasificacion,
            prioridad=10,
        )
        movimiento = MovimientoBancario.objects.create(
            empresa=self.empresa,
            cuenta_financiera=financiera,
            fecha=date(2026, 4, 12),
            descripcion="DB COMISION ACH 40.00",
            debito=Decimal("40.00"),
            saldo=Decimal("1000.00"),
        )

        self.client.login(username="contador", password="pass12345")
        response = self.client.get(reverse("aplicar_reglas_bancarias", args=[self.empresa.slug]))

        self.assertRedirects(response, reverse("movimientos_bancarios", args=[self.empresa.slug]))
        movimiento.refresh_from_db()
        self.assertEqual(movimiento.clasificacion, clasificacion)
        self.assertEqual(movimiento.estado, "clasificado")

    def test_crear_regla_bancaria_desde_interfaz(self):
        cuenta_gasto = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="610216",
            nombre="Telefonia Reglas",
            tipo="gasto",
        )
        clasificacion = ClasificacionMovimientoBanco.objects.create(
            empresa=self.empresa,
            nombre="Telefonia Reglas",
            tipo="egreso",
            cuenta_contable=cuenta_gasto,
        )

        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("crear_regla_bancaria", args=[self.empresa.slug]),
            {
                "nombre": "Detectar Tigo",
                "texto_busqueda": "TIGO",
                "tipo_movimiento": "egreso",
                "clasificacion": str(clasificacion.id),
                "prioridad": "20",
                "activa": "on",
            },
        )

        self.assertRedirects(response, reverse("reglas_bancarias", args=[self.empresa.slug]))
        self.assertTrue(ReglaClasificacionBanco.objects.filter(empresa=self.empresa, nombre="Detectar Tigo").exists())

    def test_puede_editar_movimiento_bancario_importado_antes_de_contabilizar(self):
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110213",
            nombre="Banco Edicion",
            tipo="activo",
        )
        cuenta_gasto = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="610213",
            nombre="Gasto Bancario Edicion",
            tipo="gasto",
        )
        financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Edicion",
            tipo="banco",
            cuenta_contable=cuenta_banco,
        )
        clasificacion = ClasificacionMovimientoBanco.objects.create(
            empresa=self.empresa,
            nombre="Comision Bancaria Edicion",
            tipo="egreso",
            cuenta_contable=cuenta_gasto,
        )
        movimiento = MovimientoBancario.objects.create(
            empresa=self.empresa,
            cuenta_financiera=financiera,
            fecha=date(2026, 4, 10),
            descripcion="COMISION MAL LEIDA",
            debito=Decimal("400.00"),
            saldo=Decimal("1000.00"),
            origen_importacion="estado.pdf",
        )

        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("editar_movimiento_bancario", args=[self.empresa.slug, movimiento.id]),
            {
                "fecha": "2026-04-11",
                "descripcion": "COMISION CORREGIDA",
                "referencia": "REF-01",
                "debito": "40.00",
                "credito": "0.00",
                "saldo": "960.00",
                "clasificacion": str(clasificacion.id),
            },
        )

        self.assertRedirects(response, reverse("movimientos_bancarios", args=[self.empresa.slug]))
        movimiento.refresh_from_db()
        self.assertEqual(movimiento.fecha, date(2026, 4, 11))
        self.assertEqual(movimiento.descripcion, "COMISION CORREGIDA")
        self.assertEqual(movimiento.debito, Decimal("40.00"))
        self.assertEqual(movimiento.saldo, Decimal("960.00"))
        self.assertEqual(movimiento.clasificacion, clasificacion)
        self.assertEqual(movimiento.estado, "clasificado")

    def test_no_permite_editar_movimiento_bancario_contabilizado(self):
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110214",
            nombre="Banco Bloqueado",
            tipo="activo",
        )
        financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Bloqueado",
            tipo="banco",
            cuenta_contable=cuenta_banco,
        )
        asiento = AsientoContable.objects.create(
            empresa=self.empresa,
            numero="ASI-BLOQ",
            fecha=date(2026, 4, 10),
            descripcion="Movimiento ya contabilizado",
            estado="contabilizado",
        )
        movimiento = MovimientoBancario.objects.create(
            empresa=self.empresa,
            cuenta_financiera=financiera,
            fecha=date(2026, 4, 10),
            descripcion="NO EDITAR",
            debito=Decimal("100.00"),
            saldo=Decimal("900.00"),
            estado="contabilizado",
            asiento=asiento,
        )

        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("editar_movimiento_bancario", args=[self.empresa.slug, movimiento.id]),
            {
                "fecha": "2026-04-11",
                "descripcion": "EDITADO",
                "referencia": "",
                "debito": "50.00",
                "credito": "0.00",
                "saldo": "950.00",
                "clasificacion": "",
            },
        )

        self.assertRedirects(response, reverse("movimientos_bancarios", args=[self.empresa.slug]))
        movimiento.refresh_from_db()
        self.assertEqual(movimiento.descripcion, "NO EDITAR")
        self.assertEqual(movimiento.debito, Decimal("100.00"))

    def test_puede_crear_cuenta_contable(self):
        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("crear_cuenta_contable", args=[self.empresa.slug]),
            {
                "codigo": "1101",
                "nombre": "Caja General",
                "tipo": "activo",
                "descripcion": "Cuenta de efectivo",
                "acepta_movimientos": "on",
                "activa": "on",
            },
        )
        self.assertRedirects(response, reverse("catalogo_cuentas", args=[self.empresa.slug]))
        self.assertTrue(CuentaContable.objects.filter(empresa=self.empresa, codigo="1101").exists())

    def test_puede_cargar_catalogo_base_honduras_sin_duplicar_existentes(self):
        CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1101",
            nombre="Caja Existente",
            tipo="activo",
            acepta_movimientos=True,
        )

        self.client.login(username="contador", password="pass12345")
        response = self.client.post(reverse("cargar_catalogo_base_honduras", args=[self.empresa.slug]))

        self.assertRedirects(response, reverse("catalogo_cuentas", args=[self.empresa.slug]))
        caja = CuentaContable.objects.get(empresa=self.empresa, codigo="1101")
        cuenta_isv_pagado = CuentaContable.objects.get(empresa=self.empresa, codigo="113001")
        cuenta_activo = CuentaContable.objects.get(empresa=self.empresa, codigo="1")
        self.assertEqual(caja.nombre, "Caja Existente")
        self.assertEqual(caja.cuenta_padre.codigo, "11")
        self.assertEqual(cuenta_isv_pagado.cuenta_padre.codigo, "1130")
        self.assertFalse(cuenta_activo.acepta_movimientos)
        self.assertTrue(
            CuentaFinanciera.objects.filter(
                empresa=self.empresa,
                nombre="Caja General",
                tipo="caja",
                cuenta_contable=caja,
            ).exists()
        )
        self.assertTrue(
            CuentaFinanciera.objects.filter(
                empresa=self.empresa,
                nombre="Banco Principal HNL",
                tipo="banco",
                cuenta_contable__codigo="110201",
            ).exists()
        )
        self.assertTrue(
            ClasificacionMovimientoBanco.objects.filter(
                empresa=self.empresa,
                nombre="Comisiones Bancarias",
                tipo="egreso",
                cuenta_contable__codigo="610201",
            ).exists()
        )
        self.assertTrue(
            ClasificacionMovimientoBanco.objects.filter(
                empresa=self.empresa,
                nombre="Cobro de Clientes",
                tipo="ingreso",
                cuenta_contable__codigo="1110",
            ).exists()
        )

        total_cuentas = CuentaContable.objects.filter(empresa=self.empresa).count()
        total_cuentas_financieras = CuentaFinanciera.objects.filter(empresa=self.empresa).count()
        total_clasificaciones = ClasificacionMovimientoBanco.objects.filter(empresa=self.empresa).count()
        response = self.client.post(reverse("cargar_catalogo_base_honduras", args=[self.empresa.slug]))
        self.assertRedirects(response, reverse("catalogo_cuentas", args=[self.empresa.slug]))
        self.assertEqual(CuentaContable.objects.filter(empresa=self.empresa).count(), total_cuentas)
        self.assertEqual(CuentaFinanciera.objects.filter(empresa=self.empresa).count(), total_cuentas_financieras)
        self.assertEqual(ClasificacionMovimientoBanco.objects.filter(empresa=self.empresa).count(), total_clasificaciones)

    def test_puede_crear_subcuenta_contable(self):
        cuenta_padre = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1100",
            nombre="Activo Corriente",
            tipo="activo",
            acepta_movimientos=False,
        )
        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("crear_cuenta_contable", args=[self.empresa.slug]),
            {
                "codigo": "1101",
                "nombre": "Caja General",
                "tipo": "activo",
                "cuenta_padre": str(cuenta_padre.id),
                "descripcion": "Subcuenta de efectivo",
                "acepta_movimientos": "on",
                "activa": "on",
            },
        )
        self.assertRedirects(response, reverse("catalogo_cuentas", args=[self.empresa.slug]))
        cuenta = CuentaContable.objects.get(empresa=self.empresa, codigo="1101")
        self.assertEqual(cuenta.cuenta_padre, cuenta_padre)
        self.assertEqual(cuenta.nivel, 1)

    def test_cuenta_contable_no_permite_ciclo(self):
        cuenta_padre = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1100",
            nombre="Activo Corriente",
            tipo="activo",
            acepta_movimientos=False,
        )
        subcuenta = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1101",
            nombre="Caja General",
            tipo="activo",
            cuenta_padre=cuenta_padre,
        )
        cuenta_padre.cuenta_padre = subcuenta
        with self.assertRaises(ValidationError):
            cuenta_padre.full_clean()

    def test_importar_catalogo_cuentas_desde_excel(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["codigo", "nombre", "tipo", "codigo_padre", "acepta_movimientos", "activa", "descripcion"])
        sheet.append(["1100", "Activo Corriente", "activo", "", "no", "si", "Cuenta agrupadora"])
        sheet.append(["1101", "Caja General", "activo", "1100", "si", "si", "Efectivo disponible"])
        sheet.append(["5100", "Costo de Ventas", "costo", "", "si", "si", "Costo operativo"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        archivo = SimpleUploadedFile(
            "catalogo.xlsx",
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("importar_catalogo_cuentas", args=[self.empresa.slug]),
            {"actualizar_existentes": "on", "archivo": archivo},
        )

        self.assertRedirects(response, reverse("catalogo_cuentas", args=[self.empresa.slug]))
        cuenta_padre = CuentaContable.objects.get(empresa=self.empresa, codigo="1100")
        subcuenta = CuentaContable.objects.get(empresa=self.empresa, codigo="1101")
        self.assertFalse(cuenta_padre.acepta_movimientos)
        self.assertEqual(subcuenta.cuenta_padre, cuenta_padre)
        self.assertTrue(CuentaContable.objects.filter(empresa=self.empresa, codigo="5100", tipo="costo").exists())

    def test_importar_catalogo_jerarquico_tipo_brisas_del_mar(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "CCORIGINAL"
        sheet.append([None, None, None, None, None, None, None, None, "BRISAS DEL MAR MANAGMENT S.A."])
        sheet.append([None])
        sheet.append(["Titulo"])
        sheet.append([1, 2, 3, 4, 5, 6, 7])
        sheet.append([1, None, None, None, None, None, None, "1", "ACTIVOS"])
        sheet.append([1, 1, None, None, None, None, None, "11", "ACTIVOS CORRIENTES"])
        sheet.append([1, 1, "01", None, None, None, None, "1101", "EFECTIVO Y EQUIVALENTES"])
        sheet.append([1, 1, "01", "01", None, None, None, "110101", "CAJA GENERAL"])
        sheet.append([4, None, None, None, None, None, None, "4", "INGRESOS"])
        sheet.append([4, 1, None, None, None, None, None, "41", "INGRESOS OPERATIVOS"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        archivo = SimpleUploadedFile(
            "catalogo_brisas.xlsx",
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("importar_catalogo_cuentas", args=[self.empresa.slug]),
            {"actualizar_existentes": "on", "archivo": archivo},
        )

        self.assertRedirects(response, reverse("catalogo_cuentas", args=[self.empresa.slug]))
        activos = CuentaContable.objects.get(empresa=self.empresa, codigo="1")
        caja = CuentaContable.objects.get(empresa=self.empresa, codigo="110101")
        ingresos = CuentaContable.objects.get(empresa=self.empresa, codigo="4")
        self.assertEqual(activos.tipo, "activo")
        self.assertFalse(activos.acepta_movimientos)
        self.assertEqual(caja.cuenta_padre.codigo, "1101")
        self.assertEqual(ingresos.tipo, "ingreso")

    def test_importar_catalogo_jerarquico_rechaza_codigos_duplicados(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([1, None, None, None, None, None, None, "1", "ACTIVOS"])
        sheet.append([1, 1, None, None, None, None, None, "11", "ACTIVOS CORRIENTES"])
        sheet.append([1, 1, "01", None, None, None, None, "1101", "CAJA"])
        sheet.append([1, 1, "02", None, None, None, None, "1101", "BANCOS"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        archivo = SimpleUploadedFile(
            "catalogo_duplicado.xlsx",
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("importar_catalogo_cuentas", args=[self.empresa.slug]),
            {"actualizar_existentes": "on", "archivo": archivo},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "codigo duplicado 1101")
        self.assertFalse(CuentaContable.objects.filter(empresa=self.empresa, codigo="1101").exists())

    def test_puede_configurar_cuentas_contables_por_empresa(self):
        cuenta_clientes = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1135",
            nombre="Clientes Especial",
            tipo="activo",
        )
        cuenta_ventas = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="4105",
            nombre="Ventas Especial",
            tipo="ingreso",
        )
        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("configuracion_contable", args=[self.empresa.slug]),
            {
                "cuenta_clientes": str(cuenta_clientes.id),
                "cuenta_ventas": str(cuenta_ventas.id),
            },
        )
        self.assertRedirects(response, reverse("contabilidad_dashboard", args=[self.empresa.slug]))
        configuracion = ConfiguracionContableEmpresa.objects.get(empresa=self.empresa)
        self.assertEqual(configuracion.cuenta_clientes, cuenta_clientes)
        self.assertEqual(configuracion.cuenta_ventas, cuenta_ventas)

    def test_puede_crear_y_contabilizar_asiento(self):
        cuenta_debe = CuentaContable.objects.create(empresa=self.empresa, codigo="1101", nombre="Caja", tipo="activo")
        cuenta_haber = CuentaContable.objects.create(empresa=self.empresa, codigo="4101", nombre="Ventas", tipo="ingreso")
        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("crear_asiento_contable", args=[self.empresa.slug]),
            {
                "fecha": "2026-04-10",
                "descripcion": "Registro inicial",
                "referencia": "DOC-001",
                "origen_modulo": "manual",
                "estado": "borrador",
                "contabilizar_ahora": "on",
                "lineas-TOTAL_FORMS": "2",
                "lineas-INITIAL_FORMS": "0",
                "lineas-MIN_NUM_FORMS": "0",
                "lineas-MAX_NUM_FORMS": "1000",
                "lineas-0-cuenta": str(cuenta_debe.id),
                "lineas-0-detalle": "Debe",
                "lineas-0-debe": "100.00",
                "lineas-0-haber": "0.00",
                "lineas-1-cuenta": str(cuenta_haber.id),
                "lineas-1-detalle": "Haber",
                "lineas-1-debe": "0.00",
                "lineas-1-haber": "100.00",
            },
        )
        asiento = AsientoContable.objects.get(empresa=self.empresa)
        self.assertRedirects(response, reverse("ver_asiento_contable", args=[self.empresa.slug, asiento.id]))
        asiento.refresh_from_db()
        self.assertEqual(asiento.estado, "contabilizado")
        self.assertTrue(asiento.numero.startswith("ASI-"))

    def test_formulario_asiento_muestra_busqueda_rapida_de_cuentas(self):
        CuentaContable.objects.create(empresa=self.empresa, codigo="1101", nombre="Caja General", tipo="activo")
        self.client.login(username="contador", password="pass12345")
        response = self.client.get(reverse("crear_asiento_contable", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Codigo o nombre de la cuenta")
        self.assertContains(response, "Caja General")

    def test_libro_diario_responde_con_partidas(self):
        cuenta_debe = CuentaContable.objects.create(empresa=self.empresa, codigo="1101", nombre="Caja", tipo="activo")
        cuenta_haber = CuentaContable.objects.create(empresa=self.empresa, codigo="4101", nombre="Ventas", tipo="ingreso")
        asiento = AsientoContable.objects.create(
            empresa=self.empresa,
            numero="ASI-00000001",
            fecha=date(2026, 4, 10),
            descripcion="Venta contabilizada",
            referencia="FAC-001",
            estado="contabilizado",
        )
        LineaAsientoContable.objects.create(asiento=asiento, cuenta=cuenta_debe, debe=Decimal("115.00"))
        LineaAsientoContable.objects.create(asiento=asiento, cuenta=cuenta_haber, haber=Decimal("115.00"))
        self.client.login(username="contador", password="pass12345")
        response = self.client.get(reverse("libro_diario", args=[self.empresa.slug]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Venta contabilizada")
        self.assertContains(response, "L. 115.00")

    def test_mayor_cuenta_responde_con_saldo(self):
        cuenta = CuentaContable.objects.create(empresa=self.empresa, codigo="1101", nombre="Caja", tipo="activo")
        asiento = AsientoContable.objects.create(
            empresa=self.empresa,
            numero="ASI-00000001",
            fecha=date(2026, 4, 10),
            descripcion="Movimiento de caja",
            referencia="MOV-001",
            estado="contabilizado",
        )
        LineaAsientoContable.objects.create(asiento=asiento, cuenta=cuenta, debe=Decimal("200.00"))
        LineaAsientoContable.objects.create(asiento=asiento, cuenta=cuenta, haber=Decimal("50.00"))
        self.client.login(username="contador", password="pass12345")
        response = self.client.get(reverse("mayor_cuenta", args=[self.empresa.slug]), {"cuenta": cuenta.id})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Movimiento de caja")
        self.assertContains(response, "L. 150.00")

    def test_balance_comprobacion_responde_con_totales(self):
        cuenta_debe = CuentaContable.objects.create(empresa=self.empresa, codigo="1101", nombre="Caja", tipo="activo")
        cuenta_haber = CuentaContable.objects.create(empresa=self.empresa, codigo="4101", nombre="Ventas", tipo="ingreso")
        asiento = AsientoContable.objects.create(
            empresa=self.empresa,
            numero="ASI-00000001",
            fecha=date(2026, 4, 10),
            descripcion="Balance inicial",
            referencia="BAL-001",
            estado="contabilizado",
        )
        LineaAsientoContable.objects.create(asiento=asiento, cuenta=cuenta_debe, debe=Decimal("300.00"))
        LineaAsientoContable.objects.create(asiento=asiento, cuenta=cuenta_haber, haber=Decimal("300.00"))
        self.client.login(username="contador", password="pass12345")
        response = self.client.get(reverse("balance_comprobacion", args=[self.empresa.slug]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Caja")
        self.assertContains(response, "Ventas")
        self.assertContains(response, "Cuadrado")
        self.assertContains(response, "L. 300.00")

    def test_estado_resultados_responde_con_utilidad(self):
        cuenta_ingreso = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="4101",
            nombre="Ventas",
            tipo="ingreso",
        )
        cuenta_gasto = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="5101",
            nombre="Gastos Administrativos",
            tipo="gasto",
        )
        cuenta_costo = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="5001",
            nombre="Costo de Ventas",
            tipo="costo",
        )
        asiento = AsientoContable.objects.create(
            empresa=self.empresa,
            numero="ASI-00000001",
            fecha=date(2026, 4, 10),
            descripcion="Resultado del periodo",
            referencia="ER-001",
            estado="contabilizado",
        )
        LineaAsientoContable.objects.create(asiento=asiento, cuenta=cuenta_ingreso, haber=Decimal("500.00"))
        LineaAsientoContable.objects.create(asiento=asiento, cuenta=cuenta_costo, debe=Decimal("125.00"))
        LineaAsientoContable.objects.create(asiento=asiento, cuenta=cuenta_gasto, debe=Decimal("75.00"))
        self.client.login(username="contador", password="pass12345")
        response = self.client.get(reverse("estado_resultados", args=[self.empresa.slug]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ventas")
        self.assertContains(response, "Costo de Ventas")
        self.assertContains(response, "Gastos Administrativos")
        self.assertContains(response, "L. 300.00")

    def test_balance_general_responde_cuadrado(self):
        cuenta_activo = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1101",
            nombre="Caja",
            tipo="activo",
        )
        cuenta_pasivo = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="2101",
            nombre="Prestamo",
            tipo="pasivo",
        )
        cuenta_patrimonio = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="3101",
            nombre="Capital",
            tipo="patrimonio",
        )
        asiento = AsientoContable.objects.create(
            empresa=self.empresa,
            numero="ASI-00000001",
            fecha=date(2026, 4, 10),
            descripcion="Balance general",
            referencia="BG-001",
            estado="contabilizado",
        )
        LineaAsientoContable.objects.create(asiento=asiento, cuenta=cuenta_activo, debe=Decimal("500.00"))
        LineaAsientoContable.objects.create(asiento=asiento, cuenta=cuenta_pasivo, haber=Decimal("300.00"))
        LineaAsientoContable.objects.create(asiento=asiento, cuenta=cuenta_patrimonio, haber=Decimal("200.00"))
        self.client.login(username="contador", password="pass12345")
        response = self.client.get(reverse("balance_general", args=[self.empresa.slug]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Caja")
        self.assertContains(response, "Prestamo")
        self.assertContains(response, "Capital")
        self.assertContains(response, "Cuadrado")

    def test_exportar_reportes_contables_excel(self):
        cuenta_activo = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1101",
            nombre="Caja",
            tipo="activo",
        )
        cuenta_ingreso = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="4101",
            nombre="Ventas",
            tipo="ingreso",
        )
        asiento = AsientoContable.objects.create(
            empresa=self.empresa,
            numero="ASI-00000001",
            fecha=date(2026, 4, 10),
            descripcion="Reporte exportable",
            referencia="EXP-001",
            estado="contabilizado",
        )
        LineaAsientoContable.objects.create(asiento=asiento, cuenta=cuenta_activo, debe=Decimal("100.00"))
        LineaAsientoContable.objects.create(asiento=asiento, cuenta=cuenta_ingreso, haber=Decimal("100.00"))
        self.client.login(username="contador", password="pass12345")
        urls = [
            "exportar_libro_diario_excel",
            "exportar_balance_comprobacion_excel",
            "exportar_estado_resultados_excel",
            "exportar_balance_general_excel",
        ]
        for url_name in urls:
            response = self.client.get(reverse(url_name, args=[self.empresa.slug]))
            self.assertEqual(response.status_code, 200)
            self.assertEqual(
                response["Content-Type"],
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    def test_dashboard_bi_financiero_renderiza_lectura_ejecutiva(self):
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110555",
            nombre="Banco Bi",
            tipo="activo",
        )
        cuenta_financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Atlantida HNL",
            tipo="banco",
            cuenta_contable=cuenta_banco,
        )
        compra = CompraInventario.objects.create(
            empresa=self.empresa,
            proveedor=self.proveedor,
            proveedor_nombre=self.proveedor.nombre,
            fecha_documento=date(2026, 4, 10),
            fecha_vencimiento=date(2026, 4, 20),
            estado="aplicada",
        )
        LineaCompraInventario.objects.create(
            compra=compra,
            producto=self.producto,
            cantidad=5,
            costo_unitario=Decimal("20.00"),
        )
        PagoCompra.objects.create(
            compra=compra,
            fecha=date(2026, 4, 15),
            monto=Decimal("50.00"),
            metodo="transferencia",
            cuenta_financiera=cuenta_financiera,
        )
        MovimientoBancario.objects.create(
            empresa=self.empresa,
            cuenta_financiera=cuenta_financiera,
            fecha=date(2026, 4, 15),
            descripcion="Pago proveedor BI",
            debito=Decimal("50.00"),
            saldo=Decimal("950.00"),
            estado="contabilizado",
            conciliado=False,
        )

        self.client.login(username="contador", password="pass12345")
        response = self.client.get(reverse("dashboard_bi_financiero", args=[self.empresa.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Dashboard BI Financiero")
        self.assertContains(response, "Top proveedores por compra")
        self.assertContains(response, "Pagado vs comprado")
        self.assertContains(response, "Banco Atlantida HNL")
        self.assertContains(response, self.proveedor.nombre)

    def test_auditoria_contable_responde_y_filtra(self):
        cuenta = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1101",
            nombre="Caja",
            tipo="activo",
        )
        asiento = AsientoContable.objects.create(
            empresa=self.empresa,
            numero="ASI-00000001",
            fecha=date(2026, 4, 10),
            descripcion="Auditoria de asiento",
            referencia="AUD-001",
            origen_modulo="facturacion",
            documento_tipo="factura",
            documento_id=15,
            evento="emision",
            estado="contabilizado",
            creado_por=self.usuario,
        )
        LineaAsientoContable.objects.create(asiento=asiento, cuenta=cuenta, debe=Decimal("100.00"))
        self.client.login(username="contador", password="pass12345")
        response = self.client.get(
            reverse("auditoria_contable", args=[self.empresa.slug]),
            {"origen": "facturacion", "estado": "contabilizado"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Auditoria de asiento")
        self.assertContains(response, "factura #15")
        self.assertContains(response, "contador")

    def test_exportar_auditoria_contable_excel(self):
        asiento = AsientoContable.objects.create(
            empresa=self.empresa,
            numero="ASI-00000001",
            fecha=date(2026, 4, 10),
            descripcion="Auditoria exportable",
            referencia="AUD-EXP",
            estado="contabilizado",
            creado_por=self.usuario,
        )
        self.client.login(username="contador", password="pass12345")
        response = self.client.get(reverse("exportar_auditoria_contable_excel", args=[self.empresa.slug]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_reporte_impuestos_muestra_isv_neto(self):
        factura = Factura.objects.create(
            empresa=self.empresa,
            cliente=self.cliente,
            vendedor=self.usuario,
            fecha_emision=date(2026, 4, 10),
            fecha_vencimiento=date(2026, 4, 10),
            moneda="HNL",
            tipo_cambio=1,
            estado="borrador",
        )
        LineaFactura.objects.create(
            factura=factura,
            producto=self.producto,
            cantidad=2,
            precio_unitario=Decimal("100.00"),
            impuesto=self.impuesto,
        )
        factura.calcular_totales()
        factura.estado = "emitida"
        factura.save()

        nota = NotaCredito.objects.create(
            empresa=self.empresa,
            factura_origen=factura,
            cliente=self.cliente,
            vendedor=self.usuario,
            moneda="HNL",
            tipo_cambio=1,
            fecha_emision=date(2026, 4, 10),
            motivo="Ajuste parcial",
            estado="borrador",
        )
        LineaNotaCredito.objects.create(
            nota_credito=nota,
            producto=self.producto,
            cantidad=1,
            precio_unitario=Decimal("100.00"),
            impuesto=self.impuesto,
        )
        nota.calcular_totales()
        nota.estado = "emitida"
        nota.save()

        self.client.login(username="contador", password="pass12345")
        RegistroCompraFiscal.objects.create(
            empresa=self.empresa,
            proveedor_nombre="Proveedor Fiscal",
            numero_factura="001-001-01-00000099",
            fecha_documento=date(2026, 4, 10),
            periodo_anio=2026,
            periodo_mes=4,
            subtotal=Decimal("50.00"),
            base_15=Decimal("50.00"),
            isv_15=Decimal("7.50"),
            total=Decimal("57.50"),
        )
        response = self.client.get(
            reverse("reporte_impuestos", args=[self.empresa.slug]),
            {"fecha_inicio": "2026-04-01", "fecha_fin": "2026-04-30"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Reporte de Impuestos")
        self.assertContains(response, "ISV a pagar")
        self.assertContains(response, "15.00")

    def test_exportar_reporte_impuestos_excel(self):
        self.client.login(username="contador", password="pass12345")
        response = self.client.get(reverse("exportar_reporte_impuestos_excel", args=[self.empresa.slug]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_middleware_bloquea_sin_permiso_contable(self):
        rol = RolSistema.objects.create(nombre="Consulta", codigo="consulta")
        usuario = Usuario.objects.create_user(
            username="sinpermiso",
            password="pass12345",
            empresa=self.empresa,
            rol_sistema=rol,
        )
        self.client.login(username="sinpermiso", password="pass12345")
        response = self.client.get(reverse("contabilidad_dashboard", args=[self.empresa.slug]))
        self.assertRedirects(response, reverse("dashboard", args=[self.empresa.slug]))

    def test_middleware_bloquea_cierre_periodo_sin_permiso_de_contabilizar(self):
        periodo = PeriodoContable.objects.create(
            empresa=self.empresa,
            anio=2026,
            mes=4,
            estado="abierto",
        )
        rol = RolSistema.objects.create(
            nombre="Contabilidad Operativa",
            codigo="conta-operativa",
            puede_contabilidad=True,
        )
        usuario = Usuario.objects.create_user(
            username="contaoperativa",
            password="pass12345",
            empresa=self.empresa,
            rol_sistema=rol,
        )

        self.client.login(username="contaoperativa", password="pass12345")
        response = self.client.post(reverse("cerrar_periodo_contable", args=[self.empresa.slug, periodo.id]))

        self.assertRedirects(response, reverse("dashboard", args=[self.empresa.slug]))
        periodo.refresh_from_db()
        self.assertEqual(periodo.estado, "abierto")

    def test_emision_factura_generates_asiento_automatico(self):
        factura = Factura.objects.create(
            empresa=self.empresa,
            cliente=self.cliente,
            vendedor=self.usuario,
            fecha_emision=self.empresa.fecha_inicio_plan,
            fecha_vencimiento=self.empresa.fecha_inicio_plan,
            moneda="HNL",
            tipo_cambio=1,
            estado="borrador",
        )
        LineaFactura.objects.create(
            factura=factura,
            producto=self.producto,
            cantidad=1,
            precio_unitario=Decimal("100.00"),
            impuesto=self.impuesto,
        )
        factura.calcular_totales()
        factura.estado = "emitida"
        factura.save()
        asiento = registrar_asiento_factura_emitida(factura)
        self.assertEqual(asiento.estado, "contabilizado")
        self.assertEqual(asiento.lineas.count(), 3)

    def test_emision_factura_usa_cuenta_configurada(self):
        cuenta_ventas = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="4199",
            nombre="Ventas Personalizadas",
            tipo="ingreso",
        )
        ConfiguracionContableEmpresa.objects.create(
            empresa=self.empresa,
            cuenta_ventas=cuenta_ventas,
        )
        factura = Factura.objects.create(
            empresa=self.empresa,
            cliente=self.cliente,
            vendedor=self.usuario,
            fecha_emision=self.empresa.fecha_inicio_plan,
            fecha_vencimiento=self.empresa.fecha_inicio_plan,
            moneda="HNL",
            tipo_cambio=1,
            estado="borrador",
        )
        LineaFactura.objects.create(
            factura=factura,
            producto=self.producto,
            cantidad=1,
            precio_unitario=Decimal("100.00"),
            impuesto=self.impuesto,
        )
        factura.calcular_totales()
        factura.estado = "emitida"
        factura.save()
        asiento = registrar_asiento_factura_emitida(factura)
        self.assertTrue(
            asiento.lineas.filter(cuenta=cuenta_ventas, haber=Decimal("100.00")).exists()
        )

    def test_pago_cliente_generates_asiento_automatico(self):
        factura = Factura.objects.create(
            empresa=self.empresa,
            cliente=self.cliente,
            vendedor=self.usuario,
            fecha_emision=self.empresa.fecha_inicio_plan,
            fecha_vencimiento=self.empresa.fecha_inicio_plan,
            moneda="HNL",
            tipo_cambio=1,
            estado="borrador",
        )
        LineaFactura.objects.create(
            factura=factura,
            producto=self.producto,
            cantidad=1,
            precio_unitario=Decimal("100.00"),
            impuesto=self.impuesto,
        )
        factura.calcular_totales()
        factura.estado = "emitida"
        factura.save()
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="1103",
            nombre="Banco Cobros",
            tipo="activo",
        )
        cuenta_financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Cobros",
            tipo="banco",
            cuenta_contable=cuenta_banco,
        )
        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("registrar_pago", args=[self.empresa.slug, factura.id]),
            {
                "monto": "115.00",
                "metodo": "efectivo",
                "referencia": "RCB-01",
                "fecha": "2026-04-10",
                "cuenta_financiera": str(cuenta_financiera.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(AsientoContable.objects.filter(documento_tipo="pago_factura", evento="cobro").exists())

    def test_enlazar_deposito_bancario_a_factura_registra_pago_y_asiento(self):
        factura = Factura.objects.create(
            empresa=self.empresa,
            cliente=self.cliente,
            vendedor=self.usuario,
            fecha_emision=self.empresa.fecha_inicio_plan,
            fecha_vencimiento=self.empresa.fecha_inicio_plan,
            moneda="HNL",
            tipo_cambio=1,
            estado="borrador",
        )
        LineaFactura.objects.create(
            factura=factura,
            producto=self.producto,
            cantidad=1,
            precio_unitario=Decimal("100.00"),
            impuesto=self.impuesto,
        )
        factura.calcular_totales()
        factura.estado = "emitida"
        factura.save()
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110217",
            nombre="Banco Depositos",
            tipo="activo",
        )
        cuenta_financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Depositos",
            tipo="banco",
            cuenta_contable=cuenta_banco,
        )
        movimiento = MovimientoBancario.objects.create(
            empresa=self.empresa,
            cuenta_financiera=cuenta_financiera,
            fecha=self.empresa.fecha_inicio_plan,
            descripcion="DEPOSITO CLIENTE DEMO",
            credito=Decimal("115.00"),
            saldo=Decimal("1000.00"),
            origen_importacion="estado.pdf",
        )

        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("enlazar_movimiento_factura", args=[self.empresa.slug, movimiento.id]),
            {"factura": str(factura.id)},
        )

        self.assertRedirects(response, reverse("movimientos_bancarios", args=[self.empresa.slug]))
        movimiento.refresh_from_db()
        factura.refresh_from_db()
        pago = movimiento.pago_factura
        self.assertEqual(pago.factura, factura)
        self.assertEqual(pago.monto, Decimal("115.00"))
        self.assertEqual(pago.cuenta_financiera, cuenta_financiera)
        self.assertEqual(movimiento.estado, "contabilizado")
        self.assertIsNotNone(movimiento.asiento)
        self.assertEqual(factura.estado_pago, "pagado")
        self.assertTrue(movimiento.asiento.lineas.filter(cuenta=cuenta_banco, debe=Decimal("115.00")).exists())

    def test_enlazar_debito_bancario_a_compra_registra_pago_y_asiento(self):
        compra = CompraInventario.objects.create(
            empresa=self.empresa,
            proveedor=self.proveedor,
            proveedor_nombre=self.proveedor.nombre,
            fecha_documento=self.empresa.fecha_inicio_plan,
            estado="aplicada",
        )
        LineaCompraInventario.objects.create(
            compra=compra,
            producto=self.producto,
            cantidad=2,
            costo_unitario=Decimal("100.00"),
        )
        cuenta_banco = CuentaContable.objects.create(
            empresa=self.empresa,
            codigo="110218",
            nombre="Banco Egresos",
            tipo="activo",
        )
        cuenta_financiera = CuentaFinanciera.objects.create(
            empresa=self.empresa,
            nombre="Banco Egresos",
            tipo="banco",
            cuenta_contable=cuenta_banco,
        )
        movimiento = MovimientoBancario.objects.create(
            empresa=self.empresa,
            cuenta_financiera=cuenta_financiera,
            fecha=self.empresa.fecha_inicio_plan,
            descripcion="TRANSFERENCIA PROVEEDOR DEMO",
            debito=Decimal("200.00"),
            saldo=Decimal("800.00"),
            origen_importacion="estado.pdf",
        )

        self.client.login(username="contador", password="pass12345")
        response = self.client.post(
            reverse("enlazar_movimiento_compra", args=[self.empresa.slug, movimiento.id]),
            {"compra": str(compra.id)},
        )

        self.assertRedirects(response, reverse("movimientos_bancarios", args=[self.empresa.slug]))
        movimiento.refresh_from_db()
        pago = movimiento.pago_compra
        self.assertEqual(pago.compra, compra)
        self.assertEqual(pago.monto, Decimal("200.00"))
        self.assertEqual(pago.cuenta_financiera, cuenta_financiera)
        self.assertEqual(movimiento.estado, "contabilizado")
        self.assertIsNotNone(movimiento.asiento)
        self.assertEqual(compra.saldo_pendiente, Decimal("0.00"))
        self.assertTrue(movimiento.asiento.lineas.filter(cuenta=cuenta_banco, haber=Decimal("200.00")).exists())

    def test_aplicar_compra_generates_asiento_automatico(self):
        compra = CompraInventario.objects.create(
            empresa=self.empresa,
            proveedor=self.proveedor,
            proveedor_nombre=self.proveedor.nombre,
            fecha_documento=self.empresa.fecha_inicio_plan,
            estado="borrador",
        )
        LineaCompraInventario.objects.create(
            compra=compra,
            producto=self.producto,
            cantidad=2,
            costo_unitario=Decimal("50.00"),
        )
        self.client.login(username="contador", password="pass12345")
        response = self.client.post(reverse("aplicar_compra", args=[self.empresa.slug, compra.id]))
        self.assertEqual(response.status_code, 302)
        self.assertTrue(AsientoContable.objects.filter(documento_tipo="compra", documento_id=compra.id, evento="aplicacion").exists())

    def test_nota_credito_generates_asiento_contable(self):
        factura = Factura.objects.create(
            empresa=self.empresa,
            cliente=self.cliente,
            vendedor=self.usuario,
            fecha_emision=self.empresa.fecha_inicio_plan,
            fecha_vencimiento=self.empresa.fecha_inicio_plan,
            moneda="HNL",
            tipo_cambio=1,
            estado="borrador",
        )
        LineaFactura.objects.create(
            factura=factura,
            producto=self.producto,
            cantidad=1,
            precio_unitario=Decimal("100.00"),
            impuesto=self.impuesto,
        )
        factura.calcular_totales()
        factura.estado = "emitida"
        factura.save()
        nota = NotaCredito.objects.create(
            empresa=self.empresa,
            factura_origen=factura,
            cliente=self.cliente,
            vendedor=self.usuario,
            moneda="HNL",
            tipo_cambio=1,
            fecha_emision=self.empresa.fecha_inicio_plan,
            motivo="Devolucion total",
            estado="borrador",
        )
        from facturacion.models import LineaNotaCredito
        LineaNotaCredito.objects.create(
            nota_credito=nota,
            producto=self.producto,
            cantidad=1,
            precio_unitario=Decimal("100.00"),
            impuesto=self.impuesto,
        )
        nota.calcular_totales()
        nota.estado = "emitida"
        nota.save()
        registrar_asiento_nota_credito(nota)
        self.assertTrue(
            AsientoContable.objects.filter(
                documento_tipo="nota_credito",
                documento_id=nota.id,
                evento="emision",
            ).exists()
        )
