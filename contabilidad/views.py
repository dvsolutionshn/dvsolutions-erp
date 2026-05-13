from collections import defaultdict

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import transaction
from decimal import Decimal
from django.utils import timezone
from django.utils.dateparse import parse_date

from django.http import HttpResponse
from django.db.models import Count, Q, Sum
from django.db.models.functions import TruncMonth
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from core.models import Empresa, Usuario
from facturacion.models import CompraInventario, Factura, NotaCredito, PagoCompra, PagoFactura, RegistroCompraFiscal

from .forms import (
    AsientoContableForm,
    ClasificacionCompraFiscalForm,
    ClasificacionMovimientoBancoForm,
    ConfiguracionContableEmpresaForm,
    CuentaContableForm,
    CuentaFinancieraForm,
    EnlazarMovimientoFacturaForm,
    EnlazarMovimientoCompraForm,
    ImportarCatalogoCuentasForm,
    ImportarMovimientosBancoForm,
    LineaAsientoFormSet,
    MovimientoBancarioClasificacionForm,
    MovimientoBancarioEdicionForm,
    PeriodoContableForm,
    ReglaClasificacionBancoForm,
)
from .importadores import importar_catalogo_cuentas_desde_excel, importar_movimientos_bancarios_desde_excel
from .models import AsientoContable, ClasificacionCompraFiscal, ClasificacionMovimientoBanco, ConfiguracionContableEmpresa, CuentaContable, CuentaFinanciera, LineaAsientoContable, MovimientoBancario, PeriodoContable, ReglaClasificacionBanco
from .services import aplicar_reglas_clasificacion_bancaria, asegurar_clasificaciones_bancarias_base_honduras, cargar_catalogo_base_honduras, registrar_asiento_pago_cliente, registrar_asiento_pago_proveedor


def _empresa_desde_slug(empresa_slug):
    return get_object_or_404(Empresa, slug=empresa_slug, activa=True)


def _resumen_contabilidad(empresa):
    cuentas = CuentaContable.objects.filter(empresa=empresa)
    asientos = AsientoContable.objects.filter(empresa=empresa)
    return {
        "cuentas": cuentas.count(),
        "cuentas_activas": cuentas.filter(activa=True).count(),
        "asientos": asientos.count(),
        "contabilizados": asientos.filter(estado="contabilizado").count(),
        "borradores": asientos.filter(estado="borrador").count(),
    }


def _pendientes_cierre_periodo(empresa, periodo):
    rango = {"fecha__gte": periodo.fecha_inicio, "fecha__lte": periodo.fecha_fin}
    asientos_borrador = AsientoContable.objects.filter(
        empresa=empresa,
        estado="borrador",
        **rango,
    ).count()
    movimientos_bancarios = MovimientoBancario.objects.filter(
        empresa=empresa,
        fecha__gte=periodo.fecha_inicio,
        fecha__lte=periodo.fecha_fin,
    ).exclude(estado="contabilizado").count()
    movimientos_sin_conciliar = MovimientoBancario.objects.filter(
        empresa=empresa,
        estado="contabilizado",
        conciliado=False,
        fecha__gte=periodo.fecha_inicio,
        fecha__lte=periodo.fecha_fin,
    ).count()
    lineas = LineaAsientoContable.objects.filter(
        asiento__empresa=empresa,
        asiento__estado="contabilizado",
        asiento__fecha__gte=periodo.fecha_inicio,
        asiento__fecha__lte=periodo.fecha_fin,
    ).aggregate(debe=Sum("debe"), haber=Sum("haber"))
    total_debe = lineas["debe"] or Decimal("0.00")
    total_haber = lineas["haber"] or Decimal("0.00")
    return {
        "asientos_borrador": asientos_borrador,
        "movimientos_bancarios": movimientos_bancarios,
        "movimientos_sin_conciliar": movimientos_sin_conciliar,
        "debe": total_debe,
        "haber": total_haber,
        "balance_cuadrado": total_debe == total_haber,
        "total": asientos_borrador + movimientos_bancarios + movimientos_sin_conciliar + (0 if total_debe == total_haber else 1),
    }


def _detalle_cierre_periodo(empresa, periodo):
    return {
        "control": _pendientes_cierre_periodo(empresa, periodo),
        "asientos_borrador": AsientoContable.objects.filter(
            empresa=empresa,
            estado="borrador",
            fecha__gte=periodo.fecha_inicio,
            fecha__lte=periodo.fecha_fin,
        ).order_by("fecha", "id"),
        "movimientos_no_contabilizados": MovimientoBancario.objects.filter(
            empresa=empresa,
            fecha__gte=periodo.fecha_inicio,
            fecha__lte=periodo.fecha_fin,
        ).exclude(estado="contabilizado").select_related("cuenta_financiera", "clasificacion").order_by("fecha", "id"),
        "movimientos_sin_conciliar": MovimientoBancario.objects.filter(
            empresa=empresa,
            estado="contabilizado",
            conciliado=False,
            fecha__gte=periodo.fecha_inicio,
            fecha__lte=periodo.fecha_fin,
        ).select_related("cuenta_financiera", "asiento").order_by("fecha", "id"),
    }


def _preparar_excel(titulo):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = titulo[:31]
    sheet.freeze_panes = "A3"
    sheet.append([titulo])
    sheet["A1"].font = Font(bold=True, size=14, color="0F2742")
    sheet.append([])
    return workbook, sheet


def _aplicar_encabezado(sheet, fila):
    fill = PatternFill("solid", fgColor="0F2742")
    for cell in sheet[fila]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def _respuesta_excel(workbook, nombre_archivo):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    workbook.save(response)
    return response


def _autoajustar_columnas(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max_length + 3, 48)


@login_required
def contabilidad_dashboard(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    ultimos_asientos = AsientoContable.objects.filter(empresa=empresa).order_by("-fecha", "-id")[:8]
    context = {
        "empresa": empresa,
        "resumen": _resumen_contabilidad(empresa),
        "ultimos_asientos": ultimos_asientos,
        "cuentas_por_tipo": (
            CuentaContable.objects.filter(empresa=empresa)
            .values("tipo")
            .annotate(total=Count("id"))
            .order_by("tipo")
        ),
    }
    return render(request, "contabilidad/dashboard_contabilidad.html", context)


@login_required
def configuracion_contable(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    configuracion, _ = ConfiguracionContableEmpresa.objects.get_or_create(empresa=empresa)
    form = ConfiguracionContableEmpresaForm(
        request.POST or None,
        instance=configuracion,
        empresa=empresa,
    )
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Configuracion contable actualizada correctamente.")
        return redirect("contabilidad_dashboard", empresa_slug=empresa.slug)
    return render(request, "contabilidad/configuracion_contable.html", {
        "empresa": empresa,
        "form": form,
        "configuracion": configuracion,
        "cuentas_disponibles": CuentaContable.objects.filter(
            empresa=empresa,
            activa=True,
            acepta_movimientos=True,
        ).count(),
    })


@login_required
def periodos_contables(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    periodos = list(PeriodoContable.objects.filter(empresa=empresa).select_related("cerrado_por"))
    asientos = AsientoContable.objects.filter(empresa=empresa)
    for periodo in periodos:
        periodo.control_cierre = _pendientes_cierre_periodo(empresa, periodo)
    resumen = {
        "total": len(periodos),
        "abiertos": sum(1 for periodo in periodos if periodo.estado == "abierto"),
        "cerrados": sum(1 for periodo in periodos if periodo.estado == "cerrado"),
        "asientos": asientos.count(),
    }
    return render(request, "contabilidad/periodos_contables.html", {
        "empresa": empresa,
        "periodos": periodos,
        "resumen": resumen,
    })


@login_required
def crear_periodo_contable(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    form = PeriodoContableForm(request.POST or None, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        periodo = form.save(commit=False)
        periodo.empresa = empresa
        if periodo.estado == "cerrado":
            periodo.cerrado_por = request.user
            periodo.fecha_cierre = timezone.now()
        periodo.save()
        messages.success(request, f"Periodo {periodo.mes}/{periodo.anio} guardado correctamente.")
        return redirect("periodos_contables", empresa_slug=empresa.slug)
    return render(request, "contabilidad/periodo_form.html", {
        "empresa": empresa,
        "form": form,
        "titulo": "Nuevo Periodo Contable",
    })


@login_required
def editar_periodo_contable(request, empresa_slug, periodo_id):
    empresa = _empresa_desde_slug(empresa_slug)
    periodo = get_object_or_404(PeriodoContable, id=periodo_id, empresa=empresa)
    estado_original = periodo.estado
    form = PeriodoContableForm(request.POST or None, instance=periodo, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        periodo = form.save(commit=False)
        periodo.empresa = empresa
        if periodo.estado == "cerrado" and estado_original != "cerrado":
            periodo.cerrado_por = request.user
            periodo.fecha_cierre = timezone.now()
        if periodo.estado == "abierto":
            periodo.cerrado_por = None
            periodo.fecha_cierre = None
        periodo.save()
        messages.success(request, f"Periodo {periodo.mes}/{periodo.anio} actualizado correctamente.")
        return redirect("periodos_contables", empresa_slug=empresa.slug)
    return render(request, "contabilidad/periodo_form.html", {
        "empresa": empresa,
        "form": form,
        "periodo": periodo,
        "titulo": f"Editar Periodo {periodo.mes}/{periodo.anio}",
    })


@login_required
@require_POST
def cerrar_periodo_contable(request, empresa_slug, periodo_id):
    empresa = _empresa_desde_slug(empresa_slug)
    periodo = get_object_or_404(PeriodoContable, id=periodo_id, empresa=empresa)
    pendientes = _pendientes_cierre_periodo(empresa, periodo)
    if pendientes["total"]:
        messages.error(
            request,
            (
                f"No se puede cerrar {periodo.mes}/{periodo.anio}: "
                f"hay {pendientes['asientos_borrador']} asientos borrador y "
                f"{pendientes['movimientos_bancarios']} movimientos bancarios sin contabilizar, "
                f"{pendientes['movimientos_sin_conciliar']} movimientos sin conciliar"
                f"{' y el balance del periodo no cuadra.' if not pendientes['balance_cuadrado'] else '.'}"
            ),
        )
        return redirect("periodos_contables", empresa_slug=empresa.slug)
    periodo.estado = "cerrado"
    periodo.cerrado_por = request.user
    periodo.fecha_cierre = timezone.now()
    periodo.save(update_fields=["estado", "cerrado_por", "fecha_cierre"])
    messages.success(request, f"Periodo {periodo.mes}/{periodo.anio} cerrado correctamente.")
    return redirect("periodos_contables", empresa_slug=empresa.slug)


@login_required
def checklist_cierre_periodo(request, empresa_slug, periodo_id):
    empresa = _empresa_desde_slug(empresa_slug)
    periodo = get_object_or_404(PeriodoContable, id=periodo_id, empresa=empresa)
    detalle = _detalle_cierre_periodo(empresa, periodo)
    return render(request, "contabilidad/checklist_cierre_periodo.html", {
        "empresa": empresa,
        "periodo": periodo,
        **detalle,
    })


@login_required
@require_POST
def abrir_periodo_contable(request, empresa_slug, periodo_id):
    empresa = _empresa_desde_slug(empresa_slug)
    periodo = get_object_or_404(PeriodoContable, id=periodo_id, empresa=empresa)
    periodo.estado = "abierto"
    periodo.cerrado_por = None
    periodo.fecha_cierre = None
    periodo.save(update_fields=["estado", "cerrado_por", "fecha_cierre"])
    messages.success(request, f"Periodo {periodo.mes}/{periodo.anio} abierto nuevamente.")
    return redirect("periodos_contables", empresa_slug=empresa.slug)


@login_required
def clasificaciones_compras_fiscales(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    clasificaciones = ClasificacionCompraFiscal.objects.filter(empresa=empresa).select_related("cuenta_contable")
    return render(request, "contabilidad/clasificaciones_compras_fiscales.html", {
        "empresa": empresa,
        "clasificaciones": clasificaciones,
        "resumen": {
            "total": clasificaciones.count(),
            "activas": clasificaciones.filter(activa=True).count(),
            "inactivas": clasificaciones.filter(activa=False).count(),
        },
    })


@login_required
def crear_clasificacion_compra_fiscal(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    form = ClasificacionCompraFiscalForm(request.POST or None, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        clasificacion = form.save(commit=False)
        clasificacion.empresa = empresa
        clasificacion.full_clean()
        clasificacion.save()
        messages.success(request, f"Clasificacion {clasificacion.nombre} creada correctamente.")
        return redirect("clasificaciones_compras_fiscales", empresa_slug=empresa.slug)
    return render(request, "contabilidad/clasificacion_compra_fiscal_form.html", {
        "empresa": empresa,
        "form": form,
        "titulo": "Nueva Clasificacion de Compra",
    })


@login_required
def editar_clasificacion_compra_fiscal(request, empresa_slug, clasificacion_id):
    empresa = _empresa_desde_slug(empresa_slug)
    clasificacion = get_object_or_404(ClasificacionCompraFiscal, id=clasificacion_id, empresa=empresa)
    form = ClasificacionCompraFiscalForm(request.POST or None, instance=clasificacion, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        clasificacion = form.save(commit=False)
        clasificacion.empresa = empresa
        clasificacion.full_clean()
        clasificacion.save()
        messages.success(request, f"Clasificacion {clasificacion.nombre} actualizada correctamente.")
        return redirect("clasificaciones_compras_fiscales", empresa_slug=empresa.slug)
    return render(request, "contabilidad/clasificacion_compra_fiscal_form.html", {
        "empresa": empresa,
        "form": form,
        "clasificacion": clasificacion,
        "titulo": f"Editar Clasificacion: {clasificacion.nombre}",
    })


@login_required
def bancos_dashboard(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    cuentas = CuentaFinanciera.objects.filter(empresa=empresa).select_related("cuenta_contable")
    movimientos = MovimientoBancario.objects.filter(empresa=empresa).select_related("cuenta_financiera", "clasificacion")
    import_form = ImportarMovimientosBancoForm(empresa=empresa)
    return render(request, "contabilidad/bancos_dashboard.html", {
        "empresa": empresa,
        "cuentas": cuentas,
        "movimientos": movimientos.order_by("-fecha", "-id")[:50],
        "import_form": import_form,
        "resumen": {
            "cuentas": cuentas.count(),
            "movimientos": movimientos.count(),
            "pendientes": movimientos.filter(estado="pendiente").count(),
            "clasificados": movimientos.filter(estado="clasificado").count(),
            "contabilizados": movimientos.filter(estado="contabilizado").count(),
        },
    })


@login_required
def crear_cuenta_financiera(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    form = CuentaFinancieraForm(request.POST or None, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        cuenta = form.save(commit=False)
        cuenta.empresa = empresa
        cuenta.full_clean()
        cuenta.save()
        messages.success(request, f"Cuenta financiera {cuenta.nombre} creada correctamente.")
        return redirect("bancos_dashboard", empresa_slug=empresa.slug)
    return render(request, "contabilidad/cuenta_financiera_form.html", {
        "empresa": empresa,
        "form": form,
        "titulo": "Nueva Cuenta Financiera",
    })


@login_required
def importar_movimientos_bancarios(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    form = ImportarMovimientosBancoForm(request.POST or None, request.FILES or None, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        try:
            resultado = importar_movimientos_bancarios_desde_excel(
                empresa,
                form.cleaned_data["cuenta_financiera"],
                form.cleaned_data["archivo"],
            )
        except ValidationError as error:
            for mensaje in error.messages if hasattr(error, "messages") else [str(error)]:
                messages.error(request, mensaje)
        else:
            reglas = aplicar_reglas_clasificacion_bancaria(
                empresa,
                movimiento_ids=resultado.get("movimiento_ids", []),
            )
            messages.success(
                request,
                f"Estado importado: {resultado['creados']} movimientos creados, {resultado['duplicados']} duplicados omitidos.",
            )
            if reglas["actualizados"]:
                messages.info(request, f"{reglas['actualizados']} movimientos fueron clasificados automaticamente por reglas bancarias.")
            if resultado["omitidos"]:
                messages.warning(request, f"Se omitieron {len(resultado['omitidos'])} filas incompletas.")
    else:
        for campo, errores in form.errors.items():
            etiqueta = form.fields[campo].label if campo in form.fields else "Formulario"
            for error in errores:
                messages.error(request, f"{etiqueta}: {error}")
        if not form.errors:
            messages.error(request, "Revisa la cuenta financiera y el archivo antes de importar.")
    return redirect("bancos_dashboard", empresa_slug=empresa.slug)


@login_required
def clasificaciones_bancarias(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    clasificaciones = ClasificacionMovimientoBanco.objects.filter(empresa=empresa).select_related("cuenta_contable")
    return render(request, "contabilidad/clasificaciones_bancarias.html", {
        "empresa": empresa,
        "clasificaciones": clasificaciones,
        "resumen": {
            "total": clasificaciones.count(),
            "activas": clasificaciones.filter(activa=True).count(),
        },
    })


@login_required
def crear_clasificacion_bancaria(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    form = ClasificacionMovimientoBancoForm(request.POST or None, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        clasificacion = form.save(commit=False)
        clasificacion.empresa = empresa
        clasificacion.full_clean()
        clasificacion.save()
        messages.success(request, f"Clasificacion bancaria {clasificacion.nombre} creada correctamente.")
        return redirect("clasificaciones_bancarias", empresa_slug=empresa.slug)
    return render(request, "contabilidad/clasificacion_bancaria_form.html", {
        "empresa": empresa,
        "form": form,
        "titulo": "Nueva Clasificacion Bancaria",
    })


@login_required
def reglas_bancarias(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    reglas = ReglaClasificacionBanco.objects.filter(empresa=empresa).select_related("clasificacion", "clasificacion__cuenta_contable")
    return render(request, "contabilidad/reglas_bancarias.html", {
        "empresa": empresa,
        "reglas": reglas,
        "resumen": {
            "total": reglas.count(),
            "activas": reglas.filter(activa=True).count(),
        },
    })


@login_required
def crear_regla_bancaria(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    asegurar_clasificaciones_bancarias_base_honduras(empresa)
    form = ReglaClasificacionBancoForm(request.POST or None, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        regla = form.save(commit=False)
        regla.empresa = empresa
        regla.full_clean()
        regla.save()
        messages.success(request, f"Regla bancaria {regla.nombre} creada correctamente.")
        return redirect("reglas_bancarias", empresa_slug=empresa.slug)
    return render(request, "contabilidad/regla_bancaria_form.html", {
        "empresa": empresa,
        "form": form,
        "titulo": "Nueva Regla Bancaria",
    })


@login_required
def editar_regla_bancaria(request, empresa_slug, regla_id):
    empresa = _empresa_desde_slug(empresa_slug)
    regla = get_object_or_404(ReglaClasificacionBanco, id=regla_id, empresa=empresa)
    form = ReglaClasificacionBancoForm(request.POST or None, instance=regla, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        regla = form.save(commit=False)
        regla.empresa = empresa
        regla.full_clean()
        regla.save()
        messages.success(request, f"Regla bancaria {regla.nombre} actualizada correctamente.")
        return redirect("reglas_bancarias", empresa_slug=empresa.slug)
    return render(request, "contabilidad/regla_bancaria_form.html", {
        "empresa": empresa,
        "form": form,
        "titulo": f"Editar Regla: {regla.nombre}",
    })


@login_required
def aplicar_reglas_bancarias_view(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    resultado = aplicar_reglas_clasificacion_bancaria(empresa)
    if resultado["actualizados"]:
        messages.success(request, f"{resultado['actualizados']} movimientos clasificados automaticamente.")
    else:
        messages.info(request, "No se encontraron movimientos pendientes que coincidan con las reglas activas.")
    return redirect("movimientos_bancarios", empresa_slug=empresa.slug)


@login_required
def movimientos_bancarios(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    asegurar_clasificaciones_bancarias_base_honduras(empresa)
    cuenta_id = request.GET.get("cuenta") or ""
    estado = request.GET.get("estado") or ""
    tipo = request.GET.get("tipo") or ""
    q = request.GET.get("q") or ""
    movimientos = MovimientoBancario.objects.filter(empresa=empresa).select_related("cuenta_financiera", "clasificacion", "asiento").order_by("-fecha", "-id")
    if cuenta_id:
        movimientos = movimientos.filter(cuenta_financiera_id=cuenta_id)
    if estado:
        movimientos = movimientos.filter(estado=estado)
    if tipo == "ingreso":
        movimientos = movimientos.filter(credito__gt=0)
    elif tipo == "egreso":
        movimientos = movimientos.filter(debito__gt=0)
    if q:
        movimientos = movimientos.filter(
            Q(descripcion__icontains=q) |
            Q(referencia__icontains=q) |
            Q(cuenta_financiera__nombre__icontains=q)
        )
    clasificaciones = ClasificacionMovimientoBanco.objects.filter(empresa=empresa, activa=True).order_by("nombre")
    resumen = {
        "total": movimientos.count(),
        "pendientes": movimientos.filter(estado="pendiente").count(),
        "clasificados": movimientos.filter(estado="clasificado").count(),
        "contabilizados": movimientos.filter(estado="contabilizado").count(),
        "debitos": movimientos.aggregate(total=Sum("debito"))["total"] or Decimal("0.00"),
        "creditos": movimientos.aggregate(total=Sum("credito"))["total"] or Decimal("0.00"),
    }
    return render(request, "contabilidad/movimientos_bancarios.html", {
        "empresa": empresa,
        "movimientos": movimientos,
        "cuentas": CuentaFinanciera.objects.filter(empresa=empresa).order_by("nombre"),
        "clasificaciones": clasificaciones,
        "resumen": resumen,
        "filtros": {"cuenta": cuenta_id, "estado": estado, "tipo": tipo, "q": q},
    })


@login_required
def clasificar_movimientos_bancarios_lote(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    redirect_url = request.POST.get("next") or redirect("movimientos_bancarios", empresa_slug=empresa.slug).url
    if request.method != "POST":
        return redirect(redirect_url)

    ids = request.POST.getlist("movimiento_ids")
    movimientos = MovimientoBancario.objects.filter(empresa=empresa, id__in=ids).exclude(estado="contabilizado")
    clasificaciones_validas = set(
        ClasificacionMovimientoBanco.objects.filter(empresa=empresa, activa=True).values_list("id", flat=True)
    )
    actualizados = 0

    with transaction.atomic():
        for movimiento in movimientos:
            valor = request.POST.get(f"clasificacion_{movimiento.id}") or ""
            if valor:
                try:
                    clasificacion_id = int(valor)
                except ValueError:
                    continue
                if clasificacion_id not in clasificaciones_validas:
                    continue
                if movimiento.clasificacion_id != clasificacion_id or movimiento.estado != "clasificado":
                    movimiento.clasificacion_id = clasificacion_id
                    movimiento.estado = "clasificado"
                    movimiento.save(update_fields=["clasificacion", "estado"])
                    actualizados += 1
            elif movimiento.clasificacion_id:
                movimiento.clasificacion = None
                movimiento.estado = "pendiente"
                movimiento.save(update_fields=["clasificacion", "estado"])
                actualizados += 1

    if actualizados:
        messages.success(request, f"{actualizados} movimientos actualizados correctamente.")
    else:
        messages.info(request, "No hubo cambios en las clasificaciones.")
    return redirect(redirect_url)


@login_required
def clasificar_movimiento_bancario(request, empresa_slug, movimiento_id):
    empresa = _empresa_desde_slug(empresa_slug)
    movimiento = get_object_or_404(MovimientoBancario, id=movimiento_id, empresa=empresa)
    form = MovimientoBancarioClasificacionForm(request.POST or None, instance=movimiento, empresa=empresa)
    redirect_url = request.POST.get("next") or redirect("movimientos_bancarios", empresa_slug=empresa.slug).url
    if request.method == "POST" and form.is_valid():
        movimiento = form.save(commit=False)
        movimiento.estado = "clasificado" if movimiento.clasificacion_id else "pendiente"
        movimiento.save(update_fields=["clasificacion", "estado"])
        messages.success(request, "Movimiento clasificado correctamente.")
    else:
        messages.error(request, "Selecciona una clasificacion valida.")
    return redirect(redirect_url)


@login_required
def editar_movimiento_bancario(request, empresa_slug, movimiento_id):
    empresa = _empresa_desde_slug(empresa_slug)
    movimiento = get_object_or_404(
        MovimientoBancario.objects.select_related("cuenta_financiera", "clasificacion", "asiento"),
        id=movimiento_id,
        empresa=empresa,
    )
    if movimiento.estado == "contabilizado" or movimiento.asiento_id:
        messages.error(request, "Este movimiento ya fue contabilizado y no se puede editar. Reversa el asiento si necesitas corregirlo.")
        return redirect("movimientos_bancarios", empresa_slug=empresa.slug)

    form = MovimientoBancarioEdicionForm(request.POST or None, instance=movimiento, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        movimiento = form.save(commit=False)
        movimiento.estado = "clasificado" if movimiento.clasificacion_id else "pendiente"
        try:
            movimiento.full_clean()
            movimiento.save()
        except ValidationError as exc:
            if hasattr(exc, "message_dict"):
                for campo, errores in exc.message_dict.items():
                    for error in errores:
                        form.add_error(campo if campo != "__all__" else None, error)
            else:
                form.add_error(None, exc.messages[0] if exc.messages else str(exc))
        else:
            messages.success(request, "Movimiento bancario actualizado correctamente.")
            return redirect("movimientos_bancarios", empresa_slug=empresa.slug)

    return render(request, "contabilidad/movimiento_bancario_form.html", {
        "empresa": empresa,
        "movimiento": movimiento,
        "form": form,
    })


@login_required
def enlazar_movimiento_factura(request, empresa_slug, movimiento_id):
    empresa = _empresa_desde_slug(empresa_slug)
    movimiento = get_object_or_404(
        MovimientoBancario.objects.select_related("cuenta_financiera", "pago_factura", "asiento"),
        id=movimiento_id,
        empresa=empresa,
    )
    if not movimiento.es_ingreso:
        messages.error(request, "Solo los creditos bancarios se pueden enlazar como pago de factura.")
        return redirect("movimientos_bancarios", empresa_slug=empresa.slug)
    if movimiento.estado == "contabilizado" or movimiento.asiento_id or movimiento.pago_factura_id:
        messages.error(request, "Este movimiento ya fue contabilizado o enlazado a un pago.")
        return redirect("movimientos_bancarios", empresa_slug=empresa.slug)

    form = EnlazarMovimientoFacturaForm(request.POST or None, empresa=empresa, movimiento=movimiento)
    if request.method == "POST" and form.is_valid():
        factura = form.cleaned_data["factura"]
        with transaction.atomic():
            pago = PagoFactura.objects.create(
                factura=factura,
                fecha=movimiento.fecha,
                monto=movimiento.credito,
                metodo="transferencia",
                referencia=movimiento.referencia or movimiento.descripcion[:100],
                cuenta_financiera=movimiento.cuenta_financiera,
            )
            asiento = registrar_asiento_pago_cliente(pago)
            movimiento.pago_factura = pago
            movimiento.asiento = asiento
            movimiento.estado = "contabilizado" if asiento else "clasificado"
            movimiento.save(update_fields=["pago_factura", "asiento", "estado"])
        messages.success(request, f"Deposito enlazado a factura {factura.numero_factura or factura.id} y pago registrado correctamente.")
        return redirect("movimientos_bancarios", empresa_slug=empresa.slug)

    return render(request, "contabilidad/enlazar_movimiento_factura.html", {
        "empresa": empresa,
        "movimiento": movimiento,
        "form": form,
    })


@login_required
def enlazar_movimiento_compra(request, empresa_slug, movimiento_id):
    empresa = _empresa_desde_slug(empresa_slug)
    movimiento = get_object_or_404(
        MovimientoBancario.objects.select_related("cuenta_financiera", "pago_compra", "asiento"),
        id=movimiento_id,
        empresa=empresa,
    )
    if movimiento.es_ingreso:
        messages.error(request, "Solo los debitos bancarios se pueden enlazar como pago de compra.")
        return redirect("movimientos_bancarios", empresa_slug=empresa.slug)
    if movimiento.estado == "contabilizado" or movimiento.asiento_id or movimiento.pago_compra_id:
        messages.error(request, "Este movimiento ya fue contabilizado o enlazado a un pago.")
        return redirect("movimientos_bancarios", empresa_slug=empresa.slug)

    form = EnlazarMovimientoCompraForm(request.POST or None, empresa=empresa, movimiento=movimiento)
    if request.method == "POST" and form.is_valid():
        compra = form.cleaned_data["compra"]
        with transaction.atomic():
            pago = PagoCompra.objects.create(
                compra=compra,
                fecha=movimiento.fecha,
                monto=movimiento.debito,
                metodo="transferencia",
                referencia=movimiento.referencia or movimiento.descripcion[:100],
                cuenta_financiera=movimiento.cuenta_financiera,
                observacion=f"Generado desde movimiento bancario {movimiento.id}",
            )
            asiento = registrar_asiento_pago_proveedor(pago)
            movimiento.pago_compra = pago
            movimiento.asiento = asiento
            movimiento.estado = "contabilizado" if asiento else "clasificado"
            movimiento.save(update_fields=["pago_compra", "asiento", "estado"])
        messages.success(request, f"Debito bancario enlazado a compra {compra.numero_compra or compra.id} y pago registrado correctamente.")
        return redirect("movimientos_bancarios", empresa_slug=empresa.slug)

    return render(request, "contabilidad/enlazar_movimiento_compra.html", {
        "empresa": empresa,
        "movimiento": movimiento,
        "form": form,
    })


@login_required
@require_POST
def contabilizar_movimiento_bancario(request, empresa_slug, movimiento_id):
    empresa = _empresa_desde_slug(empresa_slug)
    redirect_url = request.POST.get("next") or redirect("movimientos_bancarios", empresa_slug=empresa.slug).url
    movimiento = get_object_or_404(
        MovimientoBancario.objects.select_related("cuenta_financiera", "clasificacion__cuenta_contable"),
        id=movimiento_id,
        empresa=empresa,
    )
    if movimiento.estado == "contabilizado":
        messages.warning(request, "Este movimiento ya fue contabilizado.")
        return redirect(redirect_url)
    if not movimiento.clasificacion_id:
        messages.error(request, "Clasifica el movimiento antes de contabilizarlo.")
        return redirect(redirect_url)
    if PeriodoContable.fecha_bloqueada(empresa, movimiento.fecha):
        messages.error(request, "No se puede contabilizar porque el periodo contable esta cerrado.")
        return redirect(redirect_url)

    with transaction.atomic():
        asiento = AsientoContable.objects.create(
            empresa=empresa,
            fecha=movimiento.fecha,
            descripcion=f"Movimiento bancario: {movimiento.descripcion}",
            referencia=movimiento.referencia or str(movimiento.id),
            origen_modulo="bancos",
            documento_tipo="movimiento_bancario",
            documento_id=movimiento.id,
            evento="contabilizacion",
            creado_por=request.user,
            estado="borrador",
        )
        if movimiento.es_ingreso:
            lineas = [
                (movimiento.cuenta_financiera.cuenta_contable, "Ingreso en banco", movimiento.credito, Decimal("0.00")),
                (movimiento.clasificacion.cuenta_contable, movimiento.clasificacion.nombre, Decimal("0.00"), movimiento.credito),
            ]
        else:
            lineas = [
                (movimiento.clasificacion.cuenta_contable, movimiento.clasificacion.nombre, movimiento.debito, Decimal("0.00")),
                (movimiento.cuenta_financiera.cuenta_contable, "Salida de banco/tarjeta", Decimal("0.00"), movimiento.debito),
            ]
        for cuenta, detalle, debe, haber in lineas:
            LineaAsientoContable.objects.create(asiento=asiento, cuenta=cuenta, detalle=detalle, debe=debe, haber=haber)
        asiento.generar_numero()
        asiento.estado = "contabilizado"
        asiento.save(update_fields=["numero", "estado"])
        movimiento.asiento = asiento
        movimiento.estado = "contabilizado"
        movimiento.save(update_fields=["asiento", "estado"])
    messages.success(request, f"Movimiento contabilizado en asiento {asiento.numero}.")
    return redirect(redirect_url)


@login_required
def conciliacion_bancaria(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    cuenta_id = request.GET.get("cuenta") or ""
    estado = request.GET.get("estado") or ""
    conciliado = request.GET.get("conciliado") or ""
    fecha_desde = parse_date(request.GET.get("desde") or "")
    fecha_hasta = parse_date(request.GET.get("hasta") or "")

    movimientos = MovimientoBancario.objects.filter(empresa=empresa).select_related(
        "cuenta_financiera",
        "clasificacion",
        "asiento",
        "conciliado_por",
    ).order_by("-fecha", "-id")

    if cuenta_id:
        movimientos = movimientos.filter(cuenta_financiera_id=cuenta_id)
    if estado:
        movimientos = movimientos.filter(estado=estado)
    if conciliado == "si":
        movimientos = movimientos.filter(conciliado=True)
    elif conciliado == "no":
        movimientos = movimientos.filter(conciliado=False)
    if fecha_desde:
        movimientos = movimientos.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        movimientos = movimientos.filter(fecha__lte=fecha_hasta)

    resumen_base = MovimientoBancario.objects.filter(empresa=empresa)
    resumen_filtrado = movimientos.aggregate(
        debitos=Sum("debito"),
        creditos=Sum("credito"),
    )
    resumen_cuentas = []
    for cuenta in CuentaFinanciera.objects.filter(empresa=empresa).select_related("cuenta_contable").order_by("nombre"):
        movimientos_cuenta = MovimientoBancario.objects.filter(empresa=empresa, cuenta_financiera=cuenta)
        ult_movimiento = movimientos_cuenta.order_by("-fecha", "-id").first()
        totales_cuenta = movimientos_cuenta.aggregate(debitos=Sum("debito"), creditos=Sum("credito"))
        lineas_cuenta = LineaAsientoContable.objects.filter(
            asiento__empresa=empresa,
            asiento__estado="contabilizado",
            cuenta=cuenta.cuenta_contable,
        ).aggregate(debe=Sum("debe"), haber=Sum("haber"))
        saldo_contable = (lineas_cuenta["debe"] or Decimal("0.00")) - (lineas_cuenta["haber"] or Decimal("0.00"))
        saldo_banco = ult_movimiento.saldo if ult_movimiento else Decimal("0.00")
        resumen_cuentas.append({
            "cuenta": cuenta,
            "movimientos": movimientos_cuenta.count(),
            "debitos": totales_cuenta["debitos"] or Decimal("0.00"),
            "creditos": totales_cuenta["creditos"] or Decimal("0.00"),
            "saldo_banco": saldo_banco,
            "saldo_contable": saldo_contable,
            "diferencia": saldo_banco - saldo_contable,
            "sin_conciliar": movimientos_cuenta.filter(estado="contabilizado", conciliado=False).count(),
            "pendientes": movimientos_cuenta.exclude(estado="contabilizado").count(),
        })

    return render(request, "contabilidad/conciliacion_bancaria.html", {
        "empresa": empresa,
        "movimientos": movimientos,
        "cuentas": CuentaFinanciera.objects.filter(empresa=empresa).order_by("nombre"),
        "clasificaciones": ClasificacionMovimientoBanco.objects.filter(empresa=empresa, activa=True).order_by("nombre"),
        "resumen_cuentas": resumen_cuentas,
        "filtros": {
            "cuenta": cuenta_id,
            "estado": estado,
            "conciliado": conciliado,
            "desde": request.GET.get("desde") or "",
            "hasta": request.GET.get("hasta") or "",
        },
        "resumen": {
            "pendientes": resumen_base.filter(estado="pendiente").count(),
            "clasificados": resumen_base.filter(estado="clasificado").count(),
            "contabilizados": resumen_base.filter(estado="contabilizado").count(),
            "conciliados": resumen_base.filter(conciliado=True).count(),
            "sin_conciliar": resumen_base.filter(estado="contabilizado", conciliado=False).count(),
            "debitos": resumen_filtrado["debitos"] or Decimal("0.00"),
            "creditos": resumen_filtrado["creditos"] or Decimal("0.00"),
        },
    })


@login_required
def conciliar_movimiento_bancario(request, empresa_slug, movimiento_id):
    empresa = _empresa_desde_slug(empresa_slug)
    movimiento = get_object_or_404(MovimientoBancario, id=movimiento_id, empresa=empresa)
    redirect_url = request.POST.get("next") or redirect("conciliacion_bancaria", empresa_slug=empresa.slug).url

    if request.method != "POST":
        return redirect(redirect_url)
    if movimiento.estado != "contabilizado":
        messages.error(request, "Solo se pueden conciliar movimientos ya contabilizados.")
        return redirect(redirect_url)

    movimiento.conciliado = True
    movimiento.fecha_conciliacion = timezone.now()
    movimiento.conciliado_por = request.user
    movimiento.save(update_fields=["conciliado", "fecha_conciliacion", "conciliado_por"])
    messages.success(request, "Movimiento conciliado correctamente.")
    return redirect(redirect_url)


@login_required
def desconciliar_movimiento_bancario(request, empresa_slug, movimiento_id):
    empresa = _empresa_desde_slug(empresa_slug)
    movimiento = get_object_or_404(MovimientoBancario, id=movimiento_id, empresa=empresa)
    redirect_url = request.POST.get("next") or redirect("conciliacion_bancaria", empresa_slug=empresa.slug).url

    if request.method != "POST":
        return redirect(redirect_url)

    movimiento.conciliado = False
    movimiento.fecha_conciliacion = None
    movimiento.conciliado_por = None
    movimiento.save(update_fields=["conciliado", "fecha_conciliacion", "conciliado_por"])
    messages.success(request, "Movimiento marcado como no conciliado.")
    return redirect(redirect_url)


@login_required
def catalogo_cuentas(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    cuentas = CuentaContable.objects.filter(empresa=empresa).select_related("cuenta_padre").order_by("codigo", "nombre")
    context = {
        "empresa": empresa,
        "cuentas": cuentas,
        "resumen": {
            "total": cuentas.count(),
            "activas": cuentas.filter(activa=True).count(),
            "movimiento": cuentas.filter(acepta_movimientos=True).count(),
            "agrupadoras": cuentas.filter(acepta_movimientos=False).count(),
        },
    }
    return render(request, "contabilidad/catalogo_cuentas.html", context)


@login_required
def cargar_catalogo_base_honduras_view(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    if request.method != "POST":
        return redirect("catalogo_cuentas", empresa_slug=empresa.slug)

    resultado = cargar_catalogo_base_honduras(empresa)
    if resultado["creadas"] or resultado["cuentas_financieras_creadas"]:
        messages.success(
            request,
            f"Catalogo base Honduras cargado: {resultado['creadas']} cuentas creadas, {resultado['existentes']} ya existentes, {resultado['cuentas_financieras_creadas']} cuentas financieras listas y {resultado['clasificaciones_bancarias_creadas']} clasificaciones bancarias listas.",
        )
    else:
        messages.info(request, "El catalogo base Honduras ya estaba cargado para esta empresa.")
    return redirect("catalogo_cuentas", empresa_slug=empresa.slug)


@login_required
def crear_cuenta_contable(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    form = CuentaContableForm(request.POST or None, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        cuenta = form.save(commit=False)
        cuenta.empresa = empresa
        cuenta.full_clean()
        cuenta.save()
        messages.success(request, f"Cuenta {cuenta.codigo} creada correctamente.")
        return redirect("catalogo_cuentas", empresa_slug=empresa.slug)
    return render(request, "contabilidad/cuenta_form.html", {
        "empresa": empresa,
        "form": form,
        "titulo": "Nueva Cuenta Contable",
    })


@login_required
def importar_catalogo_cuentas(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    form = ImportarCatalogoCuentasForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        try:
            resultado = importar_catalogo_cuentas_desde_excel(
                empresa,
                form.cleaned_data["archivo"],
                actualizar_existentes=form.cleaned_data["actualizar_existentes"],
            )
        except ValidationError as error:
            mensajes = error.messages if hasattr(error, "messages") else [str(error)]
            for mensaje in mensajes:
                messages.error(request, mensaje)
        else:
            messages.success(
                request,
                f"Catalogo importado: {resultado['creadas']} creadas, {resultado['actualizadas']} actualizadas.",
            )
            return redirect("catalogo_cuentas", empresa_slug=empresa.slug)
    return render(request, "contabilidad/importar_catalogo_cuentas.html", {
        "empresa": empresa,
        "form": form,
    })


@login_required
def editar_cuenta_contable(request, empresa_slug, cuenta_id):
    empresa = _empresa_desde_slug(empresa_slug)
    cuenta = get_object_or_404(CuentaContable, id=cuenta_id, empresa=empresa)
    form = CuentaContableForm(request.POST or None, instance=cuenta, empresa=empresa)
    if request.method == "POST" and form.is_valid():
        cuenta = form.save(commit=False)
        cuenta.full_clean()
        cuenta.save()
        messages.success(request, f"Cuenta {cuenta.codigo} actualizada correctamente.")
        return redirect("catalogo_cuentas", empresa_slug=empresa.slug)
    return render(request, "contabilidad/cuenta_form.html", {
        "empresa": empresa,
        "form": form,
        "titulo": f"Editar Cuenta: {cuenta.codigo}",
    })


@login_required
def asientos_contables(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    asientos = AsientoContable.objects.filter(empresa=empresa).prefetch_related("lineas").order_by("-fecha", "-id")
    q = request.GET.get("q", "").strip()
    estado = request.GET.get("estado", "").strip()
    fecha_inicio = request.GET.get("fecha_inicio", "").strip()
    fecha_fin = request.GET.get("fecha_fin", "").strip()
    if q:
        asientos = asientos.filter(
            Q(numero__icontains=q)
            | Q(descripcion__icontains=q)
            | Q(referencia__icontains=q)
            | Q(origen_modulo__icontains=q)
        )
    if estado:
        asientos = asientos.filter(estado=estado)
    if fecha_inicio:
        asientos = asientos.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        asientos = asientos.filter(fecha__lte=fecha_fin)
    context = {
        "empresa": empresa,
        "asientos": asientos,
        "resumen": _resumen_contabilidad(empresa),
        "filtros": {
            "q": q,
            "estado": estado,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
        },
    }
    return render(request, "contabilidad/asientos_contables.html", context)


@login_required
def crear_asiento_contable(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    asiento = AsientoContable(empresa=empresa, creado_por=request.user)
    form = AsientoContableForm(request.POST or None, instance=asiento)
    cuentas_disponibles = CuentaContable.objects.filter(
        empresa=empresa,
        activa=True,
        acepta_movimientos=True,
    ).order_by("codigo")
    formset = LineaAsientoFormSet(
        request.POST or None,
        instance=asiento,
        form_kwargs=None,
    )
    for child_form in formset.forms:
        if "cuenta" in child_form.fields:
            child_form.fields["cuenta"].queryset = cuentas_disponibles
    if hasattr(formset, "empty_form") and "cuenta" in formset.empty_form.fields:
        formset.empty_form.fields["cuenta"].queryset = cuentas_disponibles

    if request.method == "POST" and form.is_valid() and formset.is_valid():
        if PeriodoContable.fecha_bloqueada(empresa, form.cleaned_data["fecha"]):
            messages.error(request, "No se puede crear el asiento porque el periodo contable esta cerrado.")
            return render(request, "contabilidad/asiento_form.html", {
                "empresa": empresa,
                "form": form,
                "formset": formset,
                "titulo": "Nuevo Asiento Contable",
            })
        with transaction.atomic():
            asiento = form.save(commit=False)
            asiento.empresa = empresa
            asiento.creado_por = request.user
            asiento.save()
            formset.instance = asiento
            lineas = formset.save(commit=False)
            for deleted in formset.deleted_objects:
                deleted.delete()
            for linea in lineas:
                linea.asiento = asiento
                linea.full_clean()
                linea.save()
            if not asiento.lineas.exists():
                messages.error(request, "Debes agregar al menos una linea al asiento.")
                return render(request, "contabilidad/asiento_form.html", {
                    "empresa": empresa,
                    "form": form,
                    "formset": formset,
                    "titulo": "Nuevo Asiento Contable",
                })
            if form.cleaned_data.get("contabilizar_ahora"):
                if asiento.esta_balanceado:
                    asiento.generar_numero()
                    asiento.estado = "contabilizado"
                    asiento.save(update_fields=["numero", "estado"])
                    messages.success(request, f"Asiento {asiento.numero} contabilizado correctamente.")
                else:
                    messages.warning(request, "El asiento se guardo en borrador porque no esta balanceado.")
            else:
                messages.success(request, "Asiento guardado correctamente.")
            return redirect("ver_asiento_contable", empresa_slug=empresa.slug, asiento_id=asiento.id)

    return render(request, "contabilidad/asiento_form.html", {
        "empresa": empresa,
        "form": form,
        "formset": formset,
        "titulo": "Nuevo Asiento Contable",
        "cuentas_busqueda": [
            {
                "id": cuenta.id,
                "codigo": cuenta.codigo,
                "nombre": cuenta.nombre,
            }
            for cuenta in cuentas_disponibles
        ],
    })


@login_required
def ver_asiento_contable(request, empresa_slug, asiento_id):
    empresa = _empresa_desde_slug(empresa_slug)
    asiento = get_object_or_404(AsientoContable.objects.prefetch_related("lineas__cuenta"), id=asiento_id, empresa=empresa)
    return render(request, "contabilidad/ver_asiento.html", {
        "empresa": empresa,
        "asiento": asiento,
    })


@login_required
@require_POST
def duplicar_asiento_contable(request, empresa_slug, asiento_id):
    empresa = _empresa_desde_slug(empresa_slug)
    asiento_origen = get_object_or_404(
        AsientoContable.objects.prefetch_related("lineas"),
        id=asiento_id,
        empresa=empresa,
    )
    fecha = timezone.now().date()
    if PeriodoContable.fecha_bloqueada(empresa, fecha):
        messages.error(request, "No se puede duplicar el asiento porque el periodo actual esta cerrado.")
        return redirect("ver_asiento_contable", empresa_slug=empresa.slug, asiento_id=asiento_origen.id)
    with transaction.atomic():
        nuevo = AsientoContable.objects.create(
            empresa=empresa,
            fecha=fecha,
            descripcion=f"Copia de {asiento_origen.descripcion}",
            referencia=asiento_origen.referencia,
            origen_modulo=asiento_origen.origen_modulo or "manual",
            documento_tipo="asiento_duplicado",
            documento_id=asiento_origen.id,
            evento="duplicado",
            creado_por=request.user,
            estado="borrador",
        )
        for linea in asiento_origen.lineas.all():
            LineaAsientoContable.objects.create(
                asiento=nuevo,
                cuenta=linea.cuenta,
                detalle=linea.detalle,
                debe=linea.debe,
                haber=linea.haber,
            )
    messages.success(request, "Asiento duplicado como borrador para revision.")
    return redirect("ver_asiento_contable", empresa_slug=empresa.slug, asiento_id=nuevo.id)


@login_required
@require_POST
def reversar_asiento_contable(request, empresa_slug, asiento_id):
    empresa = _empresa_desde_slug(empresa_slug)
    asiento_origen = get_object_or_404(
        AsientoContable.objects.prefetch_related("lineas"),
        id=asiento_id,
        empresa=empresa,
    )
    if asiento_origen.estado != "contabilizado":
        messages.error(request, "Solo se pueden reversar asientos contabilizados.")
        return redirect("ver_asiento_contable", empresa_slug=empresa.slug, asiento_id=asiento_origen.id)
    if AsientoContable.objects.filter(
        empresa=empresa,
        documento_tipo="asiento_contable",
        documento_id=asiento_origen.id,
        evento="reversion",
        estado="contabilizado",
    ).exists():
        messages.warning(request, "Este asiento ya tiene una reversion contabilizada.")
        return redirect("ver_asiento_contable", empresa_slug=empresa.slug, asiento_id=asiento_origen.id)
    fecha = timezone.now().date()
    if PeriodoContable.fecha_bloqueada(empresa, fecha):
        messages.error(request, "No se puede reversar el asiento porque el periodo actual esta cerrado.")
        return redirect("ver_asiento_contable", empresa_slug=empresa.slug, asiento_id=asiento_origen.id)
    with transaction.atomic():
        reversion = AsientoContable.objects.create(
            empresa=empresa,
            fecha=fecha,
            descripcion=f"Reversion de {asiento_origen.numero or asiento_origen.descripcion}",
            referencia=asiento_origen.referencia,
            origen_modulo=asiento_origen.origen_modulo or "manual",
            documento_tipo="asiento_contable",
            documento_id=asiento_origen.id,
            evento="reversion",
            creado_por=request.user,
            estado="borrador",
        )
        for linea in asiento_origen.lineas.all():
            LineaAsientoContable.objects.create(
                asiento=reversion,
                cuenta=linea.cuenta,
                detalle=f"Reversion {linea.detalle or asiento_origen.descripcion}",
                debe=linea.haber,
                haber=linea.debe,
            )
        reversion.generar_numero()
        reversion.estado = "contabilizado"
        reversion.save(update_fields=["numero", "estado"])
    messages.success(request, f"Reversion contabilizada en asiento {reversion.numero}.")
    return redirect("ver_asiento_contable", empresa_slug=empresa.slug, asiento_id=reversion.id)


@login_required
@require_POST
def contabilizar_asiento_contable(request, empresa_slug, asiento_id):
    empresa = _empresa_desde_slug(empresa_slug)
    asiento = get_object_or_404(AsientoContable, id=asiento_id, empresa=empresa)
    if asiento.estado != "borrador":
        messages.warning(request, "Solo los asientos en borrador pueden contabilizarse.")
        return redirect("ver_asiento_contable", empresa_slug=empresa.slug, asiento_id=asiento.id)
    if not asiento.esta_balanceado:
        messages.error(request, "El asiento no esta balanceado. Revisa debe y haber antes de contabilizar.")
        return redirect("ver_asiento_contable", empresa_slug=empresa.slug, asiento_id=asiento.id)
    if PeriodoContable.fecha_bloqueada(empresa, asiento.fecha):
        messages.error(request, "No se puede contabilizar porque el periodo contable esta cerrado.")
        return redirect("ver_asiento_contable", empresa_slug=empresa.slug, asiento_id=asiento.id)
    asiento.generar_numero()
    asiento.estado = "contabilizado"
    asiento.save(update_fields=["numero", "estado"])
    messages.success(request, f"Asiento {asiento.numero} contabilizado correctamente.")
    return redirect("ver_asiento_contable", empresa_slug=empresa.slug, asiento_id=asiento.id)


@login_required
def reportes_contables(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    asientos = AsientoContable.objects.filter(empresa=empresa, estado="contabilizado")
    cuentas = CuentaContable.objects.filter(empresa=empresa)
    context = {
        "empresa": empresa,
        "resumen": {
            "asientos_contabilizados": asientos.count(),
            "debe_total": sum((asiento.total_debe for asiento in asientos), 0),
            "haber_total": sum((asiento.total_haber for asiento in asientos), 0),
            "cuentas_activas": cuentas.filter(activa=True).count(),
        },
        "cuentas_por_tipo": (
            cuentas.values("tipo").annotate(total=Count("id")).order_by("tipo")
        ),
    }
    return render(request, "contabilidad/reportes_contables.html", context)


def _aplicar_filtros_fecha_documentos(queryset, campo_fecha, fecha_inicio, fecha_fin):
    if fecha_inicio:
        queryset = queryset.filter(**{f"{campo_fecha}__gte": fecha_inicio})
    if fecha_fin:
        queryset = queryset.filter(**{f"{campo_fecha}__lte": fecha_fin})
    return queryset


def _variacion_porcentual(actual, anterior):
    actual = actual or Decimal("0.00")
    anterior = anterior or Decimal("0.00")
    if anterior == 0:
        return Decimal("100.00") if actual > 0 else Decimal("0.00")
    return ((actual - anterior) / anterior) * Decimal("100.00")


@login_required
def dashboard_bi_financiero(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    hoy = timezone.now().date()
    inicio_mes_actual = hoy.replace(day=1)
    if inicio_mes_actual.month == 1:
        inicio_mes_anterior = inicio_mes_actual.replace(year=inicio_mes_actual.year - 1, month=12)
    else:
        inicio_mes_anterior = inicio_mes_actual.replace(month=inicio_mes_actual.month - 1)

    compras_qs = (
        CompraInventario.objects.filter(empresa=empresa)
        .exclude(estado="anulada")
        .select_related("proveedor")
        .prefetch_related("lineas")
        .order_by("fecha_documento", "id")
    )
    compras = list(compras_qs)
    pagos_qs = (
        PagoCompra.objects.filter(compra__empresa=empresa)
        .select_related("compra__proveedor", "cuenta_financiera")
        .order_by("fecha", "id")
    )
    pagos = list(pagos_qs)
    movimientos_qs = (
        MovimientoBancario.objects.filter(empresa=empresa)
        .select_related("cuenta_financiera")
        .order_by("fecha", "id")
    )
    movimientos = list(movimientos_qs)

    compras_mes_actual = [compra for compra in compras if compra.fecha_documento >= inicio_mes_actual]
    compras_mes_anterior = [
        compra for compra in compras
        if inicio_mes_anterior <= compra.fecha_documento < inicio_mes_actual
    ]
    pagos_mes_actual = [pago for pago in pagos if pago.fecha >= inicio_mes_actual]
    pagos_mes_anterior = [
        pago for pago in pagos
        if inicio_mes_anterior <= pago.fecha < inicio_mes_actual
    ]
    movimientos_mes_actual = [mov for mov in movimientos if mov.fecha >= inicio_mes_actual]

    total_compras_actual = sum((compra.total_documento for compra in compras_mes_actual), Decimal("0.00"))
    total_compras_anterior = sum((compra.total_documento for compra in compras_mes_anterior), Decimal("0.00"))
    total_pagos_actual = sum((pago.monto for pago in pagos_mes_actual), Decimal("0.00"))
    total_pagos_anterior = sum((pago.monto for pago in pagos_mes_anterior), Decimal("0.00"))
    saldo_total = sum((compra.saldo_pendiente for compra in compras), Decimal("0.00"))
    total_documentos_actual = len(compras_mes_actual)
    ticket_promedio = (
        total_compras_actual / Decimal(str(total_documentos_actual))
        if total_documentos_actual else Decimal("0.00")
    )
    tasa_pago_actual = (
        (total_pagos_actual / total_compras_actual) * Decimal("100.00")
        if total_compras_actual > 0 else Decimal("0.00")
    )

    compras_por_mes_map = defaultdict(lambda: {"comprado": Decimal("0.00"), "pagado": Decimal("0.00"), "documentos": 0})
    for compra in compras:
        clave = compra.fecha_documento.replace(day=1)
        compras_por_mes_map[clave]["comprado"] += compra.total_documento
        compras_por_mes_map[clave]["documentos"] += 1
    for pago in pagos:
        clave = pago.fecha.replace(day=1)
        compras_por_mes_map[clave]["pagado"] += pago.monto

    compras_por_mes = []
    if compras_por_mes_map:
        max_comprado = max((item["comprado"] for item in compras_por_mes_map.values()), default=Decimal("0.00"))
        max_pagado = max((item["pagado"] for item in compras_por_mes_map.values()), default=Decimal("0.00"))
        max_valor = max(max_comprado, max_pagado, Decimal("1.00"))
        for mes in sorted(compras_por_mes_map.keys())[-6:]:
            item = compras_por_mes_map[mes]
            compras_por_mes.append({
                "mes": mes,
                "etiqueta": mes.strftime("%b %Y"),
                "comprado": item["comprado"],
                "pagado": item["pagado"],
                "documentos": item["documentos"],
                "alto_comprado": float((item["comprado"] / max_valor) * Decimal("100.00")) if max_valor else 0,
                "alto_pagado": float((item["pagado"] / max_valor) * Decimal("100.00")) if max_valor else 0,
            })

    top_proveedores_map = defaultdict(lambda: {"total": Decimal("0.00"), "documentos": 0})
    for compra in compras:
        nombre = (compra.proveedor.nombre if compra.proveedor_id else compra.proveedor_nombre) or "Proveedor sin nombre"
        top_proveedores_map[nombre]["total"] += compra.total_documento
        top_proveedores_map[nombre]["documentos"] += 1
    top_proveedores = [
        {"nombre": nombre, "total": data["total"], "documentos": data["documentos"]}
        for nombre, data in top_proveedores_map.items()
    ]
    top_proveedores.sort(key=lambda item: item["total"], reverse=True)
    top_proveedores = top_proveedores[:6]
    max_proveedor = max((item["total"] for item in top_proveedores), default=Decimal("1.00"))
    for item in top_proveedores:
        item["ancho"] = float((item["total"] / max_proveedor) * Decimal("100.00")) if max_proveedor else 0

    proveedores_vencidos = []
    for compra in compras:
        if not compra.esta_vencida:
            continue
        proveedores_vencidos.append({
            "nombre": (compra.proveedor.nombre if compra.proveedor_id else compra.proveedor_nombre) or "Proveedor sin nombre",
            "saldo": compra.saldo_pendiente,
            "dias_vencido": (hoy - compra.fecha_control_cxp).days,
            "numero_compra": compra.numero_compra or f"COM-{compra.id:08d}",
        })
    proveedores_vencidos.sort(key=lambda item: item["saldo"], reverse=True)
    proveedores_vencidos = proveedores_vencidos[:6]
    max_vencido = max((item["saldo"] for item in proveedores_vencidos), default=Decimal("1.00"))
    for item in proveedores_vencidos:
        item["ancho"] = float((item["saldo"] / max_vencido) * Decimal("100.00")) if max_vencido else 0

    egresos_por_banco_map = defaultdict(lambda: {"total": Decimal("0.00"), "operaciones": 0, "tipo": "Caja"})
    for pago in pagos:
        nombre = "Caja operativa"
        tipo = "Caja"
        if pago.cuenta_financiera_id:
            nombre = pago.cuenta_financiera.nombre
            tipo = pago.cuenta_financiera.get_tipo_display()
        egresos_por_banco_map[nombre]["total"] += pago.monto
        egresos_por_banco_map[nombre]["operaciones"] += 1
        egresos_por_banco_map[nombre]["tipo"] = tipo
    egresos_por_banco = [
        {"nombre": nombre, "total": data["total"], "operaciones": data["operaciones"], "tipo": data["tipo"]}
        for nombre, data in egresos_por_banco_map.items()
    ]
    egresos_por_banco.sort(key=lambda item: item["total"], reverse=True)
    egresos_por_banco = egresos_por_banco[:6]
    max_banco = max((item["total"] for item in egresos_por_banco), default=Decimal("1.00"))
    for item in egresos_por_banco:
        item["ancho"] = float((item["total"] / max_banco) * Decimal("100.00")) if max_banco else 0

    resumen_movimientos = []
    estados_mov = [
        ("pendiente", "Pendiente", "#f59e0b"),
        ("clasificado", "Clasificado", "#2968f2"),
        ("contabilizado", "Contabilizado", "#14b8a6"),
    ]
    total_movimientos = len(movimientos)
    pendientes_conciliar = sum(1 for mov in movimientos if mov.estado == "contabilizado" and not mov.conciliado)
    conciliados = sum(1 for mov in movimientos if mov.conciliado)
    for codigo, etiqueta, color in estados_mov:
        cantidad = sum(1 for mov in movimientos if mov.estado == codigo)
        monto = sum((mov.monto for mov in movimientos if mov.estado == codigo), Decimal("0.00"))
        porcentaje = (Decimal(cantidad) / Decimal(total_movimientos) * Decimal("100.00")) if total_movimientos else Decimal("0.00")
        resumen_movimientos.append({
            "codigo": codigo,
            "etiqueta": etiqueta,
            "cantidad": cantidad,
            "monto": monto,
            "color": color,
            "porcentaje": porcentaje,
        })

    total_volumen_bancario = sum((mov.monto for mov in movimientos_mes_actual), Decimal("0.00"))

    context = {
        "empresa": empresa,
        "compras": compras,
        "inicio_mes_actual": inicio_mes_actual,
        "kpis_bi": {
            "total_compras_actual": total_compras_actual,
            "variacion_compras": _variacion_porcentual(total_compras_actual, total_compras_anterior),
            "documentos_actual": total_documentos_actual,
            "pagos_actual": total_pagos_actual,
            "variacion_pagos": _variacion_porcentual(total_pagos_actual, total_pagos_anterior),
            "saldo_total": saldo_total,
            "ticket_promedio": ticket_promedio,
            "tasa_pago_actual": tasa_pago_actual,
            "movimientos_mes": len(movimientos_mes_actual),
            "conciliados": conciliados,
            "pendientes_conciliar": pendientes_conciliar,
            "total_volumen_bancario": total_volumen_bancario,
        },
        "compras_por_mes": compras_por_mes,
        "top_proveedores": top_proveedores,
        "proveedores_vencidos": proveedores_vencidos,
        "egresos_por_banco": egresos_por_banco,
        "resumen_movimientos": resumen_movimientos,
    }
    return render(request, "contabilidad/dashboard_bi_financiero.html", context)


def _resumen_fiscal_documentos(documentos):
    resumen = {
        "subtotal": Decimal("0.00"),
        "descuento_total": Decimal("0.00"),
        "base_15": Decimal("0.00"),
        "base_18": Decimal("0.00"),
        "base_exento": Decimal("0.00"),
        "base_exonerado": Decimal("0.00"),
        "isv_15": Decimal("0.00"),
        "isv_18": Decimal("0.00"),
        "impuesto": Decimal("0.00"),
        "total": Decimal("0.00"),
        "documentos": documentos.count(),
    }
    for documento in documentos:
        fiscal = documento.resumen_fiscal()
        resumen["subtotal"] += documento.subtotal or Decimal("0.00")
        resumen["descuento_total"] += fiscal.get("descuento_total") or Decimal("0.00")
        resumen["base_15"] += fiscal.get("base_15") or Decimal("0.00")
        resumen["base_18"] += fiscal.get("base_18") or Decimal("0.00")
        resumen["base_exento"] += fiscal.get("base_exento") or Decimal("0.00")
        resumen["base_exonerado"] += fiscal.get("base_exonerado") or Decimal("0.00")
        resumen["isv_15"] += fiscal.get("isv_15") or Decimal("0.00")
        resumen["isv_18"] += fiscal.get("isv_18") or Decimal("0.00")
        resumen["impuesto"] += documento.impuesto or Decimal("0.00")
        resumen["total"] += documento.total or Decimal("0.00")
    return resumen


def _reporte_impuestos_data(empresa, fecha_inicio="", fecha_fin=""):
    facturas = Factura.objects.filter(
        empresa=empresa,
        estado="emitida",
    ).select_related("cliente").prefetch_related("lineas__impuesto").order_by("fecha_emision", "numero_factura", "id")
    notas_credito = NotaCredito.objects.filter(
        empresa=empresa,
        estado="emitida",
    ).select_related("cliente", "factura_origen").prefetch_related("lineas__impuesto").order_by("fecha_emision", "numero_nota", "id")
    compras = RegistroCompraFiscal.objects.filter(
        empresa=empresa,
        estado="registrada",
    ).select_related("proveedor").order_by("fecha_documento", "numero_factura", "id")

    facturas = _aplicar_filtros_fecha_documentos(facturas, "fecha_emision", fecha_inicio, fecha_fin)
    notas_credito = _aplicar_filtros_fecha_documentos(notas_credito, "fecha_emision", fecha_inicio, fecha_fin)
    compras = _aplicar_filtros_fecha_documentos(compras, "fecha_documento", fecha_inicio, fecha_fin)

    ventas = _resumen_fiscal_documentos(facturas)
    creditos = _resumen_fiscal_documentos(notas_credito)
    compras_resumen = compras.aggregate(
        base_15=Sum("base_15"),
        isv_15=Sum("isv_15"),
        base_18=Sum("base_18"),
        isv_18=Sum("isv_18"),
        base_exento=Sum("exento"),
        base_exonerado=Sum("exonerado"),
        total=Sum("total"),
    )
    compras_resumen = {clave: valor or Decimal("0.00") for clave, valor in compras_resumen.items()}
    compras_resumen["documentos"] = compras.count()
    compras_resumen["impuesto"] = compras_resumen["isv_15"] + compras_resumen["isv_18"]
    neto = {
        "base_15": ventas["base_15"] - creditos["base_15"],
        "base_18": ventas["base_18"] - creditos["base_18"],
        "base_exento": ventas["base_exento"] - creditos["base_exento"],
        "base_exonerado": ventas["base_exonerado"] - creditos["base_exonerado"],
        "isv_15": ventas["isv_15"] - creditos["isv_15"],
        "isv_18": ventas["isv_18"] - creditos["isv_18"],
        "impuesto": ventas["impuesto"] - creditos["impuesto"],
        "total": ventas["total"] - creditos["total"],
    }
    isv_a_pagar = neto["impuesto"] - compras_resumen["impuesto"]
    return {
        "facturas": facturas,
        "notas_credito": notas_credito,
        "compras": compras,
        "ventas": ventas,
        "creditos": creditos,
        "compras_resumen": compras_resumen,
        "neto": neto,
        "isv_a_pagar": isv_a_pagar,
        "filtros": {
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
        },
    }


@login_required
def reporte_impuestos(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    data = _reporte_impuestos_data(
        empresa,
        request.GET.get("fecha_inicio") or "",
        request.GET.get("fecha_fin") or "",
    )
    data["empresa"] = empresa
    return render(request, "contabilidad/reporte_impuestos.html", data)


@login_required
def exportar_reporte_impuestos_excel(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    data = _reporte_impuestos_data(
        empresa,
        request.GET.get("fecha_inicio") or "",
        request.GET.get("fecha_fin") or "",
    )
    workbook, sheet = _preparar_excel("Reporte de Impuestos")

    sheet.append(["Resumen", "Documentos", "Base 15%", "ISV 15%", "Base 18%", "ISV 18%", "Exento", "Exonerado", "ISV total", "Total"])
    _aplicar_encabezado(sheet, 3)
    sheet.append(["Facturas emitidas", data["ventas"]["documentos"], data["ventas"]["base_15"], data["ventas"]["isv_15"], data["ventas"]["base_18"], data["ventas"]["isv_18"], data["ventas"]["base_exento"], data["ventas"]["base_exonerado"], data["ventas"]["impuesto"], data["ventas"]["total"]])
    sheet.append(["Notas de credito", data["creditos"]["documentos"], data["creditos"]["base_15"], data["creditos"]["isv_15"], data["creditos"]["base_18"], data["creditos"]["isv_18"], data["creditos"]["base_exento"], data["creditos"]["base_exonerado"], data["creditos"]["impuesto"], data["creditos"]["total"]])
    sheet.append(["Neto fiscal", "", data["neto"]["base_15"], data["neto"]["isv_15"], data["neto"]["base_18"], data["neto"]["isv_18"], data["neto"]["base_exento"], data["neto"]["base_exonerado"], data["neto"]["impuesto"], data["neto"]["total"]])
    sheet.append(["Compras acreditables", data["compras_resumen"]["documentos"], data["compras_resumen"]["base_15"], data["compras_resumen"]["isv_15"], data["compras_resumen"]["base_18"], data["compras_resumen"]["isv_18"], data["compras_resumen"]["base_exento"], data["compras_resumen"]["base_exonerado"], data["compras_resumen"]["impuesto"], data["compras_resumen"]["total"]])
    sheet.append(["ISV estimado a pagar", "", "", "", "", "", "", "", data["isv_a_pagar"], ""])
    sheet.append([])
    sheet.append(["Facturas emitidas"])
    sheet.append(["Fecha", "Numero", "Cliente", "Base 15%", "ISV 15%", "Base 18%", "ISV 18%", "Exento", "Exonerado", "ISV", "Total"])
    _aplicar_encabezado(sheet, sheet.max_row)
    for factura in data["facturas"]:
        fiscal = factura.resumen_fiscal()
        sheet.append([
            factura.fecha_emision,
            factura.numero_factura or "Sin numero",
            factura.cliente.nombre if factura.cliente else "",
            fiscal["base_15"],
            fiscal["isv_15"],
            fiscal["base_18"],
            fiscal["isv_18"],
            fiscal["base_exento"],
            fiscal["base_exonerado"],
            factura.impuesto,
            factura.total,
        ])
    sheet.append([])
    sheet.append(["Notas de credito"])
    sheet.append(["Fecha", "Numero", "Factura origen", "Cliente", "Base 15%", "ISV 15%", "Base 18%", "ISV 18%", "Exento", "Exonerado", "ISV", "Total"])
    _aplicar_encabezado(sheet, sheet.max_row)
    for nota in data["notas_credito"]:
        fiscal = nota.resumen_fiscal()
        sheet.append([
            nota.fecha_emision,
            nota.numero_nota or "Sin numero",
            nota.factura_origen.numero_factura if nota.factura_origen else "",
            nota.cliente.nombre if nota.cliente else "",
            fiscal["base_15"],
            fiscal["isv_15"],
            fiscal["base_18"],
            fiscal["isv_18"],
            fiscal["base_exento"],
            fiscal["base_exonerado"],
            nota.impuesto,
            nota.total,
        ])
    sheet.append([])
    sheet.append(["Compras fiscales"])
    sheet.append(["Fecha", "Factura", "Proveedor", "Base 15%", "ISV 15%", "Base 18%", "ISV 18%", "Exento", "Exonerado", "ISV", "Total"])
    _aplicar_encabezado(sheet, sheet.max_row)
    for compra in data["compras"]:
        sheet.append([
            compra.fecha_documento,
            compra.numero_factura,
            compra.proveedor_nombre,
            compra.base_15,
            compra.isv_15,
            compra.base_18,
            compra.isv_18,
            compra.exento,
            compra.exonerado,
            compra.impuesto_total,
            compra.total,
        ])
    _autoajustar_columnas(sheet)
    return _respuesta_excel(workbook, f"Reporte_Impuestos_{empresa.slug}.xlsx")


def _auditoria_asientos_queryset(empresa, filtros):
    asientos = AsientoContable.objects.filter(empresa=empresa).select_related("creado_por").prefetch_related("lineas")
    if filtros.get("fecha_inicio"):
        asientos = asientos.filter(fecha__gte=filtros["fecha_inicio"])
    if filtros.get("fecha_fin"):
        asientos = asientos.filter(fecha__lte=filtros["fecha_fin"])
    if filtros.get("estado"):
        asientos = asientos.filter(estado=filtros["estado"])
    if filtros.get("origen"):
        asientos = asientos.filter(origen_modulo=filtros["origen"])
    if filtros.get("documento_tipo"):
        asientos = asientos.filter(documento_tipo=filtros["documento_tipo"])
    if filtros.get("usuario"):
        asientos = asientos.filter(creado_por_id=filtros["usuario"])
    return asientos.order_by("-fecha", "-id")


@login_required
def auditoria_contable(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    filtros = {
        "fecha_inicio": request.GET.get("fecha_inicio") or "",
        "fecha_fin": request.GET.get("fecha_fin") or "",
        "estado": request.GET.get("estado") or "",
        "origen": request.GET.get("origen") or "",
        "documento_tipo": request.GET.get("documento_tipo") or "",
        "usuario": request.GET.get("usuario") or "",
    }
    asientos = _auditoria_asientos_queryset(empresa, filtros)
    base_asientos = AsientoContable.objects.filter(empresa=empresa)
    usuarios_ids = base_asientos.exclude(creado_por__isnull=True).values_list("creado_por_id", flat=True).distinct()
    context = {
        "empresa": empresa,
        "asientos": asientos,
        "filtros": filtros,
        "estados": AsientoContable.ESTADO_CHOICES,
        "origenes": base_asientos.exclude(origen_modulo__isnull=True).exclude(origen_modulo__exact="").values_list("origen_modulo", flat=True).distinct().order_by("origen_modulo"),
        "documentos": base_asientos.exclude(documento_tipo__isnull=True).exclude(documento_tipo__exact="").values_list("documento_tipo", flat=True).distinct().order_by("documento_tipo"),
        "usuarios": Usuario.objects.filter(id__in=usuarios_ids).order_by("username"),
        "resumen": {
            "total": asientos.count(),
            "contabilizados": asientos.filter(estado="contabilizado").count(),
            "borradores": asientos.filter(estado="borrador").count(),
            "anulados": asientos.filter(estado="anulado").count(),
        },
    }
    return render(request, "contabilidad/auditoria_contable.html", context)


@login_required
def exportar_auditoria_contable_excel(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    filtros = {
        "fecha_inicio": request.GET.get("fecha_inicio") or "",
        "fecha_fin": request.GET.get("fecha_fin") or "",
        "estado": request.GET.get("estado") or "",
        "origen": request.GET.get("origen") or "",
        "documento_tipo": request.GET.get("documento_tipo") or "",
        "usuario": request.GET.get("usuario") or "",
    }
    asientos = _auditoria_asientos_queryset(empresa, filtros)
    workbook, sheet = _preparar_excel("Auditoria Contable")
    sheet.append(["Fecha", "Numero", "Estado", "Origen", "Documento", "Evento", "Referencia", "Usuario", "Debe", "Haber", "Creado"])
    _aplicar_encabezado(sheet, 3)
    for asiento in asientos:
        sheet.append([
            asiento.fecha,
            asiento.numero or "Sin numero",
            asiento.get_estado_display(),
            asiento.origen_modulo or "",
            f"{asiento.documento_tipo or ''} {asiento.documento_id or ''}".strip(),
            asiento.evento or "",
            asiento.referencia or "",
            asiento.creado_por.username if asiento.creado_por else "",
            asiento.total_debe,
            asiento.total_haber,
            asiento.fecha_creacion.replace(tzinfo=None) if asiento.fecha_creacion else "",
        ])
    _autoajustar_columnas(sheet)
    return _respuesta_excel(workbook, f"Auditoria_Contable_{empresa.slug}.xlsx")


@login_required
def libro_diario(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    fecha_inicio = request.GET.get("fecha_inicio") or ""
    fecha_fin = request.GET.get("fecha_fin") or ""
    cuenta_id = request.GET.get("cuenta") or ""

    lineas = LineaAsientoContable.objects.filter(
        asiento__empresa=empresa,
        asiento__estado="contabilizado",
    ).select_related("asiento", "cuenta").order_by("asiento__fecha", "asiento__numero", "id")

    if fecha_inicio:
        lineas = lineas.filter(asiento__fecha__gte=fecha_inicio)
    if fecha_fin:
        lineas = lineas.filter(asiento__fecha__lte=fecha_fin)
    if cuenta_id:
        lineas = lineas.filter(cuenta_id=cuenta_id)

    totales = lineas.aggregate(debe=Sum("debe"), haber=Sum("haber"))
    context = {
        "empresa": empresa,
        "lineas": lineas,
        "cuentas": CuentaContable.objects.filter(empresa=empresa, activa=True).order_by("codigo", "nombre"),
        "filtros": {
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "cuenta": cuenta_id,
        },
        "resumen": {
            "debe_total": totales["debe"] or Decimal("0.00"),
            "haber_total": totales["haber"] or Decimal("0.00"),
            "lineas": lineas.count(),
        },
    }
    return render(request, "contabilidad/libro_diario.html", context)


@login_required
def exportar_libro_diario_excel(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    fecha_inicio = request.GET.get("fecha_inicio") or ""
    fecha_fin = request.GET.get("fecha_fin") or ""
    cuenta_id = request.GET.get("cuenta") or ""

    lineas = LineaAsientoContable.objects.filter(
        asiento__empresa=empresa,
        asiento__estado="contabilizado",
    ).select_related("asiento", "cuenta").order_by("asiento__fecha", "asiento__numero", "id")
    if fecha_inicio:
        lineas = lineas.filter(asiento__fecha__gte=fecha_inicio)
    if fecha_fin:
        lineas = lineas.filter(asiento__fecha__lte=fecha_fin)
    if cuenta_id:
        lineas = lineas.filter(cuenta_id=cuenta_id)

    workbook, sheet = _preparar_excel("Libro Diario")
    sheet.append(["Fecha", "Asiento", "Cuenta", "Detalle", "Referencia", "Debe", "Haber"])
    _aplicar_encabezado(sheet, 3)
    for linea in lineas:
        sheet.append([
            linea.asiento.fecha,
            linea.asiento.numero or "Sin numero",
            f"{linea.cuenta.codigo} - {linea.cuenta.nombre}",
            linea.detalle or linea.asiento.descripcion,
            linea.asiento.referencia or "",
            linea.debe,
            linea.haber,
        ])
    _autoajustar_columnas(sheet)
    return _respuesta_excel(workbook, f"Libro_Diario_{empresa.slug}.xlsx")


@login_required
def mayor_cuenta(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    fecha_inicio = request.GET.get("fecha_inicio") or ""
    fecha_fin = request.GET.get("fecha_fin") or ""
    cuenta_id = request.GET.get("cuenta") or ""
    cuenta = None
    lineas = LineaAsientoContable.objects.none()
    movimientos = []
    saldo = Decimal("0.00")

    if cuenta_id:
        cuenta = get_object_or_404(CuentaContable, id=cuenta_id, empresa=empresa)
        lineas = LineaAsientoContable.objects.filter(
            asiento__empresa=empresa,
            asiento__estado="contabilizado",
            cuenta=cuenta,
        ).select_related("asiento", "cuenta").order_by("asiento__fecha", "asiento__numero", "id")

        if fecha_inicio:
            lineas = lineas.filter(asiento__fecha__gte=fecha_inicio)
        if fecha_fin:
            lineas = lineas.filter(asiento__fecha__lte=fecha_fin)

        for linea in lineas:
            saldo += linea.debe - linea.haber
            movimientos.append({
                "linea": linea,
                "saldo": saldo,
            })

    totales = lineas.aggregate(debe=Sum("debe"), haber=Sum("haber"))
    context = {
        "empresa": empresa,
        "cuenta": cuenta,
        "movimientos": movimientos,
        "cuentas": CuentaContable.objects.filter(empresa=empresa, activa=True).order_by("codigo", "nombre"),
        "filtros": {
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "cuenta": cuenta_id,
        },
        "resumen": {
            "debe_total": totales["debe"] or Decimal("0.00"),
            "haber_total": totales["haber"] or Decimal("0.00"),
            "saldo": saldo,
            "movimientos": len(movimientos),
        },
    }
    return render(request, "contabilidad/mayor_cuenta.html", context)


@login_required
def balance_comprobacion(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    fecha_inicio = request.GET.get("fecha_inicio") or ""
    fecha_fin = request.GET.get("fecha_fin") or ""
    tipo = request.GET.get("tipo") or ""
    incluir_ceros = request.GET.get("incluir_ceros") == "1"

    cuentas = CuentaContable.objects.filter(empresa=empresa).order_by("codigo", "nombre")
    if tipo:
        cuentas = cuentas.filter(tipo=tipo)

    movimientos_base = LineaAsientoContable.objects.filter(
        asiento__empresa=empresa,
        asiento__estado="contabilizado",
    )
    if fecha_inicio:
        movimientos_base = movimientos_base.filter(asiento__fecha__gte=fecha_inicio)
    if fecha_fin:
        movimientos_base = movimientos_base.filter(asiento__fecha__lte=fecha_fin)

    movimientos_por_cuenta = {
        item["cuenta_id"]: item
        for item in movimientos_base.values("cuenta_id").annotate(debe=Sum("debe"), haber=Sum("haber"))
    }

    filas = []
    total_debe = Decimal("0.00")
    total_haber = Decimal("0.00")
    saldo_deudor_total = Decimal("0.00")
    saldo_acreedor_total = Decimal("0.00")

    for cuenta in cuentas:
        movimiento = movimientos_por_cuenta.get(cuenta.id, {})
        debe = movimiento.get("debe") or Decimal("0.00")
        haber = movimiento.get("haber") or Decimal("0.00")
        saldo = debe - haber
        saldo_deudor = saldo if saldo > 0 else Decimal("0.00")
        saldo_acreedor = abs(saldo) if saldo < 0 else Decimal("0.00")

        if not incluir_ceros and debe == 0 and haber == 0 and saldo == 0:
            continue

        total_debe += debe
        total_haber += haber
        saldo_deudor_total += saldo_deudor
        saldo_acreedor_total += saldo_acreedor
        filas.append({
            "cuenta": cuenta,
            "debe": debe,
            "haber": haber,
            "saldo_deudor": saldo_deudor,
            "saldo_acreedor": saldo_acreedor,
        })

    context = {
        "empresa": empresa,
        "filas": filas,
        "tipos_cuenta": CuentaContable.TIPO_CHOICES,
        "filtros": {
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "tipo": tipo,
            "incluir_ceros": incluir_ceros,
        },
        "resumen": {
            "cuentas": len(filas),
            "debe_total": total_debe,
            "haber_total": total_haber,
            "saldo_deudor_total": saldo_deudor_total,
            "saldo_acreedor_total": saldo_acreedor_total,
            "cuadrado": total_debe == total_haber,
        },
    }
    return render(request, "contabilidad/balance_comprobacion.html", context)


@login_required
def exportar_balance_comprobacion_excel(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    fecha_inicio = request.GET.get("fecha_inicio") or ""
    fecha_fin = request.GET.get("fecha_fin") or ""
    tipo = request.GET.get("tipo") or ""
    incluir_ceros = request.GET.get("incluir_ceros") == "1"

    cuentas = CuentaContable.objects.filter(empresa=empresa).order_by("codigo", "nombre")
    if tipo:
        cuentas = cuentas.filter(tipo=tipo)

    movimientos_base = LineaAsientoContable.objects.filter(asiento__empresa=empresa, asiento__estado="contabilizado")
    if fecha_inicio:
        movimientos_base = movimientos_base.filter(asiento__fecha__gte=fecha_inicio)
    if fecha_fin:
        movimientos_base = movimientos_base.filter(asiento__fecha__lte=fecha_fin)

    movimientos_por_cuenta = {
        item["cuenta_id"]: item
        for item in movimientos_base.values("cuenta_id").annotate(debe=Sum("debe"), haber=Sum("haber"))
    }

    workbook, sheet = _preparar_excel("Balance de Comprobacion")
    sheet.append(["Cuenta", "Tipo", "Debe", "Haber", "Saldo deudor", "Saldo acreedor"])
    _aplicar_encabezado(sheet, 3)
    total_debe = Decimal("0.00")
    total_haber = Decimal("0.00")
    total_deudor = Decimal("0.00")
    total_acreedor = Decimal("0.00")
    for cuenta in cuentas:
        movimiento = movimientos_por_cuenta.get(cuenta.id, {})
        debe = movimiento.get("debe") or Decimal("0.00")
        haber = movimiento.get("haber") or Decimal("0.00")
        saldo = debe - haber
        saldo_deudor = saldo if saldo > 0 else Decimal("0.00")
        saldo_acreedor = abs(saldo) if saldo < 0 else Decimal("0.00")
        if not incluir_ceros and debe == 0 and haber == 0 and saldo == 0:
            continue
        total_debe += debe
        total_haber += haber
        total_deudor += saldo_deudor
        total_acreedor += saldo_acreedor
        sheet.append([f"{cuenta.codigo} - {cuenta.nombre}", cuenta.get_tipo_display(), debe, haber, saldo_deudor, saldo_acreedor])
    sheet.append(["Totales", "", total_debe, total_haber, total_deudor, total_acreedor])
    _autoajustar_columnas(sheet)
    return _respuesta_excel(workbook, f"Balance_Comprobacion_{empresa.slug}.xlsx")


@login_required
def estado_resultados(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    fecha_inicio = request.GET.get("fecha_inicio") or ""
    fecha_fin = request.GET.get("fecha_fin") or ""

    cuentas = CuentaContable.objects.filter(
        empresa=empresa,
        tipo__in=["ingreso", "costo", "gasto"],
    ).order_by("tipo", "codigo", "nombre")

    movimientos_base = LineaAsientoContable.objects.filter(
        asiento__empresa=empresa,
        asiento__estado="contabilizado",
        cuenta__tipo__in=["ingreso", "costo", "gasto"],
    )
    if fecha_inicio:
        movimientos_base = movimientos_base.filter(asiento__fecha__gte=fecha_inicio)
    if fecha_fin:
        movimientos_base = movimientos_base.filter(asiento__fecha__lte=fecha_fin)

    movimientos_por_cuenta = {
        item["cuenta_id"]: item
        for item in movimientos_base.values("cuenta_id").annotate(debe=Sum("debe"), haber=Sum("haber"))
    }

    ingresos = []
    costos = []
    gastos = []
    total_ingresos = Decimal("0.00")
    total_costos = Decimal("0.00")
    total_gastos = Decimal("0.00")

    for cuenta in cuentas:
        movimiento = movimientos_por_cuenta.get(cuenta.id, {})
        debe = movimiento.get("debe") or Decimal("0.00")
        haber = movimiento.get("haber") or Decimal("0.00")
        if debe == 0 and haber == 0:
            continue

        if cuenta.tipo == "ingreso":
            saldo = haber - debe
            total_ingresos += saldo
            ingresos.append({
                "cuenta": cuenta,
                "debe": debe,
                "haber": haber,
                "saldo": saldo,
            })
        elif cuenta.tipo == "costo":
            saldo = debe - haber
            total_costos += saldo
            costos.append({
                "cuenta": cuenta,
                "debe": debe,
                "haber": haber,
                "saldo": saldo,
            })
        elif cuenta.tipo == "gasto":
            saldo = debe - haber
            total_gastos += saldo
            gastos.append({
                "cuenta": cuenta,
                "debe": debe,
                "haber": haber,
                "saldo": saldo,
            })

    utilidad_bruta = total_ingresos - total_costos
    utilidad = utilidad_bruta - total_gastos
    context = {
        "empresa": empresa,
        "ingresos": ingresos,
        "costos": costos,
        "gastos": gastos,
        "filtros": {
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
        },
        "resumen": {
            "total_ingresos": total_ingresos,
            "total_costos": total_costos,
            "utilidad_bruta": utilidad_bruta,
            "total_gastos": total_gastos,
            "utilidad": utilidad,
            "resultado_positivo": utilidad >= 0,
            "cuentas_ingreso": len(ingresos),
            "cuentas_costo": len(costos),
            "cuentas_gasto": len(gastos),
        },
    }
    return render(request, "contabilidad/estado_resultados.html", context)


@login_required
def exportar_estado_resultados_excel(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    fecha_inicio = request.GET.get("fecha_inicio") or ""
    fecha_fin = request.GET.get("fecha_fin") or ""

    cuentas = CuentaContable.objects.filter(empresa=empresa, tipo__in=["ingreso", "costo", "gasto"]).order_by("tipo", "codigo", "nombre")
    movimientos_base = LineaAsientoContable.objects.filter(
        asiento__empresa=empresa,
        asiento__estado="contabilizado",
        cuenta__tipo__in=["ingreso", "costo", "gasto"],
    )
    if fecha_inicio:
        movimientos_base = movimientos_base.filter(asiento__fecha__gte=fecha_inicio)
    if fecha_fin:
        movimientos_base = movimientos_base.filter(asiento__fecha__lte=fecha_fin)
    movimientos_por_cuenta = {
        item["cuenta_id"]: item
        for item in movimientos_base.values("cuenta_id").annotate(debe=Sum("debe"), haber=Sum("haber"))
    }

    workbook, sheet = _preparar_excel("Estado de Resultados")
    sheet.append(["Tipo", "Cuenta", "Debe", "Haber", "Saldo"])
    _aplicar_encabezado(sheet, 3)
    total_ingresos = Decimal("0.00")
    total_costos = Decimal("0.00")
    total_gastos = Decimal("0.00")
    for cuenta in cuentas:
        movimiento = movimientos_por_cuenta.get(cuenta.id, {})
        debe = movimiento.get("debe") or Decimal("0.00")
        haber = movimiento.get("haber") or Decimal("0.00")
        if debe == 0 and haber == 0:
            continue
        if cuenta.tipo == "ingreso":
            saldo = haber - debe
            total_ingresos += saldo
        else:
            saldo = debe - haber
            if cuenta.tipo == "costo":
                total_costos += saldo
            else:
                total_gastos += saldo
        sheet.append([cuenta.get_tipo_display(), f"{cuenta.codigo} - {cuenta.nombre}", debe, haber, saldo])
    sheet.append([])
    sheet.append(["Total ingresos", "", "", "", total_ingresos])
    sheet.append(["Total costos", "", "", "", total_costos])
    sheet.append(["Utilidad bruta", "", "", "", total_ingresos - total_costos])
    sheet.append(["Total gastos", "", "", "", total_gastos])
    sheet.append(["Utilidad neta", "", "", "", total_ingresos - total_costos - total_gastos])
    _autoajustar_columnas(sheet)
    return _respuesta_excel(workbook, f"Estado_Resultados_{empresa.slug}.xlsx")


@login_required
def balance_general(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    fecha_corte = request.GET.get("fecha_corte") or ""

    cuentas = CuentaContable.objects.filter(
        empresa=empresa,
        tipo__in=["activo", "pasivo", "patrimonio"],
    ).order_by("tipo", "codigo", "nombre")

    movimientos_base = LineaAsientoContable.objects.filter(
        asiento__empresa=empresa,
        asiento__estado="contabilizado",
        cuenta__tipo__in=["activo", "pasivo", "patrimonio"],
    )
    if fecha_corte:
        movimientos_base = movimientos_base.filter(asiento__fecha__lte=fecha_corte)

    movimientos_por_cuenta = {
        item["cuenta_id"]: item
        for item in movimientos_base.values("cuenta_id").annotate(debe=Sum("debe"), haber=Sum("haber"))
    }

    activos = []
    pasivos = []
    patrimonio = []
    total_activos = Decimal("0.00")
    total_pasivos = Decimal("0.00")
    total_patrimonio = Decimal("0.00")

    for cuenta in cuentas:
        movimiento = movimientos_por_cuenta.get(cuenta.id, {})
        debe = movimiento.get("debe") or Decimal("0.00")
        haber = movimiento.get("haber") or Decimal("0.00")
        if debe == 0 and haber == 0:
            continue

        if cuenta.tipo == "activo":
            saldo = debe - haber
            total_activos += saldo
            activos.append({"cuenta": cuenta, "debe": debe, "haber": haber, "saldo": saldo})
        elif cuenta.tipo == "pasivo":
            saldo = haber - debe
            total_pasivos += saldo
            pasivos.append({"cuenta": cuenta, "debe": debe, "haber": haber, "saldo": saldo})
        elif cuenta.tipo == "patrimonio":
            saldo = haber - debe
            total_patrimonio += saldo
            patrimonio.append({"cuenta": cuenta, "debe": debe, "haber": haber, "saldo": saldo})

    total_pasivo_patrimonio = total_pasivos + total_patrimonio
    diferencia = total_activos - total_pasivo_patrimonio
    context = {
        "empresa": empresa,
        "activos": activos,
        "pasivos": pasivos,
        "patrimonio": patrimonio,
        "filtros": {
            "fecha_corte": fecha_corte,
        },
        "resumen": {
            "total_activos": total_activos,
            "total_pasivos": total_pasivos,
            "total_patrimonio": total_patrimonio,
            "total_pasivo_patrimonio": total_pasivo_patrimonio,
            "diferencia": diferencia,
            "cuadrado": diferencia == 0,
        },
    }
    return render(request, "contabilidad/balance_general.html", context)


@login_required
def exportar_balance_general_excel(request, empresa_slug):
    empresa = _empresa_desde_slug(empresa_slug)
    fecha_corte = request.GET.get("fecha_corte") or ""
    cuentas = CuentaContable.objects.filter(empresa=empresa, tipo__in=["activo", "pasivo", "patrimonio"]).order_by("tipo", "codigo", "nombre")
    movimientos_base = LineaAsientoContable.objects.filter(
        asiento__empresa=empresa,
        asiento__estado="contabilizado",
        cuenta__tipo__in=["activo", "pasivo", "patrimonio"],
    )
    if fecha_corte:
        movimientos_base = movimientos_base.filter(asiento__fecha__lte=fecha_corte)
    movimientos_por_cuenta = {
        item["cuenta_id"]: item
        for item in movimientos_base.values("cuenta_id").annotate(debe=Sum("debe"), haber=Sum("haber"))
    }

    workbook, sheet = _preparar_excel("Balance General")
    sheet.append(["Tipo", "Cuenta", "Debe", "Haber", "Saldo"])
    _aplicar_encabezado(sheet, 3)
    total_activos = Decimal("0.00")
    total_pasivos = Decimal("0.00")
    total_patrimonio = Decimal("0.00")
    for cuenta in cuentas:
        movimiento = movimientos_por_cuenta.get(cuenta.id, {})
        debe = movimiento.get("debe") or Decimal("0.00")
        haber = movimiento.get("haber") or Decimal("0.00")
        if debe == 0 and haber == 0:
            continue
        if cuenta.tipo == "activo":
            saldo = debe - haber
            total_activos += saldo
        else:
            saldo = haber - debe
            if cuenta.tipo == "pasivo":
                total_pasivos += saldo
            else:
                total_patrimonio += saldo
        sheet.append([cuenta.get_tipo_display(), f"{cuenta.codigo} - {cuenta.nombre}", debe, haber, saldo])
    total_pasivo_patrimonio = total_pasivos + total_patrimonio
    sheet.append([])
    sheet.append(["Total activos", "", "", "", total_activos])
    sheet.append(["Total pasivos", "", "", "", total_pasivos])
    sheet.append(["Total patrimonio", "", "", "", total_patrimonio])
    sheet.append(["Pasivo + patrimonio", "", "", "", total_pasivo_patrimonio])
    sheet.append(["Diferencia", "", "", "", total_activos - total_pasivo_patrimonio])
    _autoajustar_columnas(sheet)
    return _respuesta_excel(workbook, f"Balance_General_{empresa.slug}.xlsx")
